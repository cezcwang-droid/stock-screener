import re
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import backtest_engine
import json
import os
import time

# ============ 代理屏蔽（彻底禁用代理，避免VPN/系统代理干扰akshare） ============
for _k in list(os.environ.keys()):
    if 'proxy' in _k.lower():
        os.environ.pop(_k, None)
os.environ['NO_PROXY'] = '*'
os.environ['no_proxy'] = '*'

# ============ 日志控制 ============
import logging
logging.getLogger("uvicorn").setLevel(logging.WARNING)
_VERBOSE = False  # 设为 True 可恢复详细终端输出

def _patch_requests_no_proxy():
    """Force all requests.Session to disable system proxy."""
    try:
        import requests
        # Patch 1: Force trust_env=False on every Session creation
        _orig_init = requests.Session.__init__
        def _no_proxy_init(self, *args, **kwargs):
            _orig_init(self, *args, **kwargs)
            self.trust_env = False
            self.proxies = {'http': '', 'https': ''}
        requests.Session.__init__ = _no_proxy_init
        # Patch 2: merge_environment_settings returns empty proxies
        _orig_merge = requests.Session.merge_environment_settings
        def _no_proxy_merge(self, url, proxies, stream, verify, cert):
            settings = _orig_merge(self, url, {}, stream, verify, cert)
            settings['proxies'] = {}
            return settings
        requests.Session.merge_environment_settings = _no_proxy_merge
    except Exception as _e:
        pass

_patch_requests_no_proxy()

def _patch_push2_http_fallback():
    """push2 HTTPS 直连不稳定（远端主动断开），强制将 push2 请求降级为 HTTP。
    同时禁用代理，避免系统代理干扰。
    """
    try:
        import requests

        # Patch Session.send: 强制无代理 + HTTPS→HTTP 降级
        _orig_send = requests.Session.send

        def _patched_send(self, request, **kwargs):
            # 强制禁用代理
            self.trust_env = False
            self.proxies = {'http': '', 'https': ''}
            kwargs.pop('proxies', None)

            # push2 HTTPS 降级为 HTTP
            url_lower = request.url.lower()
            if 'push2.eastmoney.com' in url_lower and url_lower.startswith('https://'):
                request.url = request.url.replace('https://', 'http://', 1)

            return _orig_send(self, request, **kwargs)

        requests.Session.send = _patched_send

        # Patch requests.get/post/head 等顶层函数（它们通过 with Session() 调用）
        _orig_request = requests.Session.request

        def _patched_request(self, method, url, **kwargs):
            self.trust_env = False
            self.proxies = {'http': '', 'https': ''}
            kwargs.pop('proxies', None)
            if 'push2.eastmoney.com' in url.lower() and isinstance(url, str) and url.lower().startswith('https://'):
                url = url.replace('https://', 'http://', 1)
            return _orig_request(self, method, url, **kwargs)

        requests.Session.request = _patched_request
    except Exception:
        pass

_patch_push2_http_fallback()


# ===================================================================
#  动态推荐数量：根据市场状态自动调整每日推荐股票数量
# ===================================================================

def calculate_dynamic_recommend_count(default=10):
    """
    根据上证指数(000001)的市场状态，动态计算当日推荐股票数量。

    判断指标：
    - 20日涨跌幅：指数最近20个交易日的涨跌
    - 量能趋势：最近5日均量 vs 20日均量
    - 均线排列：5日均线是否在20日均线上方
    - 连续下跌天数：最近5日中有多少日收跌

    市场状态 → 推荐数量：
    - 强势牛市（涨>5%，多头排列，放量） → 12-15只
    - 偏强（涨>3%，多头排列） → 10-12只
    - 震荡偏强（涨>0% 或 微跌+放量） → 9只
    - 震荡（-3% ~ 0%） → 8只
    - 偏弱（-5% ~ -3%） → 5-7只
    - 熊市（跌超5%） → 3-5只
    - 极端行情（连跌5日+跌超10%） → 0只

    兜底：市场数据获取失败时返回 default (10只)
    """
    try:
        kline = get_stock_kline('000001', days=60)
        if kline is None or len(kline) < 25:
            return default

        close_col = '收盘' if '收盘' in kline.columns else 'close'
        volume_col = '成交量' if '成交量' in kline.columns else 'volume'

        closes = kline[close_col].values.astype(float)
        volumes = kline[volume_col].values.astype(float)
        n = len(closes)

        # 1. 20日涨跌幅
        chg_20d = (closes[-1] / closes[max(0, n - 21)] - 1) * 100

        # 2. 量能趋势 (5日均量 vs 20日均量)
        vol_5 = np.mean(volumes[-5:]) if n >= 5 else np.mean(volumes)
        vol_20 = np.mean(volumes[-20:]) if n >= 20 else np.mean(volumes)
        vol_ratio = vol_5 / vol_20 if vol_20 > 0 else 1.0

        # 3. 均线多头 (5日 > 20日)
        ma5 = np.mean(closes[-5:]) if n >= 5 else closes[-1]
        ma20 = np.mean(closes[-20:]) if n >= 20 else closes[-1]
        ma_bull = ma5 > ma20

        # 4. 连续下跌检测
        down_days = 0
        for i in range(max(0, n - 5), n - 1):
            if closes[i + 1] < closes[i]:
                down_days += 1

        # 极端行情: 连续5日下跌且20日跌幅>10%
        if down_days >= 5 and chg_20d < -10:
            return 0

        # 市场状态判定 → 基础推荐数量
        if chg_20d > 5 and ma_bull and vol_ratio > 1.1:
            base = 12
        elif chg_20d > 3 and ma_bull:
            base = 11
        elif chg_20d > 0 or (chg_20d > -1 and vol_ratio > 1.0):
            base = 9
        elif chg_20d >= -3:
            base = 8
        elif chg_20d >= -5:
            base = 6
        else:
            base = 4

        # 量能微调：放量+1，缩量-1
        if vol_ratio > 1.2:
            base += 1
        elif vol_ratio < 0.8:
            base -= 1

        # 反弹初期加成：均线多头但指数还在跌 → 多给推荐
        if ma_bull and chg_20d < 0:
            base += 1

        return max(0, min(15, base))

    except Exception:
        return default


# ===================================================================
import struct
import glob
from datetime import datetime, timedelta
from collections import defaultdict
import traceback
import io
import pickle
from concurrent.futures import ThreadPoolExecutor, as_completed

# ==================== 可选依赖 ====================
try:
    import akshare as ak
    AKSHARE_AVAILABLE = True
except ImportError:
    AKSHARE_AVAILABLE = False

try:
    import pytdx
    TDX_PYTORCH_AVAILABLE = True
except ImportError:
    TDX_PYTORCH_AVAILABLE = False

# 导入 TDX 统一数据提供模块
from tdx_provider import (
    tdx_available,
    fetch_all_quotes_tdx, fetch_kline_tdx, get_today_quote_single,
    fetch_sector_data_tdx, get_kline_with_today as tdx_get_kline_with_today,
    resolve_market, find_tdx_vipdoc_dir as tdx_find_vipdoc,
    get_full_name_map as tdx_get_full_name_map
)

# ==================== 共振模型：本地K线自算资金流向（无需外部付费API）====================

# ==================== 路径配置 ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WATCHLIST_JSON = os.path.join(BASE_DIR, "watchlist.json")
CACHE_DATA_JSON = os.path.join(BASE_DIR, "cache_data.json")
CACHE_DIR = os.path.join(BASE_DIR, "resonance_cache")
os.makedirs(CACHE_DIR, exist_ok=True)

# ==================== 通达信本地数据配置 ====================
TDX_COMMON_PATHS = [
    r"C:\zd_cjzq\vipdoc",
    r"C:\new_tdx\vipdoc",
    r"C:\tdx\vipdoc",
    r"D:\new_tdx\vipdoc",
    r"D:\tdx\vipdoc",
]

def find_tdx_vipdoc():
    """自动查找通达信vipdoc目录"""
    for p in TDX_COMMON_PATHS:
        if os.path.isdir(p):
            # 验证内部结构
            if os.path.isdir(os.path.join(p, "sh", "lday")) or os.path.isdir(os.path.join(p, "sz", "lday")):
                return p
    return None

TDX_VIPDOC_DIR = find_tdx_vipdoc()
TDX_AVAILABLE = TDX_VIPDOC_DIR is not None

def read_tdx_day_file(stock_code):
    """
    读取通达信本地日线数据文件
    stock_code: 6位代码 如 '000001'
    返回 DataFrame[日期,开盘,最高,最低,收盘,成交量,成交额] 或 None
    """
    if not TDX_VIPDOC_DIR:
        return None
    try:
        code = str(stock_code).zfill(6)
        # 判断沪/深/北
        if code.startswith(('6', '9')):
            filepath = os.path.join(TDX_VIPDOC_DIR, "sh", "lday", f"sh{code}.day")
        elif code.startswith(('0', '3', '2')):
            filepath = os.path.join(TDX_VIPDOC_DIR, "sz", "lday", f"sz{code}.day")
        elif code.startswith(('4', '8')):
            filepath = os.path.join(TDX_VIPDOC_DIR, "bj", "lday", f"bj{code}.day")
        else:
            return None

        if not os.path.exists(filepath):
            return None

        record_size = 32  # 每条记录32字节
        data = []
        with open(filepath, 'rb') as f:
            while True:
                raw = f.read(record_size)
                if len(raw) < record_size:
                    break
                # 通达信day文件标准格式 (32字节/条):
                # date(uint32) open(uint32) high(uint32) low(uint32) close(uint32) amount(float32) volume(uint32) reserved(uint32)
                # 价格字段存储为 实际价格×100 的整数
                date_int, open_i, high_i, low_i, close_i, amount, volume, _ = struct.unpack('<IIIIIfII', raw[:32])
                # 价格除以100还原（通达信标准存储方式）
                open_p = open_i / 100.0
                high_p = high_i / 100.0
                low_p = low_i / 100.0
                close_p = close_i / 100.0
                # 校验日期范围
                if date_int < 19900101 or date_int > 20991231:
                    continue
                # 校验价格合理性
                if close_p <= 0 or close_p > 100000:
                    continue
                data.append({
                    '日期': pd.Timestamp(str(date_int)),
                    '开盘': open_p,
                    '最高': high_p,
                    '最低': low_p,
                    '收盘': close_p,
                    '成交量': volume,
                    '成交额': amount,
                })
        if not data:
            return None
        df = pd.DataFrame(data)
        df = df.sort_values('日期').reset_index(drop=True)
        return df
    except Exception:
        return None

# ==================== 共振模型数据获取（东方财富DDE数据）====================
# DDE数据文件路径
DDE_DATA_FILE = os.path.join(BASE_DIR, "DDE_data", "Table.xlsx")

# ==================== 新浪财经资金流向备选链路 ====================
# DDE Excel 依赖东方财富 push2 API（已被 _patch_push2_fail_fast 拦截），
# AkShare 的 stock_individual_fund_flow_rank 同样走 push2 必然失败。
# 新浪财经 MoneyFlow API 走独立域名 vip.stock.finance.sina.com.cn，不受影响。

_SINA_MONEYFLOW_URL = (
    "https://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/"
    "MoneyFlow.ssl_bkzj_ssggzj?num=5000&sort=netamount&asc=0"
)

# 新浪 symbol 前缀 → 市场代码（用于去除前缀恢复纯数字代码）
_SINA_PREFIX_MAP = {"sh": "", "sz": "", "bj": ""}


@st.cache_data(ttl=1800)
def _load_dde_from_sina_cached():
    """
    新浪财经资金流向缓存层：通过 urllib.request 直接调用新浪接口，
    返回包含 symbol/name/trade/changeratio/r0_ratio/r3_ratio/ratioamount 等
    字段的 dict-list，缓存 30 分钟。
    
    使用 urllib.request 而非 requests 库，完全绕过 push2 fail-fast patch。
    """
    try:
        import urllib.request, json

        req = urllib.request.Request(_SINA_MONEYFLOW_URL, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Accept': '*/*',
        })
        r = urllib.request.urlopen(req, timeout=20)
        raw = r.read().decode('gbk')
        data = json.loads(raw)

        if not data or len(data) == 0:
            if _VERBOSE:
                print("[共振模型] 新浪资金流向接口返回空数据")
            return None

        if _VERBOSE:
            r3_ok = sum(1 for d in data if abs(float(d.get('r3_net', 0))) > 0)
            print(f"[共振模型] 新浪资金流向: {len(data)} 只, r3_net非零 {r3_ok}")

        return data

    except Exception as e:
        if _VERBOSE:
            print(f"[共振模型] 新浪资金流向接口失败: {e}")
        return None


def _load_dde_from_sina(quotes_df=None):
    """
    新浪财经资金流向备选：当 DDE Excel 不可用时，
    通过新浪 MoneyFlow API 获取资金流向数据，
    并映射为与 DDE Excel 兼容的 DataFrame 格式。

    字段映射策略（新浪 → DDE）：
      - symbol (sz000001) → 代码 (000001)
      - name → 名称
      - trade → 最新
      - changeratio (小数) → 涨幅 (%)
      - ratioamount (总净流入占比) → DDX ≈ ratioamount / 10
      - r3_ratio (大单净占比) → 特大单净比 / 大单净比（新浪不区分大单和特大单，两者填同一值）
      - DDY / DDZ / 连续 / 买卖方向 → 填 0（新浪不提供 Level-2 数据）
    """
    raw = _load_dde_from_sina_cached()
    if raw is None or len(raw) == 0:
        return None

    import pandas as pd

    # ---- 解析新浪 JSON 到 DataFrame ----
    records = []
    for item in raw:
        sym = str(item.get('symbol', '')).strip()
        # 去除 sh/sz/bj 前缀，恢复纯数字代码
        if len(sym) >= 2 and sym[:2] in _SINA_PREFIX_MAP:
            code = sym[2:]
        else:
            code = sym

        # 只保留 6 位数字代码
        if not code.isdigit() or len(code) != 6:
            continue

        try:
            trade = float(item.get('trade', 0) or 0)
            changeratio = float(item.get('changeratio', 0) or 0) * 100  # 小数→百分比
            ratioamount = float(item.get('ratioamount', 0) or 0)       # 总净流入占比(%)
            r3_ratio = float(item.get('r3_ratio', 0) or 0)            # 大单净占比(%)
        except (ValueError, TypeError):
            continue

        if trade == 0:
            continue

        records.append({
            '代码': code,
            '名称': str(item.get('name', '')).strip(),
            '最新': trade,
            '涨幅': changeratio,
            'DDX': max(-3, min(3, ratioamount * 10)),
            'DDY': 0.0,
            'DDZ': 0.0,
            '5日DDX': max(-3, min(3, ratioamount * 10)),
            '5日DDY': 0.0,
            '10日DDX': max(-3, min(3, ratioamount * 10)),
            '10日DDY': 0.0,
            '连续': 0,
            '5日内': 0,
            '10日内': 0,
            '特大买入': 0.0,
            '特大卖出': 0.0,
            '特大单净比': r3_ratio * 100 if r3_ratio > 0 else ratioamount * 100,
            '大单买入': 0.0,
            '大单卖出': 0.0,
            '大单净比': r3_ratio * 100 if r3_ratio > 0 else ratioamount * 100,
        })

    base = pd.DataFrame(records)

    # ---- 过滤无效行 ----
    skip = base['代码'].isin(['000000']) | (base['最新'] == 0)
    base = base[~skip].copy()

    # ---- 与行情数据做交集 ----
    if quotes_df is not None and len(quotes_df) > 0:
        valid_codes = set(base['代码']) & set(quotes_df['代码'])
        base = base[base['代码'].isin(valid_codes)]

    # ---- 确保列顺序与 DDE Excel 一致 ----
    needed = ['代码', '名称', '最新', '涨幅', 'DDX', 'DDY', 'DDZ',
              '5日DDX', '5日DDY', '10日DDX', '10日DDY',
              '连续', '5日内', '10日内',
              '特大买入', '特大卖出', '特大单净比',
              '大单买入', '大单卖出', '大单净比']
    base = base[needed].copy()

    if _VERBOSE:
        print(f"[共振模型] 新浪备选链路返回: {len(base)} 只股票")

    return base if len(base) > 0 else None


def _get_dde_or_fallback(quotes_df=None):
    """
    DDE 数据双层降级入口：
    1. 优先读取本地 DDE Excel 文件（东方财富 DDE 数据）
    2. Excel 不可用时自动走新浪财经资金流向备选链路
    3. 全部失败返回 None
    """
    # 第一层：DDE Excel
    dde_df = _load_dde_data()
    if dde_df is not None:
        if _VERBOSE:
            print(f"[共振模型] DDE Excel 加载成功: {len(dde_df)} 只")
        return dde_df

    # 第一层失败时诊断
    if not os.path.exists(DDE_DATA_FILE):
        if _VERBOSE:
            print(f"[共振模型] DDE Excel 文件不存在: {DDE_DATA_FILE}")
    else:
        if _VERBOSE:
            print(f"[共振模型] DDE Excel 文件存在但解析失败")

    # 第二层：新浪财经资金流向
    if _VERBOSE:
        print("[共振模型] DDE Excel 不可用，尝试新浪财经资金流向备选...")

    dde_df = _load_dde_from_sina(quotes_df)
    if dde_df is not None:
        if _VERBOSE:
            print(f"[共振模型] 新浪备选链路成功: {len(dde_df)} 只 (数据源: sina)")
        return dde_df

    # 第二层失败时诊断
    if _VERBOSE:
        print("[共振模型] 新浪资金流向接口也失败 — 可能是网络问题，可稍后重试")

    # 第三层：全部失败
    return None


def _load_dde_data():
    """加载东方财富 DDE 数据文件（Excel格式），返回 DataFrame"""
    if not os.path.exists(DDE_DATA_FILE):
        if _VERBOSE: print(f"[共振模型] DDE数据文件不存在: {DDE_DATA_FILE}")
        return None
    
    try:
        import pandas as pd
        
        df = pd.read_excel(DDE_DATA_FILE)
        
        # 列名映射：东方财富Excel列名 → 标准列名
        # 实际列: 序/代码/名称/最新/涨幅%/当日资金流(DDX)/(DDY)/(DDZ)/5日资金流(5日DDX)/(5日DDY)/10日资金流(10日DDX)/(10日DDY)/DDX飘红天数(连续)/(5日内)/(10日内)/特大买入%/特大卖出%/特大单净比%/大单买入%/大单卖出%/大单净比%
        col_rename = {}
        for c in df.columns:
            c_str = str(c).strip()
            if c_str == '序':
                col_rename[c] = '序号'
            elif c_str == '代码':
                col_rename[c] = '代码'
            elif c_str == '名称':
                col_rename[c] = '名称'
            elif c_str == '最新':
                col_rename[c] = '最新'
            elif c_str == '涨幅%':
                col_rename[c] = '涨幅'
            elif c_str == '当日资金流':
                col_rename[c] = 'DDX'
            elif c_str == '5日资金流':
                col_rename[c] = '5日DDX'
            elif c_str == '10日资金流':
                col_rename[c] = '10日DDX'
            elif c_str == 'DDX飘红天数':
                col_rename[c] = '连续'
            elif c_str == '特大买入%':
                col_rename[c] = '特大买入'
            elif c_str == '特大卖出%':
                col_rename[c] = '特大卖出'
            elif c_str == '特大单净比%':
                col_rename[c] = '特大单净比'
            elif c_str == '大单买入%':
                col_rename[c] = '大单买入'
            elif c_str == '大单卖出%':
                col_rename[c] = '大单卖出'
            elif c_str == '大单净比%':
                col_rename[c] = '大单净比'
            elif 'Unnamed' in c_str:
                # Unnamed 列为 Excel 二级表头（DDY/DDZ等），列序号对应实际含义
                col_idx = df.columns.get_loc(c)
                unamed_map = {6: 'DDY', 7: 'DDZ', 9: '5日DDY', 11: '10日DDY', 13: '5日内', 14: '10日内'}
                if col_idx in unamed_map:
                    col_rename[c] = unamed_map[col_idx]
        
        df = df.rename(columns=col_rename)
        
        # 过滤掉嵌入的子标题行（第一行含"DDX"字符串的非数值行）
        if len(df) > 0:
            first_code = str(df.iloc[0].get('代码', '')).strip()
            if first_code == 'nan' or first_code == '' or first_code == 'DDX':
                df = df.iloc[1:].copy()
        
        # 需要的标准列
        needed = ['代码', '名称', '最新', '涨幅', 'DDX', 'DDY', 'DDZ',
                  '5日DDX', '5日DDY', '10日DDX', '10日DDY',
                  '连续', '5日内', '10日内',
                  '特大买入', '特大卖出', '特大单净比',
                  '大单买入', '大单卖出', '大单净比']
        
        for col in needed:
            if col not in df.columns:
                if _VERBOSE: print(f"[共振模型] 缺失列: {col}，实际列: {df.columns.tolist()}")
                return None
        
        df = df[needed].copy()
        
        # 转换数值列
        numeric_cols = ['最新', '涨幅', 'DDX', 'DDY', 'DDZ',
                        '5日DDX', '5日DDY', '10日DDX', '10日DDY',
                        '连续', '5日内', '10日内',
                        '特大买入', '特大卖出', '特大单净比',
                        '大单买入', '大单卖出', '大单净比']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # 代码格式化：去除小数点，左补零到6位
        df['代码'] = df['代码'].apply(lambda x: str(int(float(x))).zfill(6) if str(x).replace('.','').replace('-','').isdigit() else str(x).zfill(6))
        
        skip = df['代码'].isin(['000000']) | (df['最新'] == 0)
        df = df[~skip]
        
        if _VERBOSE: print(f"[共振模型] 加载DDE数据: {len(df)} 只股票 (Excel解析成功)")
        return df
    except Exception as e:
        if _VERBOSE: print(f"[共振模型] DDE数据加载失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_resonance_data(quotes_df):
    """
    基于东方财富DDE数据计算共振模型四维评分：
    1. 资金流向 (30分)：特大单净比 + 大单净比
    2. DDE决策   (20分)：DDX + DDY + DDZ + 连续天数
    3. K线结构   (25分)：由 calculate_resonance_score 从行情数据补充
    4. 板块热度   (25分)：由 calculate_resonance_score 从板块数据补充

    quotes_df: fetch_all_a_stocks() 返回的行情 DataFrame
    """
    dde_df = _get_dde_or_fallback(quotes_df)
    if dde_df is None:
        return None
    
    if quotes_df is None or len(quotes_df) == 0:
        return None
    
    # 过滤 ST/退市
    quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)].copy()
    
    # 交集：DDE数据 ∩ 行情数据
    valid_codes = set(dde_df['代码']) & set(quotes_df['代码'])
    if _VERBOSE: print(f"[共振模型] DDE+行情交集: {len(valid_codes)} 只")
    
    if len(valid_codes) == 0:
        return None
    
    # 用 DDX 全局排名归一化（DDX 是 DDE 核心指标）
    dde_sorted = dde_df.set_index('代码')
    
    # 分数归一化辅助函数
    def normalize_rank(series, max_score, center=0.5):
        """按百分位排名映射到 [0, max_score]"""
        ranked = series.rank(pct=True)
        return (ranked * max_score).clip(0, max_score)
    
    # ----- 1. 资金流向 (30分) -----
    # 特大单净比 + 大单净比 = 主力资金净流向
    # P0: 持续性过滤 — 特大单净比>0 且大单净比>0 同时满足才给满分
    dde_sorted['资金流向_raw'] = (dde_sorted['特大单净比'] + dde_sorted['大单净比']).fillna(0)
    dde_sorted['资金流向'] = normalize_rank(dde_sorted['资金流向_raw'], 30)
    # 持续性标记：双条件同时满足
    dde_sorted['_both_positive'] = ((dde_sorted['特大单净比'] > 0) & (dde_sorted['大单净比'] > 0)).astype(float)
    dde_sorted['_one_positive'] = ((dde_sorted['特大单净比'] > 0) | (dde_sorted['大单净比'] > 0)).astype(float)
    # 只有一项满足给一半分（相对于满分30的比例）
    dde_sorted['资金流向_persist'] = dde_sorted['资金流向'] * (
        dde_sorted['_both_positive'] * 1.0 + (1 - dde_sorted['_both_positive']) * dde_sorted['_one_positive'] * 0.5
    )
    
    # ----- 2. DDE决策 (20分) -----
    # P0: DDX 平滑 — 用近3日 DDX 均值替代单日 DDX（用 5日DDX/5 近似3日均值）
    dde_sorted['DDX_3d_mean'] = dde_sorted['5日DDX'].fillna(dde_sorted['DDX']) / 5
    # DDX (权重0.4) + DDY (0.2) + DDZ (0.2) + 连续天数加成 (0.2)
    dde_sorted['DDX_norm'] = normalize_rank(dde_sorted['DDX_3d_mean'].fillna(0), 8)
    dde_sorted['DDY_norm'] = normalize_rank(dde_sorted['DDY'].fillna(0), 4)
    dde_sorted['DDZ_norm'] = normalize_rank(dde_sorted['DDZ'].fillna(0), 4)
    
    # 连续天数：正值 DDX 时连续天数越多越好，负值 DDX 时连续越少越好
    dde_sorted['连续加成'] = dde_sorted.apply(
        lambda r: min(4, max(0, r['连续'] * 0.5)) if r.get('DDX', 0) >= 0 else max(0, 4 - r['连续'] * 0.5), axis=1
    )
    dde_sorted['DDE决策'] = (
        dde_sorted['DDX_norm'] + dde_sorted['DDY_norm'] + dde_sorted['DDZ_norm'] + dde_sorted['连续加成']
    ).clip(0, 20)
    
    # ----- 构建结果 -----
    # P0: 资金流向改用持续性过滤后的分数
    result = {}
    for code in valid_codes:
        row = dde_sorted.loc[code]
        result[code] = {
            'money_flow_score': round(float(row['资金流向_persist']), 1),
            'dde_proxy_score': round(float(row['DDE决策']), 1),
            'kline_structure_raw': 0,      # K线原始分，由 calculate_resonance_score 补充
            'sector_heat_raw': 5,           # 板块原始分，由 calculate_resonance_score 补充
            'kline_structure_score': 0,     # 归一化后
            'sector_heat_score': 5,         # 归一化后
        }
    
    if _VERBOSE: print(f"[共振模型] 生成评分数据: {len(result)} 只")
    return result if result else None


def calculate_resonance_score(resonance_data, quotes_df=None):
    """
    计算四维共振综合评分（满分100）
    资金流向 30分 + DDE决策 20分 + K线结构 25分 + 板块热度 25分
    P1: K线结构和板块热度使用百分位归一化
    """
    if resonance_data is None:
        return {}
    
    sector_data = _get_cached_sector_data()
    codes = list(resonance_data.keys())
    
    # ===== 第一遍：计算所有股票的 K线和板块 原始分 =====
    kl_raw = {}
    sh_raw = {}
    for code in codes:
        rd = resonance_data[code]
        
        # ----- K线结构原始分 -----
        kl_score = 10.0
        if quotes_df is not None:
            row = quotes_df[quotes_df['代码'] == code]
            if len(row) > 0:
                r = row.iloc[0]
                chg5 = float(r.get('5日涨幅', 0) or 0)
                chg5_score = min(15, max(0, 7.5 + chg5 * 0.75))
                vol_ratio_q = float(r.get('量比', 1) or 1)
                vol_score_k = min(10, max(0, vol_ratio_q * 5))
                kl_score = chg5_score + vol_score_k
        kl_raw[code] = kl_score
        
        # ----- 板块热度原始分 -----
        sh_score = 10.0
        sector_found = False
        if quotes_df is not None and sector_data:
            row = quotes_df[quotes_df['代码'] == code]
            if len(row) > 0:
                sector = str(row.iloc[0].get('板块', ''))
                if sector and sector in sector_data:
                    sd = sector_data[sector]
                    sec_chg = float(sd.get('涨幅', 0) or 0)
                    sec_chg_score = min(15, max(0, 7.5 + sec_chg * 0.5))
                    zt_count = int(sd.get('涨停数', 0) or 0)
                    zt_score = min(10, zt_count * 1.5)
                    sh_score = sec_chg_score + zt_score
                    sector_found = True

        # 🔧 P0修复: 板块数据不可用时，用个股自身表现替代（避免全部12.5无区分度）
        if not sector_found and quotes_df is not None:
            row = quotes_df[quotes_df['代码'] == code]
            if len(row) > 0:
                r = row.iloc[0]
                chg5 = float(r.get('5日涨幅', 0) or 0)
                chg20 = float(r.get('涨跌幅_20d', 0) or 0)
                vol_ratio = float(r.get('量比', 1) or 1)
                # 用个股5日+20日涨幅和量比近似估计板块热度
                proxy_chg = (chg5 * 0.6 + chg20 * 0.4)  # 加权涨跌幅
                proxy_score = min(15, max(0, 7.5 + proxy_chg * 0.5))
                proxy_vol = min(10, max(0, vol_ratio * 3))
                sh_score = proxy_score + proxy_vol
        sh_raw[code] = sh_score
    
    # ===== P1: 百分位归一化（K线和板块） =====
    kl_arr = np.array([kl_raw[c] for c in codes])
    sh_arr = np.array([sh_raw[c] for c in codes])
    
    # 纯 numpy 实现 rank → percentile → 0~max_score
    def _percentile_norm(raw_arr, max_score):
        if len(raw_arr) <= 1 or np.all(raw_arr == raw_arr[0]):
            return np.full_like(raw_arr, max_score * 0.5, dtype=float)
        ranks = np.argsort(np.argsort(raw_arr)) + 1  # 最小=1, 最大=N
        percentiles = ranks / len(raw_arr)
        return (percentiles * max_score).clip(0, max_score)
    
    kl_norm = _percentile_norm(kl_arr, 25)
    sh_norm = _percentile_norm(sh_arr, 25)
    
    # ===== 组装最终得分 =====
    scores = {}
    for i, code in enumerate(codes):
        rd = resonance_data[code]
        mf_score = rd.get('money_flow_score', 0)
        dde_score = rd.get('dde_proxy_score', 0)
        kl_final = round(float(kl_norm[i]), 1)
        sh_final = round(float(sh_norm[i]), 1)
        
        rd['kline_structure_raw'] = round(kl_raw[code], 1)
        rd['sector_heat_raw'] = round(sh_raw[code], 1)
        rd['kline_structure_score'] = kl_final
        rd['sector_heat_score'] = sh_final
        
        total = mf_score + dde_score + kl_final + sh_final
        scores[code] = {
            'total': round(total, 1),
            'money_flow': round(mf_score, 1),
            'dde_proxy': round(dde_score, 1),
            'kline_structure': kl_final,
            'sector_heat': sh_final,
        }
    
    return scores


def _classify_resonance_style(code, resonance_data, quotes_df):
    """
    轻量分类：基于行情快照+DDE数据判断共振标的当前状态
    - 盘中走强：涨跌幅 ≥ 2% 且 量比 ≥ 1.0 且 DDE决策 ≥ 10 → 今日已启动
    - 超跌待反转：涨跌幅 ≤ 0% 且 换手率 ≤ 10% 且 DDE决策 ≤ 10 → 回调末端
    - 蓄势待发：不满足以上任一 → 未启动，等待催化
    注意：标签描述当前状态，综合评分描述未来潜力，两者独立。
    """
    if quotes_df is None:
        return '蓄势待发'
    
    row = quotes_df[quotes_df['代码'] == code]
    if len(row) == 0:
        return '蓄势待发'
    
    r = row.iloc[0]
    chg = float(r.get('涨跌幅', 0) or 0)
    vol_ratio = float(r.get('量比', 1) or 1)
    turnover = float(r.get('换手率', 5) or 5)
    
    rd = resonance_data.get(code, {}) if resonance_data else {}
    dde_score = float(rd.get('dde_proxy_score', 10))
    
    if chg >= 2 and vol_ratio >= 1.0 and dde_score >= 10:
        return '盘中走强'
    
    if chg <= 0 and turnover <= 10 and dde_score <= 10:
        return '超跌待反转'
    
    return '蓄势待发'


@st.cache_data(ttl=1800, show_spinner=False)
def get_dde_confirmation_scores():
    """全量DDE校正分：code -> 资金流向(30)+DDE决策(20) = 满分50
    
    独立评分函数，不依赖共振模型的其他部分。
    对全部DDE覆盖股票进行 normalize_rank 排名赋予校正分，
    供追高/低吸模型的Top30表格插入 DDE资金确认列。
    """
    dde_df = _load_dde_data()
    if dde_df is None or len(dde_df) == 0:
        return {}
    
    dde_sorted = dde_df.set_index('代码')
    
    def _norm_rank(series, max_score):
        ranked = series.rank(pct=True)
        return (ranked * max_score).clip(0, max_score)
    
    dde_sorted['资金流向_raw'] = (dde_sorted['特大单净比'] + dde_sorted['大单净比']).fillna(0)
    dde_sorted['资金流向'] = _norm_rank(dde_sorted['资金流向_raw'], 30)
    
    dde_sorted['DDX_norm'] = _norm_rank(dde_sorted['DDX'].fillna(0), 8)
    dde_sorted['DDY_norm'] = _norm_rank(dde_sorted['DDY'].fillna(0), 4)
    dde_sorted['DDZ_norm'] = _norm_rank(dde_sorted['DDZ'].fillna(0), 4)
    
    dde_sorted['连续加成'] = dde_sorted.apply(
        lambda r: min(4, max(0, r['连续'] * 0.5)) if r.get('DDX', 0) >= 0
        else max(0, 4 - r['连续'] * 0.5), axis=1
    )
    dde_sorted['DDE决策'] = (dde_sorted['DDX_norm'] + dde_sorted['DDY_norm']
                             + dde_sorted['DDZ_norm'] + dde_sorted['连续加成']).clip(0, 20)
    
    dde_sorted['校正分'] = (dde_sorted['资金流向'] + dde_sorted['DDE决策']).round(1)
    
    result = {}
    for code in dde_sorted.index:
        result[code] = float(dde_sorted.loc[code, '校正分'])
    return result


st.set_page_config(
    page_title="智能选股系统",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== 自定义CSS样式 ====================
st.markdown("""
<style>
    .main { font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; }
    
    /* ===== 标题区 ===== */
    .header-container {
        background: linear-gradient(135deg, #FFF8F0 0%, #FFFFFF 100%);
        border-radius: 12px; padding: 20px 24px;
        margin-bottom: 16px; border: 1px solid #F0E6D8;
        box-shadow: 0 2px 12px rgba(196,132,45,0.08);
    }
    .main-title { font-size: 24px; font-weight: 700; color: #C4842D; margin-bottom: 8px; display: flex; align-items: center; gap: 10px; }
    .sub-title { font-size: 13px; color: #888; }

    /* ===== 数据状态提示 ===== */
    .data-status {
        background: linear-gradient(135deg, #E8F5E9 0%, #C8E6C9 100%);
        border-radius: 8px; padding: 10px 16px; margin-bottom: 16px;
        border: 1px solid #A5D6A7; display: flex; align-items: center; gap: 10px;
    }
    .data-status.warning {
        background: linear-gradient(135deg, #FFF3E0 0%, #FFE0B2 100%);
        border-color: #FFB74D;
    }

    /* ===== 精选区域 ===== */
    .top10-container {
        background: linear-gradient(135deg, #FFFBF5 0%, #FFF5E6 100%);
        border-radius: 16px; padding: 20px; margin-bottom: 20px;
        border: 2px solid #E8A838; box-shadow: 0 4px 20px rgba(196,132,45,0.15);
    }
    .top10-header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px; }
    .top10-title { font-size: 18px; font-weight: 700; color: #C4842D; display: flex; align-items: center; gap: 8px; }
    .top10-badge { background: linear-gradient(135deg, #FF6B35 0%, #F7931E 100%); color: white; padding: 4px 12px; border-radius: 20px; font-size: 11px; font-weight: 600; }
    .top10-card { background: white; border-radius: 12px; padding: 14px 16px; border: 1px solid #F0E6D8; transition: all 0.3s ease; cursor: pointer; position: relative; overflow: hidden; }
    .top10-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(196,132,45,0.2); border-color: #C4842D; }
    .top10-rank { position: absolute; top: 8px; right: 10px; font-size: 28px; font-weight: 800; color: #F0E6D8; line-height: 1; }
    .top10-rank.gold { color: #FFD700; } .top10-rank.silver { color: #C0C0C0; } .top10-rank.bronze { color: #CD7F32; }

    /* ===== 低吸卡片样式 ===== */
    .lowbuy-container {
        background: linear-gradient(135deg, #F0FFF4 0%, #E8F5E9 100%);
        border-radius: 16px; padding: 20px; margin-bottom: 20px;
        border: 2px solid #43A047; box-shadow: 0 4px 20px rgba(67,160,71,0.15);
    }
    .lowbuy-title { font-size: 18px; font-weight: 700; color: #2E7D32; display: flex; align-items: center; gap: 8px; }
    .lowbuy-badge { background: linear-gradient(135deg, #43A047 0%, #66BB6A 100%); color: white; padding: 4px 12px; border-radius: 20px; font-size: 11px; font-weight: 600; }
    .lowbuy-card { background: white; border-radius: 12px; padding: 14px 16px; border: 1px solid #C8E6C9; transition: all 0.3s ease; cursor: pointer; position: relative; overflow: hidden; }
    .lowbuy-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(67,160,71,0.2); border-color: #43A047; }

    /* ===== 统计卡片 ===== */
    .stat-card { background: #FAFAF8; border-radius: 10px; padding: 16px 20px; border: 1px solid #EDE8E0; text-align: center; transition: all 0.2s ease; }
    .stat-card:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.06); transform: translateY(-1px); }
    .stat-value { font-size: 28px; font-weight: 700; color: #333; }
    .stat-label { font-size: 12px; color: #999; margin-top: 4px; }
    .stat-value.red { color: #E74C3C; } .stat-value.green { color: #27AE60; } .stat-value.orange { color: #E67E22; }

    /* ===== 指标标签 ===== */
    .metric-badge { display: inline-block; padding: 2px 10px; border-radius: 10px; font-size: 12px; font-weight: 500; }
    .badge-buy { background: #FDECEA; color: #E74C3C; border: 1px solid #F5C6CB; }
    .badge-watch { background: #FEF5E7; color: #E67E22; border: 1px solid #F8D7DA; }
    .badge-hold { background: #E8F6EF; color: #27AE60; border: 1px solid #C3E6CB; }
    .badge-strong { background: linear-gradient(135deg, #FF6B35 0%, #E74C3C 100%); color: white; border: none; }
    .badge-attention { background: linear-gradient(135deg, #F7931E 0%, #E67E22 100%); color: white; border: none; }
    .badge-lowbuy-strong { background: linear-gradient(135deg, #43A047 0%, #2E7D32 100%); color: white; border: none; }
    .badge-lowbuy-mild { background: linear-gradient(135deg, #66BB6A 0%, #81C784 100%); color: white; border: none; }

    /* ===== 股票行悬停 ===== */
    .stock-row:hover { background-color: #FFF9F2 !important; }

    /* ===== 评分进度条 ===== */
    .score-bar { height: 6px; border-radius: 3px; background: #EEE; overflow: hidden; }
    .score-fill { height: 100%; border-radius: 3px; transition: width 0.3s ease; }

    /* ===== 详情指标卡片 ===== */
    .detail-metric-card { background: #FBFBF9; border-radius: 8px; padding: 14px 18px; border: 1px solid #EEEDE8; }
    .metric-name { font-size: 12px; color: #999; margin-bottom: 4px; }
    .metric-value { font-size: 20px; font-weight: 700; }
    .metric-value.up { color: #E74C3C; } .metric-value.down { color: #27AE60; }

    /* ===== 建议框 ===== */
    .advice-box { background: linear-gradient(135deg, #FFF8E1 0%, #FFFBF0 100%); border-left: 4px solid #FFA000; border-radius: 0 8px 8px 0; padding: 16px 20px; margin: 12px 0; }
    .advice-box.green { background: linear-gradient(135deg, #E8F5E9 0%, #F1F8E9 100%); border-left-color: #43A047; }

    /* ===== 回测卡片 ===== */
    .bt-stat-card { background: linear-gradient(145deg, #FFFFFF 0%, #FAFAF8 100%); border-radius: 10px; padding: 16px; border: 1px solid #E8E4DC; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.04); transition: all 0.2s ease; }
    .bt-stat-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,0.08); }
    .bt-label { font-size: 11px; color: #AAA; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 4px; }
    .bt-val { font-size: 24px; font-weight: 800; }
    .bt-sub { font-size: 11px; color:#888; margin-top:2px; min-height:16px; }
    .bt-val.positive { color: #E74C3C; } .bt-val.negative { color: #27AE60; } .bt-val.neutral { color: #666; }

    /* ===== 权重表格 ===== */
    .weight-table { width:100%; border-collapse:collapse; font-size:13px; }
    .weight-table th { background:#FFF8F0; color:#555; padding:10px 14px; text-align:left; border-bottom:2px solid #E8DCC8; font-weight:600;}
    .weight-table td { padding:9px 14px; border-bottom:1px solid #F0EBE4; color:#444; }
    .weight-table tr:hover td { background:#FFFAF5; }
    .range-badge { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; background:#F5F0EA; color:#997A50;}

    /* ===== 按钮样式 ===== */
    .stButton>button { border-radius: 8px !important; font-weight: 500 !important; transition: all 0.2s !important; }
    .stButton>button:hover { transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
    
    /* ===== 隐藏元素 ===== */
    #MainMenu { visibility: hidden; } footer { visibility: hidden; } header { visibility: visible; }
    
    /* ===== Tab样式 ===== */
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { border-radius: 8px 8px 0 0; padding: 10px 24px; background: #F5F5F0; border: 1px solid #E0DED8; }
    .stTabs [aria-selected="true"] { background: #FFFFFF !important; border-bottom-color: #FFFFFF !important; color: #C4842D !important; font-weight: 700; }

    /* ===== 表格交替色 ===== */
    .dataframe tbody tr:nth-child(even) { background-color: #FFFBF5 !important; }
    .dataframe tbody tr:hover { background-color: #FFF5E6 !important; }

    /* ===== 参数面板 ===== */
    .param-panel { background: #FFFDF8; border: 1px solid #F0E6D8; border-radius: 12px; padding: 16px 20px; margin-bottom: 16px; }

    /* ===== 分隔线 ===== */
    .section-divider { height: 1px; background: linear-gradient(90deg, transparent, #E8DCC8, transparent); margin: 20px 0; }

    /* ===== 加载动画 ===== */
    .loading-spinner { display: inline-block; width: 20px; height: 20px; border: 3px solid rgba(196,132,45,0.3); border-radius: 50%; border-top-color: #C4842D; animation: spin 1s ease-in-out infinite; }
    @keyframes spin { to { transform: rotate(360deg); } }
</style>
""", unsafe_allow_html=True)

# ==================== v3八维权重配置（追高模型） ====================
WEIGHT_CONFIG = {
    "趋势结构": {"desc": "MA5/10/20/60多头排列", "default": 15, "max": 35, "icon": "📐", "color": "#1A73E8", "full_score": 25},
    "动量强度": {"desc": "5日+10日涨幅",        "default": 18, "max": 35, "icon": "🚀", "color": "#E65100", "full_score": 22},
    "板块共振": {"desc": "板块涨幅/涨停/资金流",  "default": 8, "max": 25, "icon": "🌐", "color": "#00897B", "full_score": 10},
    "北向资金": {"desc": "北向资金近3日净买",     "default": 15, "max": 25, "icon": "👆", "color": "#1565C0", "full_score": 10},
    "机构净买": {"desc": "机构3日净买额",         "default": 10, "max": 25, "icon": "🏦", "color": "#C62828", "full_score": 10},
    "板块资金热度": {"desc": "板块资金排名映射",   "default": 5, "max": 15, "icon": "🔥", "color": "#E65100", "full_score": 10},
    "量价配合": {"desc": "量比/振幅/缩量新高",    "default": 14, "max": 30, "icon": "📈", "color": "#F57C00", "full_score": 12},
    "估值安全": {"desc": "PE历史分位(赛道差异化)","default": 3,  "max": 25, "icon": "🛡️", "color": "#7B1FA2", "full_score": 10},
    "筹码稳定": {"desc": "换手率接力",            "default": 6,  "max": 20, "icon": "🔒", "color": "#2E7D32", "full_score": 7},
    "情绪热度": {"desc": "热度分(主线/冷门差异化)","default": 6,  "max": 15, "icon": "🎯", "color": "#AD1457", "full_score": 4},
}
DEFAULT_WEIGHTS = {k: v["default"] for k, v in WEIGHT_CONFIG.items()}

# 赛道分类
GROWTH_SECTOR = {"半导体", "AI算力", "光模块", "EDA", "存储芯片", "云计算",
                 "半导体/芯片", "AI/人工智能", "软件/云计算", "芯片", "算力", "数据"}
CYCLE_SECTOR = {"锂电设备", "物流", "影视传媒", "电力设备", "CXO医药",
                "新能源/锂电", "化工/材料", "机械设备"}

# 仓位分层
SCORE_TIERS = [
    (85, 0.7, "💰 核心龙头·重仓追高", "主线龙头，重仓60%~80%"),
    (70, 0.3, "📈 支线趋势·轻仓试错", "支线趋势，轻仓20%~40%"),
    (0,  0.0, "⚠️ 不建议参与",       "分数不足70，放弃不参与"),
]

# ===================== 低吸模型专用参数 =====================
DEFAULT_LOWBUY_PARAMS = {
    "_params_version": 5,     # 版本号：默认值变动时递增，触发 session_state 自动刷新
    "max_results": 20,
    "pre_filter_decline": 5,
    "decline_20d_low": -40,
    "decline_20d_high": 0,
    "max_vol_ratio": 2.5,
    "no_new_low_days": 2,
    "reversal_bottom_pct": 0.2,     # 近3日均线高于15日最低点的最小百分比
    "reversal_require_uptrend": False,  # MA3今日 > MA3前日（关闭以放宽反转确认）
    "min_decline_depth": 5,
    "min_stabilization": 8,
    "min_volume_recovery": 5,
    "min_ma_support": 4,
    "min_valuation_attr": 3,
    "min_chip_settle": 3,
    "min_fund_flow": 2,       # P0: raw 0-10, 主力资金或量能替代门槛
    "min_total_score": 35,    # 0-100, 综合最低分
    "fund_weight": 0.08,      # P0: 主力资金维度权重（默认 8%）
}


def _get_lowbuy_params():
    """获取低吸参数，自动检测代码默认值更新并刷新 session_state。
    
    当 DEFAULT_LOWBUY_PARAMS._params_version 递增时，
    session_state 中缓存的旧参数会被自动替换为新默认值。
    """
    current_version = DEFAULT_LOWBUY_PARAMS.get("_params_version", 0)
    cached = st.session_state.get('lowbuy_params', None)
    if cached is None:
        st.session_state.lowbuy_params = dict(DEFAULT_LOWBUY_PARAMS)
        return dict(DEFAULT_LOWBUY_PARAMS)
    cached_version = cached.get("_params_version", 0)
    if cached_version < current_version:
        # 代码默认值已更新，完全替换为新默认值
        st.session_state.lowbuy_params = dict(DEFAULT_LOWBUY_PARAMS)
        return dict(DEFAULT_LOWBUY_PARAMS)
    return cached

LOWBUY_WEIGHT_CONFIG = {
    "下跌幅度": {"default": 23, "max": 40, "color": "#E74C3C"},
    "企稳信号": {"default": 18, "max": 35, "color": "#F39C12"},
    "量能恢复": {"default": 14, "max": 30, "color": "#3498DB"},
    "均线支撑": {"default": 14, "max": 30, "color": "#2ECC71"},
    "估值吸引": {"default": 14, "max": 30, "color": "#9B59B6"},
    "筹码沉淀": {"default":  9, "max": 25, "color": "#1ABC9C"},
    "主力资金": {"default":  8, "max": 20, "color": "#E67E22"},
}
DEFAULT_LOWBUY_WEIGHTS = {k: v["default"] for k, v in LOWBUY_WEIGHT_CONFIG.items()}

# 超跌反弹模型权重配置
ORB_WEIGHT_CONFIG = {
    "空间维度": {"default": 40, "max": 60, "color": "#E74C3C"},
    "情绪量能": {"default": 30, "max": 50, "color": "#3498DB"},
    "择时确认": {"default": 30, "max": 50, "color": "#F39C12"},
    "板块加成": {"default": 10, "max": 20, "color": "#27AE60"},
}
DEFAULT_ORB_WEIGHTS = {k: v["default"] for k, v in ORB_WEIGHT_CONFIG.items()}
DEFAULT_ORB_PARAMS = {
    "min_drawdown_pct": 30,
    "min_decline_20d": 15,
    "min_price": 3.0,
    "vol_shrink_threshold": 0.5,
}

_app_cache = {
    'raw_market_data': None,
    'raw_market_time': None,
    'dragon_tiger': None,
    'dragon_tiger_time': None,
    'sector_data': None,
    'sector_map': None,
    'sector_time': None,
    'sector_fund_flow': None,
    'sector_fund_flow_time': None,
    'stock_pool': None,
    'stock_pool_time': None,
    'stock_pool_key': None,
}

# 行业PE统计缓存（低吸模型S5维度用）
_industry_pe_cache = {"stats": None, "timestamp": 0}

def _get_cached_dragon_tiger():
    """获取龙虎榜数据（模块级缓存，1小时有效）"""
    now = time.time()
    if _app_cache['dragon_tiger'] is not None and _app_cache['dragon_tiger_time'] and (now - _app_cache['dragon_tiger_time']) < 3600:
        return _app_cache['dragon_tiger']
    data = fetch_dragon_tiger_v3()
    _app_cache['dragon_tiger'] = data
    _app_cache['dragon_tiger_time'] = now
    return data

def _get_cached_sector_data():
    """获取板块数据（模块级缓存，30分钟有效）"""
    now = time.time()
    if _app_cache['sector_data'] is not None and _app_cache['sector_time'] and (now - _app_cache['sector_time']) < 1800:
        return _app_cache['sector_data'], _app_cache['sector_map']
    data, smap = fetch_sector_board_v3()
    _app_cache['sector_data'] = data
    _app_cache['sector_map'] = smap
    _app_cache['sector_time'] = now
    return data, smap


def _get_sector_fund_flow():
    """计算板块资金流向（追高模型前置门槛用）。
    聚合 DDE 数据按板块平均净资金流比例（特大单净比+大单净比）。
    返回: {sector_name: avg_flow_ratio}，正值=净流入，负值=净流出。
    缓存30分钟。
    """
    now = time.time()
    if _app_cache['sector_fund_flow'] is not None and _app_cache['sector_fund_flow_time'] \
            and (now - _app_cache['sector_fund_flow_time']) < 1800:
        return _app_cache['sector_fund_flow']

    dde_df = _load_dde_data()
    if dde_df is None or len(dde_df) == 0:
        return {}

    _, stock_sector_map = _get_cached_sector_data()
    if not stock_sector_map:
        return {}

    # 计算每只股票的净资金流比例
    dde_df['特大单净比'] = pd.to_numeric(dde_df['特大单净比'], errors='coerce').fillna(0)
    dde_df['大单净比'] = pd.to_numeric(dde_df['大单净比'], errors='coerce').fillna(0)
    dde_df['5日DDX'] = pd.to_numeric(dde_df['5日DDX'], errors='coerce').fillna(0)
    dde_df['fund_flow_ratio'] = dde_df['特大单净比'] + dde_df['大单净比']
    # 主力净比为0时，用5日DDX作为替代
    mask_zero = dde_df['fund_flow_ratio'] == 0
    dde_df.loc[mask_zero, 'fund_flow_ratio'] = dde_df.loc[mask_zero, '5日DDX']

    # 按板块聚合
    sector_flows = defaultdict(list)
    for _, row in dde_df.iterrows():
        code = str(row['代码'])
        sector = stock_sector_map.get(code)
        if sector:
            sector_flows[sector].append(row['fund_flow_ratio'])

    # 板块平均资金流向
    result = {}
    for sector, flows in sector_flows.items():
        if flows:
            result[sector] = sum(flows) / len(flows)

    _app_cache['sector_fund_flow'] = result
    _app_cache['sector_fund_flow_time'] = now
    if _VERBOSE:
        inflow_count = sum(1 for v in result.values() if v > 0)
        outflow_count = sum(1 for v in result.values() if v <= 0)
        print(f"[板块资金] 净流入板块: {inflow_count}, 净流出板块: {outflow_count}, 总计: {len(result)}")
    return result


def init_session_state():
    """初始化会话状态"""
    defaults = dict(DEFAULT_WEIGHTS)
    for key in ['weights', 'watchlist', 'selected_stock', 'current_page',
                'backtest_result', 'bt_params', 'top10_cache', 'last_update_time',
                'data_status', 'raw_stock_data', 'current_model', 'lowbuy_params',
                'lowbuy_cache', 'lowbuy_auto_scanned', 'top10_cache_key']:
        if key not in st.session_state:
            if key == 'watchlist':
                st.session_state[key] = load_watchlist()
            elif key == 'weights':
                st.session_state[key] = defaults
            elif key == 'current_page':
                st.session_state[key] = 'screener'
            elif key == 'current_model':
                st.session_state[key] = 'chase_high'
            elif key == 'lowbuy_params':
                st.session_state[key] = dict(DEFAULT_LOWBUY_PARAMS)
            elif key == 'lowbuy_weights':
                st.session_state[key] = dict(DEFAULT_LOWBUY_WEIGHTS)
            elif key in ('top10_cache', 'lowbuy_cache', 'last_update_time'):
                st.session_state[key] = None
            elif key == 'data_status':
                st.session_state[key] = 'normal'
            elif key == 'lowbuy_auto_scanned':
                st.session_state[key] = False
            else:
                st.session_state[key] = None
    
    # 超跌反弹模型参数初始化
    if 'orb_params' not in st.session_state:
        st.session_state.orb_params = dict(DEFAULT_ORB_PARAMS)
    if 'orb_weights' not in st.session_state:
        st.session_state.orb_weights = dict(DEFAULT_ORB_WEIGHTS)
    if 'orb_results' not in st.session_state:
        st.session_state.orb_results = None
    if 'oversold_rebound_auto_scanned' not in st.session_state:
        st.session_state.oversold_rebound_auto_scanned = False

    # 从文件加载缓存（启动时恢复上次选股结果）
    if 'cache_loaded' not in st.session_state:
        cache_data = load_cache_data()
        if cache_data and is_cache_today(cache_data):
            if cache_data.get('chase_high_top10'):
                st.session_state.top10_cache = cache_data['chase_high_top10']
                st.session_state.top10_cache_key = f"top10_{datetime.now().strftime('%Y%m%d')}"
            if cache_data.get('lowbuy_top5'):
                st.session_state.lowbuy_cache = cache_data['lowbuy_top5']
            st.session_state._lb_dbg = cache_data.get('lowbuy_dbg') or {}
            st.session_state.last_update_time = datetime.fromisoformat(cache_data['timestamp'])
            st.session_state.data_status = 'cached'
            cached_lb_ver = cache_data.get('lowbuy_params_version', 0)
            current_lb_ver = DEFAULT_LOWBUY_PARAMS.get('_params_version', 0)
            if cached_lb_ver < current_lb_ver:
                st.session_state.lowbuy_cache = None
                st.session_state.last_update_time = None
                st.session_state._lb_dbg = {}
                st.session_state.data_status = 'normal'
        st.session_state.cache_loaded = True
        if '_lb_dbg' not in st.session_state: st.session_state._lb_dbg = {}

def load_watchlist():
    try:
        if os.path.exists(WATCHLIST_JSON):
            with open(WATCHLIST_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data if isinstance(data, list) else []
    except Exception as e:
        print(f"加载自选股失败: {e}")
    return []

def save_watchlist(watchlist):
    try:
        with open(WATCHLIST_JSON, 'w', encoding='utf-8') as f:
            json.dump(watchlist, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存自选股失败: {e}")
        return False

def save_cache_data(data_dict):
    """保存选股结果缓存到文件"""
    try:
        cache = {
            'timestamp': datetime.now().isoformat(),
            'chase_high_top10': data_dict.get('chase_high_top10'),
            'lowbuy_top5': data_dict.get('lowbuy_top5'),
            'lowbuy_dbg': data_dict.get('lowbuy_dbg'),
            'lowbuy_params_version': DEFAULT_LOWBUY_PARAMS.get('_params_version', 0),
            'stock_pool_columns': data_dict.get('stock_pool_columns', []),
        }
        with open(CACHE_DATA_JSON, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存缓存失败: {e}")
        return False

def load_cache_data():
    """加载选股结果缓存"""
    try:
        if os.path.exists(CACHE_DATA_JSON):
            with open(CACHE_DATA_JSON, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None
    except Exception as e:
        print(f"加载缓存失败: {e}")
        return None

def is_cache_today(cache_data):
    """检查缓存是否是今天的"""
    if not cache_data or 'timestamp' not in cache_data:
        return False
    try:
        cache_date = datetime.fromisoformat(cache_data['timestamp']).date()
        return cache_date == datetime.now().date()
    except Exception:
        return False

init_session_state()


# ================================================================
#                数据获取模块（pytdx 通达信 + akshare兜底）
# ================================================================

def _supplement_tencent_quotes(df):
    """用腾讯微证券 API 补齐 pytdx 缺失的量比、换手率和名称。
    腾讯 API: qt.gtimg.cn, parts[1]=名称, parts[38]=换手率, parts[39]=PE(动态), parts[46]=量比
    分批请求（每批150只），顺序执行避免限流。
    """
    if df is None or len(df) == 0:
        return df
    try:
        import requests as _req
        codes = df['代码'].astype(str).values
        # 构建带市场前缀的代码列表
        def _to_tx(code):
            return ('sh' if code.startswith(('6','5','9')) else 'sz') + code
        tx_codes = [_to_tx(c) for c in codes]

        BATCH = 150
        results = {}  # code -> (量比, 换手率, PE, 名称)

        for batch_idx in range((len(tx_codes) + BATCH - 1) // BATCH):
            batch = tx_codes[batch_idx * BATCH : (batch_idx + 1) * BATCH]
            query = ','.join(batch)
            try:
                resp = _req.get(f'http://qt.gtimg.cn/q={query}', timeout=15,
                    headers={'User-Agent': 'Mozilla/5.0'})
                for line in resp.text.strip().split('\n'):
                    if '=\"' in line and '~' in line:
                        parts = line.split('\"')[1].split('~')
                        if len(parts) >= 47:
                            raw_code = parts[2]  # e.g. '600000'
                            # 名称
                            name = parts[1].strip() if parts[1] else ''
                            try:
                                vol_ratio = float(parts[46])
                            except (ValueError, IndexError):
                                vol_ratio = 1.0
                            try:
                                turnover = float(parts[38])
                            except (ValueError, IndexError):
                                turnover = 0.0
                            try:
                                pe = float(parts[39])
                            except (ValueError, IndexError):
                                pe = 0.0
                            results[raw_code] = (vol_ratio, turnover, pe, name)
            except Exception:
                continue  # 单批失败不阻塞整体

        # 合并到 DataFrame
        if results:
            vr_map, to_map, pe_map, name_map = {}, {}, {}, {}
            for code, (vr, to, pe, name) in results.items():
                vr_map[code] = vr
                to_map[code] = to
                if pe > 0:
                    pe_map[code] = pe
                if name:
                    name_map[code] = name
            df_codes = df['代码'].astype(str)
            df['量比'] = pd.to_numeric(df_codes.map(vr_map), errors='coerce')
            df['量比'] = df['量比'].fillna(1.0)
            df['换手率'] = pd.to_numeric(df_codes.map(to_map), errors='coerce')
            df['换手率'] = df['换手率'].fillna(0.0)
            # 补充PE：仅填充空值
            if pe_map:
                pe_mask = df['市盈率-动态'].isna() | (df['市盈率-动态'] == 0)
                pe_vals = pd.to_numeric(df_codes[pe_mask].map(pe_map), errors='coerce')
                df.loc[pe_mask, '市盈率-动态'] = pe_vals.astype(float)
            # 修复名称：用腾讯数据覆盖所有匹配到的名称（腾讯名称更可靠）
            if name_map:
                df['名称'] = df_codes.map(name_map).fillna(df['名称'])
            if _VERBOSE: print(f"[Tencent] 补充量比/换手率/PE+名称: {len(results)}/{len(df)} 只匹配")
        return df
    except Exception as _e:
        if _VERBOSE: print(f"[Tencent] 补充数据异常: {_e}")
        return df


@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_all_a_stocks_tdx():
    """通过 pytdx 从通达信服务器获取全市场A股实时行情"""
    if not tdx_available():
        print("[TDX] pytdx 不可用")
        return None
    try:
        df = fetch_all_quotes_tdx()
        if df is not None and len(df) > 0:
            if _VERBOSE: print(f"[TDX] 成功获取 {len(df)} 条股票行情数据")
            df = _supplement_tencent_quotes(df)
            return df
    except Exception as e:
        if _VERBOSE: print(f"[TDX] 行情获取异常: {e}")
    return None



def fetch_all_a_stocks():
    """获取A股全市场实时行情数据（模块级缓存30分钟）
    优先：pytdx 从通达信服务器获取快速行情快照
    兜底：akshare stock_zh_a_spot_em
    回退：本地文件缓存
    """
    # 模块级缓存优先
    if _app_cache['raw_market_data'] is not None and _app_cache['raw_market_time']:
        if (time.time() - _app_cache['raw_market_time']) < 1800:
            return _app_cache['raw_market_data']

    if not tdx_available() and not AKSHARE_AVAILABLE:
        st.error("⚠️ pytdx 和 akshare 均不可用，请检查通达信连接或安装 akshare")
        return None

    try:
        with st.spinner("📡 正在获取A股全市场行情数据..."):
            df = None

            # 第一优先级：pytdx 通达信
            if tdx_available():
                try:
                    df = _fetch_all_a_stocks_tdx()
                    if df is not None and len(df) > 0:
                        if _VERBOSE: print(f"[pytdx] 获取到 {len(df)} 只股票")
                except Exception as _tdx_e:
                    print(f"[pytdx失败] {_tdx_e}，尝试akshare...")

            # 第二优先级：akshare 兜底（带超时保护，避免push2 API阻塞）
            if df is None or len(df) == 0:
                if AKSHARE_AVAILABLE:
                    import threading
                    _ak_result = [None]
                    _ak_error = [None]
                    def _ak_fetch():
                        try:
                            _ak_result[0] = ak.stock_zh_a_spot_em()
                        except Exception as e:
                            _ak_error[0] = e
                    _ak_thread = threading.Thread(target=_ak_fetch, daemon=True)
                    _ak_thread.start()
                    _ak_thread.join(timeout=8)  # 最多等8秒
                    if _ak_thread.is_alive():
                        if _VERBOSE: print(f"[akshare] 超时跳过（push2 API不可达）")
                    elif _ak_error[0] is not None:
                        print(f"[akshare失败] {_ak_error[0]}")
                    elif _ak_result[0] is not None and len(_ak_result[0]) > 0:
                        df = _ak_result[0]
                        if _VERBOSE: print(f"[akshare] 获取到 {len(df)} 只股票")

            # 第三优先级：本地文件缓存回退
            if df is None or len(df) == 0:
                try:
                    cache_path = os.path.join(BASE_DIR, 'market_data_cache.json')
                    if os.path.exists(cache_path):
                        with open(cache_path, 'r', encoding='utf-8') as _f:
                            cached = json.load(_f)
                        if cached and 'data' in cached and len(cached['data']) > 0:
                            df = pd.DataFrame(cached['data'])
                            cache_age = time.time() - cached.get('timestamp', 0)
                            cache_hours = int(cache_age / 3600)
                            _app_cache['raw_market_data'] = df
                            _app_cache['raw_market_time'] = time.time()
                            st.warning(f"⚠️ pytdx/akshare获取失败，已使用 {cache_hours} 小时前的本地缓存数据（{len(df)} 条）")
                            if _VERBOSE: print(f"[缓存回退] 从本地缓存恢复 {len(df)} 条数据，缓存时间: {cache_hours}小时前")
                            return df
                except Exception as _cache_e:
                    print(f"[缓存回退失败] {_cache_e}")
                st.error("❌ 获取数据为空，请检查通达信连接和网络")
                return None

            # 存入模块级缓存
            _app_cache['raw_market_data'] = df
            _app_cache['raw_market_time'] = time.time()

            # 保存到本地文件作为回退
            try:
                cache_path = os.path.join(BASE_DIR, 'market_data_cache.json')
                with open(cache_path, 'w', encoding='utf-8') as _f:
                    json.dump({
                        'timestamp': time.time(),
                        'data': df.to_dict(orient='records')
                    }, _f, ensure_ascii=False)
                if _VERBOSE: print(f"[缓存] 已保存 {len(df)} 条行情数据到本地文件")
            except Exception as _e:
                if _VERBOSE: print(f"[缓存] 保存本地缓存失败: {_e}")
            return df
    except Exception as e:
        st.error(f"❌ 获取行情数据失败: {str(e)}")
        st.info("💡 提示：请检查通达信连接，或等待数据源恢复")
        return None


def preprocess_stock_data(df_raw):
    """数据预处理和过滤"""
    if df_raw is None or len(df_raw) == 0:
        return None
    df = df_raw.copy()
    
    required_cols = ['代码', '名称', '最新价', '涨跌幅']
    for col in required_cols:
        if col not in df.columns:
            st.warning(f"⚠️ 缺少必要列: {col}")
            return None
    
    # 排除ST/*ST股票
    df = df[~df['名称'].str.contains('ST|退市', case=False, na=False)]
    
    # 排除停牌股票
    if '成交量' in df.columns:
        df = df[df['成交量'] > 0]
    if '最新价' in df.columns:
        df = df[df['最新价'] > 0]
    
    # 数据类型转换
    numeric_cols = ['最新价', '涨跌幅', '涨跌额', '成交量', '成交额', '振幅', 
                    '最高', '最低', '今开', '昨收', '量比', '换手率']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # 兼容本地缓存文件的列名差异：缓存的 '市盈率' → '市盈率-动态'
    if '市盈率' in df.columns and '市盈率-动态' not in df.columns:
        df.rename(columns={'市盈率': '市盈率-动态'}, inplace=True)
    
    if '市盈率-动态' in df.columns:
        df['市盈率-动态'] = pd.to_numeric(df['市盈率-动态'], errors='coerce')
    if '市净率' in df.columns:
        df['市净率'] = pd.to_numeric(df['市净率'], errors='coerce')
    
    # 确保代码是6位字符串
    df['代码'] = df['代码'].astype(str).str.zfill(6)
    
    return df.reset_index(drop=True)


@st.cache_data(ttl=3600, show_spinner=False)
def get_stock_kline(code, days=30):
    """
    获取个股日K线数据
    优先从通达信本地数据读取（毫秒级），失败则回退 pytdx get_security_bars
    """
    # === 优先尝试通达信本地数据 ===
    if TDX_AVAILABLE:
        df = read_tdx_day_file(code)
        if df is not None and len(df) > 0:
            return df.tail(days)

    # === 回退到 pytdx ===
    if tdx_available():
        df = fetch_kline_tdx(code, days=days)
        if df is not None and len(df) > 0:
            return df

    return None



def get_kline_with_today(code, days=60):
    """Get K-line data with today's forming candle appended via pytdx real-time quote"""
    df = get_stock_kline(code, days=days)
    if df is None or len(df) == 0:
        return df, False
    now = datetime.now()
    if now.weekday() >= 5:
        return df, False
    hour_min = now.hour * 100 + now.minute
    if hour_min < 915 or hour_min > 1505:
        return df, False
    last_date_str = str(df.iloc[-1].get('\u65e5\u671f', ''))
    today_str = now.strftime('%Y-%m-%d')
    today_str2 = now.strftime('%Y%m%d')
    if today_str in last_date_str or today_str2 in last_date_str:
        return df, False
    if not tdx_available():
        return df, False
    try:
        today_quote = get_today_quote_single(code)
        if today_quote is None:
            return df, False
        today_close = today_quote.get('最新价', 0)
        today_open = today_quote.get('今开', 0)
        if today_close <= 0:
            return df, False
        today_row = {
            '\u65e5\u671f': today_str,
            '\u5f00\u76d8': today_open,
            '\u6700\u9ad8': today_quote.get('最高', today_close),
            '\u6700\u4f4e': today_quote.get('最低', today_close),
            '\u6536\u76d8': today_close,
            '\u6210\u4ea4\u91cf': today_quote.get('成交量', 0),
            '\u6210\u4ea4\u989d': today_quote.get('成交额', 0),
        }
        new_row = pd.DataFrame([today_row])
        df = pd.concat([df, new_row], ignore_index=True)
        return df, True
    except Exception as e:
        print(f"Append today K-line failed: {e}")
        return df, False


def create_candlestick_chart(kline_df, title="", is_realtime=False):
    """Create candlestick chart with volume bars"""
    df = kline_df.copy()
    if len(df) == 0:
        return None
    date_col = None
    for c in df.columns:
        if c == '\u65e5\u671f' or 'date' in c.lower():
            date_col = c
            break
    if date_col is None:
        return None
    dates = pd.to_datetime(df[date_col])
    has_volume = '\u6210\u4ea4\u91cf' in df.columns
    if has_volume:
        fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                           vertical_spacing=0.03, row_heights=[0.7, 0.3])
        fig.add_trace(go.Candlestick(
            x=dates, open=df['\u5f00\u76d8'], high=df['\u6700\u9ad8'],
            low=df['\u6700\u4f4e'], close=df['\u6536\u76d8'],
            increasing_line_color='#E74C3C', decreasing_line_color='#27AE60',
            increasing_fillcolor='#E74C3C', decreasing_fillcolor='#27AE60',
            name='K\u7ebf'
        ), row=1, col=1)
        vol_colors = ['#E74C3C' if df['\u6536\u76d8'].iloc[i] >= df['\u5f00\u76d8'].iloc[i] else '#27AE60'
                      for i in range(len(df))]
        if is_realtime and len(vol_colors) > 0:
            vol_colors[-1] = 'rgba(100,100,200,0.5)'
        fig.add_trace(go.Bar(
            x=dates, y=df['\u6210\u4ea4\u91cf'],
            marker_color=vol_colors, opacity=0.7, name='\u6210\u4ea4\u91cf',
            marker_line_width=0
        ), row=2, col=1)
    else:
        fig = go.Figure()
        fig.add_trace(go.Candlestick(
            x=dates, open=df['\u5f00\u76d8'], high=df['\u6700\u9ad8'],
            low=df['\u6700\u4f4e'], close=df['\u6536\u76d8'],
            increasing_line_color='#E74C3C', decreasing_line_color='#27AE60',
            increasing_fillcolor='#E74C3C', decreasing_fillcolor='#27AE60',
            name='K\u7ebf'
        ))
    if is_realtime:
        last_idx = len(dates) - 1
        fig.add_annotation(
            x=dates.iloc[last_idx], y=df['\u6700\u9ad8'].iloc[last_idx],
            text='\u25b6 \u672a\u6536\u76d8', showarrow=False,
            font=dict(size=10, color='#C4842D'),
            yshift=12, row=1, col=1
        )
    fig.update_layout(
        height=500 if has_volume else 350,
        paper_bgcolor='white', plot_bgcolor='rgba(250,249,245,0.5)',
        xaxis_rangeslider_visible=False,
        title_text=title if title else '',
        margin=dict(t=50, b=20, l=50, r=20),
        showlegend=False,
    )
    if has_volume:
        fig.update_yaxes(title_text='\u4ef7\u683c', row=1, col=1)
        fig.update_yaxes(title_text='\u6210\u4ea4\u91cf', row=2, col=1)
    return fig


@st.cache_data(ttl=3600, show_spinner=False)
def calculate_rsi(code, period=14):
    """计算个股RSI指标"""
    df = get_stock_kline(code, days=120)
    if df is None or len(df) < period + 1:
        return None
    delta = df['收盘'].diff()
    gain = delta.copy()
    loss = delta.copy()
    gain[gain < 0] = 0
    loss[loss > 0] = 0
    loss = abs(loss)
    avg_gain = gain.rolling(window=period).mean()
    avg_loss = loss.rolling(window=period).mean()
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return round(rsi.iloc[-1], 1) if not pd.isna(rsi.iloc[-1]) else None


def batch_calculate_rsi(codes, max_count=300):
    """批量计算RSI（多线程并行）"""
    rsi_dict = {}
    codes_to_calc = codes[:max_count]
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def _calc_one_rsi(code):
        return code, calculate_rsi(code)
    
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(_calc_one_rsi, code) for code in codes_to_calc]
        done_count = 0
        for future in as_completed(futures):
            done_count += 1
            if done_count % 30 == 0:
                status_text.text(f"正在计算RSI... ({done_count}/{len(codes_to_calc)})")
            progress_bar.progress(done_count / len(codes_to_calc))
            try:
                code, rsi = future.result(timeout=15)
                if rsi is not None:
                    rsi_dict[code] = rsi
            except Exception:
                pass
    
    progress_bar.empty()
    status_text.empty()
    return rsi_dict

# ==================== 共振模型缓存 ====================
def get_resonance_cache(date_str=None):
    """获取共振模型缓存数据"""
    if date_str is None:
        date_str = datetime.now().strftime('%Y%m%d')
    cache_file = os.path.join(CACHE_DIR, f'resonance_{date_str}.pkl')
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'rb') as f:
                return pickle.load(f)
        except:
            pass
    return None

def save_resonance_cache(data, date_str=None):
    """保存共振模型缓存数据"""
    if date_str is None:
        date_str = datetime.now().strftime('%Y%m%d')
    cache_file = os.path.join(CACHE_DIR, f'resonance_{date_str}.pkl')
    try:
        with open(cache_file, 'wb') as f:
            pickle.dump(data, f)
    except:
        pass


def _get_resonance_cross_ref():
    """获取共振模型交叉评分字典 {code: 综合评分} 和 {code: 当前走势}"""
    scores = {}; styles = {}
    # 优先从 session_state
    rr = st.session_state.get('resonance_results', [])
    if not rr:
        # 回退到今日缓存文件
        cache_file = os.path.join(CACHE_DIR, f"resonance_{datetime.now().strftime('%Y%m%d')}.pkl")
        if os.path.exists(cache_file):
            try:
                with open(cache_file, 'rb') as f:
                    cache = pickle.load(f)
                rr = cache.get('results', [])
            except:
                pass
    for r in rr:
        code = str(r.get('代码', '')).zfill(6)
        scores[code] = float(r.get('综合评分', 0))
        styles[code] = r.get('当前走势', '-')
    return scores, styles


# ================================================================
#              v3八维评分模块（追高模型）
#     每个维度独立打分（满分=权重值），加权归一到0~100
# ================================================================

@st.cache_data(ttl=3600)
def fetch_dragon_tiger_v3():
    """获取龙虎榜数据，返回 {code: {inst_net_buy_3d, north_net_buy, pure_hot_money_only}}"""
    try:
        import akshare as ak
        today_str = datetime.now().strftime('%Y%m%d')
        cache_key = f"lhb_v3_{today_str}"
        if st.session_state.get(cache_key):
            return st.session_state[cache_key]
        result = {}
        for days_ago in range(5):
            try:
                dt = (datetime.now() - timedelta(days=days_ago)).strftime('%Y%m%d')
                df = ak.stock_lhb_detail_em(start_date=dt, end_date=dt)
                if df is None or len(df) == 0:
                    continue
                for _, r in df.iterrows():
                    code = str(r.get('代码', ''))
                    if not code or len(code) != 6:
                        continue
                    if code not in result:
                        result[code] = {'inst_net_buy_3d': 0, 'north_net_buy': 0,
                                        'pure_hot_money_only': True, 'inst_sell_2d': False,
                                        'inst_net_sell_2d': False}
                    buyer_str = str(r.get('买方', '')) + str(r.get('解读', ''))
                    if '机构' in buyer_str or '机构专用' in buyer_str:
                        result[code]['pure_hot_money_only'] = False
                        result[code]['inst_net_buy_3d'] += float(r.get('买入额', 0) or 0) - float(r.get('卖出额', 0) or 0)
                    if '北向' in buyer_str or '沪股通' in buyer_str or '深股通' in buyer_str:
                        result[code]['north_net_buy'] += float(r.get('买入额', 0) or 0)
            except:
                continue
        st.session_state[cache_key] = result
        return result
    except Exception as e:
        print(f"龙虎榜数据获取失败: {e}")
    return {}

@st.cache_data(ttl=1800)
def fetch_sector_board_v3():
    """获取板块行情数据（通过 pytdx 通达信板块数据）
    返回: ({板块名: {daily_gain, limit_up_count, net_inflow}}, {stock_code: sector_name})
    """
    if not tdx_available():
        return {}, {}
    try:
        sector_board, stock_sector_map = fetch_sector_data_tdx()
        if _VERBOSE: print(f"[pytdx] 板块数据: {len(sector_board)} 个板块, {len(stock_sector_map)} 只成分股映射")
        return sector_board, stock_sector_map
    except Exception as e:
        print(f"板块数据获取失败: {e}")
    return {}, {}

def _classify_stock_sector(stock_code, stock_name, sector_board_data, stock_sector_map=None):
    """根据股票代码匹配所属板块（优先使用成分股映射）"""
    import re
    code = str(stock_code)
    # 优先：通过成分股反向映射精确匹配
    if stock_sector_map and code in stock_sector_map:
        sec = stock_sector_map[code]
        # 🔧 P0修复: 过滤 pytdx 内部非中文编码（如 W300881W3、20175920 等）
        if re.search(r'[\u4e00-\u9fff]', sec):
            return sec
    # 备选：关键词匹配
    name = str(stock_name)
    kw_map = {'半导体': '半导体', '芯片': '半导体', 'AI': 'AI算力',
              '算力': 'AI算力', '光模块': 'AI算力', '云': '云计算',
              '软件': '云计算', '新能源': '新能源', '锂电': '新能源',
              '光伏': '新能源', '医药': '医药', '医疗': '医药',
              '白酒': '白酒', '食品': '食品饮料', '消费': '食品饮料'}
    for kw, sec in kw_map.items():
        if kw in name and sec in sector_board_data:
            return sec
    return None  # 未匹配到有效板块

def _estimate_pe_percentile(pe_val):
    """PE历史分位估算（0~100）"""
    if pd.isna(pe_val) or pe_val <= 0:
        return 50
    if pe_val < 15: return 15
    elif pe_val < 25: return 35
    elif pe_val < 40: return 55
    elif pe_val < 60: return 70
    elif pe_val < 80: return 82
    elif pe_val < 120: return 90
    else: return 96

def _calc_macd_divergence(kline_df):
    """检测MACD顶背离（简化版）"""
    if kline_df is None or len(kline_df) < 35:
        return False
    close = kline_df['收盘']
    # 计算MACD
    ema12 = close.ewm(span=12).mean()
    ema26 = close.ewm(span=26).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9).mean()
    macd_bar = (dif - dea) * 2
    # 检查最近10日是否有价格新高但MACD柱走低
    if len(close) >= 10:
        recent = close.tail(10)
        recent_macd = macd_bar.tail(10)
        price_high_idx = recent.idxmax()
        if price_high_idx == recent.index[-1]:  # 最新价是近10日最高
            if len(recent_macd) >= 2 and recent_macd.iloc[-1] < recent_macd.iloc[-3] if len(recent_macd) >= 3 else False:
                return True
    return False

def _build_stock_data(row, kline_df, lhb_data, sector_data, stock_sector_map=None):
    """将stock_screener数据转换为v3评分函数所需的stock_data字典"""
    code = str(row.get('代码', ''))
    name = str(row.get('名称', ''))

    # 从K线计算MA
    ma5 = ma10 = ma20 = ma60 = 0
    ma5_prev = ma10_prev = ma20_prev = ma60_prev = 0
    new_high_2d = False
    new_high_today = False
    vol_today = 1
    ma5_vol = 1
    close = float(row.get('最新价', 10))
    amp = float(row.get('振幅', 5))

    if kline_df is not None and len(kline_df) >= 60:
        c = kline_df['收盘']
        ma5 = c.rolling(5).mean().iloc[-1]
        ma10 = c.rolling(10).mean().iloc[-1]
        ma20 = c.rolling(20).mean().iloc[-1]
        ma60 = c.rolling(60).mean().iloc[-1]
        if len(c) >= 6:
            ma5_prev = c.rolling(5).mean().iloc[-2]
            ma10_prev = c.rolling(10).mean().iloc[-2]
            ma20_prev = c.rolling(20).mean().iloc[-2]
            ma60_prev = c.rolling(60).mean().iloc[-2]
        close = float(c.iloc[-1])
        # 新高判断
        if len(c) >= 3:
            new_high_today = c.iloc[-1] >= c.tail(20).max()
            new_high_2d = c.iloc[-1] >= c.tail(20).max() and c.iloc[-2] >= c.tail(21).iloc[:-1].max()
        # 成交量
        v = kline_df['成交量']
        vol_today = float(v.iloc[-1]) if len(v) > 0 else 1
        ma5_vol = float(v.rolling(5).mean().iloc[-1]) if len(v) >= 5 else vol_today
        # 振幅
        if '振幅' in kline_df.columns and len(kline_df) > 0:
            amp = float(kline_df['振幅'].iloc[-1]) if pd.notna(kline_df['振幅'].iloc[-1]) else amp
        elif '最高' in kline_df.columns and '最低' in kline_df.columns:
            high = float(kline_df['最高'].iloc[-1])
            low = float(kline_df['最低'].iloc[-1])
            if low > 0:
                amp = (high - low) / low * 100
    else:
        # K线不足，用最新价估算MA；量价数据不可用标记N/A
        ma5 = ma10 = ma20 = ma60 = close
        ma5_prev = ma10_prev = ma20_prev = ma60_prev = close
        vol_today = -1; ma5_vol = -1; amp = -1

    # 涨幅
    g3 = float(row.get('涨跌幅_3d', 0) or 0)
    g5 = float(row.get('涨跌幅_5d', 0) or 0)
    g10 = float(row.get('涨跌幅_10d', 0) or 0)

    # 6日涨幅（从K线计算）
    g6 = 0
    if kline_df is not None and len(kline_df) >= 7:
        c6 = kline_df['收盘']
        g6 = (float(c6.iloc[-1]) / float(c6.iloc[-7]) - 1) * 100

    # 赛道
    sector = _classify_stock_sector(code, name, sector_data, stock_sector_map)
    sec_info = sector_data.get(sector, {}) if sector else {}

    # 龙虎榜
    lhb = lhb_data.get(code, {})

    # PE分位
    pe_val = float(row.get('市盈率-动态', 30) or 30)
    pe_pct = _estimate_pe_percentile(pe_val)

    # 换手率
    turnover = float(row.get('换手率', 3) or 3)

    # 热度估算（用RSI+涨幅综合）— N/A标记：RSI不可用时hot=-1
    rsi_raw = row.get('RSI', None)
    if rsi_raw is None or (isinstance(rsi_raw, float) and pd.isna(rsi_raw)):
        hot = -1
    else:
        rsi = float(rsi_raw or 50)
        hot = min(10, max(0, (rsi - 30) / 7 + max(0, g5) / 5))

    # MACD背离
    macd_div = _calc_macd_divergence(kline_df)

    stock_data = {
        "sector": sector,
        "3d_gain": g3, "5d_gain": g5, "10d_gain": g10, "6d_gain": g6,
        "close": close, "ma60": ma60, "ma60_prev": ma60_prev,
        "ma5": ma5, "ma5_prev": ma5_prev,
        "ma10": ma10, "ma10_prev": ma10_prev,
        "ma20": ma20, "ma20_prev": ma20_prev,
        "vol_today": vol_today, "ma5_vol": ma5_vol,
        "amplitude": amp,
        "new_high_2d": new_high_2d, "new_high_today": new_high_today,
        "inst_net_sell_2d": lhb.get('inst_sell_2d', False),
        "macd_top_divergence": macd_div,
        "sector_daily_gain": sec_info.get('daily_gain', 0),
        "sector_limit_up_count_count": sec_info.get('limit_up_count', 0),
        "sector_net_inflow": sec_info.get('net_inflow', 0),
        "inst_net_buy_3d": -1 if not lhb else lhb.get('inst_net_buy_3d', 0),
        "north_net_buy": -1 if not lhb else lhb.get('north_net_buy', 0),
        "pure_hot_money_only": lhb.get('pure_hot_money_only', True),
        "pe_hist_percent": pe_pct,
        "pe_raw": pe_val,
        "turnover_rate": turnover,
        "stock_hot_score": hot,
        "sector_fund_outflow": False,
    }

    # 填充板块资金流向判定（追高模型：板块净流出则淘汰）
    if stock_data["sector"]:
        fund_flows = _get_sector_fund_flow()
        sector_flow = fund_flows.get(stock_data["sector"], None)
        if sector_flow is not None:
            stock_data["sector_fund_outflow"] = (sector_flow <= 0)
    return stock_data


# ===================== 前置硬过滤 =====================
def hard_filter_v3(stock_data):
    """满足任意一条直接淘汰（7条硬过滤规则）。return: (True=保留, msg)
    1. 3日涨幅 > 15% → 淘汰（短期过热）
    2. 距MA60涨幅 > 80% → 淘汰（中长期乖离过大）
    3. 缩量新高顶背离 → 淘汰
    4. 机构连续2日净卖出 → 淘汰
    5. MACD高位顶背离 → 淘汰
    6. 连续6交易日累计涨幅 > 25% → 淘汰（中期过热）
    7. 所属板块资金净流出 → 淘汰（板块无主力资金支持）
    """
    if stock_data["3d_gain"] > 15:
        return False, f"3日涨幅{stock_data['3d_gain']:.1f}%>15%"
    if stock_data["ma60"] > 0 and (stock_data["close"] / stock_data["ma60"] - 1) > 0.8:
        return False, f"距MA60涨幅>{80}%"
    if stock_data["vol_today"] < stock_data["ma5_vol"] * 0.8 and stock_data["new_high_2d"]:
        return False, "缩量新高顶背离"
    if stock_data["inst_net_sell_2d"]:
        return False, "机构连续2日净卖出"
    if stock_data["macd_top_divergence"]:
        return False, "MACD高位顶背离"
    if stock_data.get("6d_gain", 0) > 25:
        return False, f"6日累计涨幅{stock_data['6d_gain']:.1f}%>25%"
    if stock_data.get("sector_fund_outflow", False):
        return False, f"板块{stock_data.get('sector', '')}资金净流出"
    return True, "过滤通过"

# ===================== 维度1：趋势结构 满分25 =====================
def score_trend_struct_v3(stock_data):
    # N/A 检测：K线不足时所有MA被设为收盘价(flat pattern)，无法判断趋势
    if (stock_data["ma5"] == stock_data["ma60"] 
        and stock_data["ma5"] == stock_data["ma5_prev"]
        and stock_data["ma10"] == stock_data["ma10_prev"]
        and stock_data["ma20"] == stock_data["ma20_prev"]):
        return -1  # N/A：K线数据不足，无法评分

    ma5_up = stock_data["ma5"] > stock_data["ma5_prev"]
    ma10_up = stock_data["ma10"] > stock_data["ma10_prev"]
    ma20_up = stock_data["ma20"] > stock_data["ma20_prev"]
    ma60_up = stock_data["ma60"] > stock_data["ma60_prev"]
    close_on_ma5 = stock_data["close"] > stock_data["ma5"]
    # 分级计分：满足的条件数 × 5，替代原 all-or-nothing 三档
    met_count = sum([ma5_up, ma10_up, ma20_up, ma60_up, close_on_ma5])
    base_score = met_count * 5
    # 高位乖离惩罚：距MA60涨幅>80%仍扣分
    if stock_data["ma60"] > 0 and (stock_data["close"] / stock_data["ma60"] - 1) > 0.8:
        base_score -= 10
    return max(base_score, 0)

# ===================== 维度2：动量强度 满分22 =====================
def score_momentum_v3(stock_data):
    # N/A 检测：3d/5d/10d 涨幅全为0且MA flat pattern → 数据不可用
    if (stock_data["3d_gain"] == 0 and stock_data["5d_gain"] == 0 and stock_data["10d_gain"] == 0
        and stock_data["ma5"] == stock_data["ma60"] and stock_data["ma5"] == stock_data["ma5_prev"]):
        return -1  # N/A：涨幅数据缺失，无法评分

    gain5 = stock_data["5d_gain"]
    gain10 = stock_data["10d_gain"]
    score5 = 0
    if 8 <= gain5 <= 25: score5 = 13.2     # 8%+ 5日涨幅已是强势
    elif 3 <= gain5 < 8: score5 = 8         # 温和小涨
    elif 0 < gain5 < 3: score5 = 4          # 微涨也有分
    elif -3 < gain5 <= 0: score5 = 2        # 横盘/微跌也给基础分
    score10 = 0
    if 15 <= gain10 <= 50: score10 = 8.8    # 15%+ 10日趋势确立
    elif 5 <= gain10 < 15: score10 = 5       # 温和趋势
    elif 0 < gain10 < 5: score10 = 2         # 微涨趋势
    elif -5 < gain10 <= 0: score10 = 1       # 横盘/微跌基础分
    total = score5 + score10
    if stock_data["3d_gain"] > 15:
        total -= 12
    return max(total, 0)

# ===================== 维度3：板块共振 满分10 =====================
def score_sector_resonance_v3(stock_data):
    sector = stock_data.get("sector")
    if not sector:
        return 0  # 未匹配到板块，不给分也不扣分
    sec_gain = stock_data.get("sector_daily_gain", 0)
    sec_limit = stock_data.get("sector_limit_up_count_count", 0)
    sec_flow = stock_data.get("sector_net_inflow", 0)

    # 板块数据全部为0（数据不可用）时，用个股涨跌幅+量比做fallback
    if sec_gain == 0 and sec_limit == 0 and sec_flow == 0:
        stock_gain = stock_data.get("3d_gain", 0)
        vol_ratio = stock_data.get("volume_ratio", 1.0)
        if stock_gain > 5 and vol_ratio > 1.5:
            return 6  # 个股强势放量上涨
        elif stock_gain > 2 and vol_ratio > 1.2:
            return 4
        elif stock_gain > 0:
            return 2
        else:
            return 0  # 个股弱势

    is_growth = sector in GROWTH_SECTOR
    # 评分逻辑（基于日涨幅）
    if is_growth and sec_gain > 2 and sec_limit >= 10 and sec_flow > 0:
        return 10  # 成长板块强势共振 (10只涨停即算)
    elif sec_gain > 2 and sec_limit >= 8:
        return 8   # 板块整体强势
    elif sec_gain > 1 and sec_limit >= 5:
        return 6   # 板块温和上涨
    elif sec_gain > 0 and sec_limit >= 3:
        return 4   # 板块小幅上涨
    elif sec_gain > 0:
        return 2   # 板块微涨
    else:
        return 0   # 板块下跌

# ===================== 维度4-5-6：资金信号分层（P0：拆分为北向/机构/板块资金） =====================
def score_north_capital_v3(stock_data):
    """北向资金：近3日净流入方向，满分10。无LHB数据返回-1(N/A)"""
    north_net = stock_data.get("north_net_buy", 0)
    if north_net < 0:
        return -1  # N/A：无LHB数据
    if north_net > 50000:
        return 10
    elif north_net > 20000:
        return 8
    elif north_net > 5000:
        return 6
    elif north_net > 0:
        return 4
    elif north_net > -5000:
        return 2
    else:
        return 0

def score_inst_capital_v3(stock_data):
    """机构净买：近3日净买入，满分10。无LHB数据返回-1(N/A)"""
    inst_net = stock_data.get("inst_net_buy_3d", 0)
    if inst_net < 0:
        return -1  # N/A：无LHB数据
    pure_hot = stock_data.get("pure_hot_money_only", False)
    if inst_net > 30000:
        base = 10
    elif inst_net > 10000:
        base = 7
    elif inst_net > 0:
        base = 5
    else:
        base = 2
    if pure_hot:
        base = max(base - 4, 0)
    return base

def score_sector_fund_heat_v3(stock_data):
    """板块资金热度：基于板块在板块反转模型排名中的位置映射，满分10"""
    sector_rank = stock_data.get("sector_resonance_rank", -1)
    if sector_rank < 0:
        # 没有板块排名时，使用板块净流入做替代
        sec_flow = stock_data.get("sector_net_inflow", 0)
        if sec_flow > 50000: return 10
        elif sec_flow > 20000: return 7
        elif sec_flow > 5000: return 5
        elif sec_flow > 0: return 3
        else: return -1  # N/A：板块资金数据完全不可用
    # 有排名：前20给10分，逐步递减
    if sector_rank <= 5: return 10
    elif sector_rank <= 10: return 8
    elif sector_rank <= 20: return 6
    elif sector_rank <= 30: return 4
    elif sector_rank <= 50: return 2
    else: return 1

# ===================== 维度5：龙虎榜资金（保留旧版兼容） =====================
def score_dragon_tiger_v3(stock_data):
    inst_net = stock_data["inst_net_buy_3d"]
    north_net = stock_data["north_net_buy"]
    pure_hot = stock_data["pure_hot_money_only"]
    if inst_net > 30000 and north_net > 0:
        base = 10
    elif inst_net > 0 or north_net > 0:
        base = 6
    else:
        base = 2
    if pure_hot:
        base -= 6
    return max(base, 0)

# ===================== 维度5：量价配合 满分12 =====================
def score_volume_price_v3(stock_data):
    vr = stock_data["vol_today"] / max(stock_data["ma5_vol"], 1)
    amp = stock_data["amplitude"]
    # N/A：成交量或振幅数据不可用
    if stock_data.get("vol_today", 0) <= 0 or stock_data.get("ma5_vol", 0) <= 0 or amp <= 0:
        return -1
    if 1.0 <= vr <= 1.8 and 5 <= amp <= 12:
        base = 12
    elif 0.8 <= vr < 1.0:
        base = 7
    elif vr < 0.8:
        base = 3
    else:
        base = 5
    if stock_data["new_high_today"] and vr < 0.8:
        base -= 8
    return max(base, 0)

# ===================== 维度6：估值安全 满分10 =====================
def score_valuation_v3(stock_data):
    sector = stock_data["sector"]
    pe_pct = stock_data["pe_hist_percent"]
    # N/A：PE分位数据不可用或板块未识别
    if pe_pct is None or (isinstance(pe_pct, float) and pd.isna(pe_pct)) or pe_pct < 0:
        return -1
    if sector in GROWTH_SECTOR:
        if pe_pct <= 80: return 10
        elif pe_pct <= 95: return 7
        else: return 3
    elif sector in CYCLE_SECTOR:
        if pe_pct <= 70: return 10
        elif pe_pct <= 90: return 4
        else: return 0
    else:
        # 非成长/非周期板块：通用估值打分，PE分位越低估值越安全
        if pe_pct <= 60: return 10
        elif pe_pct <= 75: return 7
        elif pe_pct <= 85: return 5
        elif pe_pct <= 95: return 3
        else: return 0

# ===================== 维度7：筹码稳定 满分7 =====================
def score_chip_v3(stock_data):
    t = stock_data["turnover_rate"]
    if 10 <= t <= 20: return 7
    elif 5 <= t < 10 or 20 < t <= 30: return 4
    else: return 1

# ===================== 维度8：情绪热度 满分4 =====================
def score_sentiment_v3(stock_data):
    hot = stock_data["stock_hot_score"]
    # N/A：热度数据不可用
    if hot < 0:
        return -1
    sector = stock_data["sector"]
    base = 4
    if hot >= 8:
        if sector in GROWTH_SECTOR:
            base -= 2
        else:
            base -= 4
    return max(base, 0)

# ===================== 总打分主函数 =====================


# ===================== 低吸维度1: 下跌幅度 满分30 =====================
def lb_score_decline_depth(stock_data, kline_df):
    """下跌幅度越深, 得分越高"""
    closes = kline_df["收盘"].values.astype(float)
    n = len(closes)
    if n < 20:
        return 0
    high_20 = max(closes[-20:])
    low_20 = min(closes[-20:])
    current = closes[-1]
    if high_20 <= 0:
        return 0
    drop_pct = (high_20 - current) / high_20 * 100
    if drop_pct >= 20: return 30
    elif drop_pct >= 15: return 25
    elif drop_pct >= 10: return 20
    elif drop_pct >= 7: return 15
    elif drop_pct >= 5: return 10
    elif drop_pct >= 3: return 5
    return 0

# ===================== 低吸维度2: 企稳信号 满分25 =====================
def lb_score_stabilization(stock_data, kline_df):
    """下跌放缓 + 波动收窄 + 近端止跌回升 = 企稳"""
    closes = kline_df["收盘"].values.astype(float)
    n = len(closes)
    if n < 15:
        return 0
    score = 0

    # 条件A: 近5日 vs 前5日——斜率从下跌转为走平/微升
    recent = closes[-5:]
    past = closes[-10:-5]
    try:
        s1 = np.polyfit(range(len(past)), past, 1)[0]
        s2 = np.polyfit(range(len(recent)), recent, 1)[0]
        # s2 需要从下跌转为至少不跌（反转信号）
        if s2 > 0 and s1 < 0:
            score += 14  # 从跌转涨，明确反转
        elif s2 >= 0 and s1 < 0:
            score += 10  # 从跌转平
        elif s2 > s1 and s1 < 0:
            score += 6   # 跌势放缓
        elif s2 > s1:
            score += 3   # 改善但非反转
    except:
        pass

    # 条件B: 波动率收窄
    recent_std = np.std(recent) / np.mean(recent) * 100 if np.mean(recent) > 0 else 999
    past_std = np.std(past) / np.mean(past) * 100 if np.mean(past) > 0 else 999
    if recent_std < past_std * 0.7:
        score += 11
    elif recent_std < past_std * 0.9:
        score += 6
    elif recent_std < past_std:
        score += 3
    return min(score, 25)

# ===================== 低吸维度3: 量能恢复 满分20 =====================
def lb_score_volume_recovery(stock_data, kline_df):
    """成交量从低位恢复上升 - v2: 持续放量而非单日尖刺"""
    volumes = kline_df["成交量"].values.astype(float)
    n = len(volumes)
    if n < 15:
        return 0
    # 用近2日均量 vs 前7日均量做对比，捕捉持续放量
    avg_vol_2d = np.mean(volumes[-2:])
    avg_vol_prior7 = np.mean(volumes[-9:-2])
    if avg_vol_prior7 <= 0 or avg_vol_2d <= 0:
        return 0
    ratio = avg_vol_2d / avg_vol_prior7
    if ratio >= 1.8: return 20
    elif ratio >= 1.4: return 15
    elif ratio >= 1.15: return 10
    elif ratio >= 1.0: return 5
    return 0

# ===================== 低吸维度4: 均线支撑 满分15 =====================
def lb_score_ma_support(stock_data, kline_df):
    """股价接近或站上短期均线 + MA5拐头向上（底部金叉前兆）"""
    close = stock_data["close"]
    ma5 = stock_data["ma5"]
    ma10 = stock_data["ma10"]
    ma20 = stock_data["ma20"]
    score = 0

    # 价格位置分 (满分7)
    if close > ma5:
        score += 4
    elif close > ma5 * 0.98:
        score += 2
    if close > ma10:
        score += 3
    elif close > ma10 * 0.98:
        score += 1

    # MA5 拐头确认 (满分8): 从K线数据计算MA5斜率变化
    closes_all = kline_df["收盘"].values.astype(float)
    n = len(closes_all)
    if n >= 8:
        ma5_now = np.mean(closes_all[-5:])
        ma5_3d_ago = np.mean(closes_all[-8:-3])
        ma5_change = (ma5_now - ma5_3d_ago) / ma5_3d_ago * 100 if ma5_3d_ago > 0 else 0
        if ma5_change > 0.3:
            score += 8   # MA5 明确拐头向上
        elif ma5_change >= 0:
            score += 4   # 走平或微升
        elif ma5_change > -0.3:
            score += 2   # 跌势趋缓
    return min(score, 15)

# ===================== 行业PE统计（低吸S5辅助） =====================
def _get_industry_pe_stats(df_market, stock_sector_map, force_refresh=False):
    """按行业分组计算PE中位数/P25/P75。模块级缓存30分钟。"""
    now = time.time()
    if not force_refresh and _industry_pe_cache["stats"] and (now - _industry_pe_cache["timestamp"]) < 1800:
        return _industry_pe_cache["stats"]

    industry_pes = {}
    for _, row in df_market.iterrows():
        code = str(row.get('代码', ''))
        pe = row.get('市盈率-动态', None)
        if pe is None or pd.isna(pe) or pe <= 0:
            continue
        sector = stock_sector_map.get(code, '其他') or '其他'
        industry_pes.setdefault(sector, []).append(float(pe))

    stats = {}
    for sector, pes in industry_pes.items():
        n = len(pes)
        if n < 5:
            continue
        arr = np.array(pes)
        stats[sector] = {
            "count": n,
            "median": float(np.median(arr)),
            "p25": float(np.percentile(arr, 25)),
            "p75": float(np.percentile(arr, 75)),
        }

    _industry_pe_cache["stats"] = stats
    _industry_pe_cache["timestamp"] = now
    return stats

# ===================== 低吸维度5: 估值吸引 满分10 =====================
def lb_score_valuation(stock_data):
    """行业相对PE打分 + 亏损拦截 + 景气修正系数。满分10。"""
    sector = stock_data.get("sector", "")
    pe_raw = float(stock_data.get("pe_raw", 0) or 0)

    # ===== 亏损拦截（优先级最高） =====
    if pe_raw <= 0:
        return 0

    stats = _industry_pe_cache.get("stats", {}) or {}
    info = stats.get(sector) if sector else None

    if info:
        median = info["median"]
        p25 = info["p25"]
        p75 = info["p75"]
        if pe_raw <= p25:
            base = 10
        elif pe_raw <= median:
            base = 8
        elif pe_raw <= p75:
            base = 5
        else:
            base = 2
    else:
        # fallback: absolute PE percentile
        pe_pct = float(stock_data.get("pe_hist_percent", 50) or 50)
        if pe_pct <= 20:
            base = 10
        elif pe_pct <= 40:
            base = 8
        elif pe_pct <= 60:
            base = 6
        elif pe_pct <= 80:
            base = 3
        else:
            base = 0

    # ===== 景气修正系数（TODO: 接入 akshare stock_yjbb_em 净利润同比/环比） =====
    coeff = _get_earnings_growth_coeff(stock_data)
    return int(base * coeff)

# ===== 景气修正系数函数 =====
def _get_earnings_growth_coeff(stock_data):
    """
    净利润景气修正系数（0~1.2），乘到估值基础分。
    TODO: 接入 akshare stock_yjbb_em(date="20260630") 获取:
      - 净利润同比增长率 (YoY)
      - 单季度净利润环比 (QoQ)
    规则:
      - 全年亏损（已由 pe_raw≤0 拦截） → 0
      - YoY↓ + QoQ↓ → 0.5
      - 一平一降 → 0.7
      - 双平 → 1.0
      - 一正一平 → 1.1
      - YoY↑ + QoQ↑ → 1.2
    当前默认 1.0（中性，不对基础分做修正）。
    """
    return 1.0

# ===================== 低吸维度6: 筹码沉淀 满分10 =====================
def lb_score_chip(stock_data, kline_df):
    """低换手+缩量=筹码锁定"""
    t = float(stock_data.get("turnover_rate", 5) or 5)
    volumes = kline_df["成交量"].values.astype(float)
    vol_ratio = volumes[-1] / np.mean(volumes[-10:]) if len(volumes) >= 10 and np.mean(volumes[-10:]) > 0 else 1
    score = 0
    if 3 <= t <= 6:
        score += 6
    elif 1 <= t < 3 or 6 < t <= 8:
        score += 3
    if vol_ratio < 0.8:
        score += 4
    elif vol_ratio < 1.1:
        score += 2
    return min(score, 10)


# ===================== P0: 低吸维度7: 主力资金 满分10 =====================
def lb_score_fund_flow(stock_data, kline_df):
    """
    主力资金信号：
    - 优先：近3日主力净流入 / 流通市值
    - 替代：近3日成交额均值 / 近20日成交额均值（量能活跃度）
    """
    # 尝试获取主力资金
    inst_net = stock_data.get("inst_net_buy_3d", 0)
    north_net = stock_data.get("north_net_buy", 0)
    main_net = inst_net + north_net

    if main_net != 0:
        cap = stock_data.get("float_cap", 1e9)
        ratio = abs(main_net) / max(cap, 1e7) * 100
        if main_net > 0:
            if ratio > 0.5: return 10
            elif ratio > 0.2: return 7
            elif ratio > 0.1: return 5
            else: return 3
        else:
            return 0
    else:
        # 替代：近3日成交额均值 / 近20日成交额均值
        if kline_df is None or len(kline_df) < 20:
            return 0
        try:
            amounts = kline_df["成交额"].values.astype(float) if "成交额" in kline_df.columns else kline_df["成交量"].values.astype(float)
        except:
            return 0
        if len(amounts) < 20:
            return 0
        avg_3d = np.mean(amounts[-3:])
        avg_20d = np.mean(amounts[-20:])
        if avg_20d <= 0:
            return 0
        ratio = avg_3d / avg_20d
        if ratio >= 1.8: return 10
        elif ratio >= 1.4: return 7
        elif ratio >= 1.1: return 5
        elif ratio >= 0.8: return 3
        else: return 1


# ===================== 低吸综合评分（P1: 支持百分位归一化） =====================
def calculate_lowbuy_score(stock_data, kline_df, params=None, weights=None):
    """低吸模型专用评分函数 - P1: 返回原始分供外部归一化"""
    if params is None:
        params = _get_lowbuy_params()
    if weights is None:
        weights = st.session_state.get("lowbuy_weights", DEFAULT_LOWBUY_WEIGHTS)
    s1 = lb_score_decline_depth(stock_data, kline_df)     # 0~30
    s2 = lb_score_stabilization(stock_data, kline_df)     # 0~25
    s3 = lb_score_volume_recovery(stock_data, kline_df)   # 0~20
    s4 = lb_score_ma_support(stock_data, kline_df)        # 0~15
    s5 = lb_score_valuation(stock_data)                   # 0~10
    s6 = lb_score_chip(stock_data, kline_df)              # 0~10
    s7 = lb_score_fund_flow(stock_data, kline_df)         # 0~10

    # 加权计算（使用归一化前的原始分 × 权重）
    n1 = s1 / 30 * 100
    n2 = s2 / 25 * 100
    n3 = s3 / 20 * 100
    n4 = s4 / 15 * 100
    n5 = s5 / 10 * 100
    n6 = s6 / 10 * 100
    n7 = s7 / 10 * 100

    total = int(
        n1 * weights.get("下跌幅度", 23) / 100 +
        n2 * weights.get("企稳信号", 18) / 100 +
        n3 * weights.get("量能恢复", 14) / 100 +
        n4 * weights.get("均线支撑", 14) / 100 +
        n5 * weights.get("估值吸引", 14) / 100 +
        n6 * weights.get("筹码沉淀",  9) / 100 +
        n7 * weights.get("主力资金",  8) / 100
    )

    # 7-8月二季报窗口：估值权重上浮、筹码权重下调（不影响UI权重滑块）
    current_month = datetime.now().month
    if current_month in [7, 8]:
        total = int(
            n1 * weights.get("下跌幅度", 23) / 100 +
            n2 * weights.get("企稳信号", 18) / 100 +
            n3 * weights.get("量能恢复", 14) / 100 +
            n4 * weights.get("均线支撑", 14) / 100 +
            n5 * 17 / 100 +
            n6 * 6 / 100 +
            n7 * weights.get("主力资金", 8) / 100
        )
    return {
        "pass": True,
        "综合评分": total,
        "下跌幅度": int(n1),
        "企稳信号": int(n2),
        "量能恢复": int(n3),
        "均线支撑": int(n4),
        "估值吸引": int(n5),
        "筹码沉淀": int(n6),
        "主力资金": int(n7),
        "_raw": {"s1": s1, "s2": s2, "s3": s3, "s4": s4, "s5": s5, "s6": s6, "s7": s7},
    }


def calculate_v3_total_score(stock_data, weights=None):
    """v4十维打分主函数（P0资金信号分层 + P1百分位归一化）。返回 {各维度分(归到100), 综合评分, pass, filter_msg, position_msg}
    weights: 可选权重字典，不传则从st.session_state读取（回测引擎可传DEFAULT_WEIGHTS）"""
    # 前置硬过滤 — 五条淘汰规则
    pass_filter, filter_msg = hard_filter_v3(stock_data)
    if not pass_filter:
        return {
            "pass": False, "filter_msg": filter_msg, "综合评分": 0,
            "趋势结构": 0, "动量强度": 0, "板块共振": 0,
            "北向资金": 0, "机构净买": 0, "板块资金热度": 0,
            "量价配合": 0, "估值安全": 0, "筹码稳定": 0, "情绪热度": 0,
            "position_msg": "前置淘汰",
            "_raw": {"s1": 0, "s2": 0, "s3": 0, "s4": 0, "s5": 0, "s6": 0, "s7": 0, "s8": 0, "s9": 0, "s10": 0},
        }
    # 十维打分
    s1 = score_trend_struct_v3(stock_data)       # 满分25
    s2 = score_momentum_v3(stock_data)            # 满分22
    s3 = score_sector_resonance_v3(stock_data)    # 满分10
    s4 = score_north_capital_v3(stock_data)       # P0: 北向资金 满分10
    s5 = score_inst_capital_v3(stock_data)        # P0: 机构净买 满分10
    s6 = score_sector_fund_heat_v3(stock_data)    # P0: 板块资金热度 满分10
    s7 = score_volume_price_v3(stock_data)        # 满分12
    s8 = score_valuation_v3(stock_data)           # 满分10
    s9 = score_chip_v3(stock_data)                # 满分7
    s10 = score_sentiment_v3(stock_data)          # 满分4
    # 归一到0~100后加权（N/A值-1按0计入总分，避免拉低综合评分）
    def _safe(v):
        return 0 if v < 0 else v
    n1 = _safe(s1) / 25 * 100; n2 = _safe(s2) / 22 * 100; n3 = _safe(s3) / 10 * 100
    n4 = _safe(s4) / 10 * 100; n5 = _safe(s5) / 10 * 100; n6 = _safe(s6) / 10 * 100
    n7 = _safe(s7) / 12 * 100; n8 = _safe(s8) / 10 * 100; n9 = _safe(s9) / 7 * 100; n10 = _safe(s10) / 4 * 100
    if weights is None:
        weights = st.session_state.get('weights', DEFAULT_WEIGHTS)
    total = int(
        n1 * weights.get("趋势结构", 15) / 100 + n2 * weights.get("动量强度", 18) / 100 +
        n3 * weights.get("板块共振", 8) / 100 +
        n4 * weights.get("北向资金", 15) / 100 + n5 * weights.get("机构净买", 10) / 100 +
        n6 * weights.get("板块资金热度", 5) / 100 +
        n7 * weights.get("量价配合", 14) / 100 + n8 * weights.get("估值安全", 3) / 100 +
        n9 * weights.get("筹码稳定", 6) / 100 + n10 * weights.get("情绪热度", 6) / 100
    )
    # 仓位分层（原始评分仅供参考，P1归一化后会重新计算）
    pos_msg = "分数不足70，放弃不参与"
    if total >= 85:
        pos_msg = "主线龙头，重仓60%~80%"
    elif total >= 70:
        pos_msg = "支线趋势，轻仓20%~40%"
    return {
        "pass": True, "filter_msg": filter_msg, "综合评分": total,
        "趋势结构": int(n1), "动量强度": int(n2), "板块共振": int(n3),
        "北向资金": int(n4), "机构净买": int(n5), "板块资金热度": int(n6),
        "量价配合": int(n7), "估值安全": int(n8), "筹码稳定": int(n9), "情绪热度": int(n10),
        "position_msg": pos_msg,
        "_raw": {"s1": s1, "s2": s2, "s3": s3, "s4": s4, "s5": s5, "s6": s6, "s7": s7, "s8": s8, "s9": s9, "s10": s10},
    }

# 兼容旧接口
def calculate_five_dimensions_score(row, rsi_dict=None, weight_config=None):
    """兼容旧接口"""
    return calculate_v3_total_score(row) if isinstance(row, dict) and "close" in row else {
        "趋势结构": 50, "动量强度": 50, "板块共振": 50,
        "北向资金": 50, "机构净买": 50, "板块资金热度": 50,
        "量价配合": 50, "估值安全": 50, "筹码稳定": 50, "情绪热度": 50,
        "综合评分": 50, "pass": True, "filter_msg": "", "position_msg": ""}


def _classify_signal(score, chg5, rsi, safety, mom):
    """信号分类"""
    if score >= 70 and mom >= 65 and safety >= 55 and 30 <= rsi <= 70:
        return "强势买入", "badge-strong", "积极关注"
    if (score >= 55 and score < 70) or (rsi < 35 and score >= 45) or (safety >= 60 and mom >= 50):
        return "逢低吸纳", "badge-attention", "分批布局"
    if score < 40 or safety < 35 or rsi > 80:
        return "建议回避", "badge-watch", "谨慎对待"
    return "观望等待", "badge-hold", "等待信号"


# ================================================================
#                低吸模型 v2：八维评分硬过滤 + K线条件 + 多维指标筛选
# ================================================================

def _check_low_buy_conditions(kline_df, params):
    """
    检查单只股票是否满足低吸条件
    返回 (通过, 信号强度, 信号标签, 分析摘要)
    """
    if kline_df is None or len(kline_df) < params['decline_days'] + 5:
        return False, 0, "", "", "badge-hold"
    
    closes = kline_df['收盘'].values.astype(float)
    volumes = kline_df['成交量'].values.astype(float)
    n = len(closes)
    
    decline_days = params['decline_days']
    vol_rise_days = params['vol_rise_days']
    chg_low = params['chg_low'] / 100.0
    chg_high = params['chg_high'] / 100.0
    
    # ===== 条件5：前一日涨幅在1%~5%区间 =====
    if n < 2:
        return False, 0, "", "", "badge-hold"
    yesterday_change = (closes[-1] - closes[-2]) / closes[-2]
    if not (chg_low <= yesterday_change <= chg_high):
        return False, 0, "", "", "badge-hold"
    
    # ===== 条件1：日K线逐步下移至少N天 =====
    if n < decline_days + 2:
        return False, 0, "", "", "badge-hold"
    # 取前decline_days天的收盘价（不含最近3天的反弹）
    # 即检查在反弹之前的下跌趋势
    lookback_end = max(n - 3, decline_days + 1)
    lookback_start = lookback_end - decline_days
    if lookback_start < 0:
        return False, 0, "", "", "badge-hold"
    
    past_closes = closes[lookback_start:lookback_end]
    if len(past_closes) < 3:
        return False, 0, "", "", "badge-hold"
    
    # 线性回归斜率为负 = 下跌趋势
    x_past = np.arange(len(past_closes))
    try:
        slope_past = np.polyfit(x_past, past_closes, 1)[0]
    except:
        return False, 0, "", "", "badge-hold"
    
    if slope_past >= 0:
        return False, 0, "", "", "badge-hold"  # 没有下跌趋势
    
    # 下跌幅度检查（至少跌了decline_decline%）
    total_decline_pct = (past_closes[0] - past_closes[-1]) / past_closes[0] * 100
    if total_decline_pct < 2:  # v5.2 relaxed
        return False, 0, "", "", "badge-hold"  # 跌幅太小不算低吸
    
    # ===== 条件2：下跌趋势逐步放缓并回稳 =====
    # 比较近段和远段的跌幅
    mid = len(past_closes) // 2
    if mid < 2:
        return False, 0, "", "", "badge-hold"
    
    # 前半段跌幅 vs 后半段跌幅
    first_half_decline = (past_closes[0] - past_closes[mid]) / past_closes[0] if past_closes[0] > 0 else 0
    second_half_decline = (past_closes[mid] - past_closes[-1]) / past_closes[mid] if past_closes[mid] > 0 else 0
    
    # 后半段跌幅应小于前半段（跌势放缓）
    # 或者后半段开始走平/微涨
    decline_slowing = (second_half_decline < first_half_decline * 0.8) or (second_half_decline <= 0)
    
    # 或者用波动率收窄来判断回稳
    recent_vol_std = np.std(closes[-5:]) if n >= 5 else 999
    past_vol_std = np.std(past_closes) if len(past_closes) > 3 else 0
    vol_narrowing = recent_vol_std < past_vol_std * 0.8
    
    if not (decline_slowing or vol_narrowing):
        return False, 0, "", "", "badge-hold"
    
    # ===== 条件3：出现N日及以上交易量上升趋势 =====
    if n < vol_rise_days + 1:
        return False, 0, "", "", "badge-hold"
    
    recent_volumes = volumes[-vol_rise_days:]
    vol_increasing = all(recent_volumes[i] > recent_volumes[i-1] for i in range(1, len(recent_volumes)))
    if not vol_increasing:
        return False, 0, "", "", "badge-hold"
    
    # ===== 条件4：股价缓慢上升（近3日趋势向上但幅度小） =====
    if n < 3:
        return False, 0, "", "", "badge-hold"
    recent_3 = closes[-3:]
    x_3 = np.arange(3)
    try:
        slope_3 = np.polyfit(x_3, recent_3, 1)[0]
    except:
        return False, 0, "", "", "badge-hold"
    
    if slope_3 < -0.001 * np.mean(recent_3):  # v5.2 relaxed
        return False, 0, "", "", "badge-hold"
    
    avg_price = np.mean(recent_3)
    relative_slope = slope_3 / avg_price if avg_price > 0 else 0
    if relative_slope > 0.02:  # 上升太快不算低吸
        return False, 0, "", "", "badge-hold"
    
    # ===== 全部通过，计算信号强度 =====
    # 信号强度基于：量能回升幅度、偏离均线距离、跌幅深度
    vol_increase_ratio = (volumes[-1] / volumes[-vol_rise_days-1] - 1) * 100 if volumes[-vol_rise_days-1] > 0 else 0
    
    # 偏离20日均线程度
    ma20 = np.mean(closes[-20:]) if n >= 20 else np.mean(closes)
    deviation_from_ma = (closes[-1] - ma20) / ma20 * 100 if ma20 > 0 else 0
    
    # 综合信号强度 (0-100)
    signal_strength = 50
    signal_strength += min(20, max(0, vol_increase_ratio * 3))  # 量能回升贡献
    signal_strength += min(15, max(0, abs(deviation_from_ma) * 2))  # 偏离均线贡献（越偏离底部越高）
    signal_strength += min(15, max(0, total_decline_pct))  # 跌幅深度贡献
    signal_strength = int(max(30, min(95, signal_strength)))
    
    # 信号标签
    if signal_strength >= 75:
        label = "强烈低吸"
        badge_class = "badge-lowbuy-strong"
    elif signal_strength >= 55:
        label = "适度低吸"
        badge_class = "badge-lowbuy-mild"
    else:
        label = "轻度低吸"
        badge_class = "badge-lowbuy-mild"
    
    summary = f"跌幅{total_decline_pct:.1f}% | 量能连升{vol_rise_days}日(+{vol_increase_ratio:.0f}%) | 昨日涨{yesterday_change*100:.1f}%"
    
    return True, signal_strength, label, summary, badge_class



def screen_low_buy_stocks(df_market, params=None):
    """
    低吸模型主筛选函数 (v6 - 完全独立评分体系)
    """
    # [DEBUG] 统计初始化 — 放最前面，确保任何提前返回都能保存
    _dbg = {"total": 0, "no_kline": 0, "no_new_low": 0, "reversal1": 0, "reversal2": 0, "reversal3": 0,
            "build_err": 0, "score_fail": 0, "decline": 0, "stabil": 0, "vol_rec": 0,
            "ma_sup": 0, "val": 0, "chip": 0, "fund": 0, "total_score": 0, "passed": 0,
            "decline_20d_fail": 0}
    
    if params is None:
        params = _get_lowbuy_params()
    
    # [DEBUG] 确认运行时参数
    _dbg["_params_debug"] = {k: params.get(k) for k in [
        'min_stabilization', 'min_volume_recovery', 'reversal_bottom_pct',
        'reversal_require_uptrend', 'min_total_score', '_params_version'
    ]}
    
    if df_market is None or len(df_market) == 0:
        st.session_state._lb_dbg = _dbg
        return pd.DataFrame()
    
    df = df_market.copy()
    
    # 获取辅助数据
    lhb_data = {}
    sector_data = {}
    stock_sector_map = {}
    try:
        lhb_data = _get_cached_dragon_tiger()
        sector_data, stock_sector_map = _get_cached_sector_data()
    except Exception:
        pass

    # 预计算行业PE分布（低吸S5维度用）
    _get_industry_pe_stats(df_market, stock_sector_map)
    
    # ===== 预筛 =====
    pre_db = [f"入参: decline_20d=[{params.get('decline_20d_low',-50)},{params.get('decline_20d_high',0)}] max_vol_ratio={params.get('max_vol_ratio',3)}"]
    pre_db.append(f"原始: {len(df)}只")
    
    if '涨跌幅' in df.columns:
        df = df[df['涨跌幅'] < 9.5]
        pre_db.append(f"涨幅<9.5%: {len(df)}只")
    if '最新价' in df.columns:
        df = df[(df['最新价'] >= 3) & (df['最新价'] <= 200)]
        pre_db.append(f"价格3-200: {len(df)}只")
    
    decline_low = params.get('decline_20d_low', -50)
    decline_high = params.get('decline_20d_high', 0)
    # 10日跌幅过滤延后到逐只K线循环中执行（避免追高模型fillna(0)假数据污染）
    # 此处只做轻量预筛：如果有真实20d数据且明显不符合，提前剔除
    if '涨跌幅_20d' in df.columns:
        df['涨跌幅_20d'] = pd.to_numeric(df['涨跌幅_20d'], errors='coerce')
        has_real = df['涨跌幅_20d'].notna() & (df['涨跌幅_20d'] != 0)  # 0=fillna填充，非真实数据
        pre_db.append(f"20d有真实数据: {has_real.sum()}/{len(df)}只")
        if has_real.any():
            # 仅对真实数据应用跌幅过滤；无真实数据的放过（在K线循环中再查）
            in_range = has_real & ((df['涨跌幅_20d'] >= decline_low) & (df['涨跌幅_20d'] <= decline_high))
            df = df[in_range | ~has_real]
            pre_db.append(f"20d跌幅[{decline_low},{decline_high}]: {in_range.sum()}只通过")
    
    max_vol_ratio = params.get('max_vol_ratio', 3.0)
    if '量比' in df.columns:
        df['量比'] = pd.to_numeric(df['量比'], errors='coerce').fillna(1.0)
        df = df[df['量比'] <= max_vol_ratio]
        pre_db.append(f"量比<={max_vol_ratio}: {len(df)}只")
    
    sample_n = min(len(df), 928)
    pre_db.append(f"预筛后: {len(df)}只 取前{sample_n}")
    _dbg["_pre"] = " | ".join(pre_db)
    
    if len(df) == 0:
        st.session_state._lb_dbg = _dbg
        return pd.DataFrame()
    
    df = df.head(sample_n)
    
    # ===== 逐只验证 =====
    results = []
    progress_bar = st.progress(0)
    status_text = st.empty()
    total = len(df)
    no_new_low_days = params.get('no_new_low_days', 1)
    
    for i, (idx, row) in enumerate(df.iterrows()):
        _dbg["total"] += 1
        if i % 20 == 0:
            status_text.text(f"\U0001f50d 低吸筛选中... ({i+1}/{total}) 已找到 {len(results)} 只")
        progress_bar.progress((i + 1) / total)
        
        code = str(row.get('代码', ''))
        kline = get_stock_kline(code, days=max(60, no_new_low_days + 15))
        if kline is None or len(kline) < 10:
            _dbg["no_kline"] += 1; continue
        
        closes = kline['收盘'].values.astype(float)
        n = len(closes)
        
        # 10日跌幅预检（使用K线真实数据，避免fillna(0)假数据）
        if n >= 12:
            kline_chg_10d = (closes[-1] / closes[-12] - 1) * 100
            if kline_chg_10d < decline_low or kline_chg_10d > decline_high:
                _dbg["decline_20d_fail"] += 1; continue
        else:
            kline_chg_10d = None
        
        # K线条件: 不创新低
        if n >= no_new_low_days:
            recent_lows = closes[-no_new_low_days:]
            if closes[-1] <= min(recent_lows) * 0.998:
                _dbg["no_new_low"] += 1; continue
        else:
            _dbg["no_new_low"] += 1; continue
        
        # ===== 硬性反转确认: 底部已过 + 近端止跌回升 =====
        if n < 15:
            _dbg["reversal1"] += 1; continue
        # 找到近15日最低点位置
        low_idx_15 = np.argmin(closes[-15:])
        low_pos = 15 - low_idx_15  # 最低点距今多少天（1=昨天, >=15表示超过15天）
        if low_pos < 1:
            _dbg["reversal1"] += 1; continue  # 最低点就是今天，还没确认反转
        # 近3日均线必须高于最低点一定百分比（确认已脱离底部）
        avg_close_3d = np.mean(closes[-3:])
        min_close_15d = np.min(closes[-15:])
        rev_bottom_pct = params.get('reversal_bottom_pct', 0.2) / 100.0
        if min_close_15d > 0 and (avg_close_3d - min_close_15d) / min_close_15d < rev_bottom_pct:
            _dbg["reversal2"] += 1; continue
        # MA3今日 > MA3前日: 今日均价高于前日（确认企稳，可通过参数关闭）
        if params.get('reversal_require_uptrend', False):
            if np.mean(closes[-3:]) <= np.mean(closes[-4:-1]):
                _dbg["reversal3"] += 1; continue  # MA3未走强
        
        # 机构数据
        lhb_info = lhb_data.get(code, {})
        inst_net_3d = lhb_info.get('inst_net_buy_3d', 0)
        
        # 构建评分数据
        try:
            stock_data = _build_stock_data(row, kline, lhb_data, sector_data, stock_sector_map)
        except Exception:
            _dbg["build_err"] += 1; continue
        
        # 低吸专用评分
        try:
            score_result = calculate_lowbuy_score(stock_data, kline, params)
        except Exception:
            _dbg["score_fail"] += 1; continue
        
        if not score_result.get('pass', False):
            _dbg["score_fail"] += 1; continue
        
        # 六维→七维门槛 — 用原始分(_raw), 而非归一化值(0-100)
        _raw = score_result.get('_raw', {})
        s_decline = _raw.get('s1', 0)     # raw 0-30
        s_stabil = _raw.get('s2', 0)      # raw 0-25
        s_vol_rec = _raw.get('s3', 0)     # raw 0-20
        s_ma = _raw.get('s4', 0)          # raw 0-15
        s_val = _raw.get('s5', 0)         # raw 0-10
        s_chip = _raw.get('s6', 0)        # raw 0-10
        s_fund = _raw.get('s7', 0)        # raw 0-10
        s_total = score_result.get('综合评分', 0)  # 0-100 归一化总分
        
        if s_decline < params.get('min_decline_depth', 8):
            _dbg["decline"] += 1; continue
        if s_stabil < params.get('min_stabilization', 10):
            _dbg["stabil"] += 1; continue
        if s_vol_rec < params.get('min_volume_recovery', 8):
            _dbg["vol_rec"] += 1; continue
        if s_ma < params.get('min_ma_support', 4):
            _dbg["ma_sup"] += 1; continue
        if s_val < params.get('min_valuation_attr', 4):
            _dbg["val"] += 1; continue
        if s_chip < params.get('min_chip_settle', 3):
            _dbg["chip"] += 1; continue
        if s_fund < params.get('min_fund_flow', 2):
            _dbg.setdefault("fund", 0); _dbg["fund"] += 1; continue
        if s_total < params.get('min_total_score', 20):
            _dbg["total_score"] += 1; continue
        
        _dbg["passed"] += 1
        
        # 信号分级
        # 强烈低吸：≥55分 + 机构净买 + 估值原始分≥7（二季报窗口强化基本面约束）
        if s_total >= 55 and inst_net_3d > 0 and s_val >= 7:
            label = "强烈低吸"
            badge_class = "badge-lowbuy-strong"
        elif s_total >= 40:
            label = "标准低吸"
            badge_class = "badge-lowbuy-mild"
        else:
            label = "谨慎低吸"
            badge_class = "badge-lowbuy-mild"
        
        chg_10d = kline_chg_10d if kline_chg_10d is not None else float(row.get('涨跌幅_20d', 0) or 0)
        sec_name = _classify_stock_sector(code, row.get('名称', ''), sector_data, stock_sector_map) or '其他'
        summary = f"10日跌{chg_10d:.1f}% | 综合{s_total}分 | 机构3日净买{inst_net_3d:.0f}万"
        # 谨慎低吸 + 估值维度弱 → 单只仓位上限压缩至2成
        if label == "谨慎低吸" and s_val <= 5:
            summary += " | 限仓2成"
        
        results.append({
            "代码": code,
            "名称": row.get('名称', ''),
            "板块": sec_name,
            "最新价": row.get('最新价', 0),
            "涨跌幅": row.get('涨跌幅', 0),
            "信号强度": s_total,
            "信号": label,
            "信号类": badge_class,
            "分析摘要": summary,
            "综合评分": s_total,
            "下跌幅度": s_decline,
            "企稳信号": s_stabil,
            "量能恢复": s_vol_rec,
            "均线支撑": s_ma,
            "估值吸引": s_val,
            "筹码沉淀": s_chip,
            "主力资金": s_fund,
            "_raw": {"s1": s_decline, "s2": s_stabil, "s3": s_vol_rec, 
                     "s4": s_ma, "s5": s_val, "s6": s_chip, "s7": s_fund,
                     "inst_net_3d": inst_net_3d, "chg_10d": chg_10d, "sec_name": sec_name},
            "龙虎榜": inst_net_3d,
            "换手率": row.get('换手率', 0),
            "量比": row.get('量比', 0),
            "成交额": row.get('成交额', 0),
            "20日涨幅": f"{chg_10d:+.1f}%",
        })
        
        if len(results) >= 50:
            break
    
    progress_bar.empty()
    status_text.empty()
    
    # [DEBUG] 保存统计到 session_state，在外层显示
    st.session_state._lb_dbg = _dbg
    
    if not results:
        return pd.DataFrame()
    
    # ===== P1: 百分位归一化（七维分别做） =====
    n_res = len(results)
    if n_res > 1:
        def _pct_norm_vals(vals, full_score):
            arr = np.array(vals, dtype=float)
            if np.all(arr == arr[0]):
                return np.full_like(arr, full_score * 0.5)
            ranks = np.argsort(np.argsort(arr)) + 1
            return (ranks / n_res * full_score).clip(0, full_score)
        
        s1_arr = _pct_norm_vals([r["_raw"]["s1"] for r in results], 30)
        s2_arr = _pct_norm_vals([r["_raw"]["s2"] for r in results], 25)
        s3_arr = _pct_norm_vals([r["_raw"]["s3"] for r in results], 20)
        s4_arr = _pct_norm_vals([r["_raw"]["s4"] for r in results], 15)
        s5_arr = _pct_norm_vals([r["_raw"]["s5"] for r in results], 10)
        s6_arr = _pct_norm_vals([r["_raw"]["s6"] for r in results], 10)
        s7_arr = _pct_norm_vals([r["_raw"]["s7"] for r in results], 10)
        
        weights = st.session_state.get("lowbuy_weights", DEFAULT_LOWBUY_WEIGHTS)
        w1, w2, w3 = weights.get("下跌幅度", 23), weights.get("企稳信号", 18), weights.get("量能恢复", 14)
        w4, w5, w6 = weights.get("均线支撑", 14), weights.get("估值吸引", 14), weights.get("筹码沉淀", 9)
        w7 = weights.get("主力资金", 8)
        
        for i, r in enumerate(results):
            n1 = s1_arr[i] / 30 * 100
            n2 = s2_arr[i] / 25 * 100
            n3 = s3_arr[i] / 20 * 100
            n4 = s4_arr[i] / 15 * 100
            n5 = s5_arr[i] / 10 * 100
            n6 = s6_arr[i] / 10 * 100
            n7 = s7_arr[i] / 10 * 100
            total_n = int(n1 * w1/100 + n2 * w2/100 + n3 * w3/100 + n4 * w4/100 + n5 * w5/100 + n6 * w6/100 + n7 * w7/100)
            
            r["综合评分"] = total_n
            r["信号强度"] = total_n
            r["下跌幅度"] = int(s1_arr[i])
            r["企稳信号"] = int(s2_arr[i])
            r["量能恢复"] = int(s3_arr[i])
            r["均线支撑"] = int(s4_arr[i])
            r["估值吸引"] = int(s5_arr[i])
            r["筹码沉淀"] = int(s6_arr[i])
            r["主力资金"] = int(s7_arr[i])
            
            # 重新做信号分级（使用归一化后的综合评分）
            inst_net_3d = r["_raw"]["inst_net_3d"]
            if total_n >= 55 and inst_net_3d > 0:
                r["信号"] = "强烈低吸"
                r["信号类"] = "badge-lowbuy-strong"
            elif total_n >= 40:
                r["信号"] = "标准低吸"
                r["信号类"] = "badge-lowbuy-mild"
            else:
                r["信号"] = "谨慎低吸"
                r["信号类"] = "badge-lowbuy-mild"
            r["分析摘要"] = f"10日跌{r['_raw']['chg_10d']:.1f}% | 综合{total_n}分 | 机构3日净买{inst_net_3d:.0f}万"
    
    result_df = pd.DataFrame(results)
    result_df = result_df.sort_values('综合评分', ascending=False).head(50)
    
    result_df['涨跌幅_显示'] = result_df['涨跌幅'].apply(lambda x: f"{x:+.2f}%")
    result_df['换手率_显示'] = result_df['换手率'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A")
    result_df['量比_显示'] = result_df['量比'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
    
    return result_df.reset_index(drop=True)



def _get_sector(code):
    """根据代码判断板块"""
    if code.startswith('6'):
        return '主板'
    elif code.startswith('000') or code.startswith('001'):
        return '主板'
    elif code.startswith('002'):
        return '中小板'
    elif code.startswith('300'):
        return '创业板'
    elif code.startswith('688'):
        return '科创板'
    elif code.startswith('8') or code.startswith('4'):
        return '北交所'
    return '其他'


# ================================================================
#              追高模型：股票池生成（基于真实数据）
# ================================================================

@st.cache_data(ttl=1800)
def get_stock_pool_chase_high(sample_source="全市场A股"):
    """追高模型股票池"""
    df_raw = fetch_all_a_stocks()
    if df_raw is None:
        return pd.DataFrame()
    
    # 🌟 样本来源过滤
    if sample_source == "热门板块":
        hot_codes = get_hot_concept_stocks(6)
        if hot_codes:
            df_raw['代码'] = df_raw['代码'].astype(str).str.zfill(6)
            df_raw = df_raw[df_raw['代码'].isin(hot_codes)]
        else:
            st.warning("未能获取热门板块数据，回退为全市场扫描")
    elif sample_source == "量价":
        vp_codes = get_volprice_sectors(6)
        if vp_codes:
            df_raw['代码'] = df_raw['代码'].astype(str).str.zfill(6)
            df_raw = df_raw[df_raw['代码'].isin(vp_codes)]
        else:
            st.warning("未能获取量价反转板块数据，回退为全市场扫描")
    
    st.session_state.data_status = 'normal'
    st.session_state.last_update_time = datetime.now()
    
    df = preprocess_stock_data(df_raw)
    if df is None or len(df) == 0:
        return pd.DataFrame()
    
    st.session_state.raw_stock_data = df
    
    # 批量计算涨跌幅
    df['涨跌幅'] = pd.to_numeric(df['涨跌幅'], errors='coerce').fillna(0)
    df = df[df['涨跌幅'] < 20]  # v5: pre-filter daily gain < 20%

    # 防御性补列：pytdx 不提供量比/换手率/市盈率/市净率，akshare 也不保证一定有
    for _col, _default in [('量比', 1), ('换手率', 0), ('市盈率-动态', None), ('市净率', None)]:
        if _col in df.columns:
            if _default is None:
                df[_col] = pd.to_numeric(df[_col], errors='coerce')  # 保留 NaN，显示为 N/A
            else:
                df[_col] = pd.to_numeric(df[_col], errors='coerce').fillna(_default)
        else:
            df[_col] = float(_default) if _default is not None else np.nan
    
    df['初筛分'] = (
        df['涨跌幅'].clip(-10, 10) * 3 +
        df['量比'].clip(0, 4) * 5 +
        (100 - df['换手率'].clip(0, 15)) * 0.5
    )
    df = df.sort_values('初筛分', ascending=False).reset_index(drop=True)
    
    # 计算历史涨跌幅（前500只）
    top500_codes = df.head(200)['代码'].tolist()  # 优化：200只足够覆盖Top30展示
    price_changes = {}
    status_text = st.empty()
    progress_bar = st.progress(0)
    
    def _fetch_kline_pc(code):
        kline = get_stock_kline(code, days=80)
        if kline is not None and len(kline) >= 4:
            pc = {
                3: round((kline['收盘'].iloc[-1] / kline['收盘'].iloc[-4] - 1) * 100, 2) if len(kline) >= 4 else None,
                5: round((kline['收盘'].iloc[-1] / kline['收盘'].iloc[-6] - 1) * 100, 2) if len(kline) >= 6 else None,
                10: round((kline['收盘'].iloc[-1] / kline['收盘'].iloc[-11] - 1) * 100, 2) if len(kline) >= 11 else None,
                20: round((kline['收盘'].iloc[-1] / kline['收盘'].iloc[-21] - 1) * 100, 2) if len(kline) >= 21 else None,
            }
            pc['_kline'] = kline
            return code, pc
        return code, None

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(_fetch_kline_pc, code) for code in top500_codes]
        done_count = 0
        for future in as_completed(futures):
            done_count += 1
            if done_count % 50 == 0:
                status_text.text(f"📊 计算涨跌幅数据... ({done_count}/{len(top500_codes)})")
            progress_bar.progress(done_count / len(top500_codes))
            try:
                code, pc = future.result(timeout=15)
                if pc is not None:
                    price_changes[code] = pc
            except Exception:
                pass
    
    progress_bar.empty()
    status_text.empty()
    
    df['涨跌幅_3d'] = df['代码'].map(lambda x: price_changes.get(x, {}).get(3, None))
    df['涨跌幅_5d'] = df['代码'].map(lambda x: price_changes.get(x, {}).get(5, None))
    df['涨跌幅_10d'] = df['代码'].map(lambda x: price_changes.get(x, {}).get(10, None))
    df['涨跌幅_20d'] = df['代码'].map(lambda x: price_changes.get(x, {}).get(20, None))
    df['涨跌幅_3d'] = df['涨跌幅_3d'].fillna(df['涨跌幅'] * 2)
    df['涨跌幅_5d'] = df['涨跌幅_5d'].fillna(df['涨跌幅'] * 3)
    df['涨跌幅_10d'] = df['涨跌幅_10d'].fillna(df['涨跌幅'] * 5)
    df['涨跌幅_20d'] = df['涨跌幅_20d'].fillna(0)  # 仅初筛排序用，显示前会重新格式化为N/A

    # 但 fillna(0) 导致显示 "+0.0%" 而非 "N/A"：覆盖显示列为真正N/A
    df['涨跌幅_20d_display'] = df['代码'].map(lambda x: price_changes.get(x, {}).get(20, None))

    df['_kline_cache'] = df['代码'].map(lambda x: price_changes.get(x, {}).get('_kline'))
    
    # 计算RSI（前200只）
    top200_codes = df.head(100)['代码'].tolist()  # 优化：100只足够
    rsi_dict = batch_calculate_rsi(top200_codes, max_count=100) or {}
    df['RSI'] = df['代码'].astype(str).map(lambda x: rsi_dict.get(x, None))  # 不再 fillna(50)
    
    # ====== v3: 获取龙虎榜+板块数据（使用模块级缓存） ======
    _st2 = st.empty()
    _st2.text("🐉 获取龙虎榜+板块数据...")
    lhb_data = _get_cached_dragon_tiger()
    sector_data, stock_sector_map = _get_cached_sector_data()
    _st2.empty()

    # ====== v3八维评分 ======
    score_results = []
    for idx, row in df.iterrows():
        sd = _build_stock_data(row, row.get('_kline_cache'), lhb_data, sector_data, stock_sector_map)
        result = calculate_v3_total_score(sd)
        result['代码'] = row['代码']
        result['position_msg'] = result.get('position_msg', '')
        score_results.append(result)

    score_df = pd.DataFrame(score_results)

    # ====== P1: 六维百分位归一化（追高模型）— v2 N/A感知版 ======
    if len(score_df) > 1:
        weights = st.session_state.get('weights', DEFAULT_WEIGHTS)
        
        # N/A感知百分位归一化：-1值不参与排名，返回0分（不污染排名分布）
        def _pct_norm_na(arr_raw):
            arr = np.array(arr_raw, dtype=float)
            valid = arr >= 0
            result = np.zeros_like(arr)
            if not valid.any():
                return result
            valid_arr = arr[valid]
            if np.std(valid_arr) == 0:
                result[valid] = 50.0
            else:
                ranks = np.argsort(np.argsort(valid_arr)) + 1
                result[valid] = ranks / len(valid_arr) * 100
            return result
        
        def _pct_norm_ch(series):
            """返回 0~100 的百分位得分"""
            arr = np.array(series.values, dtype=float)
            if np.std(arr) == 0:
                return np.full_like(arr, 50.0)
            ranks = np.argsort(np.argsort(arr)) + 1
            return ranks / len(arr) * 100
        
        def _safe_arr(arr):
            return np.where(arr < 0, 0, arr)
        
        # 1-3: 趋势/动量/板块共振 — 现已可能返回-1(N/A)，使用N/A感知归一化
        s1_arr = np.array([r['_raw'].get('s1', 0) for r in score_results], dtype=float)
        s2_arr = np.array([r['_raw'].get('s2', 0) for r in score_results], dtype=float)
        s3_arr = np.array([r['_raw'].get('s3', 0) for r in score_results], dtype=float)
        df['趋势得分_norm'] = _pct_norm_na(s1_arr)
        df['动量得分_norm'] = _pct_norm_na(s2_arr)
        df['板块共振_norm'] = _pct_norm_na(s3_arr)

        # s4/s5: 北向资金/机构净买 — 可能返回-1，分别N/A感知归一化后按权重合并
        s4_raw = np.array([r['_raw'].get('s4', 0) for r in score_results], dtype=float)
        s5_raw = np.array([r['_raw'].get('s5', 0) for r in score_results], dtype=float)
        s4_norm = _pct_norm_na(s4_raw)
        s5_norm = _pct_norm_na(s5_raw)
        # 资金得分展示：合并北向+机构（N/A按0处理）
        s45_arr = _safe_arr(s4_raw) + _safe_arr(s5_raw)
        df['资金得分_norm'] = _pct_norm_ch(pd.Series(s45_arr))
        
        # 5-6: 量价/情绪 — 可能返回-1，使用N/A感知归一化
        s7_raw = np.array([r['_raw'].get('s7', 0) for r in score_results], dtype=float)
        s10_raw = np.array([r['_raw'].get('s10', 0) for r in score_results], dtype=float)
        df['量价得分_norm'] = _pct_norm_na(s7_raw)
        df['情绪得分_norm'] = _pct_norm_na(s10_raw)
        
        # 非归一化维度（估值、筹码、板块资金热度）— N/A值-1按0处理
        s8_raw = np.array([r['_raw'].get('s8', 0) for r in score_results], dtype=float)
        s9_raw = np.array([r['_raw'].get('s9', 0) for r in score_results], dtype=float)
        s6_raw = np.array([r['_raw'].get('s6', 0) for r in score_results], dtype=float)
        s8_arr = _safe_arr(s8_raw)
        s9_arr = _safe_arr(s9_raw)
        s6_arr = _safe_arr(s6_raw)
        s6_norm = s6_arr / 10 * 100
        s8_norm = s8_arr / 10 * 100
        s9_norm = s9_arr / 7 * 100
        
        # ---- 动态权重重分配：N/A维度剔除，按剩余有效维度加权归一 ----
        W_TREND = weights.get("趋势结构", 15)
        W_MOM = weights.get("动量强度", 18)
        W_SEC = weights.get("板块共振", 8)
        W_NORTH = weights.get("北向资金", 15)
        W_INST = weights.get("机构净买", 10)
        W_VOL = weights.get("量价配合", 14)
        W_SENT = weights.get("情绪热度", 6)
        W_FUND = weights.get("板块资金热度", 5)
        W_VAL = weights.get("估值安全", 3)
        W_CHIP = weights.get("筹码稳定", 6)

        na_s4 = s4_raw < 0
        na_s5 = s5_raw < 0
        na_s6 = s6_raw < 0
        na_s7 = s7_raw < 0
        na_s8 = s8_raw < 0
        na_s9 = s9_raw < 0
        na_s10 = s10_raw < 0

        total_arr = np.zeros(len(score_results))
        for i in range(len(score_results)):
            valid_w = W_TREND + W_MOM + W_SEC  # s1/s2/s3 始终有效
            wsum = (df['趋势得分_norm'].iloc[i] * W_TREND +
                    df['动量得分_norm'].iloc[i] * W_MOM +
                    df['板块共振_norm'].iloc[i] * W_SEC)
            if not na_s4[i]:
                valid_w += W_NORTH; wsum += s4_norm[i] * W_NORTH
            if not na_s5[i]:
                valid_w += W_INST; wsum += s5_norm[i] * W_INST
            if not na_s7[i]:
                valid_w += W_VOL; wsum += df['量价得分_norm'].iloc[i] * W_VOL
            if not na_s10[i]:
                valid_w += W_SENT; wsum += df['情绪得分_norm'].iloc[i] * W_SENT
            if not na_s6[i]:
                valid_w += W_FUND; wsum += s6_norm[i] * W_FUND
            if not na_s8[i]:
                valid_w += W_VAL; wsum += s8_norm[i] * W_VAL
            if not na_s9[i]:
                valid_w += W_CHIP; wsum += s9_norm[i] * W_CHIP
            total_arr[i] = wsum / valid_w if valid_w > 0 else 0
        
        df['综合评分'] = total_arr.astype(int)
        # 更新结果列表中的分数 + 统一 position_msg（基于P1归一化后的综合评分）
        for i, r in enumerate(score_results):
            new_score = int(df['综合评分'].iloc[i])
            r['综合评分'] = new_score
            # position_msg 与综合评分保持同一口径
            if new_score >= 85:
                r['position_msg'] = "主线龙头，重仓60%~80%"
            elif new_score >= 70:
                r['position_msg'] = "支线趋势，轻仓20%~40%"
            else:
                r['position_msg'] = "分数不足70，放弃不参与"
            score_results[i] = r
        score_df['综合评分'] = df['综合评分'].values
        # 同步 position_msg 到 score_df
        score_df['position_msg'] = [r.get('position_msg', '') for r in score_results]
    # ====== P1 归一化结束 ======

    for col in ['趋势结构', '动量强度', '板块共振', '北向资金', '机构净买', '板块资金热度',
                '量价配合', '估值安全', '筹码稳定', '情绪热度', '综合评分', 'position_msg']:
        if col in score_df.columns:
            df[col] = score_df[col]
    
    # 信号分类
    def classify(row):
        # P0修复: 追高模型信号由综合评分驱动，不再使用低吸模型的 _classify_signal
        # 与 position_msg 保持同一口径，避免 93分→"观望等待" 的矛盾
        score = row['综合评分']
        if score >= 85:
            return ("强势买入", "badge-strong", "积极关注")
        elif score >= 70:
            return ("逢低吸纳", "badge-attention", "分批布局")
        elif score >= 55:
            return ("观望等待", "badge-hold", "等待信号")
        else:
            return ("建议回避", "badge-watch", "谨慎对待")
    
    signals = df.apply(classify, axis=1)
    df['信号'] = [s[0] for s in signals]
    df['信号类'] = [s[1] for s in signals]
    df['建议'] = [s[2] for s in signals]
    
    # 格式化
    df['3日涨幅'] = df['涨跌幅_3d'].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "N/A")
    df['5日涨幅'] = df['涨跌幅_5d'].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "N/A")
    df['10日涨幅'] = df['涨跌幅_10d'].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "N/A")
    df['20日涨幅'] = df['涨跌幅_20d_display'].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "N/A")
    df['换手率_显示'] = df['换手率'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A")
    df['PE_显示'] = df['市盈率-动态'].apply(lambda x: f"{x:.1f}" if pd.notna(x) and x > 0 else "N/A")
    df['PB_显示'] = df['市净率'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
    df['RSI_显示'] = df['RSI'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "N/A")
    df['量比_显示'] = df['量比'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
    # 🔧 用 pytdx 概念板块映射替代代码前缀分类
    _, sector_map = fetch_sector_board_v3()
    if sector_map:
        import re
        valid_map = {k: v for k, v in sector_map.items() if re.search(r'[\u4e00-\u9fff]', v)}
        df['板块'] = df['代码'].map(valid_map).fillna('')
    else:
        df['板块'] = df['代码'].apply(_get_sector)
    
    output_cols = [
        '代码', '名称', '板块', '综合评分', '趋势结构', '动量强度', '板块共振',
        '北向资金', '机构净买', '板块资金热度', '量价配合', '估值安全', '筹码稳定', '情绪热度',
        '3日涨幅', '5日涨幅', '10日涨幅', '20日涨幅',
        '换手率_显示', '量比_显示', 'RSI_显示', 'PE_显示', 'PB_显示',
        '信号', '信号类', '建议', '最新价', '涨跌幅', 'position_msg'
    ]
    output_cols = [c for c in output_cols if c in df.columns]
    # N/A维度渲染为"N/A"（_raw 中值为-1的维度标记为数据不可用）
    _dim_na_map = {
        '趋势结构': 's1', '动量强度': 's2', '板块共振': 's3',
        '北向资金': 's4', '机构净买': 's5', '板块资金热度': 's6',
        '量价配合': 's7', '估值安全': 's8', '情绪热度': 's10'
    }
    for dim_col, raw_key in _dim_na_map.items():
        if dim_col in df.columns:
            df[dim_col] = df[dim_col].astype(object)  # 先转为object以容纳"N/A"字符串
            for i, r in enumerate(score_results):
                if r['_raw'].get(raw_key, 0) < 0:
                    df.loc[i, dim_col] = "N/A"

    if len(df) == 0 or '信号' not in df.columns:
        return pd.DataFrame(columns=['代码','名称','板块','综合评分','信号','趋势结构','动量强度','板块共振','北向资金','机构净买','板块资金热度','量价配合','估值安全','筹码稳定','情绪热度'])
    return df[output_cols].sort_values('综合评分', ascending=False).reset_index(drop=True)


def get_stock_pool():
    """获取股票池（根据当前模型）"""
    model = st.session_state.get('current_model', 'chase_high')
    sample_source = st.session_state.get('chase_sample_source', '全市场A股')
    if model == 'chase_high':
        return get_stock_pool_chase_high(sample_source)
    else:
        # 低吸模型也返回追高池用于自选股和详情查看
        return get_stock_pool_chase_high(sample_source)


def get_simulated_stock_pool():
    """模拟数据（降级用）"""
    today = datetime.now().date()
    date_seed = int(today.strftime('%Y%m%d'))
    rng = np.random.RandomState(date_seed % 9999)
    
    sectors = ['半导体/芯片', 'AI/人工智能', '新能源/锂电', '医药/医疗', '消费电子',
               '软件/云计算', '金融/证券', '化工/材料', '机械设备', '军工/航空']
    
    data = []
    for i in range(200):
        code = f"{i+1:06d}"
        rng_i = np.random.RandomState(hash(code) % 10000 + date_seed)
        sector = sectors[i % len(sectors)]
        
        chg_5d = round(rng_i.randn() * 4 + 1, 1)
        chg_20d = round(chg_5d * 3 + rng_i.randn() * 8, 1)
        turnover = round(max(0.5, 2.5 + rng_i.randn() * 2), 1)
        vol_ratio = round(max(0.3, min(4.0, 1.0 + rng_i.randn() * 0.8)), 2)
        pe = round(max(10, min(80, 30 + rng_i.randn() * 20)), 1)
        pb = round(max(0.5, min(8, pe / 15 + rng_i.randn())), 2)
        rsi = round(max(25, min(80, 50 + rng_i.randn() * 20)), 1)
        
        momentum = int(max(30, min(90, 50 + chg_5d * 4 + rng_i.randn() * 10)))
        vol_price = int(max(30, min(90, 50 + (vol_ratio - 1) * 15 + rng_i.randn() * 10)))
        safety = int(max(30, min(90, 50 + (40 - pe) / 2 if 15 <= pe <= 60 else 30 + rng_i.randn() * 10)))
        chip = int(max(30, min(90, 50 + (5 - turnover) * 3 + rng_i.randn() * 10)))
        sentiment = int(max(25, min(90, 50 + (rsi - 50) * 0.8 + rng_i.randn() * 10)))
        trend = int(max(25, min(90, 50 + chg_20d * 2 + rng_i.randn() * 15)))
        lhb_s = int(max(25, min(90, 50 + rng_i.randn() * 20)))
        sec_s = int(max(25, min(90, 50 + rng_i.randn() * 15)))

        weights = st.session_state.get('weights', DEFAULT_WEIGHTS)
        total = int(
            momentum * weights.get("动量强度", 18)/100 + vol_price * weights.get("量价配合", 14)/100 +
            safety * weights.get("估值安全", 3)/100 + chip * weights.get("筹码稳定", 6)/100 +
            sentiment * weights.get("情绪热度", 6)/100 + trend * weights.get("趋势结构", 15)/100 +
            lhb_s * weights.get("北向资金", 15)/100 + lhb_s * weights.get("机构净买", 10)/100 +
            lhb_s * weights.get("板块资金热度", 5)/100 + sec_s * weights.get("板块共振", 8)/100
        )
        
        signal, sig_cls, advice = _classify_signal(total, chg_5d, rsi, safety, momentum)
        
        data.append({
            "代码": code, "名称": f"模拟股票{i+1:03d}", "板块": sector,
            "综合评分": total, "动量强度": momentum, "量价配合": vol_price,
            "估值安全": safety, "筹码稳定": chip, "情绪热度": sentiment,
            "趋势结构": trend, "北向资金": lhb_s, "机构净买": lhb_s,
            "板块资金热度": lhb_s, "板块共振": sec_s,
            "3日涨幅": f"{round(chg_5d*0.6,1):+.1f}%", "5日涨幅": f"{chg_5d:+.1f}%",
            "10日涨幅": f"{round(chg_5d*1.8,1):+.1f}%", "20日涨幅": f"{chg_20d:+.1f}%",
            "换手率_显示": f"{turnover:.1f}%", "量比_显示": vol_ratio,
            "RSI_显示": f"{rsi:.1f}", "PE_显示": f"{pe:.1f}", "PB_显示": f"{pb:.2f}",
            "信号": signal, "信号类": sig_cls, "建议": advice,
            "最新价": round(10 + rng.rand() * 50, 2), "涨跌幅": chg_5d
        })
    
    return pd.DataFrame(data).sort_values("综合评分", ascending=False).reset_index(drop=True)


def get_stock_detail(code):
    df = get_stock_pool()
    r = df[df["代码"] == code]
    return None if len(r) == 0 else r.iloc[0].to_dict()


def _find_stock_row(code, silent=False):
    """个股股票查找，带多级兜底。
    
    与 fetch_all_a_stocks() 不同：本函数绕过模块级缓存，在 fetch_all_a_stocks()
    返回缓存的残缺数据时，直接调用 akshare 获取全量行情来匹配。
    
    Args:
        code: 6位股票代码字符串
        silent: True 时不打印 st.warning，仅返回结果
    
    Returns:
        (row_dict_or_None, df_raw_or_None)
    """
    code = str(code).strip().zfill(6)
    
    # 第一级：fetch_all_a_stocks（有可能是过期缓存）
    df_raw = fetch_all_a_stocks()
    if df_raw is not None and len(df_raw) > 0:
        df_raw['代码'] = df_raw['代码'].astype(str).str.zfill(6)
        match = df_raw[df_raw['代码'] == code]
        if len(match) > 0:
            return match.iloc[0].to_dict(), df_raw
    
    # 第二级：直接调用 akshare（绕过 pytdx 阻塞和模块级缓存）
    if AKSHARE_AVAILABLE:
        try:
            import threading
            _ak_result = [None]
            def _fetch():
                try:
                    _ak_result[0] = ak.stock_zh_a_spot_em()
                except Exception:
                    pass
            _t = threading.Thread(target=_fetch, daemon=True)
            _t.start()
            _t.join(timeout=10)
            df_em = _ak_result[0]
            if df_em is not None and len(df_em) > 0:
                df_em['代码'] = df_em['代码'].astype(str).str.zfill(6)
                match = df_em[df_em['代码'] == code]
                if len(match) > 0:
                    row = match.iloc[0]
                    result = {k: row[k] for k in row.index}
                    # 尝试更新模块级缓存
                    try:
                        if len(df_em) > 1000:
                            _app_cache['raw_market_data'] = df_em
                            _app_cache['raw_market_time'] = time.time()
                    except Exception:
                        pass
                    return result, df_em
        except Exception:
            pass
    
    # 第三级：从本地缓存文件反查（可能比模块级缓存的版本更新）
    try:
        cache_path = os.path.join(BASE_DIR, 'market_data_cache.json')
        if os.path.exists(cache_path):
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            if cached and 'data' in cached and len(cached['data']) > 0:
                df_cache = pd.DataFrame(cached['data'])
                df_cache['代码'] = df_cache['代码'].astype(str).str.zfill(6)
                match = df_cache[df_cache['代码'] == code]
                if len(match) > 0:
                    return match.iloc[0].to_dict(), df_cache
    except Exception:
        pass
    
    return None, None


def score_single_stock(code, model):
    """个股评分：对单只股票在指定模型下计算各维度分值和总分。
    
    Args:
        code: 6位股票代码字符串
        model: 'chase_high' | 'buy_low' | 'golden_cross' | 'resonance' | 'canslim' | 'dilemma_reversal'
    
    Returns:
        dict: {'success': bool, 'model': str, 'name': str, 'close': float, 'chg': float,
               'dims': {维度名: 分值}, 'total': int, 'pass': bool, 'max_score': int,
               以及可选的 'error', 'signal', 'position_msg', 'filter_msg'}
    """
    try:
        kline250 = get_stock_kline(code, days=250)
        kline60 = get_stock_kline(code, days=60) if model != 'resonance' else None
        
        row_dict, df_raw = _find_stock_row(code)
        if row_dict is None:
            return {"success": False, "error": f"未找到股票代码 {code}"}
        row = pd.Series(row_dict).copy()
        name = str(row.get('名称', ''))
        close_val = float(row.get('最新价', 0) or 0)
        chg_val = float(row.get('涨跌幅', 0) or 0)
        
        lhb_data = _get_cached_dragon_tiger()
        sector_data, stock_sector_map = _get_cached_sector_data()
        
        # 补列
        for _col, _default in [('量比', 1), ('换手率', 0), ('市盈率-动态', 0), ('市净率', 0)]:
            if _col not in row.index:
                row[_col] = _default
            else:
                row[_col] = float(row.get(_col, _default) or 0)
        
        # K线涨跌幅（有K线就用，否则填0）
        kline_for_pc = kline250 if kline250 is not None else kline60
        if kline_for_pc is not None and len(kline_for_pc) >= 4:
            c_col = '收盘' if '收盘' in kline_for_pc.columns else 'close'
            c_arr = kline_for_pc[c_col].values.astype(float)
            n_c = len(c_arr)
            row['涨跌幅_3d'] = round((c_arr[-1] / c_arr[max(0, n_c - 4)] - 1) * 100, 2) if n_c >= 4 else 0
            row['涨跌幅_5d'] = round((c_arr[-1] / c_arr[max(0, n_c - 6)] - 1) * 100, 2) if n_c >= 6 else 0
            row['涨跌幅_10d'] = round((c_arr[-1] / c_arr[max(0, n_c - 11)] - 1) * 100, 2) if n_c >= 11 else 0
            row['涨跌幅_20d'] = round((c_arr[-1] / c_arr[max(0, n_c - 21)] - 1) * 100, 2) if n_c >= 21 else 0
        else:
            row['涨跌幅_3d'] = 0; row['涨跌幅_5d'] = 0; row['涨跌幅_10d'] = 0; row['涨跌幅_20d'] = 0
        row['RSI'] = calculate_rsi(code) or 50
        
        stock_data = _build_stock_data(row, kline250, lhb_data, sector_data, stock_sector_map)
        
        if model == 'chase_high':
            result = calculate_v3_total_score(stock_data)
            dims = {k: result.get(k, 0) for k in [
                "趋势结构","动量强度","板块共振","北向资金","机构净买",
                "板块资金热度","量价配合","估值安全","筹码稳定","情绪热度"]}
            return {"success": True, "model": "chase_high", "name": name, "close": close_val, "chg": chg_val,
                    "dims": dims, "total": result.get("综合评分", 0), "pass": result.get("pass", False),
                    "position_msg": result.get("position_msg", ""), "filter_msg": result.get("filter_msg", ""),
                    "max_score": 100}
        
        elif model == 'buy_low':
            if kline60 is None or len(kline60) < 10:
                return {"success": False, "error": "K线数据不足（需要至少10日）"}
            params = _get_lowbuy_params()
            result = calculate_lowbuy_score(stock_data, kline60, params)
            dims = {k: result.get(k, 0) for k in [
                "下跌幅度","企稳信号","量能恢复","均线支撑","估值吸引","筹码沉淀","主力资金"]}
            return {"success": True, "model": "buy_low", "name": name, "close": close_val, "chg": chg_val,
                    "dims": dims, "total": result.get("综合评分", 0), "pass": result.get("pass", False),
                    "max_score": 100}
        
        elif model == 'golden_cross':
            if kline60 is None or len(kline60) < 10:
                return {"success": False, "error": "K线数据不足（需要至少10日）"}
            info = {'代码': code, '名称': name}
            result = calculate_golden_cross_score(info, kline60)
            dims = {k: result.get(k, 0) for k in [
                "下跌形态","K线止跌","均线拐头","量能确认","MACD反转","资金确认","板块确认"]}
            return {"success": True, "model": "golden_cross", "name": name, "close": close_val, "chg": chg_val,
                    "dims": dims, "total": result.get("综合评分", 0), "pass": result.get("pass", False),
                    "signal": result.get("信号", ""), "max_score": 100}
        
        elif model == 'resonance':
            quotes_df = df_raw
            if quotes_df is None:
                return {"success": False, "error": "无法获取行情数据"}
            quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)]
            rd = get_resonance_data(quotes_df)
            if rd and code in rd:
                scores = calculate_resonance_score(rd, quotes_df)
                if code in scores:
                    sc = scores[code]
                    dims = {
                        "资金流向": sc.get('money_flow', 0),
                        "DDE决策": sc.get('dde_proxy', 0),
                        "K线结构": sc.get('kline_structure', 0),
                        "板块热度": sc.get('sector_heat', 0),
                    }
                    return {"success": True, "model": "resonance", "name": name, "close": close_val, "chg": chg_val,
                            "dims": dims, "total": sc.get('total', 0), "pass": True, "max_score": 100}
            return {"success": False, "error": "共振数据未覆盖该股票（需全市场扫描）"}
        
        elif model == 'canslim':
            if kline250 is None or len(kline250) < 60:
                return {"success": False, "error": "K线数据不足（需要至少60日）"}
            fin_data = get_financial_data(code)
            spot = get_all_spot_data()
            market_cap_raw = (spot.get(code, {}).get("总市值", 0) or 0) * 1e8 if spot else 0
            turnover_raw = spot.get(code, {}).get("换手率", 0) if spot else 0
            ctx = {'fin': fin_data, 'market_cap': market_cap_raw, 'turnover_rate': turnover_raw, 'rps': 50}
            result = calculate_canslim_score(code, kline250, ctx)
            dims = {k: result.get(k, 0) for k in [
                "C_业绩增速","A_持续增长","N_新催化","S_中小盘","L_RPS","I_流动性","M_大势"]}
            return {"success": True, "model": "canslim", "name": name, "close": close_val, "chg": chg_val,
                    "dims": dims, "total": result.get("综合评分", 0), "pass": result.get("pass", False),
                    "max_score": 100}
        
        elif model == 'dilemma_reversal':
            if kline250 is None or len(kline250) < 60:
                return {"success": False, "error": "K线数据不足（需要至少60日）"}
            fin_data = get_financial_data(code)
            ctx = {'fin': fin_data}
            result = calculate_dilemma_reversal_score(code, kline250, ctx)
            dims = {k: result.get(k, 0) for k in ["L1_拐点","L2_反转","L3_安全垫","L4_技术资金"]}
            return {"success": True, "model": "dilemma_reversal", "name": name, "close": close_val, "chg": chg_val,
                    "dims": dims, "total": result.get("综合评分", 0), "pass": result.get("pass", False),
                    "max_score": 100}
        
        elif model == 'oversold_rebound':
            if kline250 is None or len(kline250) < 60:
                return {"success": False, "error": "K线数据不足（需要至少60日）"}
            ok, filter_msg = hard_filter_oversold_rebound(kline250, stock_data)
            if not ok:
                return {"success": True, "model": "oversold_rebound", "name": name,
                        "close": close_val, "chg": chg_val, "dims": {},
                        "total": 0, "pass": False, "filter_msg": filter_msg, "max_score": 110}
            result = calculate_oversold_rebound_score(kline250, stock_data)
            dims = {k: result.get(k, 0) for k in ["空间维度","情绪量能","择时确认","板块共振"]}
            return {"success": True, "model": "oversold_rebound", "name": name, "close": close_val, "chg": chg_val,
                    "dims": dims, "total": result.get("综合评分", 0), "pass": result.get("pass", False),
                    "max_score": 110}
        
    except Exception as e:
        import traceback
        return {"success": False, "error": str(e), "traceback": traceback.format_exc(), "model": model}
    return {"success": False, "error": f"未知模型: {model}"}


def get_top10_stocks():
    """追高模型：动态推荐数量（根据市场状态自动调整）"""
    today = datetime.now().strftime('%Y%m%d')
    weights = st.session_state.get('weights', DEFAULT_WEIGHTS)
    w_str = '_'.join(f"{k}{v}" for k, v in sorted(weights.items()))
    cache_key = f"top10_{today}_{w_str}"
    if st.session_state.get('top10_cache_key') != cache_key or st.session_state.top10_cache is None:
        df = get_stock_pool()
        dyn_n = calculate_dynamic_recommend_count()
        top_n = df.head(dyn_n).copy()
        st.session_state.top10_cache = top_n.to_dict('records')
        st.session_state.top10_cache_key = cache_key
        st.session_state._chase_dynamic_n = dyn_n
        # 保存到文件缓存
        save_cache_data({
            'chase_high_top10': st.session_state.top10_cache,
            'lowbuy_top5': st.session_state.get('lowbuy_cache'),
            'lowbuy_dbg': st.session_state.get('_lb_dbg', {}),
        })
    return st.session_state.top10_cache or []


def get_lowbuy_top5(df_market=None):
    """低吸模型：Top5"""
    if df_market is None:
        df_raw = fetch_all_a_stocks()
        if df_raw is None:
            return pd.DataFrame()
        df_market = preprocess_stock_data(df_raw)
    
    params = _get_lowbuy_params()
    return screen_low_buy_stocks(df_market, params)


# ================================================================
#                     回测数据引擎
# ================================================================

@st.cache_data(ttl=300, show_spinner=False)
def run_real_backtest_cached(start_date, end_date, top_n, model='chase_high',
                stop_loss=0.05, hard_take_profit=0.12, hard_take_profit_ceiling=0.35,
                moving_stop_gain=0.15, moving_stop_drawdown=0.07, cooling_period=5,
                resonance_threshold=40, min_z_score=1.0):
    """调用真实回测引擎并转换为UI格式"""
    _trades, _daily, _sels = backtest_engine.run_backtest(
        start_date, end_date, top_n=top_n, model=model,
        stop_loss=stop_loss, hard_take_profit=hard_take_profit,
        hard_take_profit_ceiling=hard_take_profit_ceiling,
        moving_stop_gain=moving_stop_gain, moving_stop_drawdown=moving_stop_drawdown,
        cooling_period=cooling_period, resonance_threshold=resonance_threshold,
        min_z_score=min_z_score
    )
    if _daily is None or len(_daily) == 0:
        return None

    # 加载股票代码→名称映射（多源合并，确保覆盖600xxx等遗漏代码）
    import json
    code_to_name = {}
    # 源1: 本地缓存文件
    try:
        cache_path = os.path.join(BASE_DIR, 'market_data_cache.json')
        with open(cache_path, 'r', encoding='utf-8') as f:
            cache = json.load(f)
        for item in cache.get('data', []):
            code_to_name[str(item['代码']).zfill(6)] = item['名称']
    except Exception:
        pass
    # 源2: pytdx get_security_list（覆盖全量，含600xxx）
    try:
        if tdx_available():
            tdx_names = tdx_get_full_name_map()
            for c, n in tdx_names.items():
                if c not in code_to_name:
                    code_to_name[c] = n
    except Exception:
        pass
    # 源3: 模块级缓存的实时行情数据
    try:
        raw_df = _app_cache.get('raw_market_data')
        if raw_df is not None and '代码' in raw_df.columns and '名称' in raw_df.columns:
            for _, r in raw_df.iterrows():
                c = str(r['代码']).zfill(6)
                if c not in code_to_name:
                    code_to_name[c] = r['名称']
    except Exception:
        pass
    # 源4: backtest_engine 内置名称加载器（TDX文件 + pytdx多服务器兜底）
    try:
        be_names = backtest_engine.load_stock_names()
        for c, n in be_names.items():
            if c not in code_to_name:
                code_to_name[c] = n
    except Exception:
        pass

    # 使用部署资金（总投入，含卖出后循环再投）
    pv_list = _daily["portfolio_value"].values.astype(float)
    cum_total_inv_list = _daily["total_invested"].values.astype(float)
    base_equity = 10_000_000

    max_deployed = max(cum_total_inv_list) if len(cum_total_inv_list) > 0 else 10000.0 * top_n
    if max_deployed <= 0:
        max_deployed = 1.0

    # 平均资金投入：仅统计有投入的交易日
    nonzero_mask = cum_total_inv_list > 0
    avg_deployed = float(np.mean(cum_total_inv_list[nonzero_mask])) if nonzero_mask.any() else 0.0

    # 累计收益率 = (组合净值 - 初始资金) / 峰值资金（策略实际需要的资金上限）
    cum_rets = np.zeros(len(pv_list))
    for i in range(len(pv_list)):
        cum_rets[i] = (float(pv_list[i]) - backtest_engine.INITIAL_CAPITAL) / max_deployed

    daily_rets = np.zeros(len(pv_list))
    daily_rets[1:] = cum_rets[1:] - cum_rets[:-1]
    if nonzero_mask.any():
        first_idx = nonzero_mask.argmax()
        if first_idx < len(pv_list):
            daily_rets[first_idx] = cum_rets[first_idx]

    bench_rets = _daily["benchmark_return"].values.astype(float)
    bench_cum = np.cumprod(1 + bench_rets)

    peak = np.maximum.accumulate(np.maximum(pv_list, 0))
    dds = np.where(peak > 0, (peak - pv_list) / peak, 0)

    total_pnl = 0.0
    win_trades = 0
    if _trades is not None and len(_trades) > 0:
        total_pnl = float(_trades["pnl"].sum())
        win_trades = int((_trades["pnl"] > 0).sum())

    records = []
    for j in range(len(_daily)):
        d = _daily.iloc[j]
        # 持仓明细注入名称
        raw_detail = str(d.get("positions_detail", ""))
        if raw_detail and code_to_name:
            parts = []
            for seg in raw_detail.split(", "):
                if "@" in seg:
                    code, price = seg.split("@", 1)
                    name = code_to_name.get(code, "")
                    parts.append(f"{code}{name}@{price}" if name else seg)
                else:
                    parts.append(seg)
            raw_detail = ", ".join(parts)

        # 持仓收盘价明细（注入名称）
        raw_close = str(d.get("positions_close", ""))
        if raw_close and code_to_name:
            parts = []
            for seg in raw_close.split(", "):
                if "@" in seg:
                    code, price = seg.split("@", 1)
                    name = code_to_name.get(code, "")
                    parts.append(f"{code}{name}@{price}" if name else seg)
                else:
                    parts.append(seg)
            raw_close = ", ".join(parts)

        # 今日买入 / 今日卖出 / 当前持仓 注入名称
        def _fmt_bought(bought_list):
            if not bought_list:
                return ""
            parts = []
            for item in bought_list:
                c = str(item.get("code", "")).zfill(6)
                n = code_to_name.get(c, "")
                parts.append(f"{c}{n}@{item.get('buy_price',0):.2f}" if n else f"{c}@{item.get('buy_price',0):.2f}")
            return ", ".join(parts)

        def _fmt_sold(sold_list):
            if not sold_list:
                return ""
            parts = []
            for item in sold_list:
                c = str(item.get("code", "")).zfill(6)
                n = code_to_name.get(c, "")
                core = f"{item.get('buy_price',0):.2f}→{item.get('sell_price',0):.2f}"
                parts.append(f"{c}{n} {core}" if n else f"{c} {core}")
            return ", ".join(parts)

        def _fmt_current(pos_detail_str):
            if not pos_detail_str or pos_detail_str == "空仓":
                return ""
            parts = []
            for seg in pos_detail_str.split(", "):
                if "@" in seg:
                    code, price = seg.split("@", 1)
                    c = code.strip()
                    n = code_to_name.get(c, "")
                    parts.append(f"{c}{n}@{price}" if n else seg)
                else:
                    parts.append(seg)
            return ", ".join(parts)

        bought_str = _fmt_bought(d.get("bought_today", []))
        sold_str = _fmt_sold(d.get("sold_today", []))
        current_str = _fmt_current(raw_detail)

        records.append({
            "日期": pd.to_datetime(d["date"]),
            "累计投入": round(cum_total_inv_list[j], 2),
            "组合净值": round(pv_list[j] if pv_list[j] > 0 else base_equity, 2),
            "基准净值": round(base_equity * bench_cum[j], 2),
            "日收益率": round(daily_rets[j], 6),
            "基准日收益率": round(bench_rets[j], 6),
            "累计收益率": round(cum_rets[j], 6),
            "基准累计收益": round(bench_cum[j] - 1, 6),
            "超额收益": round(cum_rets[j] - (bench_cum[j] - 1), 6),
            "最大回撤": round(float(dds[j]), 6),
            "基准最大回撤": 0,
            "持仓数": int(d["num_positions"]),
            "今日买入": bought_str,
            "今日卖出": sold_str,
            "当前持仓": current_str,
            "持仓明细": raw_detail,
            "持仓收盘价": raw_close,
        })

    _bt_df = pd.DataFrame(records)
    _bt_df.attrs["max_deployed"] = max_deployed
    _bt_df.attrs["avg_deployed"] = avg_deployed
    _bt_df.attrs["total_invested"] = float(_daily["total_invested"].iloc[-1]) if "total_invested" in _daily.columns else 0.0
    _bt_df.attrs["summary"] = (
        f"策略: 止损{backtest_engine.STOP_LOSS*100:.0f}%/硬止盈{backtest_engine.HARD_TAKE_PROFIT*100:.0f}%/移动止盈{backtest_engine.MOVING_STOP_GAIN*100:.0f}%→回撤{backtest_engine.MOVING_STOP_DRAWDOWN*100:.0f}% | "
        f"交易: {len(_trades) if _trades is not None else 0}笔 | "
        f"盈亏: {total_pnl:+,.0f}元 | "
        f"胜率: {win_trades}/{max(len(_trades),1)}笔"
    )
    # 生成股票收益清单
    PER_STOCK = 10000
    stock_list = []
    # FIX: 使用 (code, buy_date) 精确匹配卖出记录，避免同一股票二次买入后被误判为已卖出
    sold_pairs = set()

    if _trades is not None and len(_trades) > 0:
        for _, t in _trades.iterrows():
            code = t['code']
            buy_date = str(t.get('buy_date', ''))[:10]
            sold_pairs.add((code, buy_date))
            name = code_to_name.get(code, code)
            buy_p = float(t['buy_price'])
            sell_p = float(t['sell_price'])
            cost = backtest_engine.TRADING_COST
            pnl_rate = (sell_p * (1 - cost)) / (buy_p * (1 + cost)) - 1
            pnl_amount = PER_STOCK * pnl_rate
            sell_date = str(t.get('sell_date', ''))[:10]
            sell_reason = t.get('sell_reason', '')
            stock_list.append({
                '买入日期': str(t.get('buy_date', ''))[:10],
                '代码': code,
                '名称': name,
                '投入金额': PER_STOCK,
                '买入价': round(buy_p, 2),
                '卖出价/收盘价': round(sell_p, 2),
                '单笔收益': round(pnl_amount, 6),
                '单笔收益率': round(pnl_rate, 6),
                '卖出原因': sell_reason,
                '卖出时间': f"{sell_date}",
            })

    # FIX: 用列表存储所有买入实例（允许同一 code 多次出现），不再用 dict 去重丢失二次买入信息
    buy_instances = []
    if _daily is not None and len(_daily) > 0:
        for _, row in _daily.iterrows():
            bt = row.get('bought_today', [])
            date = str(row['date'])[:10]
            if isinstance(bt, list):
                for b in bt:
                    buy_instances.append({
                        'code': b['code'],
                        'buy_date': date,
                        'buy_price': float(b['buy_price']),
                    })

        last_row = _daily.iloc[-1]
        pc_str = str(last_row.get('positions_close', ''))
        holdings_close = {}
        if pc_str and pc_str != '空仓':
            for item in pc_str.split(', '):
                if '@' in item:
                    parts = item.split('@')
                    try:
                        holdings_close[parts[0].strip()] = float(parts[1].strip())
                    except:
                        pass

        for bi in buy_instances:
            if (bi['code'], bi['buy_date']) in sold_pairs:
                continue
            buy_p = bi['buy_price']
            close_p = holdings_close.get(bi['code'], buy_p)
            cost = backtest_engine.TRADING_COST
            pnl_rate = close_p / (buy_p * (1 + cost)) - 1
            pnl_amount = PER_STOCK * pnl_rate
            stock_list.append({
                '买入日期': bi['buy_date'],
                '代码': bi['code'],
                '名称': code_to_name.get(bi['code'], bi['code']),
                '投入金额': PER_STOCK,
                '买入价': round(buy_p, 2),
                '卖出价/收盘价': round(close_p, 2),
                '单笔收益': round(pnl_amount, 6),
                '单笔收益率': round(pnl_rate, 6),
                '卖出原因': 'N/A',
                '卖出时间': '持仓',
            })

    stock_list.sort(key=lambda x: (x['买入日期'], x['代码']))

    # 添加期末收盘价和期末涨跌幅
    end_date_ts = pd.Timestamp(end_date)
    for s in stock_list:
        code = s['代码']
        buy_p = s['买入价']
        if s['卖出时间'] == '持仓':
            end_close = s['卖出价/收盘价']
        else:
            end_close = None
            kline_df = backtest_engine.read_tdx_day_file(code)
            if kline_df is not None and len(kline_df) > 0:
                kline_df['date'] = pd.to_datetime(kline_df['date'])
                end_rows = kline_df[kline_df['date'] == end_date_ts]
                if len(end_rows) == 0:
                    near_rows = kline_df[kline_df['date'] <= end_date_ts]
                    if len(near_rows) > 0:
                        end_close = float(near_rows['close'].iloc[-1])
                else:
                    end_close = float(end_rows['close'].iloc[0])
            if end_close is None:
                end_close = buy_p
        s['期末收盘价'] = round(end_close, 2)
        s['期末涨跌幅'] = round((end_close - buy_p) / buy_p * 100, 2)

    if stock_list:
        total_pnl_stock = sum(s['单笔收益'] for s in stock_list)
        stock_list.append({
            '买入日期': '合计',
            '代码': None,
            '名称': None,
            '投入金额': None,
            '买入价': None,
            '卖出价/收盘价': None,
            '单笔收益': round(total_pnl_stock, 6),
            '单笔收益率': None,
            '卖出原因': None,
            '卖出时间': None,
            '期末收盘价': None,
            '期末涨跌幅': None,
        })
    _bt_df.attrs['stock_profit_list'] = pd.DataFrame(stock_list)
    return _bt_df


def generate_backtest_data(start_date_str, end_date_str, top_n=5, min_score=60):
    """回测数据生成"""
    start = pd.Timestamp(start_date_str)
    end = pd.Timestamp(end_date_str)
    dates = pd.bdate_range(start=start, end=end)
    
    seed = hash((start_date_str + end_date_str + str(top_n) + str(min_score))) % (2**31)
    rng = np.random.RandomState(seed)
    
    pool_df = get_stock_pool()
    scored_pool = pool_df.sort_values("综合评分", ascending=False).reset_index(drop=True)
    candidates = scored_pool.head(top_n * 3).copy()
    n_candidates = len(candidates)
    
    results = []
    portfolio_value = 100000.0
    benchmark_value = 100000.0
    peak_val = portfolio_value
    benchmark_peak = benchmark_value
    current_holdings = []
    
    for i, dt in enumerate(dates):
        if i % 5 == 0 or i == 0:
            perturbation = rng.randn(n_candidates) * 5
            temp_scores = candidates["综合评分"].values + perturbation
            selected_indices = np.argsort(temp_scores)[-top_n:][::-1]
            current_holdings = list(selected_indices)
        
        daily_returns = []
        for idx in current_holdings:
            if idx < len(candidates):
                stock = candidates.iloc[idx]
                score_factor = (stock["综合评分"] - 50) / 500.0
                daily_ret = rng.normal(score_factor, 0.018)
                daily_returns.append({"代码": stock["代码"], "名称": stock["名称"],
                                     "评分": stock["综合评分"], "日收益率": daily_ret})
        
        port_ret = np.mean([r["日收益率"] for r in daily_returns]) if daily_returns else 0
        portfolio_value *= (1 + port_ret)
        peak_val = max(peak_val, portfolio_value)
        drawdown = (peak_val - portfolio_value) / peak_val if peak_val > 0 else 0
        
        benchmark_ret = rng.normal(0.0003, 0.015)
        benchmark_value *= (1 + benchmark_ret)
        benchmark_peak = max(benchmark_peak, benchmark_value)
        benchmark_dd = (benchmark_peak - benchmark_value) / benchmark_peak if benchmark_peak > 0 else 0
        
        results.append({
            "日期": dt.date(), "组合净值": round(portfolio_value, 2),
            "基准净值": round(benchmark_value, 2), "日收益率": round(port_ret, 5),
            "基准日收益率": round(benchmark_ret, 5),
            "累计收益率": round((portfolio_value / 100000 - 1), 5),
            "基准累计收益": round((benchmark_value / 100000 - 1), 5),
            "超额收益": round((portfolio_value / benchmark_value - 1), 5),
            "最大回撤": round(drawdown, 5), "基准最大回撤": round(benchmark_dd, 5),
            "持仓数": len(daily_returns),
            "持仓明细": [r["名称"] for r in daily_returns],
            "持仓评分": [r["评分"] for r in daily_returns],
        })
    
    return pd.DataFrame(results)


def calc_backtest_metrics(bt_df):
    """计算回测统计指标（含短期回测保护）"""
    if bt_df is None or len(bt_df) == 0:
        return None
    equity = bt_df["组合净值"].values.astype(float)
    benchmark = bt_df["基准净值"].values.astype(float)
    returns = bt_df["日收益率"].values.astype(float)
    benchmark_returns = bt_df["基准日收益率"].values.astype(float)
    total_days = len(bt_df)
    annual_factor = 245 / max(total_days, 1)
    cumret = float(bt_df["累计收益率"].iloc[-1])
    bench_cumret = float(bt_df["基准累计收益"].iloc[-1])
    
    # 保护：equity[0] 不能为 0 或 NaN
    e0 = float(equity[0]) if len(equity) > 0 and float(equity[0]) > 0 else 100000.0
    b0 = float(benchmark[0]) if len(benchmark) > 0 and float(benchmark[0]) > 0 else 100000.0
    
    # 年化收益率：短期回测（<60天）用简单线性年化，避免复利爆炸
    if total_days >= 60:
        ann_return = (float(equity[-1]) / e0) ** annual_factor - 1
        bench_ann = (float(benchmark[-1]) / b0) ** annual_factor - 1
    else:
        ann_return = cumret * annual_factor
        bench_ann = bench_cumret * annual_factor
    
    # 额外保护：截断极端年化值（绝对值超过1000视为异常，返回0）
    if abs(ann_return) > 1000:
        ann_return = 0.0
    if abs(bench_ann) > 1000:
        bench_ann = 0.0
    
    volatility = float(np.std(returns)) * np.sqrt(245)
    # 夏普保护：volatility 极小值
    if volatility < 1e-6:
        sharpe = 0.0
    else:
        sharpe = (ann_return - 0.03) / volatility
    
    alpha = ann_return - bench_ann
    # 短期回测（<60天）：超额收益和月均收益失真，显示为 N/A
    if total_days < 60:
        alpha_str = "N/A（数据不足）"
    else:
        alpha_str = f"{alpha*100:+.1f}%"
    
    max_dd = float(bt_df["最大回撤"].max())
    win_days = sum(1 for r in returns if r > 0)
    win_rate = win_days / max(len(returns), 1)
    losing_streak = 0; max_loss_streak = 0
    for r in returns:
        if r <= 0: losing_streak += 1; max_loss_streak = max(max_loss_streak, losing_streak)
        else: losing_streak = 0
    excess_wins = sum(1 for s, b in zip(returns, benchmark_returns) if s > b)
    excess_win_rate = excess_wins / max(len(returns), 1)
    bt_df_copy = bt_df.copy()
    bt_df_copy["日期"] = pd.to_datetime(bt_df_copy["日期"])
    monthly = bt_df_copy.resample('ME', on='日期').apply(lambda x: (x["组合净值"].iloc[-1] / x["组合净值"].iloc[0] - 1) if len(x) > 1 else 0)
    monthly_str = "N/A（数据不足）" if total_days < 60 else f"{monthly.mean()*100:.2f}%"
    # 卡玛保护：max_dd 极小值
    if max_dd < 1e-6:
        calmar = 0.0
    else:
        calmar = ann_return / max_dd
    
    return {
        "初始资金": 100000, "期末资金": round(float(equity[-1]), 0), "基准期末": round(float(benchmark[-1]), 0),
        "累计收益率": f"{cumret*100:+.1f}%", "基准累计收益": f"{bench_cumret*100:+.1f}%",
        "超额收益": alpha_str, "年化收益率": f"{ann_return*100:+.1f}%",
        "基准年化": f"{bench_ann*100:+.1f}%", "最大回撤": f"{max_dd*100:.1f}%",
        "夏普比率": round(sharpe, 2), "卡玛比率": round(calmar, 2),
        "年化波动率": f"{volatility*100:.1f}%", "胜率": f"{win_rate*100:.1f}%",
        "超额胜率": f"{excess_win_rate*100:.1f}%", "交易天数": total_days,
        "最大连亏天数": max_loss_streak, "月均收益": monthly_str,
        "月胜率": f"{(monthly>0).sum()/max(len(monthly),1)*100:.0f}%",
        "cumret_raw": cumret, "ann_ret_raw": ann_return, "sharpe_raw": sharpe,
        "dd_raw": max_dd, "win_rate_raw": win_rate, "alpha_raw": alpha,
        "excess_win_rate_raw": excess_win_rate,
    }


# ================================================================
#                         绘图函数
# ================================================================

def create_radar_chart(dim_data, weights):
    dims = list(dim_data.keys())
    vals = [dim_data.get(d, 50) for d in dims]
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=vals+[vals[0]], theta=dims+[dims[0]], fill='toself', name='当前评分',
        line_color='#C4842D', fillcolor='rgba(196,132,45,0.18)', line=dict(width=2.5)))
    fig.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0,100],
        tickvals=[20,40,60,80,100], tickfont=dict(size=10,color='#AAA'),
        gridcolor='#EEE', linecolor='#DDD'),
        angularaxis=dict(tickfont=dict(size=13,color='#444'), gridcolor='#F0EDE8',
        linecolor='#DDD', direction='clockwise', rotation=90), bgcolor='rgba(0,0,0,0)'),
        showlegend=True, legend=dict(orientation='h', yanchor='bottom', y=-0.12, xanchor='center', x=0.5),
        margin=dict(t=40,b=80,l=60,r=60), height=420,
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
    return fig


def create_backtest_equity_chart(bt_df, metrics=None):
    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                       vertical_spacing=0.08, row_heights=[0.45, 0.25, 0.3])
    fig.add_trace(go.Scatter(x=bt_df["日期"], y=bt_df["组合净值"],
        mode='lines', line=dict(color='#C4842D', width=2.5),
        fill='tozeroy', fillcolor='rgba(196,132,45,0.08)', name='策略净值'), row=1, col=1)
    fig.add_trace(go.Scatter(x=bt_df["日期"], y=bt_df["基准净值"],
        mode='lines', line=dict(color='#666', width=1.5, dash='dot'), name='沪深300基准'), row=1, col=1)
    colors = ['#E74C3C' if r >= 0 else '#27AE60' for r in bt_df["日收益率"]]
    fig.add_trace(go.Bar(x=bt_df["日期"], y=bt_df["日收益率"]*100,
        marker_color=colors, opacity=0.7, name='日收益率(%)', marker_line_width=0), row=2, col=1)
    fig.add_trace(go.Scatter(x=bt_df["日期"], y=-bt_df["最大回撤"]*100,
        mode='lines', fill='tozeroy', line=dict(color='#E74C3C', width=1.5),
        fillcolor='rgba(231,76,60,0.15)', name='回撤(%)'), row=3, col=1)
    fig.update_yaxes(title_text="净值(元)", row=1, col=1)
    fig.update_yaxes(title_text="收益率(%)", row=2, col=1)
    fig.update_yaxes(title_text="回撤(%)", row=3, col=1)
    m = metrics or {}
    title_text = ""
    if m.get("cumret_raw") is not None:
        cr = m["cumret_raw"]
        clr = "#E74C3C" if cr >= 0 else "#27AE60"
        title_text = f"累计收益: <span style='color:{clr};font-weight:700'>{cr*100:+.1f}%</span>"
    fig.update_layout(height=680, showlegend=True, legend=dict(orientation='h', y=-0.02),
        paper_bgcolor='white', plot_bgcolor='rgba(250,249,245,0.5)',
        title_text=title_text if title_text else "")
    return fig


def create_monthly_heatmap(bt_df):
    df = bt_df.copy()
    df["日期"] = pd.to_datetime(df["日期"])
    df['year'] = df['日期'].dt.year; df['month'] = df['日期'].dt.month
    monthly_pnl = df.groupby(['year','month']).apply(
        lambda g: (g["组合净值"].iloc[-1] / g["组合净值"].iloc[0] - 1) * 100 if len(g) > 1 else 0
    ).reset_index(name='ret')
    pivot = monthly_pnl.pivot(index='month', columns='year', values='ret')
    months_cn = ['', '1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
    pivot.index = [months_cn[m] for m in pivot.index]
    fig = go.Figure(data=go.Heatmap(
        z=pivot.values, x=[str(y) for y in pivot.columns], y=pivot.index,
        text=[[f'{v:.1f}%' if not np.isnan(v) else '' for v in row] for row in pivot.values],
        texttemplate='%{text}', textfont=dict(size=10),
        colorscale=[[0,'#27AE60'], [0.5,'white'], [1,'#E74C3C']],
        zmin=-15, zmax=15, hoverongaps=False, showscale=True,
        colorbar=dict(title='收益率%', thickness=10, len=0.85)))
    fig.update_layout(height=320, paper_bgcolor='white', plot_bgcolor='white',
        xaxis=dict(side='top'), margin=dict(l=50,r=30,t=30,b=20))
    return fig


def create_score_bar(score, width=120):
    color = '#E74C3C' if score >= 70 else '#E67E22' if score >= 50 else '#27AE60'
    return f"""<div style="display:flex;align-items:center;gap:8px;">
<div class="score-bar" style="width:{width}px;"><div class="score-fill" style="width:{score}%;background:{color};"></div></div>
<span style="font-weight:700;font-size:14px;color:#333;min-width:28px;">{score}</span></div>"""


# ================================================================
#      强势超跌金叉模型 — 评分体系（五维 + 硬过滤 + 信号分类）
# ================================================================

DEFAULT_GC_PARAMS = {
    "main_rise_pct": 20,       # 近60日主升涨幅阈值 %
    "decline_pct": 15,         # 从高点回撤阈值 %
    "volume_ratio": 1.2,       # 企稳日量能 ≥ N倍5日均量
    "steady_days": 2,           # 不创新低天数
    "sample_source": "全市场A股",  # 样本来源：全市场A股 / 热门板块
    "fund_weight": 0.10,        # P0: 资金确认维度权重（默认10%）
}

DEFAULT_GC_SAMPLE_OPTIONS = ["全市场A股", "热门板块（资金加速Top6）", "板块反转（量价+资金）"]

DEFAULT_GC_WEIGHTS = {
    "下跌形态": 28,
    "K线止跌": 18,
    "均线拐头": 18,
    "量能确认": 13,
    "MACD反转": 13,
    "资金确认": 10,
}


def get_hot_concept_stocks(n=6):
    """通过 PowerShell 子进程调用东方财富 API，按资金加速比（今日/(5日/5)）取 Top N 板块的全部成分股"""
    import subprocess, json, sys
    helper = os.path.join(os.path.dirname(__file__), "temp", "fetch_hot_sectors.ps1")
    if not os.path.exists(helper):
        if _VERBOSE: print(f"[金叉模型] helper脚本不存在: {helper}")
        return set()
    try:
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", helper, "-N", str(n)],
            capture_output=True, text=True, timeout=60
        )
        data = json.loads(result.stdout)
        if "error" in data:
            if _VERBOSE: print(f"[金叉模型] 加速板块获取失败: {data['error']}")
            return set()
        codes = data.get("codes", [])
        sectors = data.get("sectors", [])
        if _VERBOSE: print(f"[金叉模型] 资金加速Top{n}板块:")
        for s in sectors:
            tag = "反转" if s.get("reversal") else f"{s.get('accel_ratio',0):.1f}x"
            if _VERBOSE: print(f"  {s['name']}({s['code']}): 今日{s['today_inflow']}亿 / 5日{s['day5_inflow']}亿 → {tag} | {s['count']}只成分股")
        if _VERBOSE: print(f"总计 {len(codes)} 只成分股")
        return set(codes)
    except Exception as e:
        if _VERBOSE: print(f"[金叉模型] 子进程调用失败: {e}")
        return set()


def get_volprice_sectors(n=6):
    """通过量价背离 x 资金加速 综合评分选板块，返回 Top N 板块的全部成分股"""
    import subprocess, json, sys
    helper = os.path.join(os.path.dirname(__file__), "temp", "fetch_volprice_sectors.ps1")
    if not os.path.exists(helper):
        print(f"[金叉模型-量价] helper脚本不存在: {helper}")
        return set()
    try:
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", helper, "-N", str(n), "-topAccel", "50"],
            capture_output=True, text=True, timeout=90
        )
        data = json.loads(result.stdout)
        if "error" in data:
            print(f"[金叉模型-量价] 获取失败: {data['error']}")
            return set()
        codes = data.get("codes", [])
        sectors = data.get("sectors", [])
        if _VERBOSE: print(f"[金叉模型-量价] 量价+资金综合Top{n}板块:")
        for s in sectors:
            tag = "反转" if s.get("reversal") else f"{s.get('accel_ratio',0):.1f}x"
            if _VERBOSE: print(f"  {s['name']}({s['code']}): 综合={s.get('composite')} 加速={tag} 量价背离={s.get('divergence')} | {s['count']}只成分股")
        if _VERBOSE: print(f"总计 {len(codes)} 只成分股")
        return set(codes)
    except Exception as e:
        print(f"[金叉模型-量价] 子进程调用失败: {e}")
        return set()


def _get_gc_params():
    return st.session_state.get("gc_params", dict(DEFAULT_GC_PARAMS))


def _compute_macd(closes, fast=12, slow=26, sig=9):
    """返回 (DIF, DEA, MACD柱) 数组，长度与输入一致"""
    if len(closes) < slow + sig:
        return np.full(len(closes), np.nan), np.full(len(closes), np.nan), np.full(len(closes), np.nan)
    ema_fast = pd.Series(closes).ewm(span=fast, adjust=False).mean()
    ema_slow = pd.Series(closes).ewm(span=slow, adjust=False).mean()
    dif = (ema_fast - ema_slow).values
    dea = pd.Series(dif).ewm(span=sig, adjust=False).mean().values
    macd_hist = 2 * (dif - dea)
    return dif, dea, macd_hist


@st.cache_data(ttl=1800)
def _get_hot_sector_names_set():
    """返回当前热门板块名称集合，用于板块反向惩罚"""
    import subprocess, json
    helper = os.path.join(os.path.dirname(__file__), "temp", "fetch_hot_sectors.ps1")
    if not os.path.exists(helper):
        return set()
    try:
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", helper, "-N", "6"],
            capture_output=True, text=True, timeout=60
        )
        data = json.loads(result.stdout)
        return {s['name'] for s in data.get('sectors', [])}
    except Exception:
        return set()


def _gc_hard_filter(code, kline_df):
    """
    入库硬条件 — 双通道：
    通道A（零下通道）：主升涨幅 + 回撤 + MACD DIFF < 0（原有逻辑）
    通道B（水上通道）：回调洗盘后 MACD 零轴上方金叉，四条件缺一不可
      1) 价格在20日线上方但乖离率≤25%
      2) 60日线方向向上
      3) 近20日有过≥10%回撤
      4) 当日成交量≥5日均量1.2倍
    返回 (通过, 信息, 通道标识: 'zero_below'/'water_above'/None)
    """
    if kline_df is None or len(kline_df) < 40:
        return False, "K线不足40日", None
    closes = kline_df['收盘'].values.astype(float)
    n = len(closes)
    params = _get_gc_params()

    half = min(60, n)
    recent = closes[-half:]
    rise_pct = (recent.max() - recent[0]) / recent[0] * 100
    peak_idx = int(np.argmax(recent))
    peak_val = recent[peak_idx]
    current = closes[-1]
    decline = (peak_val - current) / peak_val * 100
    dif, _, _ = _compute_macd(closes)

    # ---------- 通道A：零下通道（原有逻辑） ----------
    if not np.isnan(dif[-1]) and dif[-1] < 0:
        if rise_pct < params["main_rise_pct"]:
            return False, f"主升涨幅{rise_pct:.0f}%<{params['main_rise_pct']}%", None
        if decline < params["decline_pct"]:
            return False, f"回撤{decline:.0f}%<{params['decline_pct']}%", None
        return True, f"零下 主升{rise_pct:.0f}% 回撤{decline:.0f}% DIFF={dif[-1]:.3f}", "zero_below"

    # ---------- 通道B：水上通道（新增） ----------
    # 条件1: 价格在20日线上方但乖离率≤25%
    if n < 25:
        return False, "K线不足25日 不适合水上通道", None
    ma20 = pd.Series(closes).rolling(20).mean().values
    cur_ma20 = ma20[-1]
    if np.isnan(cur_ma20) or cur_ma20 <= 0:
        return False, "MA20无效", None
    if closes[-1] < cur_ma20:
        return False, "水上通道: 股价低于20日线", None
    deviation = (closes[-1] - cur_ma20) / cur_ma20 * 100
    if deviation > 25:
        return False, f"水上通道: 乖离率{deviation:.0f}%>25% 高位风险", None

    # 条件2: 60日线方向向上
    if n < 65:
        return False, "K线不足65日 无法判断60日线方向", None
    ma60 = pd.Series(closes).rolling(60).mean().values
    if np.isnan(ma60[-1]) or np.isnan(ma60[-11]):
        return False, "MA60无效", None
    ma60_slope = (ma60[-1] - ma60[-11]) / ma60[-11] * 100 if ma60[-11] > 0 else 0
    if ma60_slope <= 0:
        return False, f"水上通道: 60日线向下({ma60_slope:.2f}%) 下跌通道风险", None

    # 条件3: 近20日有过≥10%回撤（确认是回调洗盘）
    recent20 = closes[-20:]
    recent20_peak = recent20.max()
    recent20_trough = recent20.min()
    recent20_decline = (recent20_peak - recent20_trough) / recent20_peak * 100
    if recent20_decline < 10:
        return False, f"水上通道: 近20日回撤{recent20_decline:.0f}%<10% 无充分洗盘", None

    # 条件4: 当日成交量≥5日均量1.2倍
    if '成交量' not in kline_df.columns:
        return False, "水上通道: 无成交量数据", None
    volumes = kline_df['成交量'].values.astype(float)
    if len(volumes) < 6:
        return False, "水上通道: 成交量数据不足", None
    ma5_vol = np.mean(volumes[-6:-1])
    if ma5_vol <= 0:
        return False, "水上通道: 5日均量无效", None
    today_vol = volumes[-1]
    vol_ratio = today_vol / ma5_vol
    if vol_ratio < 1.2:
        return False, f"水上通道: 量比{vol_ratio:.1f}<1.2 缩量不可靠", None

    return True, f"水上 乖离{deviation:.0f}% 回撤{recent20_decline:.0f}% MA60↑ 量比{vol_ratio:.1f}", "water_above"


# ---------- 维度1: 下跌形态（30分） ----------
def _gc_score_decline(kline_df):
    """回撤越深、急跌形态越明显，分数越高"""
    closes = kline_df['收盘'].values.astype(float)
    n = len(closes)
    half = min(60, n)
    recent = closes[-half:]
    peak_idx = int(np.argmax(recent))
    peak_val = recent[peak_idx]
    current = closes[-1]
    decline_pct = (peak_val - current) / peak_val * 100

    # 基础分：回撤深度（线性映射：30%→12分，50%→28分）
    base = min(28, max(5, (decline_pct - 25) * 0.8 + 8))

    # 急跌加分：检查从高点以来是否有连续大阴线
    post_peak = recent[peak_idx:]
    big_drops = 0
    for i in range(1, len(post_peak)):
        day_chg = (post_peak[i] - post_peak[i-1]) / post_peak[i-1] * 100
        if day_chg < -5:
            big_drops += 1
    crash_bonus = min(2, big_drops * 0.5)

    return min(30, base + crash_bonus)


# ---------- 维度2: K线止跌（20分） ----------
def _gc_score_steady(kline_df):
    """不再创新低 + 底部K线形态"""
    closes = kline_df['收盘'].values.astype(float)
    opens = kline_df['开盘'].values.astype(float) if '开盘' in kline_df.columns else closes
    highs = kline_df['最高'].values.astype(float) if '最高' in kline_df.columns else closes
    lows = kline_df['最低'].values.astype(float) if '最低' in kline_df.columns else closes
    n = len(closes)

    half = min(60, n)
    recent_lows = lows[-half:]
    stage_low = recent_lows.min()

    # 不创新低天数
    steady_days = 0
    for i in range(n - 1, max(0, n - 10) - 1, -1):
        if lows[i] > stage_low * 1.005:
            steady_days += 1
        else:
            break
    steady_score = min(14, steady_days * 3.5)

    # 底部K线形态：最近3天
    kline_bonus = 0
    for i in range(max(0, n-3), n):
        body = abs(closes[i] - opens[i])
        lower_shadow = min(opens[i], closes[i]) - lows[i]
        upper_shadow = highs[i] - max(opens[i], closes[i])
        total_range = highs[i] - lows[i] if highs[i] > lows[i] else 1
        # 长下影线
        if lower_shadow > body * 1.5 and lower_shadow > total_range * 0.4:
            kline_bonus += 2
        # 小阳线
        if closes[i] > opens[i] and body < total_range * 0.3:
            kline_bonus += 1
        # 十字星
        if body < total_range * 0.15:
            kline_bonus += 1

    # 当日收阳线 + 站上MA5/MA10
    if closes[-1] > opens[-1]:
        kline_bonus += 2
    if n >= 10:
        ma5 = pd.Series(closes).rolling(5).mean().values[-1]
        ma10 = pd.Series(closes).rolling(10).mean().values[-1]
        if closes[-1] > ma5 and closes[-1] > ma10:
            kline_bonus += 2

    return min(20, steady_score + min(6, kline_bonus))


# ---------- 维度3: 均线拐头（20分） ----------
def _gc_score_ma(kline_df):
    """5日线走平/拐头 + 收盘站上5日线 + MA20/MA60长期趋势方向"""
    closes = kline_df['收盘'].values.astype(float)
    n = len(closes)
    if n < 7:
        return 0

    ma5 = pd.Series(closes).rolling(5).mean().values
    cur_ma5 = ma5[-1] if not np.isnan(ma5[-1]) else 0
    prev_ma5 = ma5[-3] if n >= 8 and not np.isnan(ma5[-3]) else cur_ma5

    score = 0
    # 5日线斜率
    slope = (ma5[-1] - ma5[-3]) / ma5[-3] * 100 if ma5[-3] > 0 and not np.isnan(ma5[-3]) else -99
    if slope > 0.5:
        score += 8
    elif slope > -0.3:
        score += 5
    elif slope > -1.0:
        score += 2

    # 收盘站上5日线
    if closes[-1] > cur_ma5 and cur_ma5 > 0:
        score += 6
    elif closes[-1] > cur_ma5 * 0.98:
        score += 3

    # MA20方向
    if n >= 25:
        ma20 = pd.Series(closes).rolling(20).mean().values
        if not np.isnan(ma20[-1]) and not np.isnan(ma20[-6]):
            ma20_slope = (ma20[-1] - ma20[-6]) / ma20[-6] * 100 if ma20[-6] > 0 else 0
            if ma20_slope > 0.3:
                score += 4
            elif ma20_slope > 0:
                score += 2
            elif ma20_slope < -1.0:
                score -= 2  # 20日线明显向下，扣分

    # MA60方向
    if n >= 65:
        ma60 = pd.Series(closes).rolling(60).mean().values
        if not np.isnan(ma60[-1]) and not np.isnan(ma60[-11]):
            ma60_slope = (ma60[-1] - ma60[-11]) / ma60[-11] * 100 if ma60[-11] > 0 else 0
            if ma60_slope > 0.5:
                score += 2
            elif ma60_slope < -1.0:
                score -= 3  # 60日线明显向下，下跌通道扣分

    return max(0, min(20, score))


# ---------- 维度4: 量能确认（15分） ----------
def _gc_score_volume(kline_df):
    """企稳日量能 vs 5日均量，低于volume_ratio阈值直接0分"""
    volumes = kline_df['成交量'].values.astype(float)
    n = len(volumes)
    if n < 6:
        return 0
    ma5_vol = np.mean(volumes[-6:-1])
    today_vol = volumes[-1]
    if ma5_vol <= 0:
        return 0
    ratio = today_vol / ma5_vol

    params = _get_gc_params()
    min_ratio = params.get("volume_ratio", 1.0)
    if ratio < min_ratio:
        return 0

    # 下跌缩量检查（前5天均值 vs 更早5天）
    early_vol = np.mean(volumes[max(0,n-20):max(0,n-10)]) if n >= 20 else ma5_vol
    shrink_bonus = 2 if early_vol > 0 and ma5_vol < early_vol * 0.7 else 0

    # 放量评分
    if ratio >= 2.0:
        return min(15, 12 + shrink_bonus)
    elif ratio >= 1.5:
        return min(15, 9 + shrink_bonus)
    elif ratio >= min_ratio:
        return min(15, 6 + shrink_bonus)
    return 0


# ---------- 维度5: MACD反转（15分） ----------
def _gc_score_macd(kline_df):
    """DIFF走平、绿柱缩短、金叉"""
    closes = kline_df['收盘'].values.astype(float)
    dif, dea, hist = _compute_macd(closes)
    if np.isnan(dif[-1]):
        return 0
    n = len(dif)

    score = 0
    # DIFF走平（不再持续向下）
    dif_slope = dif[-1] - dif[-3] if n >= 3 and not np.isnan(dif[-3]) else -99
    if dif_slope > 0.02:
        score += 5
    elif dif_slope > -0.01:
        score += 3

    # 绿柱缩短
    if n >= 3 and not np.isnan(hist[-1]) and not np.isnan(hist[-3]):
        if hist[-1] < 0 and hist[-1] > hist[-3]:
            score += 5

    # MACD金叉（DIFF上穿DEA）
    if n >= 2 and not np.isnan(dif[-2]) and not np.isnan(dea[-2]):
        if dif[-2] <= dea[-2] and dif[-1] > dea[-1]:
            score += 5
        elif dif[-1] > dea[-1] and dif[-3] <= dea[-3]:
            score += 3

    return min(15, score)


# ---------- P0: 维度6: 资金确认（满分按权重动态） ----------
def _gc_score_fund_confirm(stock_data, kline_df):
    """
    资金确认维度：
    - 优先：近3日主力净流入方向（正流入且占比>0 则得分）
    - 替代：MACD金叉当日成交量 / 5日均量（无法获取主力资金时）
    """
    inst_net = stock_data.get("inst_net_buy_3d", 0)
    north_net = stock_data.get("north_net_buy", 0)
    main_net = inst_net + north_net

    if main_net != 0:
        if main_net > 0:
            return 10
        else:
            return 0
    else:
        if kline_df is None or len(kline_df) < 6:
            return 5
        volumes = kline_df['成交量'].values.astype(float)
        if len(volumes) < 6:
            return 5
        ma5_vol = np.mean(volumes[-6:-1]) if len(volumes) >= 6 else volumes[-1]
        today_vol = volumes[-1]
        if ma5_vol <= 0:
            return 5
        ratio = today_vol / ma5_vol
        if ratio >= 1.5:
            return 10
        elif ratio >= 1.2:
            return 7
        elif ratio >= 1.0:
            return 5
        else:
            return 2


# ---------- 维度7: 板块确认（板块反向惩罚） ----------
def _gc_score_sector(stock_data):
    """
    板块确认维度：
    - 无概念板块映射 → -5
    - 有映射但不在当前热门板块中 → -3
    - 在热门板块中 → 0
    """
    sector = stock_data.get('板块', '')
    if not sector:
        return -5
    hot = _get_hot_sector_names_set()
    if hot and sector not in hot:
        return -3
    return 0


# ---------- 主评分函数（P1: 百分位归一化） ----------
def calculate_golden_cross_score(stock_data, kline_df, weights=None):
    """
    强势超跌金叉模型主评分函数
    P1: 所有子维度先做百分位归一化，再加权
    stock_data: 股票基本数据（来自行情DataFrame的一行）
    kline_df: 该股票的K线DataFrame
    weights: 权重 dict（含 fund_weight）
    返回 dict，包含各维度分、综合评分、信号分类等
    """
    code = stock_data.get('代码', '')
    name = stock_data.get('名称', '')

    # 硬过滤
    hf_result = _gc_hard_filter(code, kline_df)
    passed, filter_info = hf_result[0], hf_result[1]
    if not passed:
        return {
            "pass": False, "filter_msg": filter_info,
            "综合评分": 0, "下跌形态": 0, "K线止跌": 0,
            "均线拐头": 0, "量能确认": 0, "MACD反转": 0, "资金确认": 0,
            "板块确认": 0,
            "信号": "不达标", "signal_class": "badge-hold",
            "_raw": {},
        }

    # 七维原始分
    s1 = _gc_score_decline(kline_df)
    s2 = _gc_score_steady(kline_df)
    s3 = _gc_score_ma(kline_df)
    s4 = _gc_score_volume(kline_df)
    s5 = _gc_score_macd(kline_df)
    s6 = _gc_score_fund_confirm(stock_data, kline_df)
    s7 = _gc_score_sector(stock_data)

    if weights is None:
        weights = st.session_state.get("gc_weights", dict(DEFAULT_GC_WEIGHTS))

    total = int(
        s1 * weights.get("下跌形态", 28) / 30 +
        s2 * weights.get("K线止跌", 18) / 20 +
        s3 * weights.get("均线拐头", 18) / 20 +
        s4 * weights.get("量能确认", 13) / 15 +
        s5 * weights.get("MACD反转", 13) / 15 +
        s6 * weights.get("资金确认", 10) / 10 +
        s7
    )

    if s1 >= 24 and s2 >= 14 and s5 >= 10 and total >= 70:
        signal = "买点A-急跌反弹"
        sig_cls = "badge-strong"
        advice = "高弹性，10%-15%仓位"
    elif s2 >= 10 and s3 >= 12 and total >= 55:
        signal = "买点B-阴跌修复"
        sig_cls = "badge-attention"
        advice = "轻仓试错，5%-8%仓位"
    elif total >= 40:
        signal = "观察"
        sig_cls = "badge-hold"
        advice = "等信号"
    else:
        signal = "放弃"
        sig_cls = "badge-hold"
        advice = "条件不充分"

    return {
        "pass": True, "filter_msg": filter_info,
        "综合评分": total,
        "下跌形态": int(s1), "K线止跌": int(s2),
        "均线拐头": int(s3), "量能确认": int(s4),
        "MACD反转": int(s5), "资金确认": int(s6),
        "板块确认": int(s7),
        "信号": signal, "signal_class": sig_cls, "建议": advice,
        "_raw": {"s1": s1, "s2": s2, "s3": s3, "s4": s4, "s5": s5, "s6": s6, "s7": s7},
    }


@st.cache_data(ttl=1800)
def _run_golden_cross_scan(codes_tuple, params_json, weights_json):
    """
    批量扫描全市场，返回结果列表（按综合评分降序）
    P1: 两遍扫描 — 第一遍收集原始分 → 百分位归一化 → 第二遍计算最终总分
    codes_tuple: 全市场股票 (代码, 名称, 板块) 列表的元组
    """
    codes = list(codes_tuple)
    weights = json.loads(weights_json) if isinstance(weights_json, str) else weights_json
    
    # P0: 获取龙虎榜资金数据供资金确认维度使用
    lhb_data = {}
    try:
        lhb_data = _get_cached_dragon_tiger()
    except Exception:
        pass
    
    total = len(codes)
    progress = st.progress(0, text="强势超跌金叉模型扫描中...")
    step = max(1, total // 100)
    
    # ===== 第一遍：收集所有通过硬过滤的原始分 =====
    raw_results = []  # [(code, name, sector, close, raw_scores_dict, filter_info, kline_np)]
    for i, (code, name, sector) in enumerate(codes):
        if i % step == 0:
            progress.progress(min(i / total, 1.0),
                              text=f"金叉模型扫描 {i}/{total} — {name}")
        try:
            kline_df = get_stock_kline(code, days=60)
            if kline_df is None or len(kline_df) < 30:
                continue
            lhb = lhb_data.get(code, {})
            stock_data = {
                "代码": code, "名称": name, "板块": sector,
                "inst_net_buy_3d": lhb.get("inst_net_buy_3d", 0),
                "north_net_buy": lhb.get("north_net_buy", 0),
            }
            score = calculate_golden_cross_score(stock_data, kline_df, weights=weights)
            if score["pass"]:
                _raw = score.get("_raw", {})
                raw_results.append({
                    "code": code, "name": name, "sector": sector,
                    "close": kline_df['收盘'].values[-1] if '收盘' in kline_df.columns else '',
                    "filter_msg": score.get("filter_msg", ""),
                    "signal": score["信号"],
                    "signal_class": score["signal_class"],
                    "advice": score["建议"],
                    "_s1": _raw.get("s1", 0),
                    "_s2": _raw.get("s2", 0),
                    "_s3": _raw.get("s3", 0),
                    "_s4": _raw.get("s4", 0),
                    "_s5": _raw.get("s5", 0),
                    "_s6": _raw.get("s6", 0),
                    "_s7": _raw.get("s7", 0),
                })
        except Exception:
            continue
    
    if not raw_results:
        progress.empty()
        return []
    
    # ===== P1: 百分位归一化（六个维度分别做） =====
    n = len(raw_results)
    def _pct_norm(vals, full_score):
        arr = np.array(vals, dtype=float)
        if n <= 1 or np.all(arr == arr[0]):
            return np.full_like(arr, full_score * 0.5)
        ranks = np.argsort(np.argsort(arr)) + 1
        percentiles = ranks / n
        return (percentiles * full_score).clip(0, full_score)
    
    s1_arr = _pct_norm([r["_s1"] for r in raw_results], 30)
    s2_arr = _pct_norm([r["_s2"] for r in raw_results], 20)
    s3_arr = _pct_norm([r["_s3"] for r in raw_results], 20)
    s4_arr = _pct_norm([r["_s4"] for r in raw_results], 15)
    s5_arr = _pct_norm([r["_s5"] for r in raw_results], 15)
    s6_arr = _pct_norm([r["_s6"] for r in raw_results], 10)
    
    # ===== 第二遍：计算归一化后的加权总分 =====
    results = []
    for i, r in enumerate(raw_results):
        s1n, s2n, s3n, s4n, s5n, s6n = s1_arr[i], s2_arr[i], s3_arr[i], s4_arr[i], s5_arr[i], s6_arr[i]
        w1 = weights.get("下跌形态", 28)
        w2 = weights.get("K线止跌", 18)
        w3 = weights.get("均线拐头", 18)
        w4 = weights.get("量能确认", 13)
        w5 = weights.get("MACD反转", 13)
        w6 = weights.get("资金确认", 10)
        total = int(s1n * w1 / 30 + s2n * w2 / 20 + s3n * w3 / 20 +
                    s4n * w4 / 15 + s5n * w5 / 15 + s6n * w6 / 10 + r["_s7"])
        
        if total < 30:
            continue
        
        results.append({
            "pass": True,
            "代码": r["code"], "名称": r["name"], "板块": r["sector"],
            "收盘": r["close"],
            "金叉评分": total,
            "下跌形态": int(s1_arr[i]), "K线止跌": int(s2_arr[i]),
            "均线拐头": int(s3_arr[i]), "量能确认": int(s4_arr[i]),
            "MACD反转": int(s5_arr[i]), "资金确认": int(s6_arr[i]),
            "板块确认": int(r["_s7"]),
            "信号": r["signal"], "signal_class": r["signal_class"],
            "建议": r["advice"], "filter_msg": r["filter_msg"],
        })
    
    progress.empty()
    results.sort(key=lambda x: x["金叉评分"], reverse=True)
    return results[:30]


# ================================================================
#              CAN SLIM模型（抓主升浪股） & 困境反转模型（抓周期股）
# ================================================================

# ----- 财务数据获取（AKShare） -----

@st.cache_data(ttl=3600)
def get_all_spot_data():
    """
    通过腾讯行情 API (qt.gtimg.cn) 拉取全市场行情快照，提取总市值和换手率，
    返回 {code: {"总市值": float(yi), "换手率": float}}。失败返回空字典。
    使用腾讯 API 绕过 push2.eastmoney.com 阻断。
    """
    result = {}
    try:
        import requests as _req

        # 获取股票代码列表（优先从本地缓存，其次从市场数据）
        codes = _load_stock_code_list()
        if not codes:
            if _VERBOSE:
                print("[腾讯行情] 无法获取股票代码列表")
            return result

        def _to_tx(code):
            return ('sh' if code.startswith(('6', '5', '9')) else 'sz') + code

        tx_codes = [_to_tx(c) for c in codes]
        BATCH = 80  # qt.gtimg.cn 批量上限，避免超长URL

        for i in range((len(tx_codes) + BATCH - 1) // BATCH):
            batch = tx_codes[i * BATCH : (i + 1) * BATCH]
            query = ','.join(batch)
            try:
                resp = _req.get(f'http://qt.gtimg.cn/q={query}', timeout=15,
                    headers={'User-Agent': 'Mozilla/5.0'})
                for line in resp.text.strip().split('\n'):
                    if '=\"' in line and '~' in line:
                        parts = line.split('\"')[1].split('~')
                        if len(parts) >= 46:
                            raw_code = parts[2]
                            try:
                                cap_yi = float(parts[45])  # field 45: 总市值（亿元）
                            except (ValueError, IndexError):
                                cap_yi = 0.0
                            try:
                                turnover = float(parts[38])  # field 38: 换手率
                            except (ValueError, IndexError):
                                turnover = 0.0
                            if cap_yi > 0:
                                result[raw_code] = {"总市值": cap_yi, "换手率": turnover}
            except Exception:
                continue  # 单批失败不阻塞整体

            # 分批间隔，避免限流
            if i < (len(tx_codes) + BATCH - 1) // BATCH - 1:
                time.sleep(0.25)

        if _VERBOSE:
            print(f"[腾讯行情] 全市场行情: {len(result)}/{len(codes)} 只有效数据")
    except Exception as e:
        if _VERBOSE:
            print(f"[腾讯行情] 全市场行情获取失败: {e}")
    return result


def _load_stock_code_list():
    """加载A股全市场股票代码列表。
    优先级：本地代码缓存 > 本地行情缓存 > akshare stock_info_a_code_name（上交所/深交所官网API）
    """
    codes = []
    cache_dir = os.path.join(BASE_DIR, '.cache')
    os.makedirs(cache_dir, exist_ok=True)
    code_list_path = os.path.join(cache_dir, 'a_stock_code_list.json')

    # 第一优先级：本地代码缓存（7天有效）
    try:
        if os.path.exists(code_list_path):
            with open(code_list_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            if time.time() - cached.get('ts', 0) < 7 * 86400:
                codes = cached.get('codes', [])
                if codes:
                    return codes
    except Exception:
        pass

    # 第二优先级：本地行情缓存
    try:
        cache_path = os.path.join(BASE_DIR, 'market_data_cache.json')
        if os.path.exists(cache_path):
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            records = cached.get('data', [])
            codes = [str(r.get('代码', '')).zfill(6) for r in records if r.get('代码')]
            codes = list(dict.fromkeys(codes))
            if codes:
                return codes
    except Exception:
        pass

    # 第三优先级：从 fetch_all_a_stocks 获取
    try:
        df = fetch_all_a_stocks()
        if df is not None and len(df) > 0:
            codes = df['代码'].astype(str).apply(lambda x: x.zfill(6)).unique().tolist()
            if codes:
                return codes
    except Exception:
        pass

    # 第四优先级（兜底）：从 akshare 获取完整A股代码列表（上交所/深交所官网）
    try:
        if AKSHARE_AVAILABLE:
            df_info = ak.stock_info_a_code_name()
            if df_info is not None and len(df_info) > 0:
                codes = df_info['code'].astype(str).apply(lambda x: x.zfill(6)).unique().tolist()
                # 保存到本地缓存
                try:
                    with open(code_list_path, 'w', encoding='utf-8') as f:
                        json.dump({'ts': time.time(), 'codes': codes}, f)
                except Exception:
                    pass
    except Exception:
        pass

    return codes


@st.cache_data(ttl=3600)
def get_financial_data(code):
    """
    合并拉取一只股票的季度财务数据，返回字典。
    
    返回字段：
    - success: 是否成功获取到至少一项有效数据
    - eps_growth_yoy: 单季扣非净利润同比增速
    - roe, roe_ttm, eps_cagr_3y
    - np_recovery_q: 净利润恢复率 (T/ABS(T-4))
    - rev_recovery_q: 营收恢复率
    - roe_recovery_q: ROE环比恢复率 (T/ABS(T-1))
    - gm, gm_recovery_q: 毛利率及恢复率
    - ocf_recovery_q: 经营现金流恢复率
    - pb, debt_ratio, goodwill_ratio
    - market_cap: 总市值（亿）
    - turnover_rate: 换手率
    """
    EMPTY = {
        "success": False,
        "eps_growth_yoy": 0.0, "roe": 0.0, "eps_cagr_3y": None,
        "np_recovery_q": 0.0, "np_recovery_q_1": 0.0, "np_recovery_q_2": 0.0,
        "rev_recovery_q": 0.0,
        "roe_ttm": 0.0,
        "roe_recovery_q": 0.0, "roe_recovery_q_1": 0.0, "roe_recovery_q_2": 0.0,
        "gm": 0.0, "gm_recovery_q": 0.0, "ocf_recovery_q": 0.0,
        "pb": 999.0, "debt_ratio": 999.0, "goodwill_ratio": 999.0,
        "market_cap": 0.0, "turnover_rate": 0.0,
    }

    if not AKSHARE_AVAILABLE:
        return EMPTY

    data = dict(EMPTY)
    any_success = False

    # ---- 合并市值/换手率 ----
    spot_all = get_all_spot_data()
    spot = spot_all.get(code, {})
    cap = spot.get("总市值", 0)
    to = spot.get("换手率", 0)

    if cap > 0:
        data["market_cap"] = cap
        data["turnover_rate"] = to
        any_success = True
    else:
        # fallback: 单独请求腾讯行情 API
        try:
            import requests as _req
            prefix = 'sh' if code.startswith(('6', '5', '9')) else 'sz'
            resp = _req.get(f'http://qt.gtimg.cn/q={prefix}{code}', timeout=10,
                headers={'User-Agent': 'Mozilla/5.0'})
            for line in resp.text.strip().split('\n'):
                if '=\"' in line and '~' in line:
                    parts = line.split('\"')[1].split('~')
                    if len(parts) >= 46 and parts[2] == code:
                        try:
                            data["market_cap"] = float(parts[45])
                        except:
                            pass
                        try:
                            data["turnover_rate"] = float(parts[38])
                        except:
                            pass
                        any_success = True
                        break
        except:
            pass

    # ---- 季度利润表 ----
    profit_q = None
    try:
        profit_q = ak.stock_profit_sheet_by_quarterly_em(symbol=code)
    except:
        pass

    if profit_q is not None and len(profit_q) >= 5:
        try:
            df = profit_q.head(8).copy()
            df = df.sort_index()  # 按报告期升序（旧→新）
            cols = df.columns.tolist()
            
            def _col(name):
                for c in cols:
                    if name in c:
                        return c
                return None
            
            rev_col = _col('营业收入')
            np_col = _col('净利润')
            dnp_col = _col('扣非净利润')
            ocf_col = _col('经营活动') or _col('现金流')
            
            if np_col and dnp_col:
                def _safe(idx):
                    v = df.iloc[idx].get(dnp_col, 0)
                    return float(v) if v and v == v else 0.0
                
                # T期 = 最新季, T-4期 = 同比基期
                dnp_t = _safe(-1)
                dnp_t4 = _safe(-5)
                dnp_t1 = _safe(-2)
                dnp_t5 = _safe(-6)
                dnp_t2 = _safe(-3)
                dnp_t6 = _safe(-7)
                
                # 扣非净利润同比增速
                if dnp_t4 != 0:
                    data["eps_growth_yoy"] = (dnp_t - dnp_t4) / abs(dnp_t4)
                    any_success = True
                
                # 净利润恢复率（用净利润列）
                def _safe_np(idx):
                    v = df.iloc[idx].get(np_col, 0)
                    return float(v) if v and v == v else 0.0
                
                np_t = _safe_np(-1)
                np_t4 = _safe_np(-5)
                np_t1 = _safe_np(-2)
                np_t5 = _safe_np(-6)
                np_t2 = _safe_np(-3)
                np_t6 = _safe_np(-7)
                
                if abs(np_t4) > 0:
                    data["np_recovery_q"] = np_t / abs(np_t4)
                if abs(np_t5) > 0:
                    data["np_recovery_q_1"] = np_t1 / abs(np_t5)
                if abs(np_t6) > 0:
                    data["np_recovery_q_2"] = np_t2 / abs(np_t6)
            
            # 营收恢复率
            if rev_col:
                rev_t = float(df.iloc[-1].get(rev_col, 0) or 0)
                rev_t4 = float(df.iloc[-5].get(rev_col, 0) or 0)
                if abs(rev_t4) > 0:
                    data["rev_recovery_q"] = rev_t / abs(rev_t4)
                    any_success = True
            
            # 经营现金流恢复率
            if ocf_col:
                ocf_t = float(df.iloc[-1].get(ocf_col, 0) or 0)
                ocf_t4 = float(df.iloc[-5].get(ocf_col, 0) or 0)
                if abs(ocf_t4) > 0:
                    data["ocf_recovery_q"] = ocf_t / abs(ocf_t4)
                    any_success = True
            
            # 近3年净利润复合增速（用于CAGR）
            if len(df) >= 13:
                np_y1 = (float(df.iloc[-1].get(np_col, 0) or 0) +
                         float(df.iloc[-2].get(np_col, 0) or 0) +
                         float(df.iloc[-3].get(np_col, 0) or 0) +
                         float(df.iloc[-4].get(np_col, 0) or 0))
                np_y0 = (float(df.iloc[-13].get(np_col, 0) or 0) +
                         float(df.iloc[-12].get(np_col, 0) or 0) +
                         float(df.iloc[-11].get(np_col, 0) or 0) +
                         float(df.iloc[-10].get(np_col, 0) or 0))
                if np_y0 > 0 and np_y1 > 0:
                    data["eps_cagr_3y"] = (np_y1 / np_y0) ** (1/3) - 1
                    any_success = True
        except:
            pass

    # ---- 财务分析指标（ROE, 毛利率, 资产负债率）----
    def _find_col(cols, name):
        for c in cols:
            if name in c:
                return c
        return name

    try:
        fa_df = ak.stock_financial_analysis_indicator(symbol=code)
        if fa_df is not None and len(fa_df) > 0:
            latest = fa_df.iloc[-1]
            cols = fa_df.columns.tolist()
            
            def _fa(name):
                for c in cols:
                    if name in c:
                        v = latest.get(c, 0)
                        return float(v) if v and v == v else 0.0
                return 0.0
            
            data["roe"] = _fa('净资产收益率') / 100.0 if _fa('净资产收益率') > 1 else _fa('净资产收益率')
            data["roe_ttm"] = _fa('净资产收益率') / 100.0 if _fa('净资产收益率') > 1 else _fa('净资产收益率')
            data["gm"] = _fa('销售毛利率') / 100.0 if _fa('销售毛利率') > 1 else _fa('销售毛利率')
            data["debt_ratio"] = _fa('资产负债率') / 100.0 if _fa('资产负债率') > 1 else _fa('资产负债率')
            
            # 毛利率恢复率：T期 vs T-4期
            if len(fa_df) >= 5:
                gm_t4 = float(fa_df.iloc[-5].get(_find_col(cols, '销售毛利率'), 0) or 0)
                gm_t4 = gm_t4 / 100.0 if gm_t4 > 1 else gm_t4
                if abs(gm_t4) > 0 and data["gm"] > 0:
                    data["gm_recovery_q"] = data["gm"] / abs(gm_t4)
            
            # ROE 环比恢复率
            if len(fa_df) >= 2:
                roe_t1 = float(fa_df.iloc[-2].get(_find_col(cols, '净资产收益率'), 0) or 0)
                roe_t1 = roe_t1 / 100.0 if roe_t1 > 1 else roe_t1
                if abs(roe_t1) > 0 and data["roe"] != 0:
                    data["roe_recovery_q"] = data["roe"] / abs(roe_t1)
            if len(fa_df) >= 3:
                roe_t2 = float(fa_df.iloc[-3].get(_find_col(cols, '净资产收益率'), 0) or 0)
                roe_t2 = roe_t2 / 100.0 if roe_t2 > 1 else roe_t2
                roe_t1 = float(fa_df.iloc[-2].get(_find_col(cols, '净资产收益率'), 0) or 0)
                roe_t1 = roe_t1 / 100.0 if roe_t1 > 1 else roe_t1
                if abs(roe_t2) > 0 and roe_t1 != 0:
                    data["roe_recovery_q_1"] = roe_t1 / abs(roe_t2)
            if len(fa_df) >= 4:
                roe_t3 = float(fa_df.iloc[-4].get(_find_col(cols, '净资产收益率'), 0) or 0)
                roe_t3 = roe_t3 / 100.0 if roe_t3 > 1 else roe_t3
                roe_t2 = float(fa_df.iloc[-3].get(_find_col(cols, '净资产收益率'), 0) or 0)
                roe_t2 = roe_t2 / 100.0 if roe_t2 > 1 else roe_t2
                if abs(roe_t3) > 0 and roe_t2 != 0:
                    data["roe_recovery_q_2"] = roe_t2 / abs(roe_t3)
            
            any_success = True
    except:
        pass

    # ---- 资产负债表（商誉/净资产）----
    try:
        bs_df = ak.stock_balance_sheet_by_quarterly_em(symbol=code)
        if bs_df is not None and len(bs_df) > 0:
            latest = bs_df.iloc[-1]
            cols = bs_df.columns.tolist()
            
            def _bs(name):
                for c in cols:
                    if name in c:
                        v = latest.get(c, 0)
                        return float(v) if v and v == v else 0.0
                return 0.0
            
            goodwill = _bs('商誉')
            net_equity = _bs('净资产') or _bs('股东权益')
            if net_equity > 0 and goodwill > 0:
                data["goodwill_ratio"] = goodwill / net_equity
                any_success = True
            elif net_equity > 0:
                data["goodwill_ratio"] = 0.0
                any_success = True
    except:
        pass

    # ---- PB（从行情数据近似）----
    try:
        # 用总市值和净资产粗略估算PB
        if data["market_cap"] > 0:
            bs_df_check = ak.stock_balance_sheet_by_quarterly_em(symbol=code)
            if bs_df_check is not None and len(bs_df_check) > 0:
                latest_bs = bs_df_check.iloc[-1]
                for c in latest_bs.index:
                    if '净资产' in str(c) or '股东权益' in str(c):
                        net_eq = float(latest_bs.get(c, 0) or 0)
                        if net_eq > 0:
                            data["pb"] = (data["market_cap"] * 1e8) / net_eq
                            any_success = True
                            break
    except:
        pass

    data["success"] = any_success
    return data


# ----- CAN SLIM 评分引擎 -----

def compute_rps(kline_dict, codes):
    """
    计算全市场股票的250日RPS（相对价格强度）排名。
    
    kline_dict: {code: kline_df}，kline_df 含 'close' / '收盘' 列
    codes: 需要计算RPS的股票代码列表
    
    返回: {code: rps_value}，rps_value 范围 0-100
    如果某股票K线不足250日，RPS=0
    """
    rps_map = {}
    code_returns = {}
    
    for code in codes:
        df = kline_dict.get(code)
        if df is None or len(df) < 250:
            rps_map[code] = 0
            continue
        close_col = '收盘' if '收盘' in df.columns else 'close'
        if close_col not in df.columns:
            rps_map[code] = 0
            continue
        closes = df[close_col].values.astype(float)
        if len(closes) < 250:
            rps_map[code] = 0
            continue
        ret_250 = (closes[-1] / closes[-250] - 1) * 100
        code_returns[code] = ret_250
    
    if not code_returns:
        return rps_map
    
    # 按收益排名，计算百分位
    sorted_codes = sorted(code_returns, key=code_returns.get)
    n = len(sorted_codes)
    for i, code in enumerate(sorted_codes):
        rps_map[code] = round((i + 1) / n * 100, 1)
    
    return rps_map


def calculate_canslim_score(code, kline_df, stock_pool_context=None):
    """
    CAN SLIM 七因子简化版评分（基于K线+行情数据，财务因子占位）。
    
    维度（满分100）：
    - C_业绩增速（15分）：扣非净利润同比≥25%→15, ≥15%→10, ≥5%→5, 否则0
    - A_持续增长（15分）：ROE≥17%+CAGR≥20%→15, 仅ROE→8, 仅CAGR≥15%→5
    - N_新催化（15分）：距52周高点≤15%→15，≤25%→8，否则0
    - S_中小盘（10分）：总市值<500亿→10，<1000亿→5，否则0
    - L_RPS（20分）：RPS≥80→20，≥60→10，否则0
    - I_流动性（10分）：换手率2%-15%→10，1%-20%→5，否则0
    - M_大势（15分）：多头排列+>MA200→15，任一满足→7，否则0
    
    kline_df: 含 'close'/'收盘', 'open'/'开盘', 'high'/'最高', 'low'/'最低', 'volume'/'成交量' 等列
    stock_pool_context: dict，含 'rps', 'market_cap', 'turnover_rate' 等
    """
    close_col = '收盘' if '收盘' in kline_df.columns else 'close'
    high_col = '最高' if '最高' in kline_df.columns else 'high'
    volume_col = '成交量' if '成交量' in kline_df.columns else 'volume'
    
    closes = kline_df[close_col].values.astype(float)
    highs = kline_df[high_col].values.astype(float)
    volumes = kline_df[volume_col].values.astype(float)
    n = len(closes)
    
    if n < 60:
        return {"pass": False, "综合评分": 0,
                "C_业绩增速": 0, "A_持续增长": 0, "N_新催化": 0,
                "S_中小盘": 0, "L_RPS": 0, "I_流动性": 0, "M_大势": 0}
    
    # --- C_业绩增速（15分）：扣非净利润同比 ---
    c_score = -1  # -1 = N/A: 财务数据不可用
    fin_cs = stock_pool_context.get('fin') if stock_pool_context else None
    if fin_cs and fin_cs.get('success'):
        g = fin_cs.get('eps_growth_yoy', 0)
        if g >= 0.25:
            c_score = 15
        elif g >= 0.15:
            c_score = 10
        elif g >= 0.05:
            c_score = 5
        else:
            c_score = 0
    
    # --- A_持续增长（15分）：ROE + 3年净利润CAGR ---
    a_score = -1  # -1 = N/A: 财务数据不可用
    if fin_cs and fin_cs.get('success'):
        roe = fin_cs.get('roe', 0)
        cagr = fin_cs.get('eps_cagr_3y')
        ok_roe = roe >= 0.17
        ok_cagr = cagr is not None and cagr >= 0.15
        if ok_roe and (cagr is not None and cagr >= 0.20):
            a_score = 15
        elif ok_roe:
            a_score = 8
        elif ok_cagr:
            a_score = 5
        else:
            a_score = 0
    
    # --- N_新催化：股价距52周高点 ---
    n_score = 0
    lookback = min(250, n)
    high_52w = np.max(highs[-lookback:])
    current_close = closes[-1]
    if high_52w > 0:
        dist_from_high = (high_52w - current_close) / high_52w * 100
        if dist_from_high <= 15:
            n_score = 15
        elif dist_from_high <= 25:
            n_score = 8
    
    # --- S_中小盘（优先财务数据，fallback 上下文） ---
    s_score = -1  # -1 = N/A: 无市值数据
    cap_yi = None
    if fin_cs and fin_cs.get('success') and fin_cs.get('market_cap', 0) > 0:
        cap_yi = fin_cs['market_cap']
    elif stock_pool_context:
        market_cap = stock_pool_context.get('market_cap', 0)
        if market_cap > 0:
            cap_yi = market_cap / 1e8  # 元转亿
    if cap_yi is not None and cap_yi > 0:
        if cap_yi < 500:
            s_score = 10
        elif cap_yi < 1000:
            s_score = 5
        else:
            s_score = 0
    
    # --- L_RPS ---
    l_score = 0
    if stock_pool_context:
        rps = stock_pool_context.get('rps', 0)
        if rps >= 80:
            l_score = 20
        elif rps >= 60:
            l_score = 10
    
    # --- I_流动性（优先财务数据，fallback 上下文） ---
    i_score = -1  # -1 = N/A: 无换手率数据（与 C/A 维度一致）
    turnover = 0
    if fin_cs and fin_cs.get('success') and fin_cs.get('turnover_rate', 0) > 0:
        turnover = fin_cs['turnover_rate']
    elif stock_pool_context:
        raw_t = stock_pool_context.get('turnover_rate', 0)
        if raw_t > 0:
            turnover = raw_t
    if turnover > 0:
        if 2 <= turnover <= 15:
            i_score = 10
        elif 1 <= turnover <= 20:
            i_score = 5
        else:
            i_score = 0
    
    # --- M_大势：均线多头排列 + 价格>MA200 ---
    m_score = 0
    cond1 = False  # 均线多头排列 MA5>MA20>MA60
    cond2 = False  # 收盘价 > MA200
    
    if n >= 60:
        ma5 = np.mean(closes[-5:])
        ma20 = np.mean(closes[-20:])
        ma60 = np.mean(closes[-60:])
        if ma5 > ma20 > ma60:
            cond1 = True
    
    if n >= 200:
        ma200 = np.mean(closes[-200:])
        if current_close > ma200:
            cond2 = True
    
    if cond1 and cond2:
        m_score = 15
    elif cond1 or cond2:
        m_score = 7
    
    # 综合评分：N/A（-1）维度按0计入总分
    def _safe(v):
        return 0 if v < 0 else v
    total = _safe(c_score) + _safe(a_score) + _safe(n_score) + _safe(s_score) + _safe(l_score) + _safe(i_score) + _safe(m_score)
    
    return {
        "pass": True,
        "综合评分": total,
        "C_业绩增速": c_score,
        "A_持续增长": a_score,
        "N_新催化": n_score,
        "S_中小盘": s_score,
        "L_RPS": l_score,
        "I_流动性": i_score,
        "M_大势": m_score,
    }


def calculate_dilemma_reversal_score(code, kline_df, stock_pool_context=None):
    """
    困境反转模型（抓周期股）四层简化版评分（Layer 4 完整实现，Layers 1-3 占位）。
    
    维度（满分100）：
    - L1_拐点（25分）：ROE环比.>1.0→6+NPR>1.20→7+NPR-1>0.70→4+NPR-2>0→3+RevR>1.0→5
    - L2_反转（15分）：毛利率恢复率>0.90→7+经营现金流恢复率>0.90→8
    - L3_安全垫（20分）：PB<2.0→8+负债率<70%→7+商誉/净资产<20%→5
    - L4_技术资金（40分）：
      * 突破均线（10分）：>MA250→10，否则0
      * MACD信号（10分）：标准金叉→10，DIF加速转正→5
      * 主力放量（10分）：5日均量≥60日均量1.2倍→10，否则0
      * 低位启动（10分）：250日位置≤50%→10，否则0
    
    kline_df: 含 'close'/'收盘', 'open'/'开盘', 'high'/'最高', 'low'/'最低', 'volume'/'成交量' 等列
    """
    close_col = '收盘' if '收盘' in kline_df.columns else 'close'
    open_col = '开盘' if '开盘' in kline_df.columns else 'open'
    high_col = '最高' if '最高' in kline_df.columns else 'high'
    low_col = '最低' if '最低' in kline_df.columns else 'low'
    volume_col = '成交量' if '成交量' in kline_df.columns else 'volume'
    
    closes = kline_df[close_col].values.astype(float)
    opens = kline_df[open_col].values.astype(float)
    highs = kline_df[high_col].values.astype(float)
    lows = kline_df[low_col].values.astype(float)
    volumes = kline_df[volume_col].values.astype(float)
    n = len(closes)
    
    if n < 60:
        return {"pass": False, "综合评分": 0,
                "L1_拐点": 0, "L2_反转": 0, "L3_安全垫": 0, "L4_技术资金": 0}
    
    # --- L1_拐点（25分）：财务恢复率（优先）→ 技术面近似（fallback） ---
    l1_score = 0
    fin_dr = stock_pool_context.get('fin') if stock_pool_context else None
    fin_ok = bool(fin_dr and fin_dr.get('success'))
    # 🔧 财务数据空值检测：API返回success=True但关键字段全为0/None → 视为无效，回退技术面
    if fin_ok:
        _fin_key_fields = ['roe_ttm', 'roe_recovery_q', 'np_recovery_q', 'np_recovery_q_1',
                           'np_recovery_q_2', 'rev_recovery_q', 'gm_recovery_q', 'ocf_recovery_q',
                           'pb', 'debt_ratio', 'goodwill_ratio']
        fin_ok = any(fin_dr.get(f) for f in _fin_key_fields if fin_dr.get(f) is not None)
    if fin_ok:
        # roe_ttm > 0 且 ROE环比恢复率 > 1.0（环比改善）
        if fin_dr.get('roe_ttm', 0) > 0 and fin_dr.get('roe_recovery_q', 0) > 1.0:
            l1_score += 6
        # 净利润恢复率 T/ABS(T-4) > 1.20
        if fin_dr.get('np_recovery_q', 0) > 1.20:
            l1_score += 7
        # T-1期 净利润恢复率 > 0.70
        if fin_dr.get('np_recovery_q_1', 0) > 0.70:
            l1_score += 4
        # T-2期 净利润 > 0（不亏损）
        if fin_dr.get('np_recovery_q_2', 0) > 0:
            l1_score += 3
        # 营收恢复率 > 1.0
        if fin_dr.get('rev_recovery_q', 0) > 1.0:
            l1_score += 5
    else:
        # 🔧 P0修复: 技术面fallback收紧，满分上限18（财务路径满分25，体现"无财务数据的惩罚"）
        # 原 fallback 为 10+8+7=25 过于宽松 → 回调深度分档 + 负面过滤
        l1_fallback_max = 18  # fallback满分上限

        # --- A. 回调深度分档（0~9分） ---
        if n >= 60:
            high_60 = float(np.max(highs[-60:]))
            if high_60 > 0:
                drawdown_pct = (high_60 - closes[-1]) / high_60 * 100
                # 深度回调 > 50% → 拐点概率最高
                if drawdown_pct >= 50:
                    l1_score += 9
                elif drawdown_pct >= 40:
                    l1_score += 7
                elif drawdown_pct >= 30:
                    l1_score += 5
                elif drawdown_pct >= 20:
                    l1_score += 3
                else:
                    l1_score += 1  # 浅回调，拐点确认不充分

        # --- B. 均线止跌分档（0~6分） ---
        if n >= 20:
            low_20 = float(np.min(lows[-20:]))
            low_20_pos = int(np.argmin(lows[-20:]))
            days_since_low = 19 - low_20_pos
            # 走平天数越多，止跌越确定
            if days_since_low >= 7:
                l1_score += 6   # 底部已确立超过一周
            elif days_since_low >= 4:
                l1_score += 4   # 底部确立4天以上
            elif days_since_low >= 2:
                l1_score += 2   # 底部确立但时间短

        # --- C. 量能确认（0~3分） ---
        if n >= 20:
            vol_ma5 = np.mean(volumes[-5:])
            vol_ma20 = np.mean(volumes[-20:])
            if vol_ma20 > 0:
                vol_ratio = vol_ma5 / vol_ma20
                if vol_ratio > 1.3:
                    l1_score += 3   # 放量转折
                elif vol_ratio > 1.0:
                    l1_score += 2   # 温和放量
                elif vol_ratio > 0.8:
                    l1_score += 1   # 量能平稳

        # --- D. 负面过滤（扣分项） ---
        # 近期仍有跌停 → 扣3分
        for i in range(max(-5, -n), 0):
            if lows[i] > 0 and (closes[i] - opens[i]) / lows[i] < -0.095:
                l1_score -= 3
                break
        # 近5日单日跌幅>7% → 扣2分
        for i in range(max(-5, -n), -1):
            if closes[i] > 0:
                daily_chg = (closes[i+1] - closes[i]) / closes[i] * 100
                if daily_chg < -7:
                    l1_score -= 2
                    break

        # 上限封顶
        l1_score = max(0, min(l1_score, l1_fallback_max))

    # --- L2_反转（15分）：毛利率+经营现金流（优先）→ 技术面近似（fallback） ---
    l2_score = 0
    if fin_ok:
        if fin_dr.get('gm_recovery_q', 0) > 0.90:
            l2_score += 7
        if fin_dr.get('ocf_recovery_q', 0) > 0.90:
            l2_score += 8
    else:
        # 技术面近似：MACD信号(8) + 量价配合(7)
        if n >= 35:
            ema12_d = pd.Series(closes[-60:]).ewm(span=12, adjust=False).mean().values
            ema26_d = pd.Series(closes[-60:]).ewm(span=26, adjust=False).mean().values
            dif_d = ema12_d - ema26_d
            if len(dif_d) >= 3:
                if dif_d[-2] <= 0 and dif_d[-1] > 0:
                    l2_score += 8  # MACD金叉
                elif dif_d[-1] > dif_d[-2] > dif_d[-3] and dif_d[-1] > -0.02 * closes[-1]:
                    l2_score += 4  # DIFF连续递增且接近零轴
        if n >= 4:
            up_vol_days = 0
            for i in range(-3, 0):
                if closes[i] > closes[i - 1] and volumes[i] > volumes[i - 1]:
                    up_vol_days += 1
            if up_vol_days >= 2:
                l2_score += 7  # 近3日有2日放量上涨
            elif up_vol_days >= 1:
                l2_score += 3

    # --- L3_安全垫（20分）：PB/负债率/商誉（优先）→ 技术面近似（fallback） ---
    l3_score = 0
    if fin_ok:
        if fin_dr.get('pb', 999) < 2.0:
            l3_score += 8
        if fin_dr.get('debt_ratio', 999) < 0.70:
            l3_score += 7
        if fin_dr.get('goodwill_ratio', 999) < 0.20:
            l3_score += 5
    else:
        # 技术面近似：250日位置低(10) + 回撤充分(10)
        if n >= 60:
            n250 = min(250, n)
            low_250 = float(np.min(lows[-n250:]))
            high_250 = float(np.max(highs[-n250:]))
            if high_250 > low_250:
                pos_pct = (closes[-1] - low_250) / (high_250 - low_250) * 100
                if pos_pct <= 25:
                    l3_score += 10
                elif pos_pct <= 40:
                    l3_score += 7
                elif pos_pct <= 60:
                    l3_score += 4
            n60 = min(60, n)
            high_60 = float(np.max(highs[-n60:]))
            if high_60 > 0:
                drawdown_60 = (closes[-1] / high_60 - 1.0) * 100
                if drawdown_60 <= -30:
                    l3_score += 10
                elif drawdown_60 <= -20:
                    l3_score += 7
                elif drawdown_60 <= -15:
                    l3_score += 4
    
    # --- L4_技术资金（40分）---
    l4_score = 0
    
    # 1) 突破250日线 (10分)
    if n >= 250:
        ma250 = np.mean(closes[-250:])
        if closes[-1] > ma250:
            l4_score += 10
    
    # 2) 周线MACD金叉 (10分) — 用近5根日线模拟周线趋势
    if n >= 30:
        weekly_closes = closes[-25:]
        w_close = []
        for i in range(0, len(weekly_closes), 5):
            seg = weekly_closes[i:i+5]
            w_close.append(seg[-1])
        if len(w_close) >= 9:
            ema12 = pd.Series(w_close).ewm(span=12, adjust=False).mean().values
            ema26 = pd.Series(w_close).ewm(span=26, adjust=False).mean().values
            dif = ema12 - ema26
            if len(dif) >= 3 and dif[-2] <= 0 and dif[-1] > 0:
                l4_score += 10
            elif len(dif) >= 2 and dif[-2] <= 0 and dif[-1] < dif[-2] * -1:
                # DIF 从负转正加速
                l4_score += 5
    
    # 3) 主力放量：5日均量 vs 60日均量 (10分)
    if n >= 60:
        vol_ma5 = np.mean(volumes[-5:])
        vol_ma60 = np.mean(volumes[-60:])
        if vol_ma60 > 0:
            vol_ratio = vol_ma5 / vol_ma60
            if vol_ratio >= 1.2:
                l4_score += 10
    
    # 4) 低位启动：股价在250日位置 (10分)
    if n >= 250:
        low_250 = np.min(lows[-250:])
        high_250 = np.max(highs[-250:])
        if high_250 > low_250:
            position_pct = (closes[-1] - low_250) / (high_250 - low_250) * 100
            if position_pct <= 50:
                l4_score += 10
    
    total = l1_score + l2_score + l3_score + l4_score
    
    return {
        "pass": True,
        "综合评分": total,
        "L1_拐点": l1_score,
        "L2_反转": l2_score,
        "L3_安全垫": l3_score,
        "L4_技术资金": l4_score,
    }


# ================================================================
#              超跌反弹模型（抓超跌股）
# ================================================================

def hard_filter_oversold_rebound(kline_df, stock_data=None):
    """超跌反弹硬过滤 — 满足任一即淘汰。
    规则：
    1. ST/*ST
    2. 近60日最高点回调幅度 < 15%（跌太少），但若近20日跌幅>=5%且连续下跌>=3天则放行
    3. 近5日出现过跌停（排除暴雷股）
    4. 近3日连续放量下跌（排除出货型下跌）
    5. 股价 < 1元（排除仙股）
    返回: (True=保留, filter_msg)
    """
    close_col = '收盘' if '收盘' in kline_df.columns else 'close'
    open_col = '开盘' if '开盘' in kline_df.columns else 'open'
    high_col = '最高' if '最高' in kline_df.columns else 'high'
    volume_col = '成交量' if '成交量' in kline_df.columns else 'volume'

    closes = kline_df[close_col].values.astype(float)
    opens = kline_df[open_col].values.astype(float)
    highs = kline_df[high_col].values.astype(float)
    volumes = kline_df[volume_col].values.astype(float)
    n = len(closes)

    # 5. 股价 < 1 元
    if stock_data:
        price = stock_data.get("close", closes[-1] if n > 0 else 0)
    else:
        price = closes[-1] if n > 0 else 0
    if price < 1.0:
        return False, f"股价{price:.2f}<1元"

    if n < 20:
        return False, "K线数据不足20日"

    # 2. 近60日最高点回调幅度 < 20%（例外：近20日跌幅≥10%且连续下跌≥5天放行）
    n60 = min(60, n)
    high_60 = float(np.max(highs[-n60:]))
    if high_60 > 0:
        drawdown_60 = (closes[-1] / high_60 - 1.0) * 100
    else:
        drawdown_60 = 0.0

    # 计算近20日涨跌幅 和 连续下跌天数（用于例外放行判断）
    change_20d = (closes[-1] / closes[-21] - 1) * 100 if n >= 21 else (closes[-1] / closes[0] - 1) * 100
    consecutive_down = 0
    for i in range(-1, -n, -1):
        if closes[i] < closes[i - 1]:
            consecutive_down += 1
        else:
            break

    # 例外放行：近20日跌幅>=5% 且 连续下跌>=3天
    exception_pass = (change_20d <= -5 and consecutive_down >= 3)

    if drawdown_60 > -15 and not exception_pass:
        return False, f"近60日回调{drawdown_60:.1f}%>-15%（跌幅不够，20日跌{change_20d:.1f}%，连跌{consecutive_down}天）"

    # 3. 近5日出现过跌停
    for i in range(max(-5, -n), 0):
        if n + i - 1 >= 0 and closes[i] > 0 and opens[i] > 0:
            daily_chg = (closes[i] / closes[i - 1] - 1) * 100 if i > -n else 0
            if daily_chg <= -9.9:
                return False, f"近5日内出现跌停"

    # 4. 近3日连续放量下跌（量逐日放大 + 每日收阴）
    if n >= 4:
        fail_count = 0
        for i in range(-3, 0):
            if closes[i] < opens[i] and i > -n:
                if i > -n + 1 and volumes[i] > volumes[i - 1]:
                    fail_count += 1
        if fail_count >= 2:
            return False, "近3日连续放量下跌"

    return True, "过滤通过"


def _calc_5d_gain(closes):
    """从收盘价数组计算近5日涨跌幅(%)"""
    if len(closes) < 6:
        return 0
    return (closes[-1] / closes[-6] - 1.0) * 100


def calculate_oversold_rebound_score(kline_df, stock_data=None):
    """超跌反弹模型：空间/情绪量能/择时确认/板块共振 四维评分。

    kline_df: 含 收盘/开盘/最高/最低/成交量 列（支持中英文列名）
    stock_data: 可选，含 sector 字段用于板块共振加成
    返回: {pass, 综合评分, 空间维度, 情绪量能, 择时确认, 板块共振}
    """
    close_col = '收盘' if '收盘' in kline_df.columns else 'close'
    open_col = '开盘' if '开盘' in kline_df.columns else 'open'
    high_col = '最高' if '最高' in kline_df.columns else 'high'
    low_col = '最低' if '最低' in kline_df.columns else 'low'
    volume_col = '成交量' if '成交量' in kline_df.columns else 'volume'

    closes = kline_df[close_col].values.astype(float)
    opens = kline_df[open_col].values.astype(float)
    highs = kline_df[high_col].values.astype(float)
    lows = kline_df[low_col].values.astype(float)
    volumes = kline_df[volume_col].values.astype(float)
    n = len(closes)

    if n < 20:
        return {"pass": True, "综合评分": 0, "空间维度": 0, "情绪量能": 0, "择时确认": 0, "板块共振": 0}

    # ============ 一、空间维度（满分40）============
    space_score = 0

    # 1. 跌幅达标（25分）— 近120日最高点回调幅度
    if n >= 120:
        high_120 = float(np.max(highs[-120:]))
    elif n >= 60:
        high_120 = float(np.max(highs[-n:]))
    else:
        high_120 = float(np.max(highs[-n:]))
    if high_120 > 0:
        drawdown = (closes[-1] / high_120 - 1.0) * 100  # 负值 = 下跌
        if drawdown <= -50:
            space_score += 25
        elif drawdown <= -40:
            space_score += 20
        elif drawdown <= -30:
            space_score += 15
        elif drawdown < -10:
            space_score += 5

    # 2. BIAS负乖离（15分）— 相对MA20
    if n >= 20:
        ma20 = float(np.mean(closes[-20:]))
        if ma20 > 0:
            bias = (closes[-1] / ma20 - 1.0) * 100
            if bias < -15:
                space_score += 15
            elif bias < -10:
                space_score += 10
            elif bias < -5:
                space_score += 5

    # ============ 二、情绪量能维度（满分30）============
    sentiment_score = 0

    # 1. 成交量萎缩（15分）
    if n >= 60:
        vol_ma5 = float(np.mean(volumes[-5:]))
        vol_max_60 = float(np.max(volumes[-60:]))
        if vol_max_60 > 0:
            vol_ratio = vol_ma5 / vol_max_60
            if vol_ratio <= 0.25:
                sentiment_score += 15
            elif vol_ratio <= 0.33:
                sentiment_score += 12
            elif vol_ratio <= 0.5:
                sentiment_score += 7
            else:
                sentiment_score += 2

    # 1b. 连续下跌天数加分
    consecutive_down = 0
    for i in range(-1, -n, -1):
        if closes[i] < closes[i - 1]:
            consecutive_down += 1
        else:
            break
    if consecutive_down >= 8:
        sentiment_score += 8
    elif consecutive_down >= 5:
        sentiment_score += 5
    elif consecutive_down >= 3:
        sentiment_score += 3

    # 2. K线止跌信号（15分）— 近3日
    if n >= 3:
        kline_signals = 0
        for i in range(max(-3, -n), 0):
            body = abs(closes[i] - opens[i])
            if closes[i] >= opens[i]:
                lower_shadow = opens[i] - lows[i]
            else:
                lower_shadow = closes[i] - lows[i]
            amplitude = highs[i] - lows[i]

            # 长下影线（下影线 >= 实体2倍）
            if body > 0 and lower_shadow >= body * 2:
                kline_signals = max(kline_signals, 8)

            # 十字星
            if amplitude > 0 and body / amplitude < 0.3:
                kline_signals = max(kline_signals, 5)

        # 连续2日小阳线且不创新低
        bar1_positive = closes[-1] > opens[-1]
        bar2_positive = closes[-2] > opens[-2]
        recent_low = float(np.min(lows[-3:]))
        if n >= 6:
            prev_low = float(np.min(lows[-6:-3]))
        else:
            prev_low = recent_low
        if bar1_positive and bar2_positive and recent_low >= prev_low:
            kline_signals += 7

        sentiment_score += min(kline_signals, 15)

    # ============ 三、择时确认维度（满分30）============
    timing_score = 0

    # 1. 站上MA5（10分）
    if n >= 6:
        ma5 = float(np.mean(closes[-5:]))
        ma5_prev = float(np.mean(closes[-6:-1]))
        if closes[-1] > ma5:
            if ma5 >= ma5_prev:
                timing_score += 10
            else:
                timing_score += 5

    # 2. KDJ J值拐头（5分）
    if n >= 13:
        try:
            # 计算 KDJ（9,3,3）
            k_values = [50.0, 50.0]  # 初始值
            d_values = [50.0, 50.0]
            for idx in range(8, n):
                low_9 = float(np.min(lows[idx - 8:idx + 1]))
                high_9 = float(np.max(highs[idx - 8:idx + 1]))
                rsv = ((closes[idx] - low_9) / (high_9 - low_9) * 100) if high_9 > low_9 else 50.0
                k = 2.0 / 3.0 * k_values[-1] + 1.0 / 3.0 * rsv
                d = 2.0 / 3.0 * d_values[-1] + 1.0 / 3.0 * k
                k_values.append(k)
                d_values.append(d)
            j_vals = [3.0 * k_values[i] - 2.0 * d_values[i] for i in range(len(k_values))]
            if len(j_vals) >= 3:
                j_now = j_vals[-1]
                j_prev = j_vals[-2]
                if j_now < 0 and j_now > j_prev:
                    timing_score += 5
        except Exception:
            pass

    # 3. MACD绿柱缩短/底背离（5分）
    if n >= 27:
        try:
            ema12 = pd.Series(closes).ewm(span=12, adjust=False).mean().values
            ema26 = pd.Series(closes).ewm(span=26, adjust=False).mean().values
            dif = ema12 - ema26
            dea = pd.Series(dif).ewm(span=9, adjust=False).mean().values
            macd_bar = 2.0 * (dif - dea)

            if len(macd_bar) >= 3:
                bar_now = abs(macd_bar[-1])
                bar_prev = abs(macd_bar[-2])
                # 绿柱缩短
                if macd_bar[-1] < 0 and bar_now < bar_prev:
                    timing_score += 5
        except Exception:
            pass

    # 4. 放量转强（10分）
    if n >= 5:
        vol_ma5 = float(np.mean(volumes[-5:]))
        if len(volumes) >= 2:
            if volumes[-1] > volumes[-2] * 1.2 and volumes[-1] > vol_ma5 * 1.2:
                if closes[-1] > opens[-1]:
                    timing_score += 10
                else:
                    timing_score += 5

    # ============ 四、板块共振加成（+10，不计入满分100）============
    sector_bonus = 0
    if stock_data and stock_data.get("sector"):
        sector_name = stock_data["sector"]
        fund_flows = _get_sector_fund_flow()
        sector_fallback_used = False
        if fund_flows:
            sector_flow = fund_flows.get(sector_name, None)
            if sector_flow is not None and sector_flow > 0:
                sector_bonus = 10
            else:
                sector_fallback_used = True
        else:
            sector_fallback_used = True

        # 🔧 P0修复: 板块资金流不可用时，用板块涨跌幅+涨停数做fallback
        if sector_fallback_used and sector_bonus == 0:
            try:
                sector_data, _ = _get_cached_sector_data()
                if sector_data and sector_name in sector_data:
                    sd = sector_data[sector_name]
                    # pytdx 数据字段为 daily_gain / limit_up_count，非中文键名
                    sec_gain = float(sd.get('daily_gain', 0) or 0)
                    sec_zt = int(sd.get('limit_up_count', 0) or 0)
                    # 板块涨幅>1%且涨停≥3 → +10; 涨幅>0 → +5
                    if sec_gain > 1.0 and sec_zt >= 3:
                        sector_bonus = 10
                    elif sec_gain > 0 and sec_zt >= 1:
                        sector_bonus = 5
                    elif sec_gain > 0:
                        sector_bonus = 3
                    # daily_gain/limit_up_count 为0时（pytdx未填充），尝试多级fallback
                    elif sec_gain == 0 and sec_zt == 0:
                        # 第3层fallback：从K线计算5日涨幅，侧面反映板块可能有企稳
                        _chg5 = _calc_5d_gain(closes) if n >= 6 else 0
                        if _chg5 > 5:
                            sector_bonus = 3
                elif sector_data is not None and sector_name not in sector_data:
                    # 第4层fallback：板块不在 sector_data 中，用个股自身5日涨幅
                    _chg5 = _calc_5d_gain(closes) if n >= 6 else 0
                    if _chg5 > 5:
                        sector_bonus = 3
                elif sector_data is None:
                    # sector_data为空 → 用个股5日涨幅做弱fallback
                    _chg5 = _calc_5d_gain(closes) if n >= 6 else 0
                    if _chg5 > 5:
                        sector_bonus = 3
            except Exception:
                pass

    total = space_score + sentiment_score + timing_score

    return {
        "pass": True,
        "综合评分": total,
        "空间维度": space_score,
        "情绪量能": sentiment_score,
        "择时确认": timing_score,
        "板块共振": sector_bonus,
    }


# ================================================================
#              UI: 顶部参数面板（折叠式）
# ================================================================

def render_top_params_panel():
    """页面顶部的可折叠参数调整面板"""
    model = st.session_state.current_model
    
    if model == 'chase_high':
        _render_chase_high_params(prefix="")
    elif model == 'canslim':
        _render_chase_high_params(prefix="cs_")
    elif model == 'dilemma_reversal':
        _render_chase_high_params(prefix="dr_")
    elif model == 'oversold_rebound':
        _render_orb_params()
    elif model == 'rebound_model':
        _render_gc_params()
    else:
        _render_lowbuy_params(model)


def _render_chase_high_params(prefix=""):
    """追高模型参数面板（prefix 用于避免多 Tab 渲染时的 key 冲突）"""
    with st.expander("⚙️ 追高模型 · 十维权重设置(v4)", expanded=False):
        st.markdown("""
        <table class="weight-table">
          <tr><th>维度</th><th>指标依据</th><th>默认</th><th>范围</th></tr>
          <tr><td><b>📐 趋势结构</b></td><td>MA5/10/20/60多头</td><td><b>15%</b></td><td><span class="range-badge">0~35%</span></td></tr>
          <tr><td><b>🚀 动量强度</b></td><td>5日+10日涨幅</td><td><b>18%</b></td><td><span class="range-badge">0~35%</span></td></tr>
          <tr><td><b>🌐 板块共振</b></td><td>板块涨幅/涨停/资金流</td><td><b>8%</b></td><td><span class="range-badge">0~25%</span></td></tr>
          <tr><td><b>👆 北向资金</b></td><td>北向资金近3日净买</td><td><b>15%</b></td><td><span class="range-badge">0~25%</span></td></tr>
          <tr><td><b>🏦 机构净买</b></td><td>机构3日净买额</td><td><b>10%</b></td><td><span class="range-badge">0~25%</span></td></tr>
          <tr><td><b>🔥 板块资金热度</b></td><td>板块资金排名映射</td><td><b>5%</b></td><td><span class="range-badge">0~15%</span></td></tr>
          <tr><td><b>📈 量价配合</b></td><td>量比/振幅/缩量新高</td><td><b>14%</b></td><td><span class="range-badge">0~30%</span></td></tr>
          <tr><td><b>🛡️ 估值安全</b></td><td>PE历史分位(赛道差异化)</td><td><b>3%</b></td><td><span class="range-badge">0~25%</span></td></tr>
          <tr><td><b>🔒 筹码稳定</b></td><td>换手率接力</td><td><b>6%</b></td><td><span class="range-badge">0~20%</span></td></tr>
          <tr><td><b>🎯 情绪热度</b></td><td>热度分(主线/冷门差异化)</td><td><b>6%</b></td><td><span class="range-badge">0~15%</span></td></tr>
        </table>
        """, unsafe_allow_html=True)
        
        # +/- 按钮式权重调整（每次点击改动1%）
        _items = list(WEIGHT_CONFIG.items())
        # 处理 +/- 按钮点击
        for dim, cfg in _items:
            if st.session_state.get(f"{prefix}_winc_{dim}", False):
                cur = st.session_state.weights.get(dim, cfg["default"])
                st.session_state.weights[dim] = min(cfg["max"], cur + 1)
                st.session_state[f"{prefix}_winc_{dim}"] = False
            if st.session_state.get(f"{prefix}_wdec_{dim}", False):
                cur = st.session_state.weights.get(dim, cfg["default"])
                st.session_state.weights[dim] = max(0, cur - 1)
                st.session_state[f"{prefix}_wdec_{dim}"] = False
        # 渲染权重行
        total_w = sum(st.session_state.weights.get(dim, cfg["default"]) for dim, cfg in _items)
        for dim, cfg in _items:
            cur = st.session_state.weights.get(dim, cfg["default"])
            clr = cfg.get("color", "#333")
            bc1, bc2, bc3, bc4, bc5 = st.columns([0.6, 2.5, 0.8, 0.6, 1.2])
            with bc1:
                st.markdown(f"<span style='color:{clr};font-size:13px;'>{cfg['icon']}</span>", unsafe_allow_html=True)
            with bc2:
                st.markdown(f"<span style='font-size:13px;color:#333;font-weight:600;'>{dim}</span>", unsafe_allow_html=True)
            with bc3:
                if st.button("−", key=f"{prefix}wdec_{dim}", help=f"减少1%（当前{cur}%）"):
                    st.session_state[f"{prefix}_wdec_{dim}"] = True
                    st.rerun()
            with bc4:
                st.markdown(f"<span style='font-size:16px;font-weight:800;color:{clr};'>{cur}%</span>", unsafe_allow_html=True)
            with bc5:
                if st.button("+", key=f"{prefix}winc_{dim}", help=f"增加1%（当前{cur}%，上限{cfg['max']}%）"):
                    st.session_state[f"{prefix}_winc_{dim}"] = True
                    st.rerun()
        
        wc = "#27AE60" if total_w == 100 else "#E74C3C"
        st.markdown(f"<span style='font-size:14px;color:{wc};font-weight:700;'>总权重: {total_w}% {'✅' if total_w==100 else '⚠️ 建议100%'}</span>", unsafe_allow_html=True)
        c2, c3 = st.columns(2)
        with c2:
            if st.button("✅ 应用权重", width='stretch', type="primary", key=f"{prefix}chase_apply_weights"):
                st.cache_data.clear()
                st.rerun()
        with c3:
            if st.button("🔄 重置默认", width='stretch', key=f"{prefix}chase_reset_weights"):
                st.session_state.weights = dict(DEFAULT_WEIGHTS)
                st.session_state.top10_cache = None
                st.session_state.top10_cache_key = None
                st.cache_data.clear()
                st.rerun()


def _render_gc_params():
    """强势超跌金叉模型参数面板 — 硬过滤阈值 + 五维权重"""
    with st.expander("⚙️ 金叉模型 · 筛选规则与权重", expanded=False):
        st.markdown("""
        <table class="weight-table">
          <tr><th>维度</th><th>评分依据</th><th>默认</th><th>范围</th></tr>
          <tr><td><b>📉 下跌形态</b></td><td>回撤深度+急跌识别</td><td><b>28%</b></td><td><span class="range-badge">15~40%</span></td></tr>
          <tr><td><b>🛑 K线止跌</b></td><td>不创新低+底部K线</td><td><b>18%</b></td><td><span class="range-badge">10~35%</span></td></tr>
          <tr><td><b>📐 均线拐头</b></td><td>5日线走平+站上5日线</td><td><b>18%</b></td><td><span class="range-badge">10~35%</span></td></tr>
          <tr><td><b>📊 量能确认</b></td><td>放量倍数+缩量验证</td><td><b>13%</b></td><td><span class="range-badge">5~30%</span></td></tr>
          <tr><td><b>✂️ MACD反转</b></td><td>DIFF走平+绿柱缩短+金叉</td><td><b>13%</b></td><td><span class="range-badge">5~30%</span></td></tr>
          <tr><td><b>💰 资金确认</b></td><td>主力净流入方向/量能替代</td><td><b>10%</b></td><td><span class="range-badge">0~20%</span></td></tr>
        </table>
        """, unsafe_allow_html=True)
        
        # 硬过滤阈值
        st.markdown("**硬过滤阈值**")
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            new_rise = st.number_input("主升涨幅≥%", min_value=20, max_value=100,
                                       value=st.session_state.gc_params.get("main_rise_pct", 50),
                                       step=5, key="gc_rise_pct")
        with cf2:
            new_decline = st.number_input("回撤幅度≥%", min_value=15, max_value=60,
                                          value=st.session_state.gc_params.get("decline_pct", 30),
                                          step=5, key="gc_decline_pct")
        with cf3:
            new_vol = st.number_input("放量倍数≥", min_value=0.5, max_value=5.0,
                                      value=st.session_state.gc_params.get("volume_ratio", 1.0),
                                      step=0.1, key="gc_vol_ratio")
        
        st.markdown("**五维 → 六维权重调整**")
        dims = [
            ("下跌形态", 28, "#E74C3C"),
            ("K线止跌", 18, "#F39C12"),
            ("均线拐头", 18, "#3498DB"),
            ("量能确认", 13, "#2ECC71"),
            ("MACD反转", 13, "#9B59B6"),
            ("资金确认", 10, "#E67E22"),
        ]
        for dim_name, default_w, color in dims:
            cur = st.session_state.gc_weights.get(dim_name, default_w)
            bc1, bc2, bc3, bc4, bc5 = st.columns([0.3, 2.0, 0.6, 0.6, 1.0])
            with bc1:
                st.markdown(f"<span style='color:{color};font-size:16px;'>●</span>", unsafe_allow_html=True)
            with bc2:
                st.markdown(f"<span style='font-size:13px;color:#333;font-weight:600;'>{dim_name}</span>", unsafe_allow_html=True)
            with bc3:
                if st.button("−", key=f"gc_wdec_{dim_name}"):
                    st.session_state.gc_weights[dim_name] = max(0, cur - 1)
                    st.rerun()
            with bc4:
                st.markdown(f"<span style='font-size:16px;font-weight:800;color:{color};'>{st.session_state.gc_weights.get(dim_name, default_w)}%</span>", unsafe_allow_html=True)
            with bc5:
                if st.button("+", key=f"gc_winc_{dim_name}"):
                    st.session_state.gc_weights[dim_name] = min(40, cur + 1)
                    st.rerun()
        
        total = sum(st.session_state.gc_weights.get(d, w) for d, w, _ in dims)
        wc = "#27AE60" if total == 100 else "#E74C3C"
        st.markdown(f"<span style='font-size:14px;color:{wc};font-weight:700;'>总权重: {total}% {'✅' if total==100 else '⚠️ 建议100%'}</span>", unsafe_allow_html=True)
        
        ca1, ca2 = st.columns(2)
        with ca1:
            if st.button("✅ 应用参数", width='stretch', type="primary", key="gc_apply"):
                st.session_state.gc_params["main_rise_pct"] = new_rise
                st.session_state.gc_params["decline_pct"] = new_decline
                st.session_state.gc_params["volume_ratio"] = new_vol
                st.session_state.gc_results = None
                st.session_state.gc_scanned = False
                st.cache_data.clear()
                st.rerun()
        with ca2:
            if st.button("🔄 重置默认", width='stretch', key="gc_reset"):
                st.session_state.gc_params = dict(DEFAULT_GC_PARAMS)
                st.session_state.gc_weights = dict(DEFAULT_GC_WEIGHTS)
                st.session_state.gc_results = None
                st.session_state.gc_scanned = False
                st.cache_data.clear()
                st.rerun()


def _render_orb_params():
    """超跌反弹模型参数面板 — 四维权重设置"""
    with st.expander("⚙️ 超跌反弹模型 · 四维权重设置", expanded=False):
        st.markdown("""
        <table class="weight-table">
          <tr><th>维度</th><th>指标依据</th><th>默认</th><th>范围</th></tr>
          <tr><td><b>📉 空间维度</b></td><td>高位回调幅度 + BIAS乖离</td><td><b>40%</b></td><td><span class="range-badge">0~60%</span></td></tr>
          <tr><td><b>📊 情绪量能</b></td><td>成交量萎缩 + K线止跌信号</td><td><b>30%</b></td><td><span class="range-badge">0~50%</span></td></tr>
          <tr><td><b>🎯 择时确认</b></td><td>MA5站上 + KDJ/MACD + 放量转强</td><td><b>30%</b></td><td><span class="range-badge">0~50%</span></td></tr>
          <tr><td><b>🌐 板块加成</b></td><td>板块资金净流入额外加成</td><td><b>10%</b></td><td><span class="range-badge">0~20%</span></td></tr>
        </table>
        """, unsafe_allow_html=True)

        dims = [
            ("空间维度", 40, "#E74C3C"),
            ("情绪量能", 30, "#3498DB"),
            ("择时确认", 30, "#F39C12"),
            ("板块加成", 10, "#27AE60"),
        ]
        for dim_name, default_w, color in dims:
            cur = st.session_state.orb_weights.get(dim_name, default_w)
            bc1, bc2, bc3, bc4, bc5 = st.columns([0.3, 2.0, 0.6, 0.6, 1.0])
            with bc1:
                st.markdown(f"<span style='color:{color};font-size:16px;'>●</span>", unsafe_allow_html=True)
            with bc2:
                st.markdown(f"<span style='font-size:13px;color:#333;font-weight:600;'>{dim_name}</span>", unsafe_allow_html=True)
            with bc3:
                if st.button("−", key=f"orb_wdec_{dim_name}"):
                    st.session_state.orb_weights[dim_name] = max(0, cur - 1)
                    st.rerun()
            with bc4:
                st.markdown(f"<span style='font-size:16px;font-weight:800;color:{color};'>{st.session_state.orb_weights.get(dim_name, default_w)}%</span>", unsafe_allow_html=True)
            with bc5:
                if st.button("+", key=f"orb_winc_{dim_name}"):
                    cfg = ORB_WEIGHT_CONFIG.get(dim_name, {})
                    st.session_state.orb_weights[dim_name] = min(cfg.get("max", 60), cur + 1)
                    st.rerun()

        total = sum(st.session_state.orb_weights.get(d, w) for d, w, _ in dims)
        wc = "#27AE60" if 90 <= total <= 110 else "#E74C3C"
        st.markdown(f"<span style='font-size:14px;color:{wc};font-weight:700;'>总权重: {total}% {'✅' if 90 <= total <= 110 else '⚠️ 建议前三维度合计约100%'}</span>", unsafe_allow_html=True)

        ca1, ca2 = st.columns(2)
        with ca1:
            if st.button("✅ 应用参数", width='stretch', type="primary", key="orb_apply"):
                st.session_state.orb_results = None
                st.rerun()
        with ca2:
            if st.button("🔄 重置默认", width='stretch', key="orb_reset"):
                st.session_state.orb_weights = dict(DEFAULT_ORB_WEIGHTS)
                st.session_state.orb_params = dict(DEFAULT_ORB_PARAMS)
                st.session_state.orb_results = None
                st.cache_data.clear()
                st.rerun()


def _render_lowbuy_params(model='lowbuy'):
    """低吸模型参数面板（七维评分版，model用于key隔离避免多tab冲突）"""
    prefix = f"lp_{model}"
    with st.expander("⚙️ 低吸模型 · 筛选规则", expanded=False):
        params = _get_lowbuy_params()
        
        st.markdown("""
        <div style="background:#F0FFF4;border:1px solid #C8E6C9;border-radius:8px;padding:12px 16px;margin-bottom:16px;font-size:13px;color:#2E7D32;">
        <b>低吸模型 v4 筛选逻辑：</b><br>
        📊 七维评分：下跌幅度 ≥ min · 企稳信号 ≥ min · 量能恢复 ≥ min · 均线支撑 ≥ min · 估值吸引 ≥ min · 筹码沉淀 ≥ min · 主力资金 ≥ min · 综合 ≥ min<br>
        📉 K线条件：10日跌幅(decline_low ~ decline_high) · 量比 ≤ max · no_new_low_days日不创新低<br>
        🎯 信号分级：<b>综合≥55 + 机构净买→强烈低吸</b> / <b>综合≥40→标准低吸</b> / <b>综合≥25→谨慎低吸</b>
        </div>
        """, unsafe_allow_html=True)
        
        # ---- 第一行: K线预筛条件 ----
        c1, c2, c3 = st.columns(3)
        with c1:
            new_max = st.slider("🎯 每日精选数量", 5, 30, params.get('max_results', 20),
                help="筛选出的股票数量", key=f"{prefix}_max")
            new_decline_low = st.slider("📉 10日跌幅下限(%)", -50, -5, params.get('decline_20d_low', -40),
                help="10日跌幅不低于此值", key=f"{prefix}_dec_low")
            new_decline_high = st.slider("📉 10日跌幅上限(%)", -5, 0, params.get('decline_20d_high', 0),
                help="10日跌幅不高于此值", key=f"{prefix}_dec_high")
        with c2:
            new_max_vr = st.slider("📊 量比上限", 0.5, 3.0, params.get('max_vol_ratio', 2.5),
                step=0.1, help="量比不超过此值", key=f"{prefix}_vr")
            new_fund_wt = st.slider("💰 资金维度权重", 0.0, 0.20, params.get('fund_weight', 0.08),
                step=0.01, help="主力资金维度在综合评分中的权重", key=f"{prefix}_fundwt")
            new_no_low_days = st.slider("📈 不创新低天数", 2, 20, params.get('no_new_low_days', 2),
                help="近N日不创新低", key=f"{prefix}_nolow")

        # ---- 第二行: 反转确认条件 ----
        with c3:
            new_rev_pct = st.slider("📈 脱离底部≥(%)", 0.0, 3.0, params.get('reversal_bottom_pct', 0.2),
                step=0.1, help="近3日均线高于15日最低点的最小百分比，越小越宽松", key=f"{prefix}_rev_pct")
            new_rev_trend = st.checkbox("📈 MA3今日>MA3前日", value=params.get('reversal_require_uptrend', False),
                help="要求今日3日均价高于前日3日均价", key=f"{prefix}_rev_trend")

        # ---- 第三行: 六维门槛 (每列2个) ----
        with st.container():
            c4, c5, c6 = st.columns(3)
            with c4:
                new_min_decline = st.slider("📉 下跌幅度≥", 0, 25, params.get('min_decline_depth', 5),
                    help="raw 0-30，跌幅维度门槛", key=f"{prefix}_decline")
                new_min_stabil = st.slider("🛡️ 企稳信号≥", 0, 20, params.get('min_stabilization', 3),
                    help="raw 0-25，企稳维度门槛", key=f"{prefix}_stabil")
            with c5:
                new_min_vol = st.slider("📈 量能恢复≥", 0, 18, params.get('min_volume_recovery', 2),
                    help="raw 0-20，量能维度门槛", key=f"{prefix}_vol")
                new_min_ma = st.slider("📐 均线支撑≥", 0, 14, params.get('min_ma_support', 4),
                    help="raw 0-15，均线维度门槛", key=f"{prefix}_ma")
            with c6:
                new_min_val = st.slider("💎 估值吸引≥", 0, 9, params.get('min_valuation_attr', 4),
                    help="raw 0-10，估值维度门槛", key=f"{prefix}_val_attr")
                new_min_chip = st.slider("🔒 筹码沉淀≥", 0, 9, params.get('min_chip_settle', 3),
                    help="raw 0-10，筹码维度门槛", key=f"{prefix}_chip_set")
                new_min_fund = st.slider("💰 主力资金≥", 0, 9, params.get('min_fund_flow', 2),
                    help="raw 0-10，资金维度门槛", key=f"{prefix}_fund")
        
        # ---- 第四行: 综合分 + 操作 ----
        nc1, nc2, nc3 = st.columns([2, 1, 1])
        with nc1:
            new_min_total = st.slider("🎯 综合最低分", 10, 60, params.get('min_total_score', 20),
                help="归一化综合评分(0-100)最低门槛，越低越宽松", key=f"{prefix}_total")
            st.markdown(f"""
            <div style="font-size:12px;color:#666;padding:8px 0;">
            当前：10日跌<b>{new_decline_low}%~{new_decline_high}%</b> · 量比≤<b>{new_max_vr}</b> · {new_no_low_days}日不新低 · 脱离≥<b>{new_rev_pct}%</b> · 走高={'✅' if new_rev_trend else '❌'} · 综合≥<b>{new_min_total}</b>分
            </div>
            """, unsafe_allow_html=True)
        with nc2:
            if st.button("✅ 应用参数", width='stretch', type="primary", key=f"{prefix}_apply"):
                st.session_state.lowbuy_params = {
                    '_params_version': DEFAULT_LOWBUY_PARAMS['_params_version'],
                    'max_results': new_max, 'pre_filter_decline': 5,
                    'min_decline_depth': new_min_decline,
                    'min_stabilization': new_min_stabil,
                    'min_volume_recovery': new_min_vol,
                    'min_ma_support': new_min_ma,
                    'min_valuation_attr': new_min_val,
                    'min_chip_settle': new_min_chip,
                    'min_fund_flow': new_min_fund,
                    'min_total_score': new_min_total,
                    'decline_20d_low': new_decline_low, 'decline_20d_high': new_decline_high,
                    'max_vol_ratio': new_max_vr, 'no_new_low_days': new_no_low_days,
                    'reversal_bottom_pct': new_rev_pct,
                    'reversal_require_uptrend': new_rev_trend,
                    'fund_weight': new_fund_wt,
                }
                st.session_state.lowbuy_cache = None
                st.cache_data.clear()
                st.rerun()
        with nc3:
            if st.button("🔄 重置默认", width='stretch', key=f"{prefix}_reset"):
                st.session_state.lowbuy_params = dict(DEFAULT_LOWBUY_PARAMS)
                st.session_state.lowbuy_cache = None
                st.cache_data.clear()
                st.rerun()



# ================================================================
#                      UI: 侧边栏
# ================================================================

def render_sidebar():
    """精简侧边栏 - 只保留导航和基本信息"""
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center;padding:20px 0 10px;border-bottom:1px solid #EDE8E0;margin-bottom:16px;">
        <div style="font-size:36px;">📊</div>
        <div style="font-size:18px;font-weight:700;color:#C4842D;">智能选股系统</div>
        <div style="font-size:11px;color:#BBB;margin-top:4px;">追高 · 低吸 · 双模型策略</div>
        </div>""", unsafe_allow_html=True)
        
        today = datetime.now()
        # 数据源状态指示（同时显示所有可用数据源）
        _ds_parts = []
        if TDX_AVAILABLE:
            _ds_parts.append(f'<span style="color:#2E7D32;">📂 通达信本地</span>')
        if tdx_available():
            _ds_parts.append(f'<span style="color:#1565C0;">📡 pytdx行情</span>')
        if AKSHARE_AVAILABLE:
            _ds_parts.append(f'<span style="color:#F57C00;">🌐 akshare(备用)</span>')
        if not _ds_parts:
            _ds_parts.append(f'<span style="color:#D32F2F;">⚠️ 无数据源</span>')
        _ds_status = ' | '.join(_ds_parts)
        st.markdown(f"""
        <div style="background:#FFF8F0;border-radius:8px;padding:12px;margin-bottom:16px;text-align:center;">
            <div style="font-size:11px;color:#888;">📅 {today.strftime('%Y年%m月%d日 %H:%M')}</div>
            <div style="font-size:13px;color:#C4842D;font-weight:600;">{today.strftime('%A')}</div>
            <div style="font-size:11px;margin-top:6px;">数据源: {_ds_status}</div>
        </div>""", unsafe_allow_html=True)

        page = st.radio("导航", ["🔍 选股筛选", "⭐ 我的自选", "📈 策略回测"],
            label_visibility="collapsed")
        pk = {"🔍 选股筛选": "screener", "⭐ 我的自选": "watchlist", "📈 策略回测": "backtest"}
        if st.session_state.current_page != pk[page]:
            st.session_state.current_page = pk[page]
            st.rerun()

        st.markdown("---")
        
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🔄 刷新", width='stretch', key="sidebar_refresh"):
                st.cache_data.clear(); st.rerun()
        with c2:
            if st.button("🔧 重置权重", width='stretch', key="sidebar_reset_weights"):
                st.session_state.weights = dict(DEFAULT_WEIGHTS)
                st.session_state.top10_cache = None
                st.session_state.top10_cache_key = None
                st.cache_data.clear(); st.rerun()

        wl_count = len(st.session_state.watchlist)
        if wl_count > 0:
            st.markdown(f"""
            <div style="background:#E8F6EF;border-radius:8px;padding:10px;margin-top:12px;text-align:center;">
                <span style="font-size:13px;color:#27AE60;">⭐ 自选股 <b>{wl_count}</b> 只</span>
            </div>""", unsafe_allow_html=True)

        st.markdown("---")
        st.caption("""<div style="text-align:center;color:#BBB;font-size:11px;padding-top:10px;">
        数据来源：pytdx 通达信实时行情<br>
        📊 追高模型：八维强势评分<br>
        📉 低吸模型：底部反转信号<br>
        ⚙️ 参数设置在页面顶部折叠面板</div>""", unsafe_allow_html=True)


# ================================================================
#                    UI: 筛选器 & 统计 & 列表
# ================================================================

def render_filter_bar(df):
    sectors = ["全部"] + sorted(df["板块"].unique().tolist())
    sr_options = [("全部", None), ("强势买入", ("信号", ["强势买入"])),
        ("逢低吸纳", ("信号", ["逢低吸纳"])), ("观望等待", ("信号", ["观望等待"])),
        ("建议回避", ("信号", ["建议回避"]))]
    c1, c2, c3 = st.columns([2, 2, 4])
    with c1: sec = st.selectbox("板块", sectors, label_visibility="collapsed", key="sf_sec")
    with c2:
        slabs = [o[0] for o in sr_options]
        sel_idx = slabs.index(st.selectbox("信号筛选", slabs, label_visibility="collapsed", key="sf_sig"))
        sig_filter = sr_options[sel_idx][1]
    with c3: sw = st.text_input("🔎 搜索代码/名称", "", label_visibility="collapsed", key="sf_search")
    return sec, sig_filter, sw


def render_stats_chase(df):
    t = len(df)
    if t == 0 or '信号' not in df.columns:
        cards = [("📊 0只", "筛选结果", ""), ("0只", "强势买入", ""),
            ("0只", "逢低吸纳", ""), ("0只", "建议回避", ""), ("0分", "平均评分", "")]
        cols = st.columns(5)
        for ic, (vl, lbl, cc) in zip(cols, cards):
            with ic:
                cc_s = f" {cc}" if cc else ""
                st.markdown(f"""<div class="stat-card"><div class="stat-value{cc_s}">{vl}</div><div class="stat-label">{lbl}</div></div>""", unsafe_allow_html=True)
        return
    strong = len(df[df['信号'] == '强势买入'])
    attention = len(df[df['信号'] == '逢低吸纳'])
    watch = len(df[df['信号'] == '建议回避'])
    avg_score = df['综合评分'].mean() if t > 0 else 0
    cards = [(f"📊 {t}只", "筛选结果", ""), (f"🔥 {strong}只", "强势买入", "red" if strong > 3 else ""),
        (f"📈 {attention}只", "逢低吸纳", "orange"), (f"📉 {watch}只", "建议回避", "green"),
        (f"{avg_score:.0f}分", "平均评分", "")]
    cols = st.columns(5)
    for ic, (vl, lbl, cc) in zip(cols, cards):
        with ic:
            cc_s = f" {cc}" if cc else ""
            st.markdown(f"""<div class="stat-card"><div class="stat-value{cc_s}">{vl}</div><div class="stat-label">{lbl}</div></div>""", unsafe_allow_html=True)


def render_stats_lowbuy(df):
    """低吸模型统计数据卡片"""
    t = len(df)
    if t == 0:
        return
    strong = len(df[df['信号'].str.contains('强烈', na=False)])
    cautious = len(df[df['信号'].str.contains('谨慎', na=False)])
    avg_score = df['综合评分'].mean() if '综合评分' in df.columns else 0
    # 平均10日跌幅
    if '20日涨幅' in df.columns:
        try:
            avg_decline = df['20日涨幅'].str.replace('%','').str.replace('+','').astype(float).mean()
            decline_str = f"{avg_decline:.1f}%"
        except:
            decline_str = "N/A"
    else:
        decline_str = "N/A"
    
    cards = [
        (f"📊 {t}只", "筛选结果", ""),
        (f"🔥 {strong}只", "强烈低吸", "red" if strong > 0 else ""),
        (f"📈 {cautious}只", "谨慎低吸", "orange"),
        (f"{avg_score:.0f}分", "平均评分", ""),
        (f"{decline_str}", "平均20日涨幅", "green"),
    ]
    cols = st.columns(5)
    for ic, (vl, lbl, cc) in zip(cols, cards):
        with ic:
            cc_s = f" {cc}" if cc else ""
            st.markdown(f"""<div class="stat-card"><div class="stat-value{cc_s}">{vl}</div><div class="stat-label">{lbl}</div></div>""", unsafe_allow_html=True)


def _diag_stock(diag_code):
    """诊断单只股票在低吸模型中的完整打分流程"""
    code = str(diag_code).strip()
    if not code:
        st.warning("请输入股票代码"); return
    
    params = _get_lowbuy_params()
    
    # 1. 从全市场数据中找到这只股票
    df_market = _app_cache.get('raw_market_data')
    if df_market is None or len(df_market) == 0:
        df_market = fetch_all_a_stocks()
        st.error("无市场数据，请先刷新"); return
    row = df_market[df_market['代码'].astype(str).str.contains(code)]
    if len(row) == 0:
        st.error(f"未找到代码 {code} 的数据"); return
    row = row.iloc[0]
    st.markdown(f"**{row.get('名称', '')} ({code})** · 最新价: {row.get('最新价', 'N/A')} · 涨跌幅: {row.get('涨跌幅', 'N/A')}%")
    
    # 2. 获取K线
    no_new_low_days = params.get('no_new_low_days', 2)
    kline = get_stock_kline(code, days=max(60, no_new_low_days + 15))
    if kline is None or len(kline) < 10:
        st.error("❌ 无法获取K线数据（或K线不足10根）"); return
    closes = kline['收盘'].values.astype(float)
    n = len(closes)
    
    # 3. 逐项检查，显示每步结果
    steps = []
    
    # ---- 预筛检查 ----
    chg = float(row.get('涨跌幅', 99) or 99)
    price = float(row.get('最新价', 0) or 0)
    vr = float(row.get('量比', 99) or 99)
    steps.append(("预筛: 涨幅<9.5%", "✅" if chg < 9.5 else f"❌ 实际{chg}%", chg < 9.5))
    steps.append(("预筛: 价格[3,200]", "✅" if 3 <= price <= 200 else f"❌ 实际{price}", 3 <= price <= 200))
    max_vr = params.get('max_vol_ratio', 2.5)
    steps.append((f"预筛: 量比≤{max_vr}", "✅" if vr <= max_vr else f"❌ 实际{vr:.1f}", vr <= max_vr))
    
    # ---- 10日跌幅（K线真实数据）----
    decline_low = params.get('decline_20d_low', -40)
    decline_high = params.get('decline_20d_high', 0)
    if n >= 12:
        k20d = (closes[-1] / closes[-12] - 1) * 100
        in_range = decline_low <= k20d <= decline_high
        steps.append((f"10日K线跌幅 [{decline_low}%, {decline_high}%]",
                      "✅" if in_range else f"❌ 实际{k20d:.2f}% (不在此范围)", in_range))
        st.metric("10日K线跌幅", f"{k20d:.2f}%", delta=None)
    else:
        steps.append(("10日K线跌幅", f"⚠️ K线仅{n}根(需≥22)", False))
    
    # ---- 不创新低 ----
    if n >= no_new_low_days:
        recent_lows = closes[-no_new_low_days:]
        ok = not (closes[-1] <= min(recent_lows) * 0.998)
        steps.append((f"近{no_new_low_days}日不创新低",
                      "✅" if ok else f"❌ 今日收盘{closes[-1]:.2f} ≤ 近{no_new_low_days}日最低{min(recent_lows):.2f}*0.998", ok))
    else:
        steps.append((f"近{no_new_low_days}日不创新低", f"⚠️ K线不足{n}根", False))
    
    # ---- 反转条件 ----
    if n < 15:
        steps.append(("反转条件", "⚠️ K线<15根", False))
    else:
        low_idx_15 = np.argmin(closes[-15:])
        low_pos = 15 - low_idx_15
        r1_ok = low_pos >= 1
        steps.append((f"反转1: 最低点距今≥1天", "✅" if r1_ok else f"❌ 最低点就在今天(距{low_pos}天)", r1_ok))
        
        avg3d = np.mean(closes[-3:])
        min15d = np.min(closes[-15:])
        rev_pct = params.get('reversal_bottom_pct', 0.2) / 100.0
        detach_pct = (avg3d - min15d) / min15d * 100 if min15d > 0 else 0
        r2_ok = detach_pct >= rev_pct * 100
        steps.append((f"反转2: 脱离底部≥{rev_pct*100:.1f}%",
                      "✅" if r2_ok else f"❌ 实际{detach_pct:.2f}%", r2_ok))
        
        require_ut = params.get('reversal_require_uptrend', False)
        if require_ut:
            r3_ok = np.mean(closes[-3:]) > np.mean(closes[-4:-1])
            ma3_today = np.mean(closes[-3:])
            ma3_prev = np.mean(closes[-4:-1])
            trend_desc = "弱" if ma3_today <= ma3_prev else "强"
            steps.append((f"反转3: MA3今日>MA3前日", "✅" if r3_ok else f"❌ MA3今日{ma3_today:.2f}≤前日{ma3_prev:.2f}", r3_ok))
        else:
            steps.append(("反转3: MA3今日>MA3前日", "⏭️ 已关闭此检查", True))
    
    # ---- 构建评分数据并计算七维分 ----
    try:
        lhb_data = {}
        sector_data = {}
        stock_sector_map = {}
        try:
            lhb_data = _get_cached_dragon_tiger()
            sector_data, stock_sector_map = _get_cached_sector_data()
        except Exception:
            pass
        stock_data = _build_stock_data(row, kline, lhb_data, sector_data, stock_sector_map)
        score_result = calculate_lowbuy_score(stock_data, kline, params)
        
        _raw = score_result.get('_raw', {})
        dims = [
            ("下跌幅度", _raw.get('s1', 0), params.get('min_decline_depth', 5), 30),
            ("企稳信号", _raw.get('s2', 0), params.get('min_stabilization', 10), 25),
            ("量能恢复", _raw.get('s3', 0), params.get('min_volume_recovery', 8), 20),
            ("均线支撑", _raw.get('s4', 0), params.get('ma_ma_support', 8), 15),
            ("估值吸引", _raw.get('s5', 0), params.get('min_valuation_attr', 4), 10),
            ("筹码沉淀", _raw.get('s6', 0), params.get('min_chip_settle', 3), 10),
            ("主力资金", _raw.get('s7', 0), params.get('min_fund_flow', 2), 10),
        ]
        for name, raw_val, threshold, full in dims:
            # fix key name for 均线支撑
            thr = params.get('min_' + ('ma_support' if name == '均线支撑' else
                                        'decline_depth' if name == '下跌幅度' else
                                        'stabilization' if name == '企稳信号' else
                                        'volume_recovery' if name == '量能恢复' else
                                        'valuation_attr' if name == '估值吸引' else
                                        'chip_settle' if name == '筹码沉淀' else
                                        'fund_flow'), threshold)
            ok = raw_val >= thr
            steps.append((f"七维·{name} ≥{thr}(满分{full})",
                          "✅" if ok else f"❌ 实际{raw_val}/{full}", ok))
        
        s_total = score_result.get('综合评分', 0)
        min_total = params.get('min_total_score', 20)
        steps.append((f"综合总分 ≥{min_total}",
                      f"📊 **{s_total}**分" if s_total >= min_total else f"❌ **{s_total}**分 (需≥{min_total})",
                      s_total >= min_total))
        
    except Exception as e:
        steps.append(("七维评分计算", f"❌ 异常: {e}", False))
    
    # 显示结果表格
    st.markdown("---")
    c_status, c_detail = st.columns([1, 3])
    with c_status:
        all_pass = all(s[2] for s in steps)
        if all_pass:
            st.success("🎉 全部通过！应出现在结果列表中")
        else:
            fail_count = sum(1 for s in steps if not s[2])
            st.error(f"❌ 未通过 ({fail_count}/{len(steps)} 项)")
    
    with c_detail:
        for i, (desc, status, ok) in enumerate(steps):
            color = "#C8E6C9" if ok else "#FFCDD2"
            icon = "✅" if ok else ("⚠️" if "⏭" in status else "❌")
            st.markdown(f"<div style='padding:4px 10px;margin:2px 0;border-radius:4px;background:{color};font-size:13px;'>"
                       f"<b>{icon} [{i+1}]</b> {desc}: <b>{status}</b></div>", unsafe_allow_html=True)


def _export_df_to_xlsx(dframe, code_col="代码"):
    """将 DataFrame 导出为 xlsx 字节，代码列设为文本格式防止前导零丢失"""
    _illegal_xml_re = re.compile(
        '[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufffe\uffff]'
    )
    def _clean_cell(val):
        if isinstance(val, str):
            return _illegal_xml_re.sub('', val)
        return val

    df = dframe.copy()
    if code_col in df.columns:
        df[code_col] = df[code_col].astype(str).str.zfill(6)
    # 清理所有单元格中的非法XML字符（不限dtype，防止控制字符藏在混合列中）
    for col in df.columns:
        df[col] = df[col].apply(lambda x: _illegal_xml_re.sub('', x) if isinstance(x, str) else x)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        ws = writer.sheets['Sheet1']
        if code_col in df.columns:
            code_col_idx = list(df.columns).index(code_col) + 1  # openpyxl 1-based
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=code_col_idx).number_format = '@'
    return buf.getvalue()


def render_stock_buttons(stocks, prefix="bt"):
    """统一的股票操作按钮区：5列×N行，每格 序号+名称+[评分详情]+[加自选]"""
    if not stocks:
        return
    for i in range(0, len(stocks), 5):
        cols = st.columns(5)
        for j in range(5):
            if i + j >= len(stocks):
                break
            s = stocks[i + j]
            code = str(s.get('代码', ''))
            name = str(s.get('名称', ''))
            seq = i + j + 1
            wl = code in st.session_state.watchlist
            with cols[j]:
                st.caption(f"{seq}. {name}  `{code}`")
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("评分详情", key=f"{prefix}_d_{code}", width='stretch'):
                        st.session_state.selected_stock = code
                        st.rerun()
                with bc2:
                    lb = "★ 已自选" if wl else "+ 加自选"
                    if st.button(lb, key=f"{prefix}_w_{code}", width='stretch',
                                 type="secondary" if not wl else "primary"):
                        if wl:
                            st.session_state.watchlist.remove(code)
                        else:
                            st.session_state.watchlist.append(code)
                        save_watchlist(st.session_state.watchlist)
                        st.rerun()


def render_table_chase(fdf, dde_scores=None, resonance_scores=None, resonance_styles=None):
    """追高模型Top30 — st.dataframe 表格 + 5列按钮网格"""
    if fdf is None or len(fdf) == 0:
        st.info("暂无符合追高条件的标的。")
        return
    st.markdown(
        f"<div style='font-size:13px;color:#888;margin-bottom:8px;'>共 <b>{len(fdf)}</b> 只标的</div>",
        unsafe_allow_html=True
    )
    fdf = fdf.copy()
    fdf['代码'] = fdf['代码'].astype(str).str.zfill(6)
    fdf = fdf.rename(columns={'综合评分': '追高评分', '板块': '概念板块'})

    # 共振交叉评分
    if resonance_scores:
        fdf['共振评分'] = fdf['代码'].apply(lambda x: resonance_scores.get(x, None))
    if resonance_styles:
        fdf['共振评价'] = fdf['代码'].apply(lambda x: resonance_styles.get(x, '-'))

    # 确定可用列
    base_cols = ['名称', '概念板块', '趋势结构', '动量强度', '板块共振',
                 '北向资金', '机构净买', '板块资金热度', '量价配合', '估值安全', '筹码稳定', '情绪热度',
                 '5日涨幅', '换手率', 'PE', '信号', '建议', '追高评分']
    if dde_scores:
        fdf['DDE资金确认(50)'] = fdf['代码'].apply(lambda x: dde_scores.get(x, None))
        if 'DDE资金确认(50)' not in base_cols:
            base_cols.insert(-2, 'DDE资金确认(50)')
    if resonance_scores or resonance_styles:
        if resonance_scores:
            if '共振评分' not in base_cols:
                base_cols.insert(-2, '共振评分')
        if resonance_styles:
            if '共振评价' not in base_cols:
                base_cols.insert(-2, '共振评价')
    display_cols = [c for c in base_cols if c in fdf.columns]
    df_display = fdf[display_cols]

    # Arrow 序列化兼容：含 "N/A" 字符串的列是 object 类型，PyArrow 无法序列化 → 统一转 str
    for c in df_display.columns:
        if df_display[c].dtype == 'object':
            df_display[c] = df_display[c].astype(str).replace('nan', 'N/A')

    def color_score(val):
        if isinstance(val, (int, float)):
            if val >= 80: return 'background-color:#C8E6C9;font-weight:bold'
            if val >= 70: return 'background-color:#E8F5E9'
            if val >= 60: return 'background-color:#FFF9C4'
            return ''
        return ''

    styled = df_display.style.map(color_score, subset=['追高评分'])
    st.dataframe(styled, width='stretch', hide_index=True,
                 column_config={'追高评分': st.column_config.NumberColumn(format='%.0f')})

    st.markdown("---")
    st.caption("操作区 — 评分详情 | 加/取消自选")
    render_stock_buttons(fdf.to_dict('records'), prefix="ch")


def render_table_lowbuy(fdf, dde_scores=None, resonance_scores=None, resonance_styles=None):
    """低吸模型Top30 — st.dataframe 表格 + 5列按钮网格"""
    if fdf is None or len(fdf) == 0:
        st.info("暂无符合低吸条件的标的。")
        return
    st.markdown(
        f"<div style='font-size:13px;color:#888;margin-bottom:8px;'>共 <b>{len(fdf)}</b> 只标的</div>",
        unsafe_allow_html=True
    )
    fdf = fdf.copy()
    fdf['代码'] = fdf['代码'].astype(str).str.zfill(6)
    fdf = fdf.rename(columns={'综合评分': '低吸评分', '板块': '概念板块'})

    # 共振交叉评分
    if resonance_scores:
        fdf['共振评分'] = fdf['代码'].apply(lambda x: resonance_scores.get(x, None))
    if resonance_styles:
        fdf['共振评价'] = fdf['代码'].apply(lambda x: resonance_styles.get(x, '-'))

    base_cols = ['名称', '概念板块', '下跌幅度', '企稳信号', '量能恢复',
                 '均线支撑', '估值吸引', '筹码沉淀', '20日涨幅', '信号', '低吸评分']
    if dde_scores:
        fdf['DDE资金确认(50)'] = fdf['代码'].apply(lambda x: dde_scores.get(x, None))
        if 'DDE资金确认(50)' not in base_cols:
            base_cols.insert(-1, 'DDE资金确认(50)')
    if resonance_scores or resonance_styles:
        if resonance_scores:
            if '共振评分' not in base_cols:
                base_cols.insert(-1, '共振评分')
        if resonance_styles:
            if '共振评价' not in base_cols:
                base_cols.insert(-1, '共振评价')
    display_cols = [c for c in base_cols if c in fdf.columns]
    df_display = fdf[display_cols]

    def color_score(val):
        if isinstance(val, (int, float)):
            if val >= 80: return 'background-color:#C8E6C9;font-weight:bold'
            if val >= 70: return 'background-color:#E8F5E9'
            if val >= 60: return 'background-color:#FFF9C4'
            return ''
        return ''

    styled = df_display.style.map(color_score, subset=['低吸评分'])
    st.dataframe(styled, width='stretch', hide_index=True,
                 column_config={'低吸评分': st.column_config.NumberColumn(format='%.0f')})

    st.markdown("---")
    st.caption("操作区 — 评分详情 | 加/取消自选")
    render_stock_buttons(fdf.to_dict('records'), prefix="lb")


def render_top_chase_high():
    """追高模型综合动态推荐"""
    top_n = get_top10_stocks()
    if not top_n: return
    dyn_n = st.session_state.get('_chase_dynamic_n', len(top_n))
    
    st.markdown(f"""
    <div class="top10-container">
        <div class="top10-header">
            <div class="top10-title">🌟 追高综合 Top {dyn_n} <span class="top10-badge">每日更新</span></div>
        </div>
    </div>""", unsafe_allow_html=True)
    
    btn1, btn2, _ = st.columns([1, 1, 4])
    with btn1:
        if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="ch_add_all"):
            for stock in top_n:
                if stock["代码"] not in st.session_state.watchlist:
                    st.session_state.watchlist.append(stock["代码"])
            save_watchlist(st.session_state.watchlist)
            st.success(f"已将Top {dyn_n}全部加入自选！"); st.rerun()
    with btn2:
        top_n_df = pd.DataFrame(top_n)
        xlsx_data = _export_df_to_xlsx(top_n_df)
        st.download_button(f"📥 导出Top{dyn_n}", xlsx_data, f"top{dyn_n}_chase_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width='stretch', key="dl_t10")
    
    st.markdown("")
    for i in range(0, len(top_n), 5):
        row_stocks = top_n[i:i+5]
        cols = st.columns(5)
        for j, stock in enumerate(row_stocks):
            with cols[j]:
                rank = i + j + 1
                rank_class = "gold" if rank == 1 else "silver" if rank == 2 else "bronze" if rank == 3 else ""
                rank_color = "#FFD700" if rank == 1 else "#C0C0C0" if rank == 2 else "#CD7F32" if rank == 3 else "#CCC"
                chg = stock.get('5日涨幅', '0%')
                chg_color = "#E74C3C" if "+" in str(chg) else "#27AE60"
                st.markdown(f"""
                <div class="top10-card">
                    <div class="top10-rank {rank_class}" style="color:{rank_color};">{rank}</div>
                    <div style="padding-right:30px;">
                        <div style="font-weight:700;color:#333;font-size:15px;margin-bottom:2px;">{stock['名称']}</div>
                        <div style="font-size:11px;color:#888;margin-bottom:8px;">{stock['代码']} · {stock.get('概念板块', '其他')[:8]}</div>
                        <div style="display:flex;justify-content:space-between;align-items:center;">
                            <span style="font-size:20px;font-weight:800;color:#C4842D;">{stock.get('追高评分', stock.get('综合评分', ''))}</span>
                            <span style="font-size:13px;font-weight:600;color:{chg_color};">{chg}</span>
                        </div>
                        <div style="margin-top:6px;"><span class="metric-badge badge-strong">{stock['信号']}</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)
                code = stock["代码"]; in_wl = code in st.session_state.watchlist
                if st.button("⭐" if in_wl else "+自选", key=f"t10_{code}",
                    width='stretch', type="primary" if in_wl else "secondary"):
                    if in_wl: st.session_state.watchlist.remove(code)
                    else: st.session_state.watchlist.append(code)
                    save_watchlist(st.session_state.watchlist); st.rerun()


def _render_filter_dashboard(_dbg, params):
    """可视化展示各过滤阶段的拦截数量条带"""
    total = _dbg.get('total', 0)
    if total == 0:
        with st.expander("📊 过滤漏斗图 (点击展开)", expanded=True):
            st.info("尚未执行全市场扫描，请点击「刷新数据」获取过滤漏斗分析")
        return

    # 前置过滤阶段
    front_stages = [
        ("无K线",       _dbg.get('no_kline', 0),          "#999"),
        ("10d跌幅不符", _dbg.get('decline_20d_fail', 0),  "#E67E22"),
        ("创新低",      _dbg.get('no_new_low', 0),        "#E74C3C"),
        ("反转1(距今<1d)", _dbg.get('reversal1', 0),       "#E91E63"),
        ("反转2(脱离不足)", _dbg.get('reversal2', 0),      "#9C27B0"),
        ("反转3(MA3弱)",  _dbg.get('reversal3', 0),        "#673AB7"),
    ]
    # 七维门槛阶段
    seven_dim_stages = [
        ("跌幅不足",  _dbg.get('decline', 0),     "#F39C12"),
        ("企稳不足",  _dbg.get('stabil', 0),      "#E74C3C"),
        ("量能不足",  _dbg.get('vol_rec', 0),     "#3498DB"),
        ("均线不足",  _dbg.get('ma_sup', 0),      "#2ECC71"),
        ("估值不足",  _dbg.get('val', 0),         "#9B59B6"),
        ("筹码不足",  _dbg.get('chip', 0),        "#1ABC9C"),
        ("资金不足",  _dbg.get('fund', 0),        "#E67E22"),
        ("总分不足",  _dbg.get('total_score', 0), "#34495E"),
    ]
    
    passed = _dbg.get('passed', 0)
    
    max_cnt = max((c for _, c, _ in front_stages + seven_dim_stages), default=1)
    max_cnt = max(max_cnt, 1)
    
    def _bar_html(label, cnt, color, max_val):
        pct = cnt / total * 100 if total > 0 else 0
        bar_pct = cnt / max_val * 100 if max_val > 0 else 0
        bar_style = f"width:{bar_pct}%;background:{color};" if cnt > 0 else "width:0;"
        return f"""
        <div style="display:flex;align-items:center;margin:2px 0;font-size:12px;">
            <span style="width:100px;color:#666;">{label}</span>
            <span style="width:40px;text-align:right;font-weight:600;margin-right:8px;
                color:{'#E74C3C' if cnt > 0 else '#999'};">{cnt}</span>
            <div style="flex:1;background:#F0EBE4;border-radius:4px;height:14px;">
                <div style="{bar_style}height:14px;border-radius:4px;transition:width 0.3s;"></div>
            </div>
            <span style="width:45px;text-align:right;font-size:11px;color:#999;margin-left:6px;">{pct:.1f}%</span>
        </div>"""
    
    bars_html = '<div style="margin:8px 0;">'
    bars_html += '<div style="font-size:11px;color:#888;margin-bottom:4px;">▸ 前置过滤</div>'
    for label, cnt, color in front_stages:
        bars_html += _bar_html(label, cnt, color, max_cnt)
    bars_html += '<div style="font-size:11px;color:#888;margin:8px 0 4px;">▸ 七维门槛</div>'
    for label, cnt, color in seven_dim_stages:
        bars_html += _bar_html(label, cnt, color, max_cnt)
    bars_html += '<div style="font-size:11px;color:#888;margin:8px 0 4px;">▸ 结果</div>'
    bars_html += _bar_html("✅ 通过", passed, "#27AE60", max_cnt)
    bars_html += '</div>'
    
    with st.expander("📊 过滤漏斗图 (点击展开)", expanded=True):
        st.markdown(bars_html, unsafe_allow_html=True)
        st.caption(f"全市场扫描 {total} 只 → 最终通过 {passed} 只")
        # [DEBUG] 显示运行时实际参数
        pdbg = _dbg.get('_params_debug', {})
        if pdbg:
            st.caption(f"🔧 运行时参数: stabil≥{pdbg.get('min_stabilization','?')} vol≥{pdbg.get('min_volume_recovery','?')} rev_pct={pdbg.get('reversal_bottom_pct','?')}% uptrend={pdbg.get('reversal_require_uptrend','?')} total≥{pdbg.get('min_total_score','?')} ver={pdbg.get('_params_version','?')}")


def render_top_low_buy(sample_source="全市场A股"):
    """低吸模型精选Top — 动态推荐数量"""
    _lb_dyn_n = calculate_dynamic_recommend_count()
    st.markdown(f"""
    <div class="lowbuy-container">
        <div class="top10-header">
            <div class="lowbuy-title">📉 低吸精选 Top {_lb_dyn_n} <span class="lowbuy-badge">底部反转信号</span></div>
        </div>
    </div>""", unsafe_allow_html=True)
    
    # ===== 单只股票诊断（始终显示，不受缓存影响）=====
    with st.expander("🔍 查询单只股票打分", expanded=False):
        diag_code = st.text_input("输入股票代码（如 601318）", value="601318", key="lb_diag_code", max_chars=6)
        if st.button("🔬 诊断", key="lb_diag_btn", width='stretch'):
            _diag_stock(diag_code)
    
    # 检查是否有今天的缓存
    if st.session_state.get('lowbuy_cache') and st.session_state.get('last_update_time'):
        cache_time = st.session_state.last_update_time
        if cache_time.date() == datetime.now().date():
            st.info(f"📌 显示今日 {cache_time.strftime('%H:%M')} 的缓存结果，点击「刷新数据」获取最新")
            result_df = pd.DataFrame(st.session_state.lowbuy_cache)
            if len(result_df) > 0:
                # 过滤漏斗图（卡片之前，与新扫描路径一致）
                _dbg = st.session_state.get('_lb_dbg', {})
                _lp = _get_lowbuy_params()
                try:
                    _render_filter_dashboard(_dbg, _lp)
                except Exception as _e:
                    st.warning(f"过滤漏斗渲染异常: {_e}")
                btn1, btn2, _ = st.columns([1, 1, 4])
                with btn1:
                    if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="lb_add_all_cached"):
                        for _, row in result_df.iterrows():
                            if row["代码"] not in st.session_state.watchlist:
                                st.session_state.watchlist.append(row["代码"])
                        save_watchlist(st.session_state.watchlist)
                        st.success(f"已将{len(result_df)}只低吸标的加入自选！"); st.rerun()
                with btn2:
                    xlsx_data = _export_df_to_xlsx(result_df.head(_lb_dyn_n))
                    st.download_button(f"📥 导出Top{_lb_dyn_n}", xlsx_data, f"top{_lb_dyn_n}_lowbuy_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width='stretch', key="dl_lb")
                st.markdown("")
                _render_lowbuy_cards(result_df.head(_lb_dyn_n))
                return
    
    # 无缓存，运行新的筛选
    df_raw = fetch_all_a_stocks()
    if df_raw is None:
        st.warning("无法获取行情数据"); return
    
    # 🌟 样本来源过滤
    if sample_source == "热门板块":
        hot_codes = get_hot_concept_stocks(6)
        if hot_codes:
            df_raw['代码'] = df_raw['代码'].astype(str).str.zfill(6)
            df_raw = df_raw[df_raw['代码'].isin(hot_codes)]
            st.info(f"已锁定 {len(df_raw)} 只热门概念板块成分股")
        else:
            st.warning("未能获取热门板块数据，回退为全市场扫描")
    elif sample_source == "量价":
        vp_codes = get_volprice_sectors(6)
        if vp_codes:
            df_raw['代码'] = df_raw['代码'].astype(str).str.zfill(6)
            df_raw = df_raw[df_raw['代码'].isin(vp_codes)]
            st.info(f"已锁定 {len(df_raw)} 只量价反转板块成分股")
        else:
            st.warning("未能获取量价反转板块数据，回退为全市场扫描")
    
    df_market = preprocess_stock_data(df_raw)
    if df_market is None:
        st.warning("数据预处理失败"); return
    
    if sample_source == "热门板块":
        spinner_text = "正在扫描资金加速板块，筛选低吸标的…"
    elif sample_source == "量价":
        spinner_text = "正在扫描量价反转板块，筛选低吸标的…"
    else:
        spinner_text = "正在全市场扫描底部反转信号，请稍候..."
    with st.spinner(spinner_text):
        result_df = screen_low_buy_stocks(df_market)
    
    # 显示过滤漏斗图
    _dbg = st.session_state.get('_lb_dbg', {})
    _lp = _get_lowbuy_params()
    try:
        _render_filter_dashboard(_dbg, _lp)
    except Exception as _e:
        st.warning(f"过滤漏斗渲染异常: {_e}")
    
    if result_df is None or len(result_df) == 0:
        st.info("📭 今日未找到符合低吸条件的股票。可能当前市场整体处于上升趋势，或下跌个股尚未出现明确的反转信号。")
        st.markdown("""
        <div style="background:#FFF8E1;border-radius:8px;padding:12px 16px;font-size:13px;color:#888;">
        💡 <b>提示：</b>可尝试在顶部参数面板中放宽条件（如减少下跌观察天数、扩大跌幅区间）
        </div>""", unsafe_allow_html=True)
        return
    
    # 保存低吸结果到缓存
    if result_df is not None and len(result_df) > 0:
        st.session_state.lowbuy_cache = result_df.to_dict('records')
        save_cache_data({
            'chase_high_top10': st.session_state.get('top10_cache'),
            'lowbuy_top5': st.session_state.lowbuy_cache,
            'lowbuy_dbg': st.session_state.get('_lb_dbg', {}),
        })
    
    # 一键操作
    btn1, btn2, _ = st.columns([1, 1, 4])
    with btn1:
        if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="lb_add_all"):
            for _, row in result_df.iterrows():
                if row["代码"] not in st.session_state.watchlist:
                    st.session_state.watchlist.append(row["代码"])
            save_watchlist(st.session_state.watchlist)
            st.success(f"已将{len(result_df)}只低吸标的加入自选！"); st.rerun()
    with btn2:
        xlsx_data = _export_df_to_xlsx(result_df.head(_lb_dyn_n))
        st.download_button(f"📥 导出Top{_lb_dyn_n}", xlsx_data, f"top{_lb_dyn_n}_lowbuy_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width='stretch', key="dl_lb")
    
    st.markdown("")
    _render_lowbuy_cards(result_df.head(_lb_dyn_n))


def _render_lowbuy_score_table(result_df):
    """低吸模型综合评分列表（表格样式）"""
    df = result_df.head(20) if isinstance(result_df, pd.DataFrame) else pd.DataFrame(result_df).head(20)
    if len(df) == 0:
        return
    
    for idx, row in df.iterrows():
        rank = idx + 1
        score = row.get('综合评分', row.get('信号强度', 0))
        signal = row.get('信号', '标准低吸')
        sig_bg = "#E8F5E9" if "强烈" in str(signal) else "#F1F8E9"
        sig_color = "#2E7D32" if "强烈" in str(signal) else "#43A047"
        
        # 八维评分
        dims = ['下跌', '企稳', '量能', '均线', '估值', '筹码']
        dim_vals = [
            row.get('下跌幅度', 0), row.get('企稳信号', 0),
            row.get('量能恢复', 0), row.get('均线支撑', 0),
            row.get('估值吸引', 0), row.get('筹码沉淀', 0)
        ]
        
        dim_html = ""
        for d_name, d_val in zip(dims, dim_vals):
            d_color = "#E74C3C" if d_val >= 80 else "#E67E22" if d_val >= 65 else "#27AE60"
            dim_html += f'<span style="font-size:11px;color:#888;">{d_name}</span> <span style="font-size:12px;font-weight:700;color:{d_color};">{d_val}</span>  '
        
        chg_20d = row.get('20日涨幅', 'N/A')
        chg_color = "#E74C3C" if "+" in str(chg_20d) else "#27AE60"
        sector = str(row.get('概念板块', ''))[:6]
        
        bg = "#F9FFF9" if rank <= 3 else "#FFFFFF"
        rank_badge = f'<span style="display:inline-block;width:22px;height:22px;border-radius:50%;background:{"#FFD700" if rank==1 else "#C0C0C0" if rank==2 else "#CD7F32" if rank==3 else "#E0E0E0"};color:white;text-align:center;line-height:22px;font-size:11px;font-weight:700;">{rank}</span>'
        
        st.markdown(f"""<div style="display:flex;align-items:center;padding:8px 14px;border-bottom:1px solid #F0EDE8;background:{bg};border-radius:4px;margin-bottom:2px;">
<div style="width:30px;flex-shrink:0;text-align:center;">{rank_badge}</div>
<div style="width:120px;flex-shrink:0;padding:0 8px;">
<div style="font-weight:700;color:#333;font-size:13px;">{row.get('名称', '')}</div>
<div style="font-size:10px;color:#999;">{row.get('代码', '')} · {sector}</div>
</div>
<div style="width:60px;text-align:center;">
<span style="font-size:18px;font-weight:800;color:#2E7D32;">{score}</span>
</div>
<div style="width:70px;text-align:center;">
<span style="font-size:12px;font-weight:600;color:{chg_color};">{chg_20d}</span>
</div>
<div style="flex:1;padding:0 8px;">{dim_html}</div>
<div style="width:80px;text-align:center;">
<span style="display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;background:{sig_bg};color:{sig_color};">{signal}</span>
</div>
<div style="width:50px;text-align:center;">
<button style="background:none;border:1px solid #ddd;border-radius:4px;padding:2px 8px;font-size:11px;color:#666;cursor:pointer;" title="查看详情">详情</button>
</div>
</div>""", unsafe_allow_html=True)


def _render_lowbuy_cards(result_df):
    """低吸模型卡片式展示（与追高模型Top10卡片风格一致）"""
    stocks = result_df.to_dict('records') if isinstance(result_df, pd.DataFrame) else result_df
    if not stocks:
        return
    
    # 卡片式布局 - 5只/行
    for i in range(0, len(stocks), 5):
        row_stocks = stocks[i:i+5]
        cols = st.columns(5)
        for j, stock in enumerate(row_stocks):
            with cols[j]:
                rank = i + j + 1
                rank_class = "gold" if rank == 1 else "silver" if rank == 2 else "bronze" if rank == 3 else ""
                rank_color = "#FFD700" if rank == 1 else "#C0C0C0" if rank == 2 else "#CD7F32" if rank == 3 else "#CCC"
                
                chg_20d = stock.get('20日涨幅', 'N/A')
                chg_color = "#E74C3C" if "+" in str(chg_20d) else "#27AE60"
                
                score = stock.get('低吸评分', stock.get('综合评分', stock.get('信号强度', 0)))
                
                signal = stock.get('信号', '标准低吸')
                sig_class = "badge-strong" if "强烈" in signal else "badge-hold"
                
                sector = stock.get('概念板块', '其他')
                if len(sector) > 8:
                    sector = sector[:8]
                
                st.markdown(f"""
                <div class="top10-card">
                    <div class="top10-rank {rank_class}" style="color:{rank_color};">{rank}</div>
                    <div style="padding-right:30px;">
                        <div style="font-weight:700;color:#333;font-size:15px;margin-bottom:2px;">{stock.get('名称', '')}</div>
                        <div style="font-size:11px;color:#888;margin-bottom:8px;">{stock.get('代码', '')} · {sector}</div>
                        <div style="display:flex;justify-content:space-between;align-items:center;">
                            <span style="font-size:20px;font-weight:800;color:#C4842D;">{score}</span>
                            <span style="font-size:13px;font-weight:600;color:{chg_color};">{chg_20d}</span>
                        </div>
                        <div style="margin-top:6px;"><span class="metric-badge {sig_class}">{signal}</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)
                
                code = stock.get("代码", "")
                in_wl = code in st.session_state.watchlist
                if st.button("⭐" if in_wl else "+自选", key=f"lbt5_{code}",
                    width='stretch', type="primary" if in_wl else "secondary"):
                    if in_wl:
                        st.session_state.watchlist.remove(code)
                    else:
                        st.session_state.watchlist.append(code)
                    save_watchlist(st.session_state.watchlist)
                    st.rerun()


# ================================================================
#                   UI: 个股评分面板
# ================================================================

MODEL_MAX_SCORES = {
    "chase_high": 100, "buy_low": 100, "golden_cross": 100,
    "resonance": 100, "canslim": 100, "dilemma_reversal": 100,
    "oversold_rebound": 110,
}
MODEL_LABELS = {
    "chase_high": "追高模型", "buy_low": "低吸模型", "golden_cross": "金叉模型",
    "resonance": "共振模型", "canslim": "CAN SLIM", "dilemma_reversal": "困境反转",
    "oversold_rebound": "超跌反弹",
}
COLOR_SCHEME = ["#E74C3C", "#3498DB", "#2ECC71", "#F39C12", "#9B59B6", "#1ABC9C", "#E67E22"]


def render_stock_scorer():
    """个股评分面板：输入代码或名称，展示六大模型评分。"""
    st.markdown("### 个股评分")
    
    # 输入方式：代码或名称
    c1, c2 = st.columns([2, 3])
    with c1:
        code_input = st.text_input("股票代码", value="", placeholder="如 000001", key="ss_code")
    with c2:
        name_input = st.text_input("或输入名称关键词", value="", placeholder="如 平安银行", key="ss_name")
    
    if not code_input and not name_input:
        return
    
    # 名称→代码 模糊搜索
    target_code = None
    if code_input:
        target_code = code_input.strip().zfill(6)
    elif name_input and AKSHARE_AVAILABLE:
        try:
            name_df = ak.stock_info_a_code_name()
            matches = name_df[name_df['名称'].str.contains(name_input, na=False)]
            if len(matches) == 0:
                st.warning(f"未找到名称包含「{name_input}」的股票")
                return
            elif len(matches) == 1:
                target_code = str(matches.iloc[0]['代码']).zfill(6)
            else:
                # 多候选，让用户选
                candidates = []
                for _, r in matches.head(10).iterrows():
                    c = str(r['代码']).zfill(6)
                    n = str(r['名称'])
                    candidates.append(f"{c} - {n}")
                choice = st.selectbox("请选择股票", candidates, key="ss_choice")
                target_code = choice.split(" - ")[0]
        except Exception as e:
            st.warning(f"名称查询失败: {e}")
            return
    
    if not target_code:
        return
    
    # 开始评分
    model_order = ["chase_high", "buy_low", "golden_cross", "resonance", "canslim", "dilemma_reversal", "oversold_rebound"]
    
    # 先获取一次基本行情确认股票存在
    with st.spinner("正在获取行情数据..."):
        row_dict, _ = _find_stock_row(target_code)
        if row_dict is None:
            st.error(f"未找到股票 {target_code}，请检查代码是否正确或稍后重试")
            return
        stock_name = str(row_dict.get('名称', ''))
        stock_close = float(row_dict.get('最新价', 0) or 0)
        stock_chg = float(row_dict.get('涨跌幅', 0) or 0)
    
    # 股票基本信息
    chg_sign = "+" if stock_chg >= 0 else ""
    chg_color = "#E53935" if stock_chg > 0 else ("#2E7D32" if stock_chg < 0 else "#666")
    st.markdown(f"""
    <div style="background:white;border-radius:12px;padding:14px 20px;margin:8px 0;
    border:1px solid #E0E0E0;display:flex;align-items:center;gap:16px;">
    <div><span style="font-size:20px;font-weight:700;color:#333;">{stock_name}</span>
    <span style="font-size:13px;color:#999;margin-left:8px;">{target_code}</span></div>
    <div><span style="font-size:22px;font-weight:700;color:#333;">{stock_close:.2f}</span>
    <span style="font-size:15px;color:{chg_color};margin-left:6px;">{chg_sign}{stock_chg:.2f}%</span></div>
    </div>
    """, unsafe_allow_html=True)
    
    # 逐个模型评分
    with st.spinner("正在计算各模型评分..."):
        all_results = {}
        for model in model_order:
            try:
                result = score_single_stock(target_code, model)
                all_results[model] = result
            except Exception as e:
                all_results[model] = {"success": False, "error": str(e)}
    
    # 4行网格展示（7个模型：3行×2 + 1行×1）
    total_models = len(model_order)
    grid_rows = (total_models + 1) // 2
    for row_idx in range(grid_rows):
        cols = st.columns(2)
        for col_idx in range(2):
            model_idx = row_idx * 2 + col_idx
            if model_idx >= len(model_order):
                break
            model = model_order[model_idx]
            result = all_results.get(model, {"success": False, "error": "未计算"})
            label = MODEL_LABELS.get(model, model)
            color = COLOR_SCHEME[model_idx]
            
            with cols[col_idx]:
                if not result.get("success"):
                    st.metric(label=label, value="N/A", delta=result.get("error", "获取失败"))
                    continue
                
                total = result.get("total", 0)
                max_s = result.get("max_score", 100)
                pct = min(total / max_s, 1.0) if max_s > 0 else 0
                
                # 进度条
                bar_color = "#E74C3C" if pct < 0.3 else ("#F39C12" if pct < 0.6 else "#27AE60")
                st.markdown(f"""
                <div style="margin-bottom:4px;">
                <span style="font-weight:600;color:{color};font-size:15px;">{label}</span>
                <span style="float:right;font-size:24px;font-weight:700;color:#333;">{total}</span>
                <span style="float:right;font-size:12px;color:#999;margin-right:4px;margin-top:10px;">/ {max_s}</span>
                </div>
                <div style="background:#EEE;border-radius:6px;height:8px;margin-bottom:8px;">
                <div style="width:{pct*100:.0f}%;background:{bar_color};height:100%;border-radius:6px;"></div>
                </div>
                """, unsafe_allow_html=True)
                
                # 各维度明细
                dims = result.get("dims", {})
                if dims:
                    dim_lines = []
                    for dk, dv in dims.items():
                        dim_label = dk.replace("C_","").replace("A_","").replace("N_","").replace("S_","").replace("L_","").replace("I_","").replace("M_","").replace("L1_","L1 ").replace("L2_","L2 ").replace("L3_","L3 ").replace("L4_","L4 ")
                        dim_lines.append(f"{dim_label}: **{dv}**")
                    st.markdown("<span style='font-size:11px;color:#666;'>" + " &nbsp;|&nbsp; ".join(dim_lines) + "</span>", unsafe_allow_html=True)
                
                # 额外信息
                extra_parts = []
                if result.get("signal"):
                    extra_parts.append(f"信号: {result['signal']}")
                if result.get("position_msg"):
                    extra_parts.append(result["position_msg"])
                if result.get("filter_msg") and not result.get("pass", True):
                    extra_parts.append(f"过滤: {result['filter_msg']}")
                if extra_parts:
                    st.markdown("<span style='font-size:11px;color:#888;'>" + " &nbsp;|&nbsp; ".join(extra_parts) + "</span>", unsafe_allow_html=True)


# ================================================================
#                   UI: 选股筛选主页
# ================================================================

def render_screener():
    """选股筛选页面 - 含模型切换"""
    if not tdx_available() and not AKSHARE_AVAILABLE:
        st.markdown("""<div class="data-status warning">
        <span>⚠️ pytdx 和 akshare 均不可用，请检查通达信连接</span>
        <span>安装 pytdx: pip install pytdx</span>
        </div>""", unsafe_allow_html=True)
    
    # ===== 模型切换 =====
    st.markdown("""
    <style>
    button[data-baseweb="tab"] {
        font-size: 26px !important;
        padding: 12px 32px !important;
    }
    button[data-baseweb="tab"] span {
        font-size: 26px !important;
    }
    /* 共振模型（参考用途）- 蓝紫色区分 */
    div[data-baseweb="tab-list"] button:first-child {
        color: #6C5CE7 !important;
        border-bottom: 3px solid #6C5CE7 !important;
    }
    </style>
    """, unsafe_allow_html=True)
    model_names = {
    "resonance": "🎯 共振模型（多维共振参考）",
    "chase_high": "🚀 追高模型（抓强势龙头股）",
    "buy_low": "📉 低吸模型（抓超跌价值股）",
    "rebound_model": "📌 金叉模型（抓强势回调股）",
    "canslim": "📊 CAN SLIM模型（抓主升浪股）",
    "dilemma_reversal": "🔄 困境反转模型（抓周期股）",
    "oversold_rebound": "💎 超跌反弹模型（抓超跌股）"
}
    current = st.session_state.current_model
    tab_labels = list(model_names.values())
    key_order = list(model_names.keys())
    current_idx = key_order.index(current) if current in key_order else 0
    
    # 初始化新模型参数
    if "gc_params" not in st.session_state:
        st.session_state.gc_params = dict(DEFAULT_GC_PARAMS)
    if "gc_weights" not in st.session_state:
        st.session_state.gc_weights = dict(DEFAULT_GC_WEIGHTS)
    
    # ===== 个股评分入口 =====
    with st.expander("📊 个股评分（输入代码查六大模型得分）", expanded=False):
        render_stock_scorer()
    
    model_tabs = st.tabs(tab_labels)
    
    with model_tabs[0]:  # 共振模型
        st.session_state.current_model = 'resonance'
        
        # 顶部参数面板
        render_top_params_panel()
        
        st.markdown(f"""<div class="header-container"><div class="main-title">🎯 共振模型 · 多维共振参考</div>
        <div class="sub-title">资金流向 + DDE决策 + K线结构 + 板块热度 — 四维共振</div></div>""", unsafe_allow_html=True)
        
        # 刷新 / 应用按钮
        bc1, bc2, bc3 = st.columns([1, 1, 4])
        with bc1:
            refresh_clicked = st.button("🔄 刷新数据", width='stretch', key="resonance_refresh")
        with bc2:
            apply_clicked = st.button("📊 查看结果", width='stretch', key="resonance_apply")
        
        # 数据状态
        today_str = datetime.now().strftime('%Y%m%d')
        
        # 自动触发：无缓存或点击刷新时执行扫描
        if 'resonance_auto_scanned' not in st.session_state:
            st.session_state.resonance_auto_scanned = False
        
        cached_data = get_resonance_cache() if not refresh_clicked else None
        
        # 判断是否需要扫描（无缓存 或 点击刷新）
        need_scan = refresh_clicked or (cached_data is None and not st.session_state.resonance_auto_scanned)
        
        if cached_data and not refresh_clicked:
            cache_time = cached_data.get('cache_time', '')
            st.info(f"📌 显示今日 {cache_time} 的缓存结果，点击「刷新数据」获取最新")
        
        # 加载或刷新数据
        if need_scan:
            with st.spinner("正在计算共振模型（通达信K线资金流向 + DDE代理 + K线结构 + 板块热度）..."):
                # 获取行情数据（走统一缓存兜底链）
                try:
                    quotes_df = fetch_all_a_stocks()
                except Exception as e:
                    st.error(f"获取行情数据失败: {e}")
                    quotes_df = None
                
                if quotes_df is not None and len(quotes_df) > 0:
                    # 过滤 ST/退市
                    quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)]
                    
                    resonance_data = get_resonance_data(quotes_df)
                    if resonance_data:
                        # 补充K线和板块评分
                        scores = calculate_resonance_score(resonance_data, quotes_df)
                        
                        # 按总分排序取 Top30
                        sorted_codes = sorted(scores.keys(), key=lambda x: scores[x]['total'], reverse=True)[:30]
                        
                        # 获取这些股票的名称等信息
                        # 🔧 预取概念板块映射
                        _, res_sec_map = fetch_sector_board_v3()
                        if res_sec_map:
                            import re
                            res_sec_map = {k: v for k, v in res_sec_map.items() if re.search(r'[\u4e00-\u9fff]', v)}
                        
                        result_list = []
                        for code in sorted_codes:
                            sc = scores[code]
                            name = ''
                            sector = ''
                            close = ''
                            row = quotes_df[quotes_df['代码'] == code]
                            if len(row) > 0:
                                name = str(row.iloc[0].get('名称', ''))
                            sector = res_sec_map.get(code, '') if res_sec_map else _get_sector(code)
                            style_tag = _classify_resonance_style(code, resonance_data, quotes_df)
                            result_list.append({
                                '代码': code, '名称': name, '板块': sector,
                                '共振评分': sc['total'],
                                '资金流向': sc['money_flow'],
                                'DDE决策': sc['dde_proxy'],
                                'K线结构': sc['kline_structure'],
                                '板块热度': sc['sector_heat'],
                                '当前走势': style_tag,
                            })
                        
                        cache_data = {
                            'cache_time': datetime.now().strftime('%H:%M'),
                            'results': result_list,
                            'raw_scores': scores,
                        }
                        save_resonance_cache(cache_data)
                        st.session_state['resonance_results'] = result_list
                        st.session_state.resonance_auto_scanned = True
                        st.rerun()
                    else:
                        st.warning("⚠️ 共振模型数据源均不可用：DDE Excel 不存在/解析失败，新浪财经资金流向备选链路也未获取到有效数据。可稍后重试或检查网络连接。")
                        st.session_state.resonance_auto_scanned = True
                else:
                    st.warning("⚠️ 未获取到行情数据，已跳过共振模型扫描。")
                    st.session_state.resonance_auto_scanned = True
        
        # 显示结果
        resonance_results = st.session_state.get('resonance_results', None)
        if resonance_results is None and cached_data:
            resonance_results = cached_data.get('results', [])
            st.session_state['resonance_results'] = resonance_results
        
        # 兼容旧缓存：补默认风格标签 / 旧键名迁移
        if resonance_results:
            for s in resonance_results:
                s.setdefault('当前走势', '蓄势待发')
                if '综合评分' in s and '共振评分' not in s:
                    s['共振评分'] = s.pop('综合评分')
        
        if resonance_results and len(resonance_results) > 0:
            _res_dynamic_n = calculate_dynamic_recommend_count()
            top_n = resonance_results[:_res_dynamic_n]
            
            st.markdown(f"## 🏆 今日精选 Top {_res_dynamic_n}")
            
            # 导出按钮
            _, btn_export, _ = st.columns([4, 1, 4])
            with btn_export:
                top_n_df = pd.DataFrame(top_n)
                top_n_df['代码'] = top_n_df['代码'].astype(str).str.zfill(6)
                xlsx_data = _export_df_to_xlsx(top_n_df)
                st.download_button(f"📥 导出Top{_res_dynamic_n}", xlsx_data, f"top{_res_dynamic_n}_resonance_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   width='stretch', key="dl_resonance_top")
            
            # Top N 卡片
            cols = st.columns(5)
            rank_colors = {0: '#FFD700', 1: '#C0C0C0', 2: '#CD7F32'}
            style_badge = {'盘中走强': ('#E53935', '#FFEBEE'), '超跌待反转': ('#1E88E5', '#E3F2FD'), '蓄势待发': ('#757575', '#F5F5F5')}
            for i, stock in enumerate(top_n):
                with cols[i % 5]:
                    rank_color = rank_colors.get(i, '#888')
                    tag = stock.get('当前走势', '蓄势待发')
                    tc, bg = style_badge.get(tag, ('#757575', '#F5F5F5'))
                    st.markdown(f"""
                    <div style="background:white;border-radius:12px;padding:12px;margin:4px;
                    border:1px solid #E0E0E0;position:relative;min-height:130px;">
                    <div style="position:absolute;top:4px;right:8px;font-size:22px;font-weight:800;color:{rank_color};">#{i+1}</div>
                    <div style="font-size:15px;font-weight:700;color:#333;margin-top:4px;">{stock['名称']}
                    <span style="display:inline-block;font-size:10px;color:{tc};background:{bg};padding:1px 6px;border-radius:8px;margin-left:4px;vertical-align:middle;">{tag}</span></div>
                    <div style="font-size:11px;color:#999;">{stock['代码']}</div>
                    <div style="font-size:22px;font-weight:700;color:#C4842D;margin:6px 0;">{stock['共振评分']}<span style="font-size:12px;color:#999;">分</span></div>
                    <div style="font-size:10px;color:#666;">
                    资金{stock['资金流向']} | DDE{stock['DDE决策']} | K线{stock['K线结构']} | 板块{stock['板块热度']}
                    </div>
                    <div style="font-size:10px;color:#999;">{stock['板块']}</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            # Top 30 详细表格
            st.markdown("---")
            
            # 当前走势分布摘要
            tags = [s.get('当前走势', '蓄势待发') for s in resonance_results[:30]]
            qs = sum(1 for t in tags if t == '盘中走强')
            dd = sum(1 for t in tags if t == '超跌待反转')
            zx = sum(1 for t in tags if t == '蓄势待发')
            st.caption(f"当前走势分布 — 🔴 盘中走强 {qs} 只 | 🔵 超跌待反转 {dd} 只 | ⚪ 蓄势待发 {zx} 只")
            
            st.markdown("### 📋 共振模型 Top 30 详细数据")
            
            df_display = pd.DataFrame(resonance_results[:30])
            df_display = df_display.rename(columns={
                '代码': '代码', '名称': '名称', '板块': '概念板块', '当前走势': '当前走势',
                '共振评分': '共振评分', '资金流向': '资金流向(30)',
                'DDE决策': 'DDE决策(20)', 'K线结构': 'K线结构(25)',
                '板块热度': '板块热度(25)'
            })
            
            # 带颜色渲染的表格
            def color_score(val):
                if isinstance(val, (int, float)):
                    if val >= 80: return 'background-color:#C8E6C9;font-weight:bold'
                    if val >= 70: return 'background-color:#E8F5E9'
                    if val >= 60: return 'background-color:#FFF9C4'
                    return ''
                return ''
            
            def color_style(val):
                if val == '盘中走强': return 'background-color:#FFEBEE;color:#C62828;font-weight:bold'
                if val == '超跌待反转': return 'background-color:#E3F2FD;color:#1565C0;font-weight:bold'
                return 'background-color:#F5F5F5;color:#757575'
            
            styled_df = df_display.style.map(color_score, subset=['共振评分']).map(color_style, subset=['当前走势'])
            st.dataframe(styled_df, width='stretch', hide_index=True,
                        column_config={
                            '共振评分': st.column_config.NumberColumn(format='%.1f'),
                            '资金流向(30)': st.column_config.NumberColumn(format='%.1f'),
                            'DDE决策(20)': st.column_config.NumberColumn(format='%.1f'),
                            'K线结构(25)': st.column_config.NumberColumn(format='%.1f'),
                            '板块热度(25)': st.column_config.NumberColumn(format='%.1f'),
                        })
            
            # 导出
            export_df = pd.DataFrame(resonance_results)
            export_df['代码'] = export_df['代码'].astype(str).str.zfill(6)
            csv_data = export_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("📥 导出 Top30 CSV", csv_data, f"resonance_top30_{today_str}.csv",
                             "text/csv", key="dl_resonance_top30")
            
            st.markdown("---")
            st.caption("操作区 — 评分详情 | 加/取消自选")
            render_stock_buttons(resonance_results[:30], prefix="rs")
        elif not need_scan:
            st.info("💡 点击「刷新数据」获取今日共振模型选股结果（基于东方财富DDE数据计算）。")
    
    with model_tabs[1]:  # 追高模型
        st.session_state.current_model = 'chase_high'
        
        # 顶部参数面板
        render_top_params_panel()
        
        st.markdown(f"""<div class="header-container"><div class="main-title">🚀 追高模型 · 抓强势龙头股</div>
        <div class="sub-title">八维强势评分 · 全市场量化筛选</div></div>""", unsafe_allow_html=True)
        
        # 显示缓存状态
        if st.session_state.get('data_status') == 'cached' and st.session_state.get('last_update_time'):
            cache_time = st.session_state.last_update_time
            st.info(f"📌 显示今日 {cache_time.strftime('%H:%M')} 的缓存结果，点击「刷新数据」获取最新")
        
        # 样本来源选择（紧跟header，与金叉模型一致）
        cr1, cr2 = st.columns([1.8, 1])
        with cr1:
            cur_source = st.session_state.get('chase_sample_source', '全市场A股')
            idx = 0 if cur_source == '全市场A股' else (1 if cur_source == '热门板块' else 2)
            new_source = st.selectbox(
                "样本来源", DEFAULT_GC_SAMPLE_OPTIONS,
                index=idx,
                key="chase_sample_source_select"
            )
            if "全市场" in new_source:
                new_val = "全市场A股"
            elif "量价" in new_source:
                new_val = "量价"
            else:
                new_val = "热门板块"
            if new_val != cur_source:
                st.cache_data.clear()
                st.session_state.top10_cache = None
                st.session_state.top10_cache_key = None
                st.session_state.lowbuy_cache = None
                st.session_state.chase_results = None
                st.session_state.cache_loaded = False
                st.session_state.chase_sample_source = new_val
                st.rerun()
            else:
                st.session_state.chase_sample_source = new_val
        
        render_top_chase_high()
        
        st.markdown("")
        chase_sample = st.session_state.get('chase_sample_source', '全市场A股')
        col1, col2 = st.columns([2, 1])
        if chase_sample == "热门板块":
            scan_label = "🔍 加速板块扫描"
        elif chase_sample == "量价":
            scan_label = "🔍 量价扫描"
        else:
            scan_label = "🔍 全市场扫描"
        with col1:
            do_scan = st.button(scan_label, type="primary", width='stretch', key="chase_scan")
        with col2:
            if st.button("🔄 清空缓存", width='stretch', key="chase_clear"):
                st.session_state.chase_results = None
                st.rerun()

        chase_results = st.session_state.get('chase_results', None)

        if do_scan or chase_results is None:
            chase_spin = st.session_state.get('chase_sample_source', '全市场A股')
            if chase_spin == "热门板块":
                spinner_text = "正在扫描资金加速板块，计算追高评分…"
            elif chase_spin == "量价":
                spinner_text = "正在扫描量价反转板块，计算追高评分…"
            else:
                spinner_text = "正在扫描全市场，计算追高评分…"
            with st.spinner(spinner_text):
                st.cache_data.clear()
                if os.path.exists(CACHE_DATA_JSON):
                    try: os.remove(CACHE_DATA_JSON)
                    except: pass
                st.session_state.top10_cache = None
                st.session_state.top10_cache_key = None
                st.session_state.lowbuy_cache = None
                st.session_state.cache_loaded = False
                df = get_stock_pool()
                st.session_state.chase_results = df
                st.rerun()

        st.markdown("")
        df = chase_results if chase_results is not None else pd.DataFrame()
        if len(df) == 0 or '板块' not in df.columns or '信号' not in df.columns:
            st.warning("⚠️ 数据加载中或接口暂时不可用，请稍后点击扫描按钮重试。")
            st.stop()
        render_stats_chase(df)
        st.markdown("")
        
        st.markdown("---")
        title_col, btn_col = st.columns([5, 1])
        with title_col:
            st.markdown("### 📊 单项评分排行 · 显示前30只")

        # 三个选择栏排成一行：板块、信号、排序方式
        sectors = ["全部"] + sorted(df["板块"].unique().tolist())
        sr_options = [("全部", None), ("强势买入", ("信号", ["强势买入"])),
            ("逢低吸纳", ("信号", ["逢低吸纳"])), ("观望等待", ("信号", ["观望等待"])),
            ("建议回避", ("信号", ["建议回避"]))]
        fc1, fc2, fc3 = st.columns([1, 1, 1])
        with fc1:
            sec = st.selectbox("板块", sectors, label_visibility="collapsed", key="sf_sec")
        with fc2:
            slabs = [o[0] for o in sr_options]
            sel_idx = slabs.index(st.selectbox("信号筛选", slabs, label_visibility="collapsed", key="sf_sig"))
            sig_filter = sr_options[sel_idx][1]
        with fc3:
            sb = st.selectbox("排序方式", ["综合评分↓", "趋势结构↓", "动量强度↓", "板块共振↓", "北向资金↓", "量价配合↓", "5日涨幅↓"],
                label_visibility="collapsed", key="ch_sort")

        filt = df.copy()
        if sec != "全部": filt = filt[filt["板块"] == sec]
        if sig_filter is not None and sig_filter[0] == "信号": filt = filt[filt["信号"].isin(sig_filter[1])]

        if sb == "5日涨幅↓":
            filt['_chg_n'] = filt["5日涨幅"].str.replace('%','').str.replace('+','').astype(float)
            filt = filt.sort_values('_chg_n', ascending=False)
        else:
            sort_map = {"综合评分↓": ("综合评分", False), "趋势结构↓": ("趋势结构", False),
                "动量强度↓": ("动量强度", False), "板块共振↓": ("板块共振", False),
                "北向资金↓": ("北向资金", False), "量价配合↓": ("量价配合", False)}
            col_name, asc = sort_map.get(sb, ("综合评分", False))
            filt = filt.sort_values(col_name, ascending=asc)

        top30 = filt.head(30).reset_index(drop=True)
        with btn_col:
            st.markdown("")  # align with title
            xlsx_data = _export_df_to_xlsx(top30)
            st.download_button("📥 导出Top30", xlsx_data, f"top30_chase_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width='stretch', key="dl_ch_top30")
        dde_scores = get_dde_confirmation_scores()
        res_scores, res_styles = _get_resonance_cross_ref()
        render_table_chase(top30, dde_scores=dde_scores,
                           resonance_scores=res_scores, resonance_styles=res_styles)
    
    with model_tabs[2]:  # 低吸模型
        st.session_state.current_model = 'buy_low'
        
        # 顶部参数面板
        render_top_params_panel()
        
        # 样本来源选择
        sc1, sc2 = st.columns([1.8, 1])
        with sc1:
            cur_source = st.session_state.get('lowbuy_sample_source', '全市场A股')
            idx = 0 if cur_source == '全市场A股' else (1 if cur_source == '热门板块' else 2)
            new_source = st.selectbox(
                "样本来源", DEFAULT_GC_SAMPLE_OPTIONS,
                index=idx,
                key="lowbuy_sample_source_select"
            )
            if "全市场" in new_source:
                new_val = "全市场A股"
            elif "量价" in new_source:
                new_val = "量价"
            else:
                new_val = "热门板块"
            if new_val != cur_source:
                st.session_state.lowbuy_cache = None
                st.session_state.lowbuy_sample_source = new_val
                st.session_state.lowbuy_auto_scanned = False
                st.rerun()
            else:
                st.session_state.lowbuy_sample_source = new_val
        
        _lp = _get_lowbuy_params()
        st.markdown(f"""<div class="header-container" style="border-color:#C8E6C9;background:linear-gradient(135deg, #F0FFF4 0%, #FFFFFF 100%);">
        <div class="main-title" style="color:#2E7D32;">📉 低吸模型 · 抓超跌价值股</div>
        <div class="sub-title">七维评分硬过滤 + K线底部反转信号 · 10日跌{_lp['decline_20d_low']}%~-{-_lp['decline_20d_high']}% · 独立评分 · 精选{_lp['max_results']}只</div>
        </div>""", unsafe_allow_html=True)
        
        # 自动触发：首次加载时缓存为空则自动扫描
        lb_sample = st.session_state.get('lowbuy_sample_source', '全市场A股')
        lb_cache = st.session_state.get('lowbuy_cache', None)
        if lb_cache is None and not st.session_state.get('lowbuy_auto_scanned', False):
            render_top_low_buy(sample_source=lb_sample)
            st.session_state.lowbuy_auto_scanned = True
            st.rerun()
        
        render_top_low_buy(sample_source=st.session_state.get('lowbuy_sample_source', '全市场A股'))
        
        st.markdown("")
        bc1, bc2, bc3 = st.columns([1, 1, 4])
        with bc1:
            lb_source = st.session_state.get('lowbuy_sample_source', '全市场A股')
            if lb_source == "热门板块":
                lb_label = "🔄 加速扫描"
            elif lb_source == "量价":
                lb_label = "🔄 量价扫描"
            else:
                lb_label = "🔄 全市场扫描"
            if st.button(lb_label, width='stretch', key="lowbuy_refresh_data"):
                st.cache_data.clear()
                if os.path.exists(CACHE_DATA_JSON):
                    try: os.remove(CACHE_DATA_JSON)
                    except: pass
                st.session_state.top10_cache = None
                st.session_state.top10_cache_key = None
                st.session_state.lowbuy_cache = None
                st.session_state.lowbuy_auto_scanned = False
                st.session_state.cache_loaded = False
                st.rerun()

        st.markdown("")
        lb_df = pd.DataFrame(st.session_state.get('lowbuy_cache', []))
        if len(lb_df) > 0:
            render_stats_lowbuy(lb_df)
            st.markdown("")
            st.markdown("---")
            title_col, btn_col = st.columns([5, 1])
            with title_col:
                st.markdown("### 📊 单项评分排行 · 显示前30只")
            with btn_col:
                st.markdown("")
                xlsx_data = _export_df_to_xlsx(lb_df.head(30))
                st.download_button("📥 导出Top30", xlsx_data, f"top30_lowbuy_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width='stretch', key="dl_lb_top30")
            dde_scores = get_dde_confirmation_scores()
            res_scores, res_styles = _get_resonance_cross_ref()
            render_table_lowbuy(lb_df.head(30), dde_scores=dde_scores,
                                resonance_scores=res_scores, resonance_styles=res_styles)
        elif st.session_state.get('lowbuy_cache') is not None:
            st.warning("⚠️ 今日未找到符合低吸条件的股票，请尝试放宽参数后点击「刷新数据」重试。")
            st.stop()
    
    with model_tabs[3]:  # 金叉模型
        st.session_state.current_model = 'rebound_model'
        
        render_top_params_panel()
        
        st.markdown("""<div class="header-container"><div class="main-title">📌 金叉模型 · 抓强势回调股</div>
        <div class="sub-title">强势股回调反弹 — 五维评分（下跌形态·K线止跌·均线拐头·量能确认·MACD反转）</div></div>""", unsafe_allow_html=True)
        
        # 样本来源 — 独立于参数面板，选择即时生效
        sc1, sc2 = st.columns([1.8, 1])
        with sc1:
            cur_source = st.session_state.gc_params.get("sample_source", "全市场A股")
            idx = 0 if cur_source == "全市场A股" else (1 if cur_source == "热门板块" else 2)
            new_source = st.selectbox(
                "样本来源", DEFAULT_GC_SAMPLE_OPTIONS,
                index=idx,
                key="gc_sample_source_quick"
            )
            if "全市场" in new_source:
                st.session_state.gc_params["sample_source"] = "全市场A股"
            elif "量价" in new_source:
                st.session_state.gc_params["sample_source"] = "量价"
            else:
                st.session_state.gc_params["sample_source"] = "热门板块"
        
        col1, col2, col3 = st.columns([2, 1, 1])
        gc_sample = st.session_state.gc_params.get("sample_source", "全市场A股")
        if gc_sample == "热门板块":
            scan_label = "🔍 加速板块扫描"
        elif gc_sample == "量价":
            scan_label = "🔍 量价反转扫描"
        else:
            scan_label = "🔍 全市场扫描"
        with col1:
            do_scan = st.button(scan_label, type="primary", width='stretch', key="gc_scan")
        with col2:
            if st.button("🔄 刷新参数", width='stretch', key="gc_refresh_params"):
                st.session_state.gc_params = dict(DEFAULT_GC_PARAMS)
                st.session_state.gc_weights = dict(DEFAULT_GC_WEIGHTS)
                st.rerun()
        
        gc_cache = st.session_state.get('gc_results', None)
        
        if do_scan or (gc_cache is None and not st.session_state.get('gc_scanned', False)):
            if gc_sample == "热门板块":
                spinner_text = "正在扫描资金加速板块，筛选金叉标的…"
            elif gc_sample == "量价":
                spinner_text = "正在扫描量价反转板块，筛选金叉标的…"
            else:
                spinner_text = "正在扫描全市场，筛选金叉标的…"
            with st.spinner(spinner_text):
                try:
                    quotes_df = fetch_all_a_stocks()
                    if quotes_df is None or len(quotes_df) == 0:
                        st.error("获取行情数据失败")
                    else:
                        quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)]
                        
                        # 🌟 热门板块过滤
                        if gc_sample == "热门板块":
                            hot_codes = get_hot_concept_stocks(6)
                            if hot_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(hot_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只热门概念板块成分股")
                            else:
                                st.warning("未能获取热门板块数据，回退为全市场扫描")
                        elif gc_sample == "量价":
                            vp_codes = get_volprice_sectors(6)
                            if vp_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(vp_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只量价反转板块成分股")
                            else:
                                st.warning("未能获取量价反转板块数据，回退为全市场扫描")
                        
                        # 🔧 填充概念板块（数据源无板块字段，从 pytdx 板块成分股映射获取）
                        _, stock_sector_map = fetch_sector_board_v3()
                        quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                        if stock_sector_map:
                            # 过滤掉非中文板块名（block_fg.dat/block_zs.dat 含内部编码如 W300881W3）
                            import re
                            valid_map = {k: v for k, v in stock_sector_map.items() if re.search(r'[\u4e00-\u9fff]', v)}
                            quotes_df['板块'] = quotes_df['代码'].map(valid_map).fillna('')
                        else:
                            quotes_df['板块'] = ''

                        codes_list = [(str(r['代码']).zfill(6), str(r['名称']), str(r.get('板块', '')))
                                      for _, r in quotes_df.iterrows()]
                        results = _run_golden_cross_scan(tuple(codes_list),
                                                         json.dumps(st.session_state.gc_params),
                                                         json.dumps(st.session_state.gc_weights))
                        st.session_state.gc_results = results
                        st.session_state.gc_scanned = True
                        st.rerun()
                except Exception as e:
                    st.error(f"扫描出错: {e}")
        
        gc_results = st.session_state.get('gc_results', None)
        
        if gc_results is not None:
            if len(gc_results) == 0:
                st.warning("⚠️ 今日未找到符合条件的金叉标的。")
            else:
                st.markdown(f"### 📊 金叉模型 · Top {min(len(gc_results), 30)}")
                st.caption(f"共筛选出 {len(gc_results)} 只标的")
                
                # ---- Top N 精选（动态数量）----
                _gc_dyn_n = calculate_dynamic_recommend_count()
                gc_top_n = gc_results[:_gc_dyn_n]
                st.markdown(f"""
                <div class="top10-container">
                    <div class="top10-header">
                        <div class="top10-title">📌 金叉精选 Top {_gc_dyn_n} <span class="top10-badge">超跌反弹信号</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)
                btn1, btn2, _ = st.columns([1, 1, 4])
                with btn1:
                    if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="gc_add_all"):
                        for s in gc_top_n:
                            if s["代码"] not in st.session_state.watchlist:
                                st.session_state.watchlist.append(s["代码"])
                        save_watchlist(st.session_state.watchlist)
                        st.success(f"已将Top {_gc_dyn_n}全部加入自选！"); st.rerun()
                with btn2:
                    gc_top_df = pd.DataFrame(gc_top_n)
                    gc_top_df['代码'] = gc_top_df['代码'].astype(str).str.zfill(6)
                    gc_top_df = gc_top_df.rename(columns={'板块': '概念板块'})
                    export_cols = ['代码', '名称', '概念板块',
                                   '下跌形态', 'K线止跌', '均线拐头', '量能确认', 'MACD反转', '资金确认',
                                   '共振评分', '共振评价', '金叉评分', '信号', '建议']
                    gc_top_df = gc_top_df[[c for c in export_cols if c in gc_top_df.columns]]
                    xlsx_data = _export_df_to_xlsx(gc_top_df)
                    st.download_button(f"📥 导出Top{_gc_dyn_n}", xlsx_data, f"top{_gc_dyn_n}_golden_cross_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       width='stretch', key="dl_gc_t10")
                
                st.markdown("")
                for i in range(0, len(gc_top_n), 5):
                    row_stocks = gc_top_n[i:i+5]
                    cols = st.columns(5)
                    for j, stock in enumerate(row_stocks):
                        with cols[j]:
                            rank = i + j + 1
                            rank_color = "#FFD700" if rank == 1 else "#C0C0C0" if rank == 2 else "#CD7F32" if rank == 3 else "#CCC"
                            signal = stock.get('信号', '-')
                            st.markdown(f"""
                            <div class="top10-card">
                                <div class="top10-rank" style="color:{rank_color};">{rank}</div>
                                <div style="padding-right:30px;">
                                    <div style="font-weight:700;color:#333;font-size:15px;margin-bottom:2px;">{stock['名称']}</div>
                                    <div style="font-size:11px;color:#888;margin-bottom:8px;">{str(stock.get('代码','')).zfill(6)} · {stock.get('板块', '其他')[:8]}</div>
                                    <div style="display:flex;justify-content:space-between;align-items:center;">
                                        <span style="font-size:20px;font-weight:800;color:#C4842D;">{stock['金叉评分']}</span>
                                    </div>
                                    <div style="margin-top:6px;"><span class="metric-badge badge-strong">{signal}</span></div>
                                </div>
                            </div>""", unsafe_allow_html=True)
                            code = stock["代码"]; in_wl = code in st.session_state.watchlist
                            if st.button("⭐" if in_wl else "+自选", key=f"gc_t10_{code}",
                                width='stretch', type="primary" if in_wl else "secondary"):
                                if in_wl: st.session_state.watchlist.remove(code)
                                else: st.session_state.watchlist.append(code)
                                save_watchlist(st.session_state.watchlist); st.rerun()
                
                st.markdown("---")
                
                # 共振交叉评分
                res_scores, res_styles = _get_resonance_cross_ref()
                
                fdf = pd.DataFrame(gc_results[:30])
                fdf.index = range(1, len(fdf) + 1)
                fdf['代码'] = fdf['代码'].astype(str).str.zfill(6)
                fdf = fdf.rename(columns={'板块': '概念板块'})
                
                if res_scores:
                    fdf['共振评分'] = fdf['代码'].apply(lambda x: res_scores.get(x, None))
                if res_styles:
                    fdf['共振评价'] = fdf['代码'].apply(lambda x: res_styles.get(x, '-'))
                
                display_cols = ['代码', '名称', '概念板块', '信号', '建议', '金叉评分',
                                '下跌形态', 'K线止跌', '均线拐头', '量能确认', 'MACD反转', '资金确认', '板块确认',
                                '共振评分', '共振评价']
                df_display = fdf[[c for c in display_cols if c in fdf.columns]]
                
                def color_score(val):
                    if isinstance(val, (int, float)):
                        if val >= 80: return 'background-color:#C8E6C9;font-weight:bold'
                        if val >= 70: return 'background-color:#E8F5E9'
                        if val >= 60: return 'background-color:#FFF9C4'
                        return ''
                    return ''
                
                styled = df_display.style.map(color_score, subset=['金叉评分'])
                st.dataframe(styled, width='stretch',
                             column_config={
                                 '金叉评分': st.column_config.NumberColumn(format='%.0f'),
                             })
                
                st.markdown("---")
                st.caption("操作区 — 评分详情 | 加/取消自选")
                gc_for_buttons = []
                for r in gc_results[:30]:
                    gc_for_buttons.append({'代码': r.get('代码', ''), '名称': r.get('名称', '')})
                render_stock_buttons(gc_for_buttons, prefix="gc")
                
                # 导出
                export_df = pd.DataFrame(gc_results)
                export_df['代码'] = export_df['代码'].astype(str).str.zfill(6)
                export_df = export_df.rename(columns={'板块': '概念板块'})
                export_cols = ['代码', '名称', '概念板块', '信号', '建议', '金叉评分',
                               '下跌形态', 'K线止跌', '均线拐头', '量能确认', 'MACD反转', '资金确认', '板块确认',
                               '共振评分', '共振评价']
                export_df = export_df[[c for c in export_cols if c in export_df.columns]]
                csv_data = export_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 导出 Top30 CSV", csv_data, f"rebound_model_top30_{datetime.now().strftime('%Y%m%d')}.csv",
                                   "text/csv", key="dl_gc_top30")
        elif not do_scan:
            st.info("💡 点击「全市场扫描」启动金叉模型选股（基于五维评分 + 硬过滤条件）。")

    with model_tabs[4]:  # CAN SLIM模型（抓主升浪股）
        st.session_state.current_model = 'canslim'
        
        render_top_params_panel()
        
        st.markdown("""<div class="header-container"><div class="main-title">📊 CAN SLIM模型（抓主升浪股）</div>
        <div class="sub-title">七因子简化版：C业绩 · A持续增长 · N新催化 · S中小盘 · L_RPS · I流动性 · M大势</div></div>""", unsafe_allow_html=True)
        
        sc1, sc2 = st.columns([1.8, 1])
        with sc1:
            cur_source = st.session_state.get('canslim_sample_source', '全市场A股')
            idx = 0 if cur_source == '全市场A股' else (1 if cur_source == '热门板块' else 2)
            new_source = st.selectbox(
                "样本来源", DEFAULT_GC_SAMPLE_OPTIONS,
                index=idx,
                key="canslim_sample_source_select"
            )
            if "全市场" in new_source:
                new_val = "全市场A股"
            elif "量价" in new_source:
                new_val = "量价"
            else:
                new_val = "热门板块"
            if new_val != cur_source:
                st.session_state.canslim_results = None
                st.session_state.canslim_sample_source = new_val
                st.rerun()
            else:
                st.session_state.canslim_sample_source = new_val
        
        col1, col2 = st.columns([2, 1])
        cs_sample = st.session_state.get('canslim_sample_source', '全市场A股')
        if cs_sample == "热门板块":
            scan_label = "🔍 加速板块扫描"
        elif cs_sample == "量价":
            scan_label = "🔍 量价扫描"
        else:
            scan_label = "🔍 全市场扫描"
        with col1:
            do_scan = st.button(scan_label, type="primary", width='stretch', key="cs_scan")
        with col2:
            if st.button("🔄 清空缓存", width='stretch', key="cs_clear"):
                st.session_state.canslim_results = None
                st.rerun()
        
        cs_results = st.session_state.get('canslim_results', None)
        
        if do_scan or cs_results is None:
            if cs_sample == "热门板块":
                spinner_text = "正在扫描资金加速板块，计算CAN SLIM评分…"
            elif cs_sample == "量价":
                spinner_text = "正在扫描量价反转板块，计算CAN SLIM评分…"
            else:
                spinner_text = "正在扫描全市场，计算CAN SLIM评分…"
            with st.spinner(spinner_text):
                try:
                    quotes_df = fetch_all_a_stocks()
                    if quotes_df is None or len(quotes_df) == 0:
                        st.error("获取行情数据失败")
                    else:
                        quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)]
                        
                        if cs_sample == "热门板块":
                            hot_codes = get_hot_concept_stocks(6)
                            if hot_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(hot_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只热门概念板块成分股")
                            else:
                                st.warning("未能获取热门板块数据，回退为全市场扫描")
                        elif cs_sample == "量价":
                            vp_codes = get_volprice_sectors(6)
                            if vp_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(vp_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只量价反转板块成分股")
                            else:
                                st.warning("未能获取量价反转板块数据，回退为全市场扫描")
                        
                        if '量比' in quotes_df.columns:
                            quotes_df['量比'] = pd.to_numeric(quotes_df['量比'], errors='coerce').fillna(1)
                            quotes_df = quotes_df.sort_values('量比', ascending=False)
                        scan_df = quotes_df.head(300).copy()
                        
                        codes = scan_df['代码'].tolist()
                        names = dict(zip(scan_df['代码'], scan_df['名称']))
                        turnover_map = {}
                        if '换手率' in scan_df.columns:
                            turnover_map = dict(zip(scan_df['代码'],
                                pd.to_numeric(scan_df['换手率'], errors='coerce').fillna(0)))
                        # 同时提取总市值（避免ctx硬编码0导致S维度永远0分）
                        cap_map = {}
                        if '总市值' in scan_df.columns:
                            cap_map = dict(zip(scan_df['代码'],
                                pd.to_numeric(scan_df['总市值'], errors='coerce').fillna(0)))
                        
                        kline_dict = {}
                        status = st.empty()
                        bar = st.progress(0)
                        total = len(codes)
                        with ThreadPoolExecutor(max_workers=10) as ex:
                            futures = {ex.submit(get_stock_kline, c, 250): c for c in codes}
                            done = 0
                            for f in as_completed(futures):
                                done += 1
                                c = futures[f]
                                try:
                                    kline = f.result(timeout=15)
                                    if kline is not None and len(kline) >= 60:
                                        kline_dict[c] = kline
                                except:
                                    pass
                                if done % 20 == 0:
                                    bar.progress(done / total)
                                    status.text(f"📊 获取K线数据... ({done}/{total})")
                        bar.empty()
                        status.empty()
                        
                        st.text("📊 计算RPS排名...")
                        rps_map = compute_rps(kline_dict, list(kline_dict.keys()))
                        
                        results = []
                        scored_codes = list(kline_dict.keys())
                        # 并发获取财务数据（10线程）
                        fin_data_map = {}
                        fin_bar = st.progress(0)
                        fin_status = st.empty()
                        with ThreadPoolExecutor(max_workers=10) as ex:
                            futures = {ex.submit(get_financial_data, code): code for code in scored_codes}
                            fin_done = 0
                            for f in as_completed(futures):
                                fin_done += 1
                                code = futures[f]
                                try:
                                    fin_data_map[code] = f.result(timeout=30)
                                except:
                                    fin_data_map[code] = None
                                if fin_done % 30 == 0:
                                    fin_bar.progress(fin_done / len(scored_codes))
                                    fin_status.text(f"📊 获取财务数据... ({fin_done}/{len(scored_codes)})")
                        fin_bar.empty()
                        fin_status.empty()
                        
                        # 统计财务数据获取情况
                        fin_success_count = sum(1 for v in fin_data_map.values() if v and v.get('success'))
                        if fin_success_count < len(scored_codes) * 0.5:
                            st.warning(f"⚠️ 财务数据获取率仅 {fin_success_count}/{len(scored_codes)}，C/A维度可能显示N/A。请检查网络或稍后重试。")
                        
                        # 串行评分计算
                        bar2 = st.progress(0)
                        for i, code in enumerate(scored_codes):
                            kline_df = kline_dict[code]
                            fin_data = fin_data_map.get(code)
                            ctx = {
                                'rps': rps_map.get(code, 0),
                                'market_cap': cap_map.get(code, 0),
                                'turnover_rate': turnover_map.get(code, 0),
                                'fin': fin_data,
                            }
                            time.sleep(0.01)
                            sr = calculate_canslim_score(code, kline_df, stock_pool_context=ctx)
                            if sr.get('pass') and sr.get('综合评分', 0) > 0:
                                results.append({
                                    '代码': code,
                                    '名称': names.get(code, ''),
                                    '综合评分': sr['综合评分'],
                                    'C_业绩增速': sr.get('C_业绩增速', 0),
                                    'A_持续增长': sr.get('A_持续增长', 0),
                                    'N_新催化': sr.get('N_新催化', 0),
                                    'S_中小盘': sr.get('S_中小盘', 0),
                                    'L_RPS': sr.get('L_RPS', 0),
                                    'I_流动性': sr.get('I_流动性', 0),
                                    'M_大势': sr.get('M_大势', 0),
                                })
                            if (i + 1) % 50 == 0:
                                bar2.progress((i + 1) / len(scored_codes))
                        bar2.empty()
                        
                        results.sort(key=lambda x: x['综合评分'], reverse=True)
                        st.session_state.canslim_results = results
                        st.rerun()
                except Exception as e:
                    st.error(f"扫描出错: {e}")
        
        cs_results = st.session_state.get('canslim_results', None)
        
        if cs_results is not None:
            if len(cs_results) == 0:
                st.warning("⚠️ 今日未找到符合条件的CAN SLIM标的。")
            else:
                top_n = min(len(cs_results), 30)
                st.markdown(f"### 📊 CAN SLIM模型 · Top {top_n}")
                st.caption(f"共筛选出 {len(cs_results)} 只标的")
                
                _cs_dyn_n = calculate_dynamic_recommend_count()
                cs_top_n = cs_results[:_cs_dyn_n]
                st.markdown(f"""
                <div class="top10-container">
                    <div class="top10-header">
                        <div class="top10-title">📊 CAN SLIM 精选 Top {_cs_dyn_n} <span class="top10-badge">主升浪信号</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)
                btn1, btn2, _ = st.columns([1, 1, 4])
                with btn1:
                    if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="cs_add_all"):
                        for s in cs_top_n:
                            if s["代码"] not in st.session_state.watchlist:
                                st.session_state.watchlist.append(s["代码"])
                        save_watchlist(st.session_state.watchlist)
                        st.success(f"已将Top {_cs_dyn_n}全部加入自选！"); st.rerun()
                with btn2:
                    top_n_df = pd.DataFrame(cs_top_n)
                    top_n_df['代码'] = top_n_df['代码'].astype(str).str.zfill(6)
                    export_cols = ['代码', '名称',
                                   '综合评分', 'C_业绩增速', 'A_持续增长', 'N_新催化',
                                   'S_中小盘', 'L_RPS', 'I_流动性', 'M_大势']
                    top_n_df = top_n_df[[c for c in export_cols if c in top_n_df.columns]]
                    xlsx_data = _export_df_to_xlsx(top_n_df)
                    st.download_button(f"📥 导出Top{_cs_dyn_n}", xlsx_data,
                        f"top{_cs_dyn_n}_canslim_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width='stretch', key="dl_cs_t10")
                
                st.markdown("")
                for i in range(0, len(cs_top_n), 5):
                    row_stocks = cs_top_n[i:i+5]
                    cols = st.columns(5)
                    for j, stock in enumerate(row_stocks):
                        with cols[j]:
                            rank = i + j + 1
                            rank_color = "#FFD700" if rank == 1 else "#C0C0C0" if rank == 2 else "#CD7F32" if rank == 3 else "#CCC"
                            st.markdown(f"""
                            <div class="top10-card">
                                <div class="top10-rank" style="color:{rank_color};">{rank}</div>
                                <div style="padding-right:30px;">
                                    <div style="font-weight:700;color:#333;font-size:15px;margin-bottom:2px;">{stock['名称']}</div>
                                    <div style="font-size:11px;color:#888;margin-bottom:8px;">{str(stock.get('代码','')).zfill(6)}</div>
                                    <div style="display:flex;justify-content:space-between;align-items:center;">
                                        <span style="font-size:20px;font-weight:800;color:#C4842D;">{stock['综合评分']}</span>
                                    </div>
                                    <div style="margin-top:6px;"><span class="metric-badge badge-strong">CAN SLIM</span></div>
                                </div>
                            </div>""", unsafe_allow_html=True)
                            code = stock["代码"]; in_wl = code in st.session_state.watchlist
                            if st.button("⭐" if in_wl else "+自选", key=f"cs_t10_{code}",
                                width='stretch', type="primary" if in_wl else "secondary"):
                                if in_wl: st.session_state.watchlist.remove(code)
                                else: st.session_state.watchlist.append(code)
                                save_watchlist(st.session_state.watchlist); st.rerun()
                
                st.markdown("---")
                fdf = pd.DataFrame(cs_results[:top_n])
                fdf.index = range(1, len(fdf) + 1)
                fdf['代码'] = fdf['代码'].astype(str).str.zfill(6)
                display_cols = ['代码', '名称', '综合评分',
                                'C_业绩增速', 'A_持续增长', 'N_新催化',
                                'S_中小盘', 'L_RPS', 'I_流动性', 'M_大势']
                df_display = fdf[[c for c in display_cols if c in fdf.columns]].copy()
                # -1 → "N/A" 表示财务数据不可用，与真的0分区分
                for col in ['C_业绩增速', 'A_持续增长', 'S_中小盘', 'I_流动性']:
                    if col in df_display.columns:
                        df_display[col] = df_display[col].apply(lambda x: "N/A" if x == -1 else x)
                # Arrow 序列化兼容：含 "N/A" 的 object 列统一转 str
                for c in df_display.columns:
                    if df_display[c].dtype == 'object':
                        df_display[c] = df_display[c].astype(str)
                st.dataframe(df_display, width='stretch')
                
                st.markdown("---")
                st.caption("操作区 — 加/取消自选")
                cs_buttons = [{'代码': r['代码'], '名称': r['名称']} for r in cs_results[:30]]
                render_stock_buttons(cs_buttons, prefix="cs")
                
                export_df = pd.DataFrame(cs_results)
                export_df['代码'] = export_df['代码'].astype(str).str.zfill(6)
                export_cols = ['代码', '名称', '综合评分',
                               'C_业绩增速', 'A_持续增长', 'N_新催化',
                               'S_中小盘', 'L_RPS', 'I_流动性', 'M_大势']
                export_df = export_df[[c for c in export_cols if c in export_df.columns]]
                csv_data = export_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 导出 Top30 CSV", csv_data,
                    f"canslim_top30_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv", key="dl_cs_top30")
        elif not do_scan:
            st.info("💡 点击「全市场扫描」启动CAN SLIM模型选股（基于七因子简化版评分）。")

    with model_tabs[5]:  # 困境反转模型（抓周期股）
        st.session_state.current_model = 'dilemma_reversal'
        
        render_top_params_panel()
        
        st.markdown("""<div class="header-container"><div class="main-title">🔄 困境反转模型（抓周期股）</div>
        <div class="sub-title">四层评分：L1拐点 · L2反转 · L3安全垫 · L4技术资金（40分）</div></div>""", unsafe_allow_html=True)
        
        sc1, sc2 = st.columns([1.8, 1])
        with sc1:
            cur_source = st.session_state.get('dr_sample_source', '全市场A股')
            idx = 0 if cur_source == '全市场A股' else (1 if cur_source == '热门板块' else 2)
            new_source = st.selectbox(
                "样本来源", DEFAULT_GC_SAMPLE_OPTIONS,
                index=idx,
                key="dr_sample_source_select"
            )
            if "全市场" in new_source:
                new_val = "全市场A股"
            elif "量价" in new_source:
                new_val = "量价"
            else:
                new_val = "热门板块"
            if new_val != cur_source:
                st.session_state.dr_results = None
                st.session_state.dr_sample_source = new_val
                st.rerun()
            else:
                st.session_state.dr_sample_source = new_val
        
        col1, col2 = st.columns([2, 1])
        dr_sample = st.session_state.get('dr_sample_source', '全市场A股')
        if dr_sample == "热门板块":
            scan_label = "🔍 加速板块扫描"
        elif dr_sample == "量价":
            scan_label = "🔍 量价扫描"
        else:
            scan_label = "🔍 全市场扫描"
        with col1:
            do_scan = st.button(scan_label, type="primary", width='stretch', key="dr_scan")
        with col2:
            if st.button("🔄 清空缓存", width='stretch', key="dr_clear"):
                st.session_state.dr_results = None
                st.rerun()
        
        dr_results = st.session_state.get('dr_results', None)
        
        if do_scan or dr_results is None:
            if dr_sample == "热门板块":
                spinner_text = "正在扫描资金加速板块，计算困境反转评分…"
            elif dr_sample == "量价":
                spinner_text = "正在扫描量价反转板块，计算困境反转评分…"
            else:
                spinner_text = "正在扫描全市场，计算困境反转评分…"
            with st.spinner(spinner_text):
                try:
                    quotes_df = fetch_all_a_stocks()
                    if quotes_df is None or len(quotes_df) == 0:
                        st.error("获取行情数据失败")
                    else:
                        quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)]
                        
                        if dr_sample == "热门板块":
                            hot_codes = get_hot_concept_stocks(6)
                            if hot_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(hot_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只热门概念板块成分股")
                            else:
                                st.warning("未能获取热门板块数据，回退为全市场扫描")
                        elif dr_sample == "量价":
                            vp_codes = get_volprice_sectors(6)
                            if vp_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(vp_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只量价反转板块成分股")
                            else:
                                st.warning("未能获取量价反转板块数据，回退为全市场扫描")
                        
                        if '量比' in quotes_df.columns:
                            quotes_df['量比'] = pd.to_numeric(quotes_df['量比'], errors='coerce').fillna(1)
                            quotes_df = quotes_df.sort_values('量比', ascending=False)
                        scan_df = quotes_df.head(300).copy()
                        
                        codes = scan_df['代码'].tolist()
                        names = dict(zip(scan_df['代码'], scan_df['名称']))
                        
                        kline_dict = {}
                        status = st.empty()
                        bar = st.progress(0)
                        total = len(codes)
                        with ThreadPoolExecutor(max_workers=10) as ex:
                            futures = {ex.submit(get_stock_kline, c, 300): c for c in codes}
                            done = 0
                            for f in as_completed(futures):
                                done += 1
                                c = futures[f]
                                try:
                                    kline = f.result(timeout=15)
                                    if kline is not None and len(kline) >= 60:
                                        kline_dict[c] = kline
                                except:
                                    pass
                                if done % 20 == 0:
                                    bar.progress(done / total)
                                    status.text(f"📊 获取K线数据... ({done}/{total})")
                        bar.empty()
                        status.empty()
                        
                        results = []
                        scored_codes = list(kline_dict.keys())
                        _dr_diag = {"total": len(scored_codes), "fin_ok": 0, "fin_fail": 0}

                        bar2 = st.progress(0)
                        fin_status = st.empty()
                        for i, code in enumerate(scored_codes):
                            kline_df = kline_dict[code]
                            fin_data = get_financial_data(code)
                            ctx = {'fin': fin_data}
                            if fin_data and fin_data.get('success'):
                                _dr_diag["fin_ok"] += 1
                            else:
                                _dr_diag["fin_fail"] += 1
                            sr = calculate_dilemma_reversal_score(code, kline_df, stock_pool_context=ctx)
                            if sr.get('pass') and sr.get('综合评分', 0) > 0:
                                results.append({
                                    '代码': code,
                                    '名称': names.get(code, ''),
                                    '综合评分': sr['综合评分'],
                                    'L1_拐点': sr.get('L1_拐点', 0),
                                    'L2_反转': sr.get('L2_反转', 0),
                                    'L3_安全垫': sr.get('L3_安全垫', 0),
                                    'L4_技术资金': sr.get('L4_技术资金', 0),
                                })
                            if (i + 1) % 20 == 0:
                                bar2.progress((i + 1) / len(scored_codes))
                                fin_status.text(f"📊 评分中... ({i+1}/{len(scored_codes)}) | 财务数据: {_dr_diag['fin_ok']}✓ {_dr_diag['fin_fail']}✗(用技术面近似)")
                        bar2.empty()
                        fin_status.empty()
                        st.session_state._dr_diag = _dr_diag
                        
                        # P0诊断: 检查分值是否全部相同（可能暗示数据源问题）
                        if len(results) >= 2:
                            dr_scores = [r['综合评分'] for r in results]
                            if len(set(dr_scores)) == 1:
                                st.warning(f"⚠️ 困境反转: 全部{len(results)}只标的综合评分完全相同({dr_scores[0]}分)，"
                                           f"可能所有股票走同一fallback路径。财务数据可用: {_dr_diag['fin_ok']}/{_dr_diag['total']}只，"
                                           f"技术面近似: {_dr_diag['fin_fail']}只。建议检查网络/数据源。")
                        results.sort(key=lambda x: x['综合评分'], reverse=True)
                        st.session_state.dr_results = results
                        st.rerun()
                except Exception as e:
                    st.error(f"扫描出错: {e}")
        
        dr_results = st.session_state.get('dr_results', None)
        
        if dr_results is not None:
            if len(dr_results) == 0:
                st.warning("⚠️ 今日未找到符合条件的困境反转标的。")
            else:
                top_n = min(len(dr_results), 30)
                st.markdown(f"### 📊 困境反转模型 · Top {top_n}")
                st.caption(f"共筛选出 {len(dr_results)} 只标的")
                _dr_diag = st.session_state.get('_dr_diag', {})
                if _dr_diag:
                    _fin_ok = _dr_diag.get('fin_ok', 0)
                    _fin_fail = _dr_diag.get('fin_fail', 0)
                    _tot = _dr_diag.get('total', _fin_ok + _fin_fail)
                    if _fin_fail > 0:
                        st.info(f"📋 诊断: 共{_tot}只 | 真实财务数据{_fin_ok}只 | 技术面近似{_fin_fail}只（L1拐点/L2反转/L3安全垫用K线指标替代）")
                
                _dr_dyn_n = calculate_dynamic_recommend_count()
                dr_top_n = dr_results[:_dr_dyn_n]
                st.markdown(f"""
                <div class="top10-container">
                    <div class="top10-header">
                        <div class="top10-title">🔄 困境反转 精选 Top {_dr_dyn_n} <span class="top10-badge">周期反转信号</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)
                btn1, btn2, _ = st.columns([1, 1, 4])
                with btn1:
                    if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="dr_add_all"):
                        for s in dr_top_n:
                            if s["代码"] not in st.session_state.watchlist:
                                st.session_state.watchlist.append(s["代码"])
                        save_watchlist(st.session_state.watchlist)
                        st.success(f"已将Top {_dr_dyn_n}全部加入自选！"); st.rerun()
                with btn2:
                    top_n_df = pd.DataFrame(dr_top_n)
                    top_n_df['代码'] = top_n_df['代码'].astype(str).str.zfill(6)
                    export_cols = ['代码', '名称', '综合评分',
                                   'L1_拐点', 'L2_反转', 'L3_安全垫', 'L4_技术资金']
                    top_n_df = top_n_df[[c for c in export_cols if c in top_n_df.columns]]
                    xlsx_data = _export_df_to_xlsx(top_n_df)
                    st.download_button(f"📥 导出Top{_dr_dyn_n}", xlsx_data,
                        f"top{_dr_dyn_n}_dilemma_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width='stretch', key="dl_dr_t10")
                
                st.markdown("")
                for i in range(0, len(dr_top_n), 5):
                    row_stocks = dr_top_n[i:i+5]
                    cols = st.columns(5)
                    for j, stock in enumerate(row_stocks):
                        with cols[j]:
                            rank = i + j + 1
                            rank_color = "#FFD700" if rank == 1 else "#C0C0C0" if rank == 2 else "#CD7F32" if rank == 3 else "#CCC"
                            st.markdown(f"""
                            <div class="top10-card">
                                <div class="top10-rank" style="color:{rank_color};">{rank}</div>
                                <div style="padding-right:30px;">
                                    <div style="font-weight:700;color:#333;font-size:15px;margin-bottom:2px;">{stock['名称']}</div>
                                    <div style="font-size:11px;color:#888;margin-bottom:8px;">{str(stock.get('代码','')).zfill(6)}</div>
                                    <div style="display:flex;justify-content:space-between;align-items:center;">
                                        <span style="font-size:20px;font-weight:800;color:#C4842D;">{stock['综合评分']}</span>
                                    </div>
                                    <div style="margin-top:6px;"><span class="metric-badge badge-strong">困境反转</span></div>
                                </div>
                            </div>""", unsafe_allow_html=True)
                            code = stock["代码"]; in_wl = code in st.session_state.watchlist
                            if st.button("⭐" if in_wl else "+自选", key=f"dr_t10_{code}",
                                width='stretch', type="primary" if in_wl else "secondary"):
                                if in_wl: st.session_state.watchlist.remove(code)
                                else: st.session_state.watchlist.append(code)
                                save_watchlist(st.session_state.watchlist); st.rerun()
                
                st.markdown("---")
                fdf = pd.DataFrame(dr_results[:top_n])
                fdf.index = range(1, len(fdf) + 1)
                fdf['代码'] = fdf['代码'].astype(str).str.zfill(6)
                display_cols = ['代码', '名称', '综合评分',
                                'L1_拐点', 'L2_反转', 'L3_安全垫', 'L4_技术资金']
                df_display = fdf[[c for c in display_cols if c in fdf.columns]]
                st.dataframe(df_display, width='stretch')
                
                st.markdown("---")
                st.caption("操作区 — 加/取消自选")
                dr_buttons = [{'代码': r['代码'], '名称': r['名称']} for r in dr_results[:30]]
                render_stock_buttons(dr_buttons, prefix="dr")
                
                export_df = pd.DataFrame(dr_results)
                export_df['代码'] = export_df['代码'].astype(str).str.zfill(6)
                export_cols = ['代码', '名称', '综合评分',
                               'L1_拐点', 'L2_反转', 'L3_安全垫', 'L4_技术资金']
                export_df = export_df[[c for c in export_cols if c in export_df.columns]]
                csv_data = export_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 导出 Top30 CSV", csv_data,
                    f"dilemma_reversal_top30_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv", key="dl_dr_top30")
        elif not do_scan:
            st.info("💡 点击「全市场扫描」启动困境反转模型选股（基于四层评分）。")

    with model_tabs[6]:  # 超跌反弹模型
        st.session_state.current_model = 'oversold_rebound'

        render_top_params_panel()

        st.markdown("""<div class="header-container"><div class="main-title">💎 超跌反弹模型（抓超跌股）</div>
        <div class="sub-title">四维评分：空间40 · 情绪量能30 · 择时确认30 · 板块共振+10</div></div>""", unsafe_allow_html=True)

        sc1, sc2 = st.columns([1.8, 1])
        with sc1:
            cur_source = st.session_state.get('orb_sample_source', '全市场A股')
            idx = 0 if cur_source == '全市场A股' else (1 if cur_source == '热门板块' else 2)
            new_source = st.selectbox(
                "样本来源", DEFAULT_GC_SAMPLE_OPTIONS,
                index=idx,
                key="orb_sample_source_select"
            )
            if "全市场" in new_source:
                new_val = "全市场A股"
            elif "量价" in new_source:
                new_val = "量价"
            else:
                new_val = "热门板块"
            if new_val != cur_source:
                st.session_state.orb_results = None
                st.session_state.orb_sample_source = new_val
                st.rerun()
            else:
                st.session_state.orb_sample_source = new_val

        col1, col2 = st.columns([2, 1])
        orb_sample = st.session_state.get('orb_sample_source', '全市场A股')
        if orb_sample == "热门板块":
            scan_label = "🔍 加速板块扫描"
        elif orb_sample == "量价":
            scan_label = "🔍 量价扫描"
        else:
            scan_label = "🔍 全市场扫描"
        with col1:
            do_scan = st.button(scan_label, type="primary", width='stretch', key="orb_scan")
        with col2:
            if st.button("🔄 清空缓存", width='stretch', key="orb_clear"):
                st.session_state.orb_results = None
                st.rerun()

        orb_results = st.session_state.get('orb_results', None)

        if do_scan or orb_results is None:
            if orb_sample == "热门板块":
                spinner_text = "正在扫描资金加速板块，计算超跌反弹评分…"
            elif orb_sample == "量价":
                spinner_text = "正在扫描量价反转板块，计算超跌反弹评分…"
            else:
                spinner_text = "正在扫描全市场，计算超跌反弹评分…"
            with st.spinner(spinner_text):
                try:
                    quotes_df = fetch_all_a_stocks()
                    if quotes_df is None or len(quotes_df) == 0:
                        st.error("获取行情数据失败")
                    else:
                        quotes_df = quotes_df[~quotes_df['名称'].str.contains('ST|退市|N|C', na=False)]

                        if orb_sample == "热门板块":
                            hot_codes = get_hot_concept_stocks(6)
                            if hot_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(hot_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只热门概念板块成分股")
                            else:
                                st.warning("未能获取热门板块数据，回退为全市场扫描")
                        elif orb_sample == "量价":
                            vp_codes = get_volprice_sectors(6)
                            if vp_codes:
                                quotes_df['代码'] = quotes_df['代码'].astype(str).str.zfill(6)
                                quotes_df = quotes_df[quotes_df['代码'].isin(vp_codes)]
                                st.info(f"已锁定 {len(quotes_df)} 只量价反转板块成分股")
                            else:
                                st.warning("未能获取量价反转板块数据，回退为全市场扫描")

                        # 超跌模型：优先找跌幅最大的股票（而非放量股）
                        if '涨跌幅' in quotes_df.columns:
                            quotes_df['涨跌幅'] = pd.to_numeric(quotes_df['涨跌幅'], errors='coerce').fillna(0)
                            quotes_df = quotes_df.sort_values('涨跌幅', ascending=True)
                        scan_df = quotes_df.head(500).copy()  # 扩大样本到500只

                        codes = scan_df['代码'].tolist()
                        names = dict(zip(scan_df['代码'], scan_df['名称']))

                        # 获取板块映射，用于板块共振加成
                        _, stock_sector_map = _get_cached_sector_data()

                        kline_dict = {}
                        status = st.empty()
                        bar = st.progress(0)
                        total = len(codes)
                        with ThreadPoolExecutor(max_workers=10) as ex:
                            futures = {ex.submit(get_stock_kline, c, 120): c for c in codes}
                            done = 0
                            for f in as_completed(futures):
                                done += 1
                                c = futures[f]
                                try:
                                    kline = f.result(timeout=15)
                                    if kline is not None and len(kline) >= 20:
                                        kline_dict[c] = kline
                                except:
                                    pass
                                if done % 20 == 0:
                                    bar.progress(done / total)
                                    status.text(f"📊 获取K线数据... ({done}/{total})")
                        bar.empty()
                        status.empty()

                        results = []
                        scored_codes = list(kline_dict.keys())
                        diag = {"total_kline": len(scored_codes), "hard_ok": 0, "score_ok": 0}
                        filter_reasons = {}

                        bar2 = st.progress(0)
                        for i, code in enumerate(scored_codes):
                            kline_df = kline_dict[code]
                            # 硬过滤：ST/跌幅不足/跌停/放量下跌/仙股
                            ok, reason = hard_filter_oversold_rebound(kline_df, None)
                            if not ok:
                                reason_key = reason.split("(")[0].strip() if "(" in reason else reason[:20]
                                filter_reasons[reason_key] = filter_reasons.get(reason_key, 0) + 1
                                continue
                            diag['hard_ok'] += 1
                            sector_name = stock_sector_map.get(code, "")
                            sr = calculate_oversold_rebound_score(
                                kline_df, stock_data={"sector": sector_name}
                            )
                            if sr.get('pass') and sr.get('综合评分', 0) > 0:
                                diag['score_ok'] += 1
                                results.append({
                                    '代码': code,
                                    '名称': names.get(code, ''),
                                    '综合评分': sr['综合评分'],
                                    '空间维度': sr.get('空间维度', 0),
                                    '情绪量能': sr.get('情绪量能', 0),
                                    '择时确认': sr.get('择时确认', 0),
                                    '板块共振': sr.get('板块共振', 0),
                                })
                            if (i + 1) % 50 == 0:
                                bar2.progress((i + 1) / len(scored_codes))
                        bar2.empty()

                        results.sort(key=lambda x: x['综合评分'], reverse=True)
                        st.session_state.orb_results = results

                        # 诊断信息
                        with st.expander("📊 扫描诊断详情", expanded=(len(results) == 0)):
                            diag_parts = [f"K线数据：{diag['total_kline']}只"]
                            diag_parts.append(f"硬滤通过：{diag['hard_ok']}只 ({diag['hard_ok']/max(diag['total_kline'],1)*100:.0f}%)")
                            diag_parts.append(f"评分通过：{diag['score_ok']}只 ({diag['score_ok']/max(diag['hard_ok'],1)*100:.0f}%)" if diag['hard_ok'] > 0 else "评分通过：0只")
                            st.caption(" · ".join(diag_parts))
                            if filter_reasons:
                                reason_items = sorted(filter_reasons.items(), key=lambda x: -x[1])
                                reason_text = " | ".join([f"{r}: {n}只" for r, n in reason_items[:8]])
                                st.caption(f"🔍 淘汰原因：{reason_text}")
                                st.info("💡 提示：硬滤要求近60日高点回调>15%（且近20日跌幅>5%+连跌≥3天例外放行），如结果太少可等待市场调整期。")

                        st.rerun()
                except Exception as e:
                    st.error(f"扫描出错: {e}")

        orb_results = st.session_state.get('orb_results', None)

        if orb_results is not None:
            if len(orb_results) == 0:
                st.warning("⚠️ 今日未找到符合条件的超跌反弹标的。")
            else:
                top_n = min(len(orb_results), 30)
                st.markdown(f"### 📊 超跌反弹模型 · Top {top_n}")
                st.caption(f"共筛选出 {len(orb_results)} 只标的")

                _orb_dyn_n = calculate_dynamic_recommend_count()
                orb_top_n = orb_results[:_orb_dyn_n]
                st.markdown(f"""
                <div class="top10-container">
                    <div class="top10-header">
                        <div class="top10-title">💎 超跌反弹 精选 Top {_orb_dyn_n} <span class="top10-badge">超跌反弹信号</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)
                btn1, btn2, _ = st.columns([1, 1, 4])
                with btn1:
                    if st.button("⭐ 一键加入自选", width='stretch', type="primary", key="orb_add_all"):
                        for s in orb_top_n:
                            if s["代码"] not in st.session_state.watchlist:
                                st.session_state.watchlist.append(s["代码"])
                        save_watchlist(st.session_state.watchlist)
                        st.success(f"已将Top {_orb_dyn_n}全部加入自选！"); st.rerun()
                with btn2:
                    top_n_df = pd.DataFrame(orb_top_n)
                    top_n_df['代码'] = top_n_df['代码'].astype(str).str.zfill(6)
                    export_cols = ['代码', '名称', '综合评分',
                                   '空间维度', '情绪量能', '择时确认', '板块共振']
                    top_n_df = top_n_df[[c for c in export_cols if c in top_n_df.columns]]
                    xlsx_data = _export_df_to_xlsx(top_n_df)
                    st.download_button(f"📥 导出Top{_orb_dyn_n}", xlsx_data,
                        f"top{_orb_dyn_n}_oversold_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width='stretch', key="dl_orb_t10")

                st.markdown("")
                for i in range(0, len(orb_top_n), 5):
                    row_stocks = orb_top_n[i:i+5]
                    cols = st.columns(5)
                    for j, stock in enumerate(row_stocks):
                        with cols[j]:
                            rank = i + j + 1
                            rank_color = "#FFD700" if rank == 1 else "#C0C0C0" if rank == 2 else "#CD7F32" if rank == 3 else "#CCC"
                            # 维度条形
                            dim_max = {"空间维度": 40, "情绪量能": 30, "择时确认": 30, "板块共振": 10}
                            dim_bars = ""
                            for dim_key, dim_label in [("空间维度","空间"), ("情绪量能","情绪"), ("择时确认","择时"), ("板块共振","板块")]:
                                dv = stock.get(dim_key, 0)
                                dm = max(dim_max.get(dim_key, 40), 1)
                                pct = min(dv / dm * 100, 100)
                                dcolor = "#E74C3C" if dim_key == "空间维度" else ("#3498DB" if dim_key == "情绪量能" else ("#F39C12" if dim_key == "择时确认" else "#27AE60"))
                                dim_bars += f'<div style="display:flex;align-items:center;margin:2px 0;font-size:10px;color:#888;"><span style="width:28px;">{dim_label}</span><div style="flex:1;height:6px;background:#EEE;border-radius:3px;margin:0 6px;"><div style="width:{pct}%;height:100%;background:{dcolor};border-radius:3px;"></div></div><span style="width:20px;text-align:right;">{dv}</span></div>'
                            st.markdown(f"""
                            <div class="top10-card">
                                <div class="top10-rank" style="color:{rank_color};">{rank}</div>
                                <div style="padding-right:30px;">
                                    <div style="font-weight:700;color:#333;font-size:15px;margin-bottom:2px;">{stock['名称']}</div>
                                    <div style="font-size:11px;color:#888;margin-bottom:8px;">{str(stock.get('代码','')).zfill(6)}</div>
                                    <div style="display:flex;justify-content:space-between;align-items:center;">
                                        <span style="font-size:20px;font-weight:800;color:#9B59B6;">{stock['综合评分']}</span>
                                    </div>
                                    {dim_bars}
                                    <div style="margin-top:6px;"><span class="metric-badge badge-strong">超跌反弹</span></div>
                                </div>
                            </div>""", unsafe_allow_html=True)
                            code = stock["代码"]; in_wl = code in st.session_state.watchlist
                            if st.button("⭐" if in_wl else "+自选", key=f"orb_t10_{code}",
                                width='stretch', type="primary" if in_wl else "secondary"):
                                if in_wl: st.session_state.watchlist.remove(code)
                                else: st.session_state.watchlist.append(code)
                                save_watchlist(st.session_state.watchlist); st.rerun()

                st.markdown("---")
                fdf = pd.DataFrame(orb_results[:top_n])
                fdf.index = range(1, len(fdf) + 1)
                fdf['代码'] = fdf['代码'].astype(str).str.zfill(6)
                display_cols = ['代码', '名称', '综合评分',
                                '空间维度', '情绪量能', '择时确认', '板块共振']
                df_display = fdf[[c for c in display_cols if c in fdf.columns]]
                st.dataframe(df_display, width='stretch')

                st.markdown("---")
                st.caption("操作区 — 加/取消自选")
                orb_buttons = [{'代码': r['代码'], '名称': r['名称']} for r in orb_results[:30]]
                render_stock_buttons(orb_buttons, prefix="orb")

                export_df = pd.DataFrame(orb_results)
                export_df['代码'] = export_df['代码'].astype(str).str.zfill(6)
                export_cols = ['代码', '名称', '综合评分',
                               '空间维度', '情绪量能', '择时确认', '板块共振']
                export_df = export_df[[c for c in export_cols if c in export_df.columns]]
                csv_data = export_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 导出 Top30 CSV", csv_data,
                    f"oversold_rebound_top30_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv", key="dl_orb_top30")
        elif not do_scan:
            st.info("💡 点击「全市场扫描」启动超跌反弹模型选股（基于四维评分）。")


# ================================================================
#                  UI: 个股详情页
# ================================================================

def render_detail(code):
    d = get_stock_detail(code)
    if not d:
        st.error("未找到数据"); return
    
    cb, ct = st.columns([1, 5])
    with cb:
        if st.button("← 返回列表", width='stretch', key="detail_back"):
            st.session_state.selected_stock = None; st.rerun()
    with ct:
        sig_color = {"强势买入":"#E74C3C","逢低吸纳":"#E67E22","观望等待":"#27AE60","建议回避":"#888"}.get(d.get('信号',''), "#666")
        st.markdown(f"""<div class="header-container" style="margin-top:0;">
        <div class="main-title">{d['名称']}
        <span style="font-size:14px;font-weight:400;color:#888;">{d['代码']} · 综合<b style='color:#C4842D;'>{d.get('综合评分','N/A')}</b>分 · {d.get('板块','')}</span>
        <span class="metric-badge" style="background:{sig_color};color:white;margin-left:10px;">{d.get('信号','')}</span>
        </div></div>""", unsafe_allow_html=True)

    mc = st.columns(3)
    met = [
        ("5日涨幅", d.get("5日涨幅","N/A"), "up" if "+" in str(d.get("5日涨幅","")) else "down"),
        ("20日涨幅", d.get("20日涨幅","N/A"), "up" if "+" in str(d.get("20日涨幅","")) else "down"),
        ("量比", d.get("量比_显示","N/A"), "neutral"),
        ("换手率", d.get("换手率_显示","N/A"), "neutral"),
        ("RSI-14", d.get("RSI_显示","N/A"), "neutral"),
        ("PE(TTM)", d.get("PE_显示","N/A"), "neutral"),
    ]
    for i, (nm, vl, mood) in enumerate(met):
        with mc[i%3]:
            cls = "up" if mood == "up" else "down" if mood == "down" else ""
            st.markdown(f"""<div class="detail-metric-card"><div class="metric-name">{nm}</div><div class="metric-value {cls}">{vl}</div></div>""", unsafe_allow_html=True)

    dims = ["趋势结构","动量强度","板块共振","北向资金","机构净买","板块资金热度","量价配合","估值安全","筹码稳定","情绪热度"]
    clrs = [WEIGHT_CONFIG.get(d, {}).get("color", "#666") for d in dims]
    
    st.markdown("<hr style='margin:20px 0;border:none;border-top:1px solid #EEE;'>", unsafe_allow_html=True)
    st.markdown("**📊 十维评分详情(v4)**")
    _dc1 = st.columns(5)
    for co, dm, cr in zip(_dc1, dims[:5], clrs[:5]):
        with co:
            sc = d.get(dm, 50); wt = st.session_state.weights.get(dm, 20)
            st.markdown(f"""<div class="detail-metric-card" style="text-align:center;">
            <div class="metric-name">{dm}</div>
            <div style="font-size:26px;font-weight:800;color:{cr};margin:4px 0;">{sc}<span style="font-size:14px;">分</span></div>
            <div style="font-size:11px;color:#999;">权重 {wt}%</div>
            <div style="margin-top:8px;"><div class="score-bar"><div class="score-fill" style="width:{sc}%;background:{cr};"></div></div></div>
            </div>""", unsafe_allow_html=True)
    _dc2 = st.columns(5)
    for co, dm, cr in zip(_dc2, dims[5:], clrs[5:]):
        with co:
            sc = d.get(dm, 50); wt = st.session_state.weights.get(dm, 20)
            st.markdown(f"""<div class="detail-metric-card" style="text-align:center;">
            <div class="metric-name">{dm}</div>
            <div style="font-size:26px;font-weight:800;color:{cr};margin:4px 0;">{sc}<span style="font-size:14px;">分</span></div>
            <div style="font-size:11px;color:#999;">权重 {wt}%</div>
            <div style="margin-top:8px;"><div class="score-bar"><div class="score-fill" style="width:{sc}%;background:{cr};"></div></div></div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<hr style='margin:20px 0;border:none;border-top:1px solid #EEE;'>", unsafe_allow_html=True)
    sv = d.get("综合评分", 50)
    if sv >= 85: lv, ds = ("💰 核心龙头·重仓追高", f"综合评分 {sv} 分（v4十维），属于核心主线龙头。信号「{d.get('信号','')}」，建议重仓60%~80%。趋势 {d.get('趋势结构','N/A')}，北向 {d.get('北向资金','N/A')}，机构 {d.get('机构净买','N/A')}，板块 {d.get('板块共振','N/A')}。")
    elif sv >= 70: lv, ds = ("📈 支线趋势·轻仓试错", f"综合评分 {sv} 分（v4十维），属于支线趋势标的。信号「{d.get('信号','')}」，建议轻仓20%~40%。趋势 {d.get('趋势结构','N/A')}。")
    elif sv >= 55: lv, ds = ("⚠️ 建议观望", f"综合评分 {sv} 分（v4十维），低于70分追高门槛，建议观望。")
    else: lv, ds = ("🚫 不建议参与", f"综合评分 {sv} 分（v3八维），分数不足70，放弃不参与。")
    
    st.markdown(f"""<div class="advice-box"><div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;"><span>⚡</span><strong>{lv}</strong></div>
    <div style="font-size:14px;color:#555;line-height:1.7;">{ds}</div></div>""", unsafe_allow_html=True)

    # === K-line chart ===
    st.markdown("**K-Line**")
    _kline_df, _kline_rt = get_kline_with_today(d.get('代码', ''), days=60)
    if _kline_df is not None and len(_kline_df) > 5:
        _kline_title = f"{d.get('名称','')} ({d.get('代码','')})"
        if _kline_rt:
            _kline_title += " [盘中]"
        _kline_fig = create_candlestick_chart(_kline_df, title=_kline_title, is_realtime=_kline_rt)
        if _kline_fig:
            st.plotly_chart(_kline_fig, width='stretch', config={'displayModeBar': False})
        if _kline_rt:
            st.caption("▶ 最右侧蜡烛为今日盘中实时数据")
    else:
        st.info("K-line loading...")

    st.markdown("**📡 八维雷达图**")
    cc, cf = st.columns([3, 2])
    with cc:
        dimd = {dd: d.get(dd, 50) for dd in dims}
        radar_fig = create_radar_chart(dimd, st.session_state.weights)
        st.plotly_chart(radar_fig, width='stretch', config={'displayModeBar': False})
    with cf:
        st.markdown("**⚖️ 快速调权**")
        for dm, cr in zip(dims, clrs):
            cw = st.session_state.weights.get(dm, 20)
            mx = WEIGHT_CONFIG[dm]["max"]
            # 处理 +/- 按钮
            if st.session_state.get(f"_dinc_{dm}", False):
                st.session_state.weights[dm] = min(mx, cw + 1)
                st.session_state[f"_dinc_{dm}"] = False
                st.rerun()
            if st.session_state.get(f"_ddec_{dm}", False):
                st.session_state.weights[dm] = max(0, cw - 1)
                st.session_state[f"_ddec_{dm}"] = False
                st.rerun()
            sc1, sc2, sc3, sc4 = st.columns([0.5, 2, 0.8, 0.8])
            with sc1:
                st.markdown(f"<span style='color:{cr};font-size:13px;'>●</span>", unsafe_allow_html=True)
            with sc2:
                st.markdown(f"<span style='font-size:12px;color:#444;'>{dm}</span> <b style='color:{cr};font-size:14px;'>{cw}%</b>", unsafe_allow_html=True)
            with sc3:
                if st.button("−", key=f"ddec_{dm}"):
                    st.session_state[f"_ddec_{dm}"] = True
                    st.rerun()
            with sc4:
                if st.button("+", key=f"dinc_{dm}"):
                    st.session_state[f"_dinc_{dm}"] = True
                    st.rerun()
        if st.button("🔄 重新计算评分", width='stretch', type="primary", key="detail_recalc"):
            st.cache_data.clear(); st.rerun()


# ================================================================
#                   UI: 自选股页面
# ================================================================

def render_watchlist():
    st.markdown("""<div class="header-container">
    <div class="main-title">⭐ 我的自选</div>
    <div class="sub-title">已自动保存至本地 · 点击股票查看详情与八维分析</div>
    </div>""", unsafe_allow_html=True)
    
    wl = st.session_state.watchlist
    if not wl:
        st.info("📭 自选股为空")
        return
    
    df = get_stock_pool()
    wl_df = df[df["代码"].isin(wl)].copy()
    if len(wl_df) == 0:
        st.warning("自选股数据加载中，请稍后刷新..."); return
    
    render_stats_chase(wl_df)
    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button("🗑️ 清空自选", width='stretch', key="wl_clear_all"):
            st.session_state.watchlist = []; save_watchlist([]); st.rerun()
    with c2:
        if st.button("📤 导出自选股", width='stretch', key="wl_export_btn"):
            csv = wl_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("下载CSV", csv, "my_watchlist.csv", "text/csv", key="dl_wl")
    
    st.markdown("")
    sec, sig_filter, sw = render_filter_bar(wl_df)
    if sec != "全部": wl_df = wl_df[wl_df["板块"] == sec]
    if sig_filter is not None and sig_filter[0] == "信号": wl_df = wl_df[wl_df["信号"].isin(sig_filter[1])]
    if sw:
        m = wl_df["代码"].str.contains(sw, False) | wl_df["名称"].str.contains(sw, False)
        wl_df = wl_df[m]
    render_table_chase(wl_df.reset_index(drop=True))


# ================================================================
#                 UI: 策略回测页面
# ================================================================

def render_backtest():
    # 模型选择 — 放大字体和图标（页面中仅此一个 radio，用 role 选择器精准命中）
    st.markdown("""
    <style>
    [role="radiogroup"] label[data-baseweb="radio"] {
        font-size: 22px !important;
        padding: 8px 20px !important;
    }
    [role="radiogroup"] label[data-baseweb="radio"] > div:first-child {
        transform: scale(1.6);
        margin-right: 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)
    model_options = {"resonance": "🎯 共振模型", "chase_high": "🚀 追高模型", "buy_low": "📉 低吸模型", "golden_cross": "📌 金叉模型", "canslim": "📊 CAN SLIM模型（抓主升浪股）", "dilemma_reversal": "🔄 困境反转模型（抓周期股）"}
    st.markdown('<p style="font-size:20px;font-weight:600;margin:0 0 4px 0;">回测模型</p>', unsafe_allow_html=True)
    selected_model = st.radio(" ", list(model_options.values()), horizontal=True, key="bt_model", label_visibility="collapsed")
    # Key映射表（emoji+中文 -> 内部key）
    _model_map = {
        "🎯 共振模型": "resonance",
        "🚀 追高模型": "chase_high",
        "📉 低吸模型": "buy_low",
        "📌 金叉模型": "golden_cross",
        "📊 CAN SLIM模型（抓主升浪股）": "canslim",
        "🔄 困境反转模型（抓周期股）": "dilemma_reversal",
    }
    model_key = _model_map.get(selected_model, "chase_high")

    if model_key == "chase_high":
        model_subtitle = "基于当前八维权重配置(v3)进行历史回测 · 含沪深300基准对比"
    elif model_key == "buy_low":
        model_subtitle = "基于七维低吸评分进行历史回测 · 含沪深300基准对比"
    elif model_key == "resonance":
        model_subtitle = "基于四维共振评分（K线结构+板块热度）进行历史回测 · 含沪深300基准对比"
    elif model_key == "canslim":
        model_subtitle = "基于CAN SLIM模型（抓主升浪股）七因子简化版（技术因子完整，财务因子占位）进行历史回测 · 含沪深300基准对比"
    elif model_key == "dilemma_reversal":
        model_subtitle = "基于困境反转模型（抓周期股）四层评分（L4技术资金完整，L1-L3占位）进行历史回测 · 含沪深300基准对比"
    else:
        model_subtitle = "基于五维金叉评分（下跌形态·K线止跌·均线拐头·量能确认·MACD反转）进行历史回测 · 含沪深300基准对比"

    st.markdown(f"""<div class="header-container"><div class="main-title">📈 策略回测中心</div>
    <div class="sub-title">{model_subtitle}</div></div>""", unsafe_allow_html=True)
    
    with st.expander("⚙️ 回测参数配置", expanded=True):
        pc1, pc2, pc3, pc4 = st.columns(4)
        with pc1: sd = st.date_input("起始日期", value=pd.Timestamp("2026-06-01").date(), key="bt_start")
        with pc2: ed = st.date_input("结束日期", value=datetime.now().date(), key="bt_end")
        with pc3: top_n = st.selectbox("持仓数量", [3,5,10,15,20], index=2, key="bt_topn")
        with pc4: min_sc = st.selectbox("最低评分", [50,55,60,65,70], index=2, key="bt_minsc")
    with st.expander("📉 五层卖出参数", expanded=False):
        sp1, sp2 = st.columns(2)
        with sp1:
            bt_stop_loss = st.slider("止损比例 (%)", min_value=1, max_value=15, value=5, step=1, key="bt_stop_loss",
                help="持仓跌幅超过此比例立即止损卖出")
            bt_hard_tp = st.slider("硬止盈低档 (%)", min_value=5, max_value=25, value=12, step=1, key="bt_hard_tp",
                help="盈利达标且共振<50时止盈卖出")
            bt_hard_tp_ceiling = st.slider("硬止盈天花板 (%)", min_value=15, max_value=70, value=35, step=5, key="bt_hard_tp_ceiling",
                help="盈利超过此比例无条件卖出")
            bt_moving_gain = st.slider("移动止盈激活线 (%)", min_value=3, max_value=25, value=15, step=1, key="bt_moving_gain",
                help="峰值盈利超过此比例后启动移动止盈监控")
        with sp2:
            bt_moving_dd = st.slider("移动止盈回撤线 (%)", min_value=1, max_value=12, value=7, step=1, key="bt_moving_dd",
                help="从峰值回撤超过此比例触发移动止盈")
            bt_cooling = st.slider("冷却期 (天)", min_value=1, max_value=10, value=5, step=1, key="bt_cooling",
                help="持仓前N天共振不参与卖出判断")
            bt_resonance = st.slider("共振卖出阈值", min_value=20, max_value=60, value=40, step=5, key="bt_resonance",
                help="冷却期后共振评分低于此值触发卖出")
            bt_min_score_pct = st.slider("得分门槛 (Z-Score)", min_value=0.0, max_value=3.0, value=1.0, step=0.1, key="bt_min_score_pct",
                help="只买入得分超过当日均值N个标准差的股票，0=不过滤。行情差时自动空仓")
    pcc1, _ = st.columns([3, 1])
    with pcc1: rebal = st.selectbox("调仓周期", ["每周(5交易日)","每两周(10交易日)","每月(20交易日)"], index=0, key="bt_rebal")

    st.markdown(f"""<div style="background:#FFFBF5;border:1px solid #E8DCC8;border-radius:10px;padding:14px 20px;margin:12px 0;">
    <span style="font-weight:700;color:#C4842D;">📋 当前权重配置：</span>""", unsafe_allow_html=True)
    ws = []; tw = 0
    if model_key == "chase_high" or model_key in ("canslim", "dilemma_reversal"):
        for dk, dv in WEIGHT_CONFIG.items():
            wv = st.session_state.weights.get(dk, dv["default"]); tw += wv
            ws.append(f"<span style='color:{dv['color']}'><b>{dk}</b> {wv}%</span>")
    elif model_key == "buy_low":
        for dk, dv in LOWBUY_WEIGHT_CONFIG.items():
            wv = st.session_state.get("lowbuy_weights", DEFAULT_LOWBUY_WEIGHTS).get(dk, dv["default"]); tw += wv
            ws.append(f"<span style='color:{dv['color']}'><b>{dk}</b> {wv}%</span>")
    elif model_key == "resonance":
        ws.append(f"<span style='color:#00897B'><b>资金流向</b> 30%</span>")
        ws.append(f"<span style='color:#1565C0'><b>DDE决策</b> 20%</span>")
        ws.append(f"<span style='color:#E65100'><b>K线结构</b> 25%</span>")
        ws.append(f"<span style='color:#6A1B9A'><b>板块热度</b> 25%</span>")
        tw = 100
    else:  # golden_cross
        for dk, dv in DEFAULT_GC_WEIGHTS.items():
            wv = st.session_state.get("gc_weights", dict(DEFAULT_GC_WEIGHTS)).get(dk, dv)
            tw += wv
            ws.append(f"<span style='color:#C4842D'><b>{dk}</b> {wv}%</span>")
    st.markdown("  ".join(ws) + f"&nbsp;&nbsp;|&nbsp;&nbsp;<b>合计:{tw}%</b></div>", unsafe_allow_html=True)
    
    btn1, _, btn3 = st.columns([1, 2, 1])
    with btn1: run_bt = st.button("▶️ 开始回测", width='stretch', type="primary", key="bt_run")
    with btn3: exp_bt = st.button("💾 导出报告", width='stretch', key="bt_export")

    if run_bt:
        with st.spinner("正在运行回测... ⏳"):
            try:
                bt_start = time.time()
                bt_df = run_real_backtest_cached(str(sd), str(ed), top_n=top_n, model=model_key,
                    stop_loss=st.session_state.get('bt_stop_loss', 5) / 100.0,
                    hard_take_profit=st.session_state.get('bt_hard_tp', 12) / 100.0,
                    hard_take_profit_ceiling=st.session_state.get('bt_hard_tp_ceiling', 35) / 100.0,
                    moving_stop_gain=st.session_state.get('bt_moving_gain', 15) / 100.0,
                    moving_stop_drawdown=st.session_state.get('bt_moving_dd', 7) / 100.0,
                    cooling_period=st.session_state.get('bt_cooling', 5),
                    resonance_threshold=st.session_state.get('bt_resonance', 40),
                    min_z_score=st.session_state.get('bt_min_score_pct', 1.0))
                bt_elapsed = time.time() - bt_start
                if bt_df is not None:
                    summary = bt_df.attrs.get('summary', '')
                    metrics = calc_backtest_metrics(bt_df)
                    # 格式化耗时
                    h = int(bt_elapsed // 3600)
                    m = int((bt_elapsed % 3600) // 60)
                    s = bt_elapsed % 60
                    if h > 0:
                        elapsed_str = f"{h}小时{m}分{int(s)}秒"
                    elif m > 0:
                        elapsed_str = f"{m}分{int(s)}秒"
                    else:
                        elapsed_str = f"{s:.1f}秒"
                    st.session_state.backtest_result = (bt_df, metrics, elapsed_str)
                    extra = bt_df.attrs.get('summary', '')
                    st.success(f"✅ 回测完成！共 {len(bt_df)} 个交易日 · 耗时 {elapsed_str} | {extra}")
            except Exception as e:
                st.error(f"❌ 回测失败: {e}")

    if st.session_state.get('backtest_result') and st.session_state.backtest_result[0] is not None:
        bt_result = st.session_state.backtest_result
        bt_df, metrics = bt_result[0], bt_result[1]
        elapsed_str = bt_result[2] if len(bt_result) > 2 else ""
        if metrics and len(bt_df) > 0:
            st.subheader("📊 核心绩效指标")
            # 时间标签
            if '日期' in bt_df.columns:
                bt_start = bt_df['日期'].min()
                bt_end = bt_df['日期'].max()
                extra_tag = f" · ⏱️ 耗时 {elapsed_str}" if elapsed_str else ""
                st.markdown(f"""<div style="text-align:center;margin-bottom:18px;">
                <span style="background:#F5F0EA;color:#8B7355;padding:6px 22px;border-radius:20px;font-size:13px;font-weight:600;">
                📅 {bt_start.strftime('%Y-%m-%d')} ~ {bt_end.strftime('%Y-%m-%d')} · {len(bt_df)}个交易日{extra_tag}
                </span></div>""", unsafe_allow_html=True)
            
            last_row = bt_df.iloc[-1]
            _sp_df = bt_df.attrs.get("stock_profit_list")
            
            # 第一行：投资回报率 — 放大醒目（ROI = 总盈利 / 峰值资金投入）
            total_profit = _sp_df[_sp_df['代码'].notna() & (_sp_df['代码'] != '合计')]['单笔收益'].sum() if _sp_df is not None and len(_sp_df) > 0 else 0.0
            max_deployed_val = bt_df.attrs.get("max_deployed", bt_df.attrs.get("avg_deployed", 1))
            roi_val_pct = (total_profit / max_deployed_val * 100) if max_deployed_val > 0 else 0
            roi_val = f"{roi_val_pct:+.2f}%"
            roi_color = "positive" if roi_val_pct >= 0 else "negative"
            st.markdown(f"""<div style="text-align:center;margin-bottom:16px;">
            <div class="bt-stat-card" style="display:inline-block;padding:20px 40px;">
            <div class="bt-label" style="font-size:16px;">投资回报率 (ROI)</div>
            <div class="bt-val {roi_color}" style="font-size:48px;line-height:1.2;">{roi_val}</div>
            <div style="font-size:12px;color:gray;margin-top:4px;">ROI = 总盈利 ÷ 峰值资金投入</div>
            </div></div>""", unsafe_allow_html=True)
            
            # 第二行：胜率 | 超额收益 | 月均收益 | 年化收益
            stat_row2 = st.columns(4)
            items2 = [
                ("胜率", metrics['胜率'], f"超额{metrics['超额胜率']}", "positive" if metrics['win_rate_raw']>=0.55 else "neutral"),
                ("超额收益", metrics['超额收益'], "vs沪深300", "positive" if metrics.get('alpha_raw',0)>=0 else "negative"),
                ("月均收益", metrics['月均收益'], "", "neutral"),
                ("年化收益", metrics['年化收益率'], "", "positive" if metrics.get('ann_ret_raw',0)>=0 else "negative"),
            ]
            for col, (lbl, val, sub, cc) in zip(stat_row2, items2):
                with col:
                    st.markdown(f"""<div class="bt-stat-card"><div class="bt-label">{lbl}</div><div class="bt-val {cc}">{val}</div><div class="bt-sub">{sub}</div></div>""", unsafe_allow_html=True)
            
            # 第三行：交易天数 | 峰值资金投入 | 平均资金投入 | 累计买入总额 | 选股数量 | 期末组合净值
            max_deployed_val = bt_df.attrs.get("max_deployed", bt_df['持仓数'].max() * 10000 * (1 + backtest_engine.TRADING_COST))
            avg_deployed_val2 = bt_df.attrs.get("avg_deployed", max_deployed_val)
            total_invested_val = bt_df.attrs.get("total_invested", 0)
            stat_row3 = st.columns(6)
            items3 = [
                ("交易天数", f"{len(bt_df)} 天", "", "neutral"),
                ("峰值资金投入", f"¥{max_deployed_val:,.2f}", "", "neutral"),
                ("平均资金投入", f"¥{avg_deployed_val2:,.2f}", "", "neutral"),
                ("累计买入总额", f"¥{total_invested_val:,.2f}", "", "neutral"),
                ("选股数量", f"{int(last_row['持仓数'])} 只", "", "neutral"),
                ("期末组合净值", f"¥{last_row['组合净值']:,.2f}", "", "neutral"),
            ]
            for col, (lbl, val, sub, cc) in zip(stat_row3, items3):
                with col:
                    st.markdown(f"""<div class="bt-stat-card"><div class="bt-label">{lbl}</div><div class="bt-val {cc}">{val}</div><div class="bt-sub">{sub}</div></div>""", unsafe_allow_html=True)
            
            # 第四行：基准收益 | 年化波动 | 夏普比率 | 卡玛比率
            stat_row4 = st.columns(4)
            items4 = [
                ("基准收益", metrics['基准累计收益'], "", "neutral"),
                ("年化波动", metrics['年化波动率'], "", "neutral"),
                ("夏普比率", str(metrics['夏普比率']), ">1优秀", "positive" if metrics['sharpe_raw']>=1 else "neutral"),
                ("卡玛比率", str(metrics['卡玛比率']), "", "neutral"),
            ]
            for col, (lbl, val, sub, cc) in zip(stat_row4, items4):
                with col:
                    st.markdown(f"""<div class="bt-stat-card"><div class="bt-label">{lbl}</div><div class="bt-val {cc}">{val}</div><div class="bt-sub">{sub}</div></div>""", unsafe_allow_html=True)

            st.markdown("")
            # 股票收益清单
            stock_profit_df = bt_df.attrs.get("stock_profit_list")
            if stock_profit_df is not None and len(stock_profit_df) > 0:
                st.subheader("📋 股票收益清单")
                # 格式化显示列
                show_df = stock_profit_df.copy()
                show_df["投入金额"] = show_df["投入金额"].apply(lambda x: f"¥{x:,.0f}" if isinstance(x, (int, float)) and x != 0 else x)
                show_df["单笔收益"] = show_df["单笔收益"].apply(lambda x: f"¥{x:+,.2f}" if isinstance(x, (int, float)) else x)
                show_df["单笔收益率"] = show_df["单笔收益率"].apply(lambda x: f"{x*100:+.2f}%" if isinstance(x, (int, float)) else x)
                st.dataframe(show_df, width='stretch', 
                    height=min(35 * len(show_df) + 38, 600), hide_index=True)
                st.markdown("")

                # 下载 Excel（三 Sheet）
                from io import BytesIO
                max_dep = bt_df.attrs.get("max_deployed", bt_df['持仓数'].max() * 10000 * (1 + backtest_engine.TRADING_COST))
                avg_dep = bt_df.attrs.get("avg_deployed", max_dep)
                total_inv = bt_df.attrs.get("total_invested", 0)
                total_profit_excel = stock_profit_df[stock_profit_df['代码'].notna() & (stock_profit_df['代码'] != '合计')]['单笔收益'].sum() if stock_profit_df is not None and len(stock_profit_df) > 0 else 0.0
                roi_excel = total_profit_excel / max_dep if max_dep > 0 else 0
                summary_for_excel = pd.DataFrame({
                    "--- 汇总统计 ---": ["回测起始日期", "回测结束日期", "盈利金额", "峰值资金投入", "平均资金投入", "累计买入总额", "ROI (总盈利÷峰值资金)", "胜率", "持仓股票数量", "交易天数"],
                    "": [
                        str(bt_start.date()),
                        str(bt_end.date()),
                        total_profit_excel,
                        max_dep,
                        avg_dep,
                        total_inv,
                        roi_excel,
                        metrics.get('胜率', 'N/A'),
                        int(last_row['持仓数']),
                        len(bt_df),
                    ],
                })
                detail_cols = ["日期","累计投入","组合净值","基准净值","日收益率",
                             "基准日收益率","累计收益率","基准累计收益","超额收益",
                             "最大回撤","持仓数"]
                daily_df = bt_df[detail_cols].copy()
                daily_df["日收益率"] = daily_df["日收益率"].apply(lambda x: f"{x*100:+.2f}%")
                daily_df["基准日收益率"] = daily_df["基准日收益率"].apply(lambda x: f"{x*100:+.2f}%")
                daily_df["累计收益率"] = daily_df["累计收益率"].apply(lambda x: f"{x*100:+.2f}%")
                daily_df["基准累计收益"] = daily_df["基准累计收益"].apply(lambda x: f"{x*100:+.2f}%")
                daily_df["超额收益"] = daily_df["超额收益"].apply(lambda x: f"{x*100:+.2f}%")
                daily_df["最大回撤"] = daily_df["最大回撤"].apply(lambda x: f"{x*100:.2f}%")

                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    summary_for_excel.to_excel(writer, sheet_name='股票收益分析汇总', index=False)
                    stock_profit_df.to_excel(writer, sheet_name='股票收益清单', index=False)
                    daily_df.to_excel(writer, sheet_name='回测日明细', index=False)

                    # --- 盈利/亏损背景色 ---
                    from openpyxl.styles import PatternFill
                    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 浅绿
                    RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红
                    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                    def _auto_width(ws, min_width=8, max_width=28):
                        for col_cells in ws.columns:
                            col_letter = col_cells[0].column_letter
                            max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
                            ws.column_dimensions[col_letter].width = max(min_width, min(max_len * 1.7 + 2, max_width))

                    # 1. 股票收益清单：格式 + 着色
                    ws_stock = writer.sheets['股票收益清单']
                    _auto_width(ws_stock)
                    cols_stock = [c.value for c in ws_stock[1]]  # header row
                    # 数字格式化
                    fmt_map = {
                        "投入金额": '#,##0.00',
                        "买入价": '0.00',
                        "卖出价/收盘价": '0.00',
                        "单笔收益": '#,##0.00',
                        "单笔收益率": '0.00%',
                    }
                    for i, h in enumerate(cols_stock):
                        if h in fmt_map:
                            for row in ws_stock.iter_rows(min_row=2, max_row=ws_stock.max_row):
                                cell = row[i]
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = fmt_map[h]
                    # 着色
                    profit_col_idx = None
                    for i, h in enumerate(cols_stock):
                        if h in ("单笔收益",):
                            profit_col_idx = i + 1  # 1-based
                            break
                    if profit_col_idx:
                        for row in ws_stock.iter_rows(min_row=2, max_row=ws_stock.max_row):
                            val = row[profit_col_idx - 1].value
                            if isinstance(val, (int, float)):
                                fill = GREEN_FILL if val > 0 else RED_FILL if val < 0 else WHITE_FILL
                                for cell in row:
                                    cell.fill = fill

                    # 2. 股票收益分析汇总：数字格式化 + 着色
                    ws_sum = writer.sheets['股票收益分析汇总']
                    _auto_width(ws_sum)
                    # 金额列格式化
                    money_labels = {"初始资金", "盈利金额", "峰值资金投入", "平均资金投入", "累计买入总额", "期末组合净值"}
                    for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row):
                        label = str(row[0].value or "")
                        data_cell = row[1]
                        if label in money_labels and isinstance(data_cell.value, (int, float)):
                            data_cell.number_format = '#,##0.00'
                        elif ("ROI" in label or "回报率" in label) and isinstance(data_cell.value, (int, float)):
                            data_cell.number_format = '0.00%'
                    # 着色
                    for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row):
                        data_cell = row[1]
                        if isinstance(data_cell.value, (int, float)):
                            label = str(row[0].value or "")
                            if "ROI" in label or "回报率" in label:
                                data_cell.fill = GREEN_FILL if data_cell.value > 0 else RED_FILL if data_cell.value < 0 else WHITE_FILL

                    # 3. 回测日明细：超额收益列着色
                    ws_daily = writer.sheets['回测日明细']
                    _auto_width(ws_daily)
                    cols_daily = [c.value for c in ws_daily[1]]
                    excess_idx = None
                    for i, h in enumerate(cols_daily):
                        if h == "超额收益":
                            excess_idx = i + 1
                            break
                    if excess_idx:
                        for row in ws_daily.iter_rows(min_row=2, max_row=ws_daily.max_row):
                            cell = row[excess_idx - 1]
                            raw = cell.value
                            if isinstance(raw, str):
                                try:
                                    raw = float(raw.replace("%", "").replace("+", ""))
                                except:
                                    raw = None
                            if isinstance(raw, (int, float)):
                                cell.fill = GREEN_FILL if raw > 0 else RED_FILL if raw < 0 else WHITE_FILL

                output.seek(0)
                _name_map = {"resonance": "共振模型", "chase_high": "追高模型", "buy_low": "低吸模型", "golden_cross": "金叉模型", "canslim": "CAN_SLIM模型", "dilemma_reversal": "困境反转模型"}
                fname = f"回测分析表_{_name_map.get(model_key, '未知模型')}.xlsx"
                st.download_button(
                    "📥 下载Excel (三Sheet)",
                    output,
                    fname,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="bt_dl_excel",
                )
            tab1, tab2, tab3 = st.tabs(["📈 净值曲线", "🗓️ 月度收益", "📋 详细记录"])
            with tab1:
                eq_fig = create_backtest_equity_chart(bt_df, metrics)
                st.plotly_chart(eq_fig, width='stretch', config={'displayModeBar': True})
            with tab2:
                hm_fig = create_monthly_heatmap(bt_df)
                st.plotly_chart(hm_fig, width='stretch', config={'displayModeBar': False})
            with tab3:
                display_cols = ["日期","累计投入","组合净值","基准净值","日收益率",
                                "基准日收益率","累计收益率","基准累计收益","超额收益",
                                "最大回撤","持仓数"]
                # 兼容新旧列名：优先用拆分列，回退到旧列
                new_trade_cols = ["今日买入","今日卖出","当前持仓"]
                for c in new_trade_cols:
                    if c in bt_df.columns:
                        display_cols.append(c)
                if not any(c in bt_df.columns for c in new_trade_cols) and "持仓明细" in bt_df.columns:
                    display_cols.append("持仓明细")
                if "持仓收盘价" in bt_df.columns:
                    display_cols.append("持仓收盘价")
                detail_df = bt_df[display_cols].copy()
                detail_df["日收益率"] = detail_df["日收益率"].apply(lambda x: f"{x*100:+.2f}%")
                detail_df["基准日收益率"] = detail_df["基准日收益率"].apply(lambda x: f"{x*100:+.2f}%")
                detail_df["累计收益率"] = detail_df["累计收益率"].apply(lambda x: f"{x*100:+.2f}%")
                detail_df["基准累计收益"] = detail_df["基准累计收益"].apply(lambda x: f"{x*100:+.2f}%")
                detail_df["超额收益"] = detail_df["超额收益"].apply(lambda x: f"{x*100:+.2f}%")
                detail_df["最大回撤"] = detail_df["最大回撤"].apply(lambda x: f"{x*100:.2f}%")
                detail_df["组合净值"] = detail_df["组合净值"].apply(lambda x: f"¥{x:,.2f}")
                detail_df["基准净值"] = detail_df["基准净值"].apply(lambda x: f"¥{x:,.2f}")
                detail_df["累计投入"] = detail_df["累计投入"].apply(lambda x: f"¥{x:,.2f}")
                st.dataframe(detail_df, width='stretch', height=450,
                    column_config={"日期": st.column_config.DateColumn("日期", format="YYYY-MM-DD", width="medium")})
                csv_out = detail_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 下载CSV", csv_out, "backtest_detail.csv", "text/csv", key="bt_dl_csv")

    st.markdown("""
    ---
    **智能选股系统 - Smart Stock Screener**
    功能：四模型选股（追高+低吸+共振+金叉）+ 个股分析 + 自选股管理 + 策略回测 + 参数可调
    数据源：pytdx通达信行情(优先) + 通达信本地数据(K线) + akshare(龙虎榜兜底) | A股全市场约6600+只股票
    """)


# ================================================================
#                       主程序入口
# ================================================================

def main():
    render_sidebar()
    if st.session_state.selected_stock:
        render_detail(st.session_state.selected_stock)
    elif st.session_state.current_page == "watchlist":
        render_watchlist()
    elif st.session_state.current_page == "backtest":
        render_backtest()
    else:
        render_screener()


if __name__ == "__main__":
    main()