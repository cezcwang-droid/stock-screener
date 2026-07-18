"""
app 层端到端测试。

使用 mock streamlit 环境和 mock 行情数据，验证渲染流程的正确性。

覆盖范围：
1. charts.py — 全部 5 个纯函数（返回 plotly Figure / HTML）
2. components.py — _export_df_to_xlsx 纯函数
3. pages.render_detail — 个股详情页渲染流程（mock st + mock data）
"""

import pytest
import pandas as pd
import numpy as np
from unittest.mock import MagicMock, patch, PropertyMock
from datetime import datetime, timedelta

# ============================================================
# Part 1: charts.py 纯函数测试
# ============================================================


class TestCreateCandlestickChart:
    """覆盖 K 线图的创建"""

    def _make_kline(self, n=60, start_price=10.0, trend="up"):
        """生成含日期列的 K 线 DataFrame"""
        from tests.conftest import make_sample_kline
        df = make_sample_kline(n=n, start_price=start_price, trend=trend)
        df = df.rename_axis("日期").reset_index()
        return df

    def test_returns_figure_with_valid_data(self):
        """正常 K 线数据返回 plotly Figure"""
        from app.charts import create_candlestick_chart
        df = self._make_kline()
        fig = create_candlestick_chart(df, title="测试K线")
        assert fig is not None
        # 应有 2 个子图（K线 + 成交量）或 1 个
        assert len(fig.data) >= 1

    def test_empty_dataframe_returns_none(self):
        """空 DataFrame 返回 None"""
        from app.charts import create_candlestick_chart
        df = pd.DataFrame(columns=["日期", "开盘", "收盘", "最高", "最低", "成交量"])
        fig = create_candlestick_chart(df)
        assert fig is None

    def test_missing_date_column_returns_none(self):
        """缺少日期列返回 None"""
        from app.charts import create_candlestick_chart
        df = pd.DataFrame({"开盘": [10], "收盘": [11], "最高": [12], "最低": [9], "成交量": [1000]})
        fig = create_candlestick_chart(df)
        assert fig is None

    def test_realtime_flag_adds_annotation(self):
        """is_realtime=True 时添加未收盘标注"""
        from app.charts import create_candlestick_chart
        df = self._make_kline(n=30)
        fig = create_candlestick_chart(df, title="实时K线", is_realtime=True)
        assert fig is not None
        # 有 annotation
        annotations = fig.layout.annotations
        assert len(annotations) >= 1
        assert "未收盘" in annotations[0].text

    def test_no_volume_no_error(self):
        """没有成交量列时仍能生成 K 线图"""
        from app.charts import create_candlestick_chart
        df = self._make_kline(n=30)
        df = df.drop(columns=["成交量"])
        fig = create_candlestick_chart(df, title="无成交量K线")
        assert fig is not None
        # 只有K线 trace，没有成交量 bar
        trace_types = [t.type for t in fig.data]
        assert "candlestick" in trace_types

    def test_volume_bar_colors_correct(self):
        """成交量颜色与涨跌一致（红涨绿跌）"""
        from app.charts import create_candlestick_chart
        df = self._make_kline(n=20, trend="up")
        fig = create_candlestick_chart(df)
        # 找到 volume trace（type='bar'）
        vol_trace = [t for t in fig.data if t.type == "bar"]
        if vol_trace:
            marker_colors = vol_trace[0].marker.color
            assert len(marker_colors) == len(df)
            # 上涨日的成交量应为红色
            for i in range(len(df)):
                close = df["收盘"].iloc[i]
                open_ = df["开盘"].iloc[i]
                expected_color = "#E74C3C" if close >= open_ else "#27AE60"
                assert marker_colors[i] == expected_color, \
                    f"第{i}日: 收盘{close}开盘{open_} 期待{expected_color} 实际{marker_colors[i]}"

    def test_realtime_volume_last_bar_has_different_color(self):
        """实时模式下最后一根成交量柱颜色不同"""
        from app.charts import create_candlestick_chart
        df = self._make_kline(n=20, trend="up")
        fig = create_candlestick_chart(df, is_realtime=True)
        vol_trace = [t for t in fig.data if t.type == "bar"]
        if vol_trace:
            colors = vol_trace[0].marker.color
            assert "rgba" in colors[-1], "最后一根应使用半透明色"


class TestCreateRadarChart:
    """覆盖雷达图创建"""

    def test_returns_figure_with_valid_data(self):
        """正常维度数据返回 plotly Figure"""
        from app.charts import create_radar_chart
        dim_data = {
            "趋势结构": 80, "动量强度": 70, "板块共振": 60,
            "北向资金": 50, "机构净买": 40,
        }
        weights = {"趋势结构": 15, "动量强度": 18}
        fig = create_radar_chart(dim_data, weights)
        assert fig is not None
        assert len(fig.data) == 1
        # 雷达图应该有 Scatterpolar trace
        assert fig.data[0].type == "scatterpolar"

    def test_all_dimensions_have_values(self):
        """所有维度的值都在 0-100 范围内"""
        from app.charts import create_radar_chart
        dim_data = {f"dim{i}": v for i, v in enumerate([0, 25, 50, 75, 100])}
        fig = create_radar_chart(dim_data, {})
        assert fig is not None
        r_values = fig.data[0].r
        assert all(0 <= v <= 100 for v in r_values)

    def test_single_dim_works(self):
        """单个维度也能生成雷达图"""
        from app.charts import create_radar_chart
        fig = create_radar_chart({"趋势结构": 80}, {})
        assert fig is not None
        assert len(fig.data) == 1
        assert len(fig.data[0].r) == 2  # 1个值+闭合点


class TestCreateBacktestEquityChart:
    """覆盖回测净值曲线"""

    def _make_bt_df(self, n=100):
        """生成模拟回测数据"""
        dates = pd.bdate_range(end=pd.Timestamp.today(), periods=n)
        rng = np.random.default_rng(42)
        daily_ret = rng.normal(0.001, 0.02, n)
        nav = pd.Series(1.0 * (1 + daily_ret).cumprod())
        benchmark = pd.Series(1.0 * (1 + rng.normal(0.0005, 0.015, n)).cumprod())
        dd = (nav.cummax() - nav) / nav.cummax()
        return pd.DataFrame({
            "日期": dates, "组合净值": nav, "基准净值": benchmark,
            "日收益率": daily_ret, "最大回撤": dd,
        })

    def test_returns_figure_with_valid_data(self):
        """正常回测数据返回 plotly Figure"""
        from app.charts import create_backtest_equity_chart
        bt_df = self._make_bt_df()
        metrics = {"cumret_raw": 0.15}
        fig = create_backtest_equity_chart(bt_df, metrics)
        assert fig is not None
        # 应有 3 个子图（净值 + 收益率 + 回撤）
        assert len(fig.data) >= 3

    def test_no_metrics_still_works(self):
        """无 metrics 参数时不崩溃"""
        from app.charts import create_backtest_equity_chart
        bt_df = self._make_bt_df()
        fig = create_backtest_equity_chart(bt_df)
        assert fig is not None

    def test_title_shows_positive_return_in_red(self):
        """正收益标题应为红色"""
        from app.charts import create_backtest_equity_chart
        bt_df = self._make_bt_df()
        fig = create_backtest_equity_chart(bt_df, {"cumret_raw": 0.25})
        assert "#E74C3C" in (fig.layout.title.text or "")

    def test_title_shows_negative_return_in_green(self):
        """负收益标题应为绿色"""
        from app.charts import create_backtest_equity_chart
        bt_df = self._make_bt_df()
        fig = create_backtest_equity_chart(bt_df, {"cumret_raw": -0.10})
        assert "#27AE60" in (fig.layout.title.text or "")


class TestCreateMonthlyHeatmap:
    """覆盖月度收益热力图"""

    def _make_bt_df(self, n=365):
        """生成一年的回测数据"""
        dates = pd.bdate_range(end=pd.Timestamp.today(), periods=n)
        np.random.seed(42)
        nav = 1.0 * (1 + np.random.normal(0.001, 0.02, n)).cumprod()
        return pd.DataFrame({"日期": dates, "组合净值": nav})

    def test_returns_figure(self):
        """正常数据返回 plotly Figure"""
        from app.charts import create_monthly_heatmap
        bt_df = self._make_bt_df()
        fig = create_monthly_heatmap(bt_df)
        assert fig is not None

    def test_monthly_labels(self):
        """月份标签为中文（1月、2月...）"""
        from app.charts import create_monthly_heatmap
        bt_df = self._make_bt_df(n=200)
        fig = create_monthly_heatmap(bt_df)
        y_labels = list(fig.data[0].y) if fig.data else []
        assert len(y_labels) > 0
        if y_labels:
            assert all("月" in str(y) for y in y_labels)

    def test_heatmap_values_are_percentages(self):
        """热力图值是百分比"""
        from app.charts import create_monthly_heatmap
        bt_df = self._make_bt_df(n=300)
        fig = create_monthly_heatmap(bt_df)
        text_values = fig.data[0].text if fig.data else []
        if text_values:
            texts_flat = [t for row in text_values for t in row if t]
            if texts_flat:
                assert all("%" in t for t in texts_flat)


class TestCreateScoreBar:
    """覆盖评分条 HTML 生成"""

    def test_returns_html_string(self):
        """返回包含 div 的 HTML 字符串"""
        from app.charts import create_score_bar
        html = create_score_bar(75)
        assert isinstance(html, str)
        assert "<div" in html
        assert "75" in html

    def test_high_score_red(self):
        """高分（>=70）使用红色"""
        from app.charts import create_score_bar
        html = create_score_bar(85)
        assert "#E74C3C" in html

    def test_medium_score_orange(self):
        """中分（50-69）使用橙色"""
        from app.charts import create_score_bar
        html = create_score_bar(60)
        assert "#E67E22" in html

    def test_low_score_green(self):
        """低分（<50）使用绿色"""
        from app.charts import create_score_bar
        html = create_score_bar(30)
        assert "#27AE60" in html

    def test_score_zero(self):
        """0 分不崩溃"""
        from app.charts import create_score_bar
        html = create_score_bar(0)
        assert "0" in html

    def test_score_100(self):
        """100 分显示完整"""
        from app.charts import create_score_bar
        html = create_score_bar(100)
        assert "100" in html

    def test_custom_width(self):
        """自定义宽度生效"""
        from app.charts import create_score_bar
        html = create_score_bar(75, width=200)
        assert "200px" in html


# ============================================================
# Part 2: components.py 纯函数测试
# ============================================================


class TestExportDfToXlsx:
    """覆盖 DataFrame 导出为 xlsx"""

    def test_returns_bytes(self):
        """导出结果为字节"""
        from app.components import _export_df_to_xlsx
        df = pd.DataFrame({
            "代码": ["000001", "600000"],
            "名称": ["平安银行", "浦发银行"],
            "综合评分": [85, 72],
        })
        result = _export_df_to_xlsx(df)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_in_memory_xlsx_has_sheets(self):
        """生成的 xlsx 至少有一个工作表"""
        from app.components import _export_df_to_xlsx
        import openpyxl
        from io import BytesIO
        df = pd.DataFrame({"代码": ["000001"], "名称": ["平安银行"], "综合评分": [85]})
        result = _export_df_to_xlsx(df)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert len(wb.sheetnames) >= 1

    def test_code_column_as_text(self):
        """代码列应设为文本格式（防止前导零丢失）"""
        from app.components import _export_df_to_xlsx
        import openpyxl
        from io import BytesIO
        df = pd.DataFrame({"代码": ["000001"], "名称": ["平安银行"]})
        result = _export_df_to_xlsx(df)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        # 代码列应为文本格式
        cell = ws.cell(row=2, column=1)
        assert cell.number_format == '@' or str(cell.value) == "000001"

    def test_custom_code_column_name(self):
        """自定义代码列名仍能设为文本"""
        from app.components import _export_df_to_xlsx
        import openpyxl
        from io import BytesIO
        df = pd.DataFrame({"stock_code": ["600000"], "名称": ["浦发银行"]})
        result = _export_df_to_xlsx(df, code_col="stock_code")
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        assert str(cell.value) == "600000"

    def test_empty_dataframe_returns_bytes(self):
        """空 DataFrame 不崩溃"""
        from app.components import _export_df_to_xlsx
        df = pd.DataFrame({"代码": pd.Series(dtype=str)})
        result = _export_df_to_xlsx(df)
        assert isinstance(result, bytes)


# ============================================================
# Part 3: pages.render_detail 端到端测试
# ============================================================

@pytest.fixture
def mock_st():
    """Mock 整个 streamlit 模块用于 app/pages 测试"""
    import sys
    import types

    # 创建 mock streamlit 模块
    mock_st = MagicMock()
    mock_st.columns = MagicMock(side_effect=lambda *args, **kwargs: tuple(MagicMock() for _ in range(args[0] if isinstance(args[0], int) else len(args[0]))))
    mock_st.session_state = MagicMock()
    mock_st.session_state.weights = {
        "趋势结构": 15, "动量强度": 18, "板块共振": 8,
        "北向资金": 15, "机构净买": 10, "量价配合": 14,
        "情绪热度": 6, "板块资金热度": 5, "估值安全": 3, "筹码稳定": 6,
    }
    mock_st.session_state.selected_stock = None
    mock_st.session_state.current_page = "screener"
    mock_st.button = MagicMock(return_value=False)

    # 注册到 sys.modules
    if "streamlit" in sys.modules:
        orig_st = sys.modules["streamlit"]
    else:
        orig_st = None
    sys.modules["streamlit"] = mock_st

    # 清除 app.pages 的模块缓存，确保 import 时拿到 mock st
    for mod in list(sys.modules.keys()):
        if mod.startswith("app.pages") or mod.startswith("app.charts"):
            del sys.modules[mod]

    yield mock_st

    # 恢复
    if orig_st:
        sys.modules["streamlit"] = orig_st
    else:
        del sys.modules["streamlit"]


class TestRenderDetail:
    """覆盖 pages.render_detail 端到端渲染流程"""

    @pytest.fixture(autouse=True)
    def _setup_mocks(self, mock_st):
        self.mock_st = mock_st

    def _make_kline(self, n=60, start_price=10.0, trend="up"):
        """生成含日期列的 K 线 DataFrame"""
        from tests.conftest import make_sample_kline
        df = make_sample_kline(n=n, start_price=start_price, trend=trend)
        df = df.rename_axis("日期").reset_index()
        return df

    def _make_detail_data(self):
        """生成模拟的个股详情数据"""
        return {
            "success": True, "model": "chase_high", "代码": "000001",
            "名称": "平安银行", "close": 12.5, "chg": 2.5,
            "综合评分": 85, "pass": True, "max_score": 100,
            "板块": "银行", "信号": "强势买入",
            "5日涨幅": "+4.0%", "20日涨幅": "+8.5%",
            "量比_显示": "1.20", "换手率_显示": "3.5%", "RSI_显示": "55.0",
            "PE_显示": "8.0", "position_msg": "主线龙头",
            "趋势结构": 25, "动量强度": 20, "板块共振": 10,
            "北向资金": 5, "机构净买": 5, "板块资金热度": 3,
            "量价配合": 8, "估值安全": 2, "筹码稳定": 3, "情绪热度": 4,
        }

    def test_render_with_valid_data(self):
        """有效数据正常渲染"""
        from app.pages import render_detail
        detail = self._make_detail_data()
        kline_df = self._make_kline(n=60)

        with patch("app.pages.get_stock_detail", return_value=detail):
            with patch("app.pages.get_kline_with_today", return_value=(kline_df, False)):
                with patch("app.pages.create_candlestick_chart") as mock_chart:
                    with patch("app.pages.create_radar_chart") as mock_radar:
                        render_detail("000001")

        # 验证关键 API 被调用
        self.mock_st.columns.assert_called()
        self.mock_st.markdown.assert_called()
        self.mock_st.button.assert_called()  # 页面中有多个 +/- 调权按钮
        mock_chart.assert_called_once()
        mock_radar.assert_called_once()

    def test_render_missing_data_shows_error(self):
        """找不到股票时显示错误"""
        from app.pages import render_detail
        with patch("app.pages.get_stock_detail", return_value=None):
            render_detail("999999")
        self.mock_st.error.assert_called_once_with("未找到数据")

    def test_render_realtime_kline(self):
        """实时 K 线模式包含 [盘中] 标注"""
        from app.pages import render_detail
        detail = self._make_detail_data()
        kline_df = self._make_kline(n=60)

        with patch("app.pages.get_stock_detail", return_value=detail):
            with patch("app.pages.get_kline_with_today", return_value=(kline_df, True)):
                with patch("app.pages.create_candlestick_chart") as mock_chart:
                    with patch("app.pages.create_radar_chart"):
                        render_detail("000001")

        # 验证 create_candlestick_chart 被调用时 is_realtime=True
        call_kwargs = mock_chart.call_args[1]
        assert call_kwargs.get("is_realtime") is True
        self.mock_st.caption.assert_called_once()

    def test_render_with_signal_colors(self):
        """不同信号显示不同颜色"""
        from app.pages import render_detail

        for signal, expected_color in [
            ("强势买入", "#E74C3C"),
            ("逢低吸纳", "#E67E22"),
            ("观望等待", "#27AE60"),
            ("建议回避", "#888"),
        ]:
            detail = self._make_detail_data()
            detail["信号"] = signal
            kline_df = self._make_kline(n=60)
            with patch("app.pages.get_stock_detail", return_value=detail):
                with patch("app.pages.get_kline_with_today", return_value=(kline_df, False)):
                    with patch("app.pages.create_candlestick_chart"):
                        with patch("app.pages.create_radar_chart"):
                            render_detail("000001")

            # 检查 markdown 中包含对应颜色
            markdown_texts = [call[0][0] for call in self.mock_st.markdown.call_args_list
                            if call[0] and isinstance(call[0][0], str)]
            has_color = any(expected_color in t for t in markdown_texts)
            assert has_color, f"信号 {signal} 应包含颜色 {expected_color}"

    def test_advice_box_based_on_score(self):
        """综合评分决定建议框内容"""
        from app.pages import render_detail

        for score, keyword in [
            (90, "核心龙头"),
            (75, "支线趋势"),
            (60, "建议观望"),
            (30, "不建议参与"),
        ]:
            detail = self._make_detail_data()
            detail["综合评分"] = score
            kline_df = self._make_kline(n=60)
            with patch("app.pages.get_stock_detail", return_value=detail):
                with patch("app.pages.get_kline_with_today", return_value=(kline_df, False)):
                    with patch("app.pages.create_candlestick_chart"):
                        with patch("app.pages.create_radar_chart"):
                            render_detail("000001")

            markdown_texts = " ".join(
                call[0][0] for call in self.mock_st.markdown.call_args_list
                if call[0] and isinstance(call[0][0], str)
            )
            assert keyword in markdown_texts, f"评分 {score} 应包含关键词 {keyword}"
